#!/usr/bin/python3
# Name:         editor.py
# Version:      v08 Omachi
# Time-stamp:   <2025.08.13-08:50:18-JST>

# editor.py
# Copyright (C) 2022-2025  Seiichiro HATA
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.


# 2022.07.21 v01 Hiroshima
# 2022.08.24 v02 Shin-Hakushima
# 2022.12.25 v03 Yokogawa
# 2023.01.07 v04 Mitaki
# 2023.03.16 v05 Aki-Nagatsuka
# 2023.06.07 v06 Shimo-Gion
# 2024.04.02 v07 Furuichibashi
# 2025.01.04 v08 Omachi

__version__ = 'v08 Omachi'


# USAGE
# from makdo.makdo_gui import Makdo
# Makdo()


######################################################################
# SETTING


import sys
import os
import subprocess
import shutil
import argparse     # Python Software Foundation License
import re
import unicodedata
import chardet      # GNU Lesser General Public License v2 or later (LGPLv2+)
import datetime     # Zope Public License
import zipfile
import tempfile
import tkinter
import tkinter.filedialog
import tkinter.simpledialog
import tkinter.messagebox
import tkinter.font
import importlib    # Python Software Foundation License
import makdo.makdo_md2docx
import makdo.makdo_docx2md
import makdo.makdo_mddiff  # MDDIFF
import openpyxl     # MIT License
import webbrowser
import threading

# To launch MS Word on Windows
if sys.platform == 'win32':
    import win32com.client  # PSF License (pip install pywin32)
# Mac doesn't support "tkinterdnd2" (drag and drop)
if sys.platform != 'darwin':
    import tkinterdnd2  # MIT License

if sys.platform == 'win32':
    CONFIG_DIR = os.getenv('APPDATA') + '\\makdo'
    CONFIG_FILE = CONFIG_DIR + '\\init.md'
elif sys.platform == 'darwin':
    CONFIG_DIR = os.getenv('HOME') + '/Library/makdo'
    CONFIG_FILE = CONFIG_DIR + '/init.md'
elif sys.platform == 'linux':
    CONFIG_DIR = os.getenv('HOME') + '/.config/makdo'
    CONFIG_FILE = CONFIG_DIR + '/init.md'
else:
    CONFIG_DIR = None
    CONFIG_FILE = None

WINDOW_SIZE = '900x600'

# MD用のフォント
BIZUD_GOTHIC_FONT = ('BIZ UDゴシック', 'BIZ UDGothic')  # 現時点で最適
BIZUD_MINCHO_FONT = ('BIZ UD明朝', 'BIZ UDMincho')
# NOTO_GOTHIC_FONT = ('Noto Sans Mono CJK JP')  # Linuxで上下に間延びする
# NOTO_MINCHO_FONT = ('Noto Serif CJK JP')
# MS_GOTHIC_FONT = ('ＭＳ ゴシック', 'MS Gothic')  # ボールドがなく幅が合わない
MS_MINCHO_FONT = ('ＭＳ 明朝', 'MS Mincho')
# IPA_GOTHIC_FONT = ('IPAゴシック', 'IPAGothic')  # ボールドがなく幅が合わない
# IPA_MINCHO_FONT = ('IPA明朝', 'IPAMincho')
YU_MINCHO_FONT = ('游明朝', 'Yu Mincho')
HIRAGINO_MINCHO_FONT = ('ヒラギノ明朝 ProN', 'Hiragino Mincho ProN')

TAB_WIDTH = 4

# DOCX用のフォント
DOCX_MINCHO_FONT = 'ＭＳ 明朝'
DOCX_ALPHANUMERIC_FONT = 'Times New Roman'

OPENAI_MODELS = [
    'o1-preview',     # $ 15.000 / 60.000
    'o1-mini',        # $  3.000 / 12.000
    'gpt-4',          # $ 30.000 / 60.000
    'gpt-4-turbo',    # $ 10.000 / 30.000
    'gpt-4o',         # $  2.500 / 10.000
    'gpt-4o-mini',    # $  0.150 /  0.075
    # 'gpt-3.5-turbo',  # $  0.500 /  1.500
]
DEFAULT_OPENAI_MODEL = 'gpt-4o-mini'
# DEFAULT_OPENAI_MODEL = 'gpt-3.5-turbo'

DEFAULT_OLLAMA_MODEL = 'gemma3:27b'

MD_TEXT_WIDTH = 68

NOT_ESCAPED = '^((?:(?:.|\n)*?[^\\\\])??(?:\\\\\\\\)*?)??'

BLACK_SPACE = ('#9191FF', '#000000', '#2323FF')        # (0.6,240),BK,(0.2,240)
WHITE_SPACE = ('#C0C000', '#FFFFFF', '#F7F700')        # (0.7, 60),WH,(0.9, 60)
LIGHTYELLOW_SPACE = ('#C0C000', '#FFFFE0', '#F7F700')  # (0.7, 60),LY,(0.9, 60)

ASCII_SYMBOLS = {'space': ' ', 'exclam': '!', 'quotedbl': '"',
                 'numbersign': '#', 'dollar': '$', 'percent': '%',
                 'ampersand': '&', 'apostrophe': "'", 'parenleft': '(',
                 'parenright': ')', 'asterisk': '*', 'plus': '+', 'comma': ',',
                 'minus': '-', 'period': '.', 'slash': '/', 'colon': ':',
                 'semicolon': ';', 'less': '<', 'equal': '=', 'greater': '>',
                 'question': '?', 'at': '@', 'bracketleft': '[',
                 'backslash': '\\', 'bracketright': ']', 'asciicircum': '^',
                 'underscore': '_', 'grave': '`', 'braceleft': '{', 'bar': '|',
                 'braceright': '}', 'asciitilde': '~'}

COLOR_SPACE = (
    # Y=   0.3        0.5        0.7        0.9
    ('#FF1C1C', '#FF5D5D', '#FF9E9E', '#FFDFDF'),  # 000 : comment, key0
    ('#DE2900', '#FF603C', '#FFA08A', '#FFDFD8'),  # 010 :
    ('#A63A00', '#FF6512', '#FFA271', '#FFE0D0'),  # 020 : del
    ('#864300', '#E07000', '#FFA64D', '#FFE1C4'),  # 030 : sect1, hnumb
    ('#714900', '#BC7A00', '#FFAC10', '#FFE3AF'),  # 040 : sect2
    ('#604E00', '#A08300', '#E0B700', '#FFE882'),  # 050 : sect3
    ('#525200', '#898900', '#C0C000', '#F7F700'),  # 060 : sect4, keyX
    ('#465600', '#758F00', '#A4C900', '#D5FF1A'),  # 070 : sect5
    ('#3A5A00', '#619500', '#88D100', '#C2FF50'),  # 080 : sect6
    ('#2F5D00', '#4E9B00', '#6DD900', '#B8FF70'),  # 090 : sect7
    ('#226100', '#38A200', '#4FE200', '#B0FF86'),  # 100 : sect8
    ('#136500', '#1FA900', '#2CED00', '#AAFF97'),  # 110 : ruby
    ('#006B00', '#00B200', '#00FA00', '#A5FFA5'),  # 120 : fontdeco, par1
    ('#006913', '#00AF20', '#00F52D', '#A1FFB2'),  # 130 :
    ('#006724', '#00AC3C', '#00F154', '#9DFFBF'),  # 140 :
    ('#006633', '#00AA55', '#00EE77', '#98FFCC'),  # 150 : length reviser
    ('#006441', '#00A76D', '#00EA99', '#94FFDA'),  # 160 : tab, par2
    ('#006351', '#00A586', '#00E7BC', '#8EFFEA'),  # 170 : fold
    ('#006161', '#00A2A2', '#00E3E3', '#87FFFF'),  # 180 : algin, keyY, par3
    ('#005F75', '#009FC3', '#21D6FF', '#B5F1FF'),  # 190 : table
    ('#005D8E', '#009AED', '#59C5FF', '#C8ECFF'),  # 200 : fsp, ins, par4
    ('#0059B2', '#1F8FFF', '#79BCFF', '#D2E9FF'),  # 210 : chap1
    ('#0053EF', '#4385FF', '#8EB6FF', '#D9E7FF'),  # 220 : chap2, par5
    ('#1F48FF', '#5F7CFF', '#9FB1FF', '#DFE5FF'),  # 230 : chap3, subsA
    ('#3F3FFF', '#7676FF', '#ADADFF', '#E4E4FF'),  # 240 : chap4, hsp, par6
    ('#5B36FF', '#8A70FF', '#B9A9FF', '#E8E2FF'),  # 250 : chap5
    ('#772EFF', '#9E6AFF', '#C5A5FF', '#ECE1FF'),  # 260 : par7
    ('#9226FF', '#B164FF', '#D0A2FF', '#EFE0FF'),  # 270 : subsC
    ('#B01DFF', '#C75DFF', '#DD9EFF', '#F4DFFF'),  # 280 : par8
    ('#D312FF', '#E056FF', '#EC9AFF', '#F9DDFF'),  # 290 : par9, subsB
    ('#FF05FF', '#FF4DFF', '#FF94FF', '#FFDBFF'),  # 300 : keyZ
    ('#FF0AD2', '#FF50DF', '#FF96EC', '#FFDCF9'),  # 310 : br, pgbr, hline
    ('#FF0EAB', '#FF53C3', '#FF98DB', '#FFDDF3'),  # 320 :
    ('#FF1188', '#FF55AA', '#FF99CC', '#FFDDEE'),  # 330 : list, fnumb
    ('#FF1566', '#FF5892', '#FF9BBE', '#FFDEE9'),  # 340 : escape
    ('#FF1843', '#FF5A79', '#FF9CAE', '#FFDEE4'),  # 350 :
    ('#4C4C4C', '#808080', '#B2B2B2', '#E6E6E6'),  # gray
)

KEYWORDS = [
    ['(加害者' +
     '|反訴' +
     '|弁護士会' +
     '|被告|本訴被告|反訴原告|別訴原告|被控訴人|被上告人' +
     '|相手方|被申立人' +
     '|被疑者|被告人|弁護人|対象弁護士|弁護士' +
     '|連帯債務者|債務者|買主|借主|賃借人|労働者|受任者|受寄者' +
     '|卑属' +
     '|乙|戊|辛)',
     'magenta'],
    ['(被害者' +
     '|本訴' +
     '|検察庁' +
     '|原告|本訴原告|反訴被告|別訴被告|控訴人|上告人' +
     '|申立人' +
     '|検察官|検察事務官|懲戒請求者' +
     '|債権者|根抵当権者|抵当権者|売主|貸主|賃貸人|使用者|委任者|寄託者' +
     '|尊属' +
     '|甲|丁|庚|癸)',
     'cyan'],
    ['(目撃者' +
     '|別訴' +
     '|裁判所' +
     '|裁判官|審判官|調停官|調停委員|司法委員|専門委員|書記官|事務官|訴外' +
     '|仲裁人' +
     '|補助参加人|利害関係人' +
     '|第三者|第三債務者|破産管財人|物上保証人|連帯保証人|保証人' +
     '|丙|己|壬)',
     'yellow']]

CONFIGURATION_SAMPLE = [
    '',
    '書題名: ',
    '文書式: 普通', '文書式: 契約', '文書式: 条文',
    '用紙サ: A3横', '用紙サ: A3縦', '用紙サ: A4横', '用紙サ: A4縦',
    '上余白: 3.5 cm',
    '下余白: 2.2 cm',
    '左余白: 3.0 cm',
    '右余白: 2.0 cm',
    '頭書き: ',
    '頁番号: 無', '頁番号: 有',
    '行番号: 無', '行番号: 有',
    '明朝体: Times New Roman / ＭＳ 明朝',
    'ゴシ体: = / ＭＳ ゴシック',
    '異字体: IPAmj明朝',
    '文字サ: 12 pt',
    '行間隔: 2.14 倍',
    '前余白: 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍',
    '後余白: 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍',
    '字間整: 無',
    '完成稿: 偽',
    '作成時: - USER',
    '更新時: - USER',
    '']

PARAGRAPH_SAMPLE = ['', '\t',
                    '<!-------q1--------q2--------q3------' +
                    '--q4--------q5--------q6--------q7-->',
                    '<!--コメント-->',
                    '# <!--タイトル-->', '## <!--第１-->', '### <!--１-->',
                    '#### <!--(1)-->', '##### <!--ア-->', '###### <!--(ｱ)-->',
                    '####### <!--ａ-->', '######## <!--(a)-->',
                    '1. <!--番号付き箇条書-->',
                    '  1. <!--番号付き箇条書-->',
                    '    1. <!--番号付き箇条書-->',
                    '      1. <!--番号付き箇条書-->',
                    '- <!--番号なし箇条書-->',
                    '  - <!--番号なし箇条書-->',
                    '    - <!--番号なし箇条書-->',
                    '      - <!--番号なし箇条書-->',
                    '        - <!--番号なし箇条書-->',
                    ': <!--左寄せ-->', ': <!--中寄せ--> :', '<!--右寄せ--> :',
                    '|<!--表のセル-->|<!--表のセル-->|',
                    '![<!--画像の名前-->](<!--画像のファイル名-->)',
                    '$ <!--第１編-->', '$$ <!--第１章-->', '$$$ <!--第１節-->',
                    '$$$$ <!--第１款-->', '$$$$$ <!--第１目-->',
                    '\\[<!--数式-->\\]', '<pgbr><!--改ページ-->',
                    '']

SCRIPT_SAMPLE = ['',
                 'sum = 0',
                 'sum += ?',
                 'x = ?; sum += x; print(x)',
                 'x = ?; sum += x; print(x, "3")',
                 'x = ?; sum += x; print(x, "4")',
                 'x = ?; sum += x; print(x, "4s")',
                 'print(sum)',
                 'print(sum, "3")',
                 'print(sum, "4")',
                 'print(sum, "4s")',
                 '']

FONT_DECORATOR_SAMPLE = ['', '\t',
                         '<!--コメント-->',
                         '*<!--斜体-->*',
                         '*<!--太字-->*',
                         '__<!--下線-->__',
                         '---<!--微字-->---',
                         '--<!--小字-->--',
                         '++<!--大字-->++',
                         '+++<!--巨字-->+++',
                         '^R^<!--字赤-->^R^',
                         '^Y^<!--字黄-->^Y^',
                         '^G^<!--字緑-->^G^',
                         '^C^<!--字シ-->^C^',
                         '^B^<!--字青-->^B^',
                         '^M^<!--字マ-->^M^',
                         '_R_<!--地赤-->_R_',
                         '_Y_<!--地黄-->_Y_',
                         '_G_<!--地緑-->_G_',
                         '_C_<!--地シ-->_C_',
                         '_B_<!--地青-->_B_',
                         '_M_<!--地マ-->_M_',
                         '@游明朝@<!--游明朝-->@游明朝@',
                         '']

ICON8_IMG = '''
iVBORw0KGgoAAAANSUhEUgAAAQAAAAEACAMAAABrrFhUAAAABGdBTUEAALGPC/xhBQAAACBjSFJNAAB
6JgAAgIQAAPoAAACA6AAAdTAAAOpgAAA6mAAAF3CculE8AAABO1BMVEUAAAAQEAAREREREREREREQEB
AREREREREREREREREREREQEBARERERIjMRIkQQM1URM2YRRHcRRJkQVaoRVLsRVbsQVbsRVbsRVakQV
HcQEBARVaoQVbsRVbsQVboRVakRVaoQEBARERERVaoQVakRVaoQVFQRVbsRVboRVKoRVKoRVbsQVGYQ
VIcRRIgRVaoRVboQVJgRRJkRIiIQVEMQRHcRERERESIRIiIRIjMRIkQRM1URRHcRRJkRRIgRVaoRVbs
iZrsRM2YRM0QRVZkiIiJVVVV3d3fu7u67u7tmZmb///+qqqpERETd3d2IiIgzMzPMzMyZmZmqu90iVb
tEd8yIqt27zO67zN2qzN13mcxViMzMzO5Ed7uZu93M3e4RM3czZrvd3e4zd7vd7u5miMxmmcxVd8yZq
t2WYN7TAAAANnRSTlMAESIzRFV3iKq7zN3u7u7dzLuqmXdmVTMiEZnu3buZiGYRZrtVMxHuqndEzBER
7qqIEe7uEd3VVqS6AAAAAWJLR0RLaQuFUAAAAAd0SU1FB+kCBgAcAqpN0CcAAA/oSURBVHja5V1pW9t
IEpbki8OzczCza4hNLIJtDGQnexkfYFqAQcZcAXMHCASY+f+/YG3M4aNb6q6uluRMfeYB6lXVW0dXV2
uah6LrRigcicZGRsfG4/HCsywVi6Uf/vbjTz//MvHrb3//R2KyJdp3Jy3VI9Gp0bE3vftlqVgqV5ZX3
iVT0+/TYcPUNPN70V0PzcRGx5ma90ixVKmuEkKSHyZmM0bWHP4PH46OcOr+agrPGFgkN5efN7JD/eVH
xgsQWSqtrbQxaEluIh0yjWHUPhwbixfgUlyvPkNgkYX3EXO4DKGtfUFWlsrL5AUCsjA9nxgWWkTRvgP
BGnkViyRTmcVhYL2Z0XgBTZbWq28YkNyHtGGawf74U+MFXCmurXZBQN5NZ4LLBnpktKBASt1G0JJUOq
i2P1ZQI8VKjxEQK5k3zMCpHx0vqJN+IyC5dGLxr6N+2wiW+xAgC6ngZEfK1e+LiM+STCeC8fmV+X6vV
AYQaDtCAJjfG/VbUt4YhGAh4i8V6KGRgndSWhlEgKTmTT+dP14o+IyARfIfv3/rd7QBkpz1JTk0pgre
S2mDhgBJmT58/vGCH1KmAkBys4t/gc//JOt0BMi08X17/5v8k4FAct4zJtA9Jv++nHCZgQDJe8QEnsZ
+Wl1QZSGQSgSa/WpooWCVhYAHXChh/ptbWBBUCFPSi8Fl/y27vllT7AStaKDUDUIyPa/tRsPe2VXsBK
36SCECYano1wKg0djbP8BAYI0NAElmlNGfXPR7AqBhf96uqXUCQtRQoT4jGf07ALQgwKCC9VUnBFRkB
HpU9n9+AaAFwaE0FSw5mgDB75zrsQIeABhUUHY0AZJCzot1hOKnGwB5KmBnxB2ZyAZN/14A5KnAxQSs
VND07wegTQUH6liAkN+NYOk/CEDD3ts6UBUI2giYgdKfAoAcFRRXXAAgE2ZA+J8NgFR2vOYGAMkHIv4
7A9AKiVAqKK26IiAfDfWZgmoA2lRQU0KDGFlxJK4eAHBIrBDlCIQRu38OAACpgMMHCJGqDUOY3V9HAG
DZMYcPkJxEf8BAnflxAaAVEoWpgMcHCHySQsc9/XADoGHbR4JUwOUD5PePfgdATgDEqcA9F+qkA1m/A
wA3AKJUsMwFAJkFESD26ScfAGLZcYUPAAgR6ujnP5wACGUF5VU+BMTPz7EJQAQAgZ5ZiRMAcRqIxP0E
gJsKOFlQnAYMBeffQgDwUkGVF4BkwscMAAJAOyRuYpTEr2dmIk4QwVf/YLveEBWOnlmFGwAy76MDHGz
v7NkNcXEvlAUA4I+FyA5Q293/DNGeiwrKAgBwx0JUB6ht7thg9d2z4xIREM6hUkwHqG3XpbR37ZkJAc
AZCWLBUt+5UBYCwPrEcys7FEczfhz1HbNjIQC42kNoNcDuDpr6HSo4QAAg9dErBqzt7zUaXgCwKoZAx
E1/E6cLhmj9Li4gCMCCW39sBufzI6vvQIKCALiNTqCEwF3kz+8YBkUBIM4mEMWIfXvIzr+L0RB4DYWO
54XGeNDM3zUVFraAnKHUAA5Qgx9qMcRxZCxvALt11M+PWg6/msCiOgPY/Izq/KgNEY5AIG0Am3veOb9
oS6y7JjIUGQAm/aM3RTlYQDYHQNQfvy3eu38goaIKwLN/BQcj/T1yE78KQNNfxdEY1+iQ3DDIriv/23
ydMSWHo1x9AalO6EHd3a55UiQ1x+ODLDBt4sbA2g6HW29jOr8EBzIioUwd7Jb/2/W2W/s1IsN3VChFg
dvO+j8n9PhDUktVMAADNChDgbt7XG7t05gc50GZRC+8Vuf4/AW/BiVZ8qnPAySyQCcCsOu77qfDKkdl
eQ9JwhIZkM1Zz/oxLM2dCsSUOIC9X3OdD1A6Lu+YCiB5gIMD9Fm29xcm+E/LwyoiwN6224SI4iszAj4
A9gCHFLBff88vTQl0BeAewGbAAf09vjbHEQdM+QNhNgPaW85DUsovTgr5ALgOYObAvfw/CIDyq7NCzV
HwiTjbAA5rTmNy6i9P8w3OTcpWwkwDqNO+r4fX5zklIRkEmQawR/3A3i1QECWBKHIIoBDgKwAerNAQD
oRQCmDmADs1FgBeLFERaQpMSlHAri3iAF6t0QFkw9AsYF/IATxapAQhAWAWcMBohNdZanqwSgvWFoqh
xkB7u6BYsBygfb/+iQOB7dBDQQPAEoQc+G1mbBLOgQf0OtjeVG0Aa3j6k+QinAMZHqDcANZXEQEgIfi
h8KE/DFBG1f/pfASWB9boMeCzYgNgrNUFNwbz4ENRRhq8r1j/FYIrKfCZ2JZQEhhU/ds3SmGDMYw6YK
c2XPqT5CQwCNDTQGYWjMN/GwRfQsBTUXohpNQDcOPfWzUAi4L0LEBhErC0pkT/VhyERcF9j2NAcZmok
TysFKJzoLo0uMSsf3LS5RCoHURPg/YOFH3+Ctv8m005DOY0UBpAr4QUUQDz8zePT07PTs8vLi2ZTGgM
LwgcKvn8LPZrXp7bz/IFDkFOAxXD9ERYQRbQ+9xeTxp/YjfsVzmBIpDUQGnAljeV4FK5yvJ+69zukVM
oAjAAqFHQ3sVWf5lJfs0Tu09OmkAA8NIA3CDw9ugqTS7tAbn0EoBDxQB0PbtLN4DzxgAAXywPAdhR2Q
x5eXjZQa7sQQDsY+8AqKkDoOiufUsuGhQALjwEoK4iD+o8vs5V81zTADgfNgCKbSk9SblSWauurPIWf
M0zGgBnzWEBoKV2ZW25Wl3Z2Fh9FrF/ug3AoJw2fY4CO5w+vlxdlaztm6e0v++hBdAzQfdMuLTO6eNu
3ewvtL9/7iEAtGLILREsVpaxmnrWDQ2AIw/zAFocdKSApfJXxJbm7R0NgGsPAaCUg07tIOe0Vly+3dM
AOPGwGBq8JkWbjHy1/SrBlWPbdwAKtUO7d+q/5pn6hDxQAQBlgsCGSL8NsL9/uaqgnX1BrcZB5WBSAw
+Kv66Js9nTX8WvSrr511QAjmEASNwXrO1u7dTrO1u7NVY/a0WF+sSi5kF3t6AhGU1ydVit5nCYoeYwh
xEFb0D9gDnE5Xn93q/m8z+1g6h5ECgRnMBYnUQ1/8qqKv3JSQMvCuZVLJBVx34OtaD9CPpladRHNDjO
8hDk6h4vCJBZvAWSXulPLqgUcAMKArmMiiXapRWV+jfPG3gcmEyAJ2X90p/hAbBE+OnGwNRw6c+IAfY
foF82jbNC0Ev96c0QYBpEfkVfpK6W/5iVILAb8jQqixoGiqr1t87oHvAA+3XzOFsk3/K/r4r1Z6TBwE
qos0YCMwwozH8dDQDqAZ1rY3jlUFm1/iwGgHrANN4q4ScCQAsATaFOADQNfLk+jcWCaJd5rItLoSwYW
gm+3BrDYkGkYd7m5SnrpPvqjnFVEVYIvWwUQ9onj+QAV0etr0wvbZtHjLu6R8D5oA8m2jrplqBEwNuT
p0z/UcgBoONBrxsUwkGJANbjTUdHKgDHDAdoQEfkrAjeQm2MFLB5efbyiWkA3J4yFlXePwL/YG4S8VE
J+RTo6vrNwikqNa9ZmzrBM5JzJt6jAtIGcHvSbeCDADRPWARgQw2ga8O0fCZQQXJ+JgAXzH01N+BB6d
DbwypjvhpAl/OzAHi4Zy4sgoYAksziPS20LqP/t+uBr/vIrT9wNIr0vbQgGQhllvrcXtwNWvcjt/73x
+C/nOlarSvpA/AcoPnwhebcPQA0L9j62ycEwwNkfQCcBB4f0bntkYf/ZXIg0r9iXsoHisD5p9s/WZ+2
C4Dba6eVnZdw3+tdqSnlA7AYaNGcfwCA41Mn/eEO0L1MT9oHlkGhz0mzx1f3v3Na2XsmdV0M7Y05yFI
DlvP3AnDl/FN3xwTLA6Rao+JJwHPR6wKA5fz5G/AcuG0Ai3jL1ZeFnf/GbdF+SzXr4dRlYzv0mlQnCz
LR1usLdoKcnf8FgMHseLANJEEA1OemoIekYoutXNz6+eMeuarfOL0lMh5g4j2xURFy/juuZ0bcX+O4u
ZIqP0O0tzZHVVNAf9ErIXeXUvqnFvGe2eGuhDnc2iv9GU/Rw7LBIicFXF3jPUN3/yCn/8JHxKe2+DiQ
0/k5v7+k/swH90CRkIcDrYcbxFcIZe3f4clFiAlwbLf74xzzEcYbWf0d3l2FmIBrEPj2J+obnKdXsvq
TBOqTm1V40Qt5iOzoVlr/POqjq85RkNHxgut/Yknr7/joqrgJOFYCx0e4DxDfPTal9Xd5eVl4cNahHc
bueAE//9mxvPouBiBeFTMB4Ch6vTd/VwMQrwgYiSBX0Suk/uklhvokmXB7fj6CkQh+Q3b+xj3O5yfWr
Jv+omflngBgXx8THJnLugIg2B71wAVa1m8h6U/tAwxILFAkaJ8+oqk/0AtHCIWKwyCq+p1V4hwSCUgi
ZNtHl4jqs/ogFB6cCkIqbN+cHDcx1SfTWU4AhJxATTFk3x093BJcyWU0bon4XA6foWvPlQKAnEBJQ+S
R4Au/A4g5gZKWmAIAeCOA8DGJkqaoAgBmxfTX9KifbXF8AN5nNVEERjDiIPBgBB2AlLD+rZpgHCkMAI
7GsAFYCGkAicTRWFCUCpAByM2bEAA4aUDB8Tg2AGlDAwlfNqBkQAI1A5rOakAxRnFJQGREBpEADQ0sI
Z58SNWQlA8lACwfUjcmhyJS+vOFAnWDkggyL6e/ps/gBkK+QhkPgFlTk0UgKtMWc6YCnmFpyQxYWv8W
Au5N0q/Q/49nXF6C/nIpBP150gG1FyZUN4EREFB9ZQZoADjfnwsBqYvjHJemfNbfHQHl1+Yg9o+ovzs
Cyi9O+ub/L2LG/L06K4roexMZALd8QPnlaTH5hK5/OyeM+3p9XkDepQ1NgeiOdYH6BQoC+W9WUyPhMV
9XaPBOwcyr0l/TQqN+LlHhpP+MplCMKT/X6HANwXzUlIoZjfu4SMld8llNseiRcf9Wabm3v5Tr3yaCE
f+WqbmcfyQ0T4TtBhU/EUimsppHokfG/FqoyCYOldGPPxooX6nJnn8wNE/FZHBhyRcErIXZrOa1MIxA
9Vpd+ucPaT6ISWcC7xHw1vt7jCAWDwACecPU/BKdmhN4iYA17dvnf/GDUT+ZcMFn9TuNknGfELByeUM
LghjRAQhUPrHxmvi//1dWC4boRqwfgiXVWXEynwiK+h02HLCCslIqzGeCpP6zI/SlBcoeWiLJ3xKLWg
BF74sIip7aSs0aWlBFD0/1esK/0Y1gel4LtOjGzGh3eoj73N5cOpTVAi96ONbFBmgPLlqp1PyiNiQy2
Y0BCgTJVEYbGvWfzKBlB6++IPvo6kIqY2hDKHpoZuSZE+HP7ubm0hktqw2r6EY4OjIW73iC6MPLVjKV
DvtY6+J5Q2gmNjoef3p6e4PLF6yWz/8yG06Yw6/9mymEItGp0bH//FBZrm4wX1+32rM9n/KzGcP8r/b
9id6CIRyJxqZ+/ul/1WSul+aTybmJfDqTMLKeevz/ASeus14TkvRxAAAAJXRFWHRkYXRlOmNyZWF0ZQ
AyMDI1LTAyLTA2VDAwOjI4OjAyKzAwOjAwvBQWOwAAACV0RVh0ZGF0ZTptb2RpZnkAMjAyNS0wMi0wN
lQwMDoyODowMiswMDowMM1JrocAAAAASUVORK5CYII=
'''

SPLASH_IMG = '''
iVBORw0KGgoAAAANSUhEUgAAAgAAAAFACAMAAAD9FjQ5AAAABGdBTUEAALGPC/xhBQAAACBjSFJNAAB
6JgAAgIQAAPoAAACA6AAAdTAAAOpgAAA6mAAAF3CculE8AAABFFBMVEV3d3f//wB3d2aqqkTd3RG7uz
Pu7gDMzCKZmVWIiGaIiFWqqjPMzBGZmURmZmZVVVVEREQzMzMzMyIiIiIRERERESIRIiIiIjMiM0QiM
1UzRGYzVXczVYhEZplVZohVd4hmd4hmd3cRIjMRIkQRM1UiRHciVZkiVaozZqoRRIgRVaoRVbszZplE
RDMRM2YRRJkiVbsRM0QRRHcRVZkiIhHu7u67u7v///+qqqrd3d2IiIjMzMyZmZkiZqpmZohEd8x3mcy
qu927zN0iZrtViMyqzN2Iqt2Zu91Ed7u7zO7MzO7M3e4RM3czZrvd3e4zd7vd7u5miMxmmcxVd8yZqt
1mZlVVZpkiERFVVURERFUzRFUiRIjWs3z3AAAAAWJLR0Q3MLi4RwAAAAd0SU1FB+kCBgAcAqpN0CcAA
BZBSURBVHja7V0LW9tItrSQZUmTHR4xeHb33mSzxjuARcCQEYZMdnkOgSUwCfu+9///j9VbLaslS63T
kozrfPkItIWNVdWn6pxuS50OAoFAIBAIBAKBQCAQiOUMRVFWOh3V+c/9qasoavC//1VzH1GCUee4nqL
o6SfR/N8mD1nPi0gQwOh09IAASkyETsdwfzCUIDz8428ToUoCSgUB6iCA6gHunmojQYCe95AzCzWXH+
7Ed+e/zssBJYBStRIHgAB1EECJZr7m/m8GBND8MdOjgTfuTEgtlAfh6M379R5Qb4wA3VDunW8c4JXvI
gKo7teuxw7DEw0JBDBAgKYIEKb+SOIdqENb4BPAfSiEJq0BkVab7m91zZmHmVEl9pSm3ouOdRNMl/kT
AtMZPq9Lv56XfjqdFffJnJ8AHR0BTNXH3eGB1o1TgVseJAgQO8UsAvR8+IzUpA9HGXyDUZYXOQTwQot
/DQQgJICh+2fUUX1D8yDpxjiWIYCbSQxj9nF2VHVnsqr6SuJ8w1Qf/uvEB8wSQPUo6ToTVQMBCAnQVV
51FW/iu1/MqCpIS8BcAvjHGjOPJ0cjifeY9puYAGraJDAE0APz4T+qgwB0BPCmvx5YQd2X/q4/ZJaVA
BbIZJZR0yZQ7UYsY2WDTwDT+0uM4MW/AwHoCKCFBDDjno8DjWH6Zq8UAVYiqWYjMRrh22XEnH1KPgE6
CQIYIIAEAmghAUy/1mf7AIbLhm44oGd37Eyd59CY0RBfT8xNXYwAkAAZBIjKfNWH2qv9AwLo7sn3DaK
W7gOEQGldrZM+IDka4huVG+UI4NqB72ACpRCAScdRTR4QwHfrK+7cX+E0atR4VocUYgmQGE0QQC0tAR
0FZSA1Abp+t99MEcB1AobzpdsLJNz5vtdTsgmwEtpIY9YCxKPuc7gpQQue13ti9jnDAzII4C1cxT0DI
FidAJrfCXjl/Ov4J9iM1F4JezB6p8O2CjM8gMqdnolRlQGdhTJ5dCYBDLUTmUAQgIYApt8LVIOunOk3
aDqx9WemtJb2+P6BYdPHnZ/fzz6cGNUCfD16BUv+CSS1uGSIy0v/a+hDdBBgWcPTky4kYHlDZZYQEMs
Yrp5owB+BQCAQCAQCkRerq2tr6+sbr/tObPqxNRj88MNvf/f7//nfN2/e/uHdH3GSXmbLwEXeA36TH1
uD4fboTz/u7Ow6PAANXta0d6DPBD4Zg+Fob2xZLg3evsOZexETvyj2USrwOeCyYP8NUsFiz/z3JcGPO
HBw6HHAmuzsIxMsKPpO2t8Uj8GRnwbcTODIAc7nUqHvp4Ht45AC1gQcWDL0AwpEWcDjALRgIXSfCH2f
Akd7Vhw7u2/hCds++QnR973AwdhiOYA00OIgnfxRDNkk4FBgH25gCXJ/IgmMEknAcQNQgiXI/TlJwFU
CUGB54HeTwLEFCiwv/G45kPSCoECbrN/rzTriKMUAawd2sA3Of7Om2P6QYoC1Cwo0nP3f9zdri+Fhmg
GTffQFGoz1GuHPYIBjBYDDi8/+uQyADrxc759mwAeLmwRQD9Q//RuA33WCY4tLASSBJZj+WdWgZwaRB
F62+sdxYPEDTuDFmv+ZnuBxBgNQDtSV/huc/t66wF4GA6x9yECb3R9Z2hiOsxgAL9hm97/WoaLAyLIg
A83EyvpPwrB17BOi2jFbBCwL1UBr5X91OrVP1yWLgFMNvANO0vCvtPDrEGA6Pfu4IbMW9GQADGhn888
jwNT+ebUvVwRgBVva/PUJ4FCAwgocjcGARev+hARwKPCpshXYyksB1gTFQAu7fzEBKKzA9hgMqBX/6r
rNEqC6FcjsCIflIDBrGf5JAlS3AvkpAAxoHf6zBHCtwIY0FwAGtA7/NAGm9llnQ1IhAAa0Dn8OAapZg
cGhBQYsEP5cAlTqDh9YYMBC1H+5BHBKQlErMJynAagGW9D/m08A1wr0pdhAhwHoCbYG/2wCCJeEo7kE
QFe4YqwS7v7LIYCgFZivAVgbrIg/5Qd/cwkg1B0uoAEOA7BDRDhWSLd/ziGAUxKWtgIFNMCydsGApgv
AYgQQsAJFNADFoLgB/KlmApS2AvN7QWBAKwqAwgQoawWOCxEApUDTBUAJApTrDhcyATCCQkH++Z+CBC
hlBbbHxRiwDzybNYClCFBiz9iwIAFgA5o2AOUIUNgKFHSBsAGlDYCED4CWIkBRK7BXkADoBjQtAGUJ4
JaEaxRLwhCBdgjAxurJtGwU2DM2KkwALAw2WAFurJ6e2dPyMX+huDgBUAs2JQD99Y8/i6BfyAoUrQMh
Ak0JQH/t1BaGf353eFiCABCB+ltA/dWTSujP3TNWhgCoBOoWABL48xeKSxEAIlCrA+yv0cCf2x0uRwB
sD6pRANZPyeD3rcBGdQJgTaA2B9j/eDadto8A8IE1JQDC7E8pAY4PBMJzEgDN9CeGn8oEOoEUkO8AKb
YBrxNPf7IyEP3AOkrA/uoZsfhTNYJQCtZRAhKnf8pWMFJADQlgg7T4I10MQgqoIQGsn5BOf8rlYKSAG
hLA2s+k4k+6IQQpQH4CWDurT/zLbglDCpDfA6C0/9SbQtELkN8EJMSffls42oHSVwHo8r+ED4YgBUhP
AGT4y/hoGBYFpVvA9bn+3y62M0zGh0OxL0B6DbhxMj+vF2kRyfl4OCpB2Qmgf1pA1hu7QAQqQdk14Lz
+v33i0qupS8TABkq3gKv5+AcN/YYuEoVKULoCrJ8VkvVmLhMHGyjdAvZPCkz/zYYuFAkbKF8B8gyAfb
I+/9PBEi8VCw2QrwBrdsH13AYuFp23PxgaQKMAeQJgf+zPvT6A1MvFQwPkK0COAMxk9vpvGJEbfwbqF
AqQUwGcrc67QojkW8bM0YC/APfqCpDTApzFv/abRkEDalCAbAeYwr/u28ahDqhBAbIdoN3Jv0iU9BtH
ohdUxzpAZg846f/TBJB961isB5SK9+QJ4FM/7zJx0m8ejW0hpWKlT50ATnjzu77bx2NNuB4LkJkAzrg
TfJVG/P0eAEECQDOwqgXIKgE4BjAiQHXxJ2gCwgSQFIGZPYDTfhYBKvR96VYBYAJmQzAlr9tlBGBzs0
Mi/mQC4JoAYF/FAnwsJQCOYnSorj82sogIABNQwQJsZGwEP8mCmez6o0QCABNQbSEgowa0VzclB5UAY
Dmgmgf8VDIBUEX1HjBcIEUbaIO/DmyvyU4AB3T4oxVUwQNmKID0BHA0piQAXKCwB/zUjAPYpsQfLlCc
AH1+DfCz5AQw/ECKP1ygcBGQ0Qb+KBn/QwsEaEcR0CnVBGwr/tgV5ITQxWEz1gFO+4uFP8oA0SKA3wb
M7ALT+L8PFgjQFgLwF4KkKsDR2JJAANSBYgt0/C6AxCbA1oEM/HEDCdEq8GPNNcDg2JITKAPW6TygvD
bwcM8CAVpFAK4HPNuQNP1H2el/MsFyUANtAP5KkCQLkDn9J+cXl1dXlxfn12gE1NwG4BcBn6RM/yz3N
7n5xb/yoG1f3QjnAewKE5q1/EawhC7A1lHW9L++Zf4G+/YaBKiTAJ16VgK3tvey1P/6c4KD9uU1CFAj
AbhVoL1ODf9xpvmb3M7kIPt2AgI0TADaImBwtJfT+rlJaZB9I9YJWnoC0O0GISTA1vDgMK/zN/kl/fJ
XYiIAAojEqczNIFvD0d6cvu8dx4Ta5yBAXQToyyPAYD76TlzwXv8CBKiNACcy+kCD4fbo+LDQms9n3u
tfggCLRYCBG0Mvtkejg73DcdEFv8kV7/WvJiDAYhDAgX10cLy3d/jhwziIktadS4B7EKDZKuC0oMYf7
40rru1P7pEBmiVAR6wTPDwqqPFz4vqvvNf/BQSojQC8xaB5jcDB6JhqU9/1A0xgswTg1YG5FmBr+wvh
ls7HJx4BPoMAtbWCOcuBeduB8tu65ePXrzwC3IIA9REgdZso3pUho9xPvaHr3CYjABaDRC8R9ym5HJv
5kRB6+C3rm03WCQQBhC8TzOaA7Pm/vSdhOze3Eyy2HAgCvBbt3PbXTsMtWZlX/xp8kbKbn9sJFlsMAg
EqrOH21zunJyennfV+1n6uQxnwW9fcPtDTIzaFikTFbTz9fs6HOaRM/6wq8EFoPwC2hUv7ON+2nOnP3
Q7k9YGEGoH4YIikT/Nsjcay8Ldup3RtABBgTcqnOSS5v5y1QPsZHw4VilUZBJD3WT53P9hXwiIABJBB
AKn4WxdcC/AgVATgAhEVGgF1Xsplzo5gYQ+INkClRkAT+GcogOCWULQB6OtAyfhn1AD231AEtKMOlI0
/fzOIYBsIBCAngFz/l7kSKLgbBJeKpS4DBrLxv77iK8A3sY+GvgP+4ncN5PX/vkjGP6MNLLgShCqQug
yQ2P/NTQCiCoClINoyYFs2/lkOQFAB4AFpXeCArACYlNoJINoGhAekdYFkN/O5vrgp1QUWXQmEBwyCi
ABEF/Od3NzbGW29u6eMWxWeW/CAjbtAIgG4u3RmOX9pd3KZca/aywk8YPMmgKQCfLz1Ov3PpQRA8PJA
sACkJoCiArh+fvAx5hLgPEMApveCl4jD7eMJW0EELcDJzVU4xXkEeMyoAAT3AsECMPG+DS2gO+bSjxx
IJ5+nxAkAFoDQBFROAI+3bIJPEyB1ZcjqCQAWgNAEjIjEP5MAWQawQgJAF4CuEKyWABjxzyLAt69Z+A
uXANgNRLgccFQF/18/p2b3c2H8hXsAWAgg1ICtCgng8eIpnd2fC+P/9U4470AByDRAvAcw+fZXnrgnC
DC5yMbfvrWgAM1rgHAT8PyS7+2eC/n/Kg4QCkCoAQPB6z89/j1rajMEePycg7+4A4QCEGqAWA14zRP/
FAHO7/PwFxcAKAChBhwLlX55yD5H8v80zYmrCrcLgwKQacDwA534Jwlwl3/U07k4/ugC0WlA+SZAsOg
7hwDX+dO/Qg8YCkC6HnBcWvwf7OlcAlx/u88/SvQ2UVgHINaAkjuB8sU/JEC6O5xuAVYwAFgJJrSBw1
JdoDmyHt4FcC780/vHCvjDAhKmgFEp8X+aFgjbnov/w10V/GEBCW1gcQswu+hbIZ5uquCPrSB8G9iXu
hJcQNbrwh+XhiJMAYOCFuDuMxn806/fKuGPGpCyEizmAQuKf8H5XxF/1ICUNrCIB7z+9mAT4n9TEX/U
gJSV4MH8E/63Xwjhnz5UxR81IGkKmFsE/Pp3Svin93cWEkCbUsCe+KKvQNiXjxYSQJtSQH4VmLHjSxz
/22sLCaBVKSB3JeD8khT+6dPzxEICkJsCXtNtB8ve8SU4/a/Oq8OPBEDcC8gkQIFF3/rTP3oA5O3AjE
ZgoUXfUvDf31DAjyYg9YoAvxH4K7H4T7/STH+sApCngFoIYF/eWTSBZUDqUrAGCXCy/4QIf+wDIC8Fp
ZtA+/752qIKlIBFYmWjPWUgKfxwgBJ8oMxGkG1f3hDCDwcoQwTktYLth9vziUUZEAAZPlDOYpD9dPnt
0aIN9ACliICU5eArcvQhALJEQMqGkGeLPiAApURgo9EtYRIIsAsBKBX/6De5KZSeALhHrCwRkLIt/Bk
CsDBrAlI+GEJOAKwByKsFZXw07BkV4OLUghI+HEpNABgAQQb81NDHw6kJ8E9gKdEISrlABAzgAhlBOZ
eIQQegFUawyCZhWReJggFckFJA3mXisAloMUoBeReKRAGwEAyQealYLAE2Xwr0K2wLy7cCRS4WXQ1/F
AC1FINSLxePArD9DJB7wwjg33oGyL5lDPBvOwMq3Ti8wE2jgH/LGSD9tnHAv+UMkH7jSODf7mpQ+q1j
gX/LGSD95tGo/9vdE5R/+3j0f9vNgO1xdd2OrMAz8G9l5K4Of6Eo3YKF4koEwPqfRAZskO0Myt0zVul
2UFj/l2oFf5LTDUouFF9UsP/AvykjsHVs0cT1hfAVwXZg/xtkAJEIuFkA9m8hjcD22Go09pH+G+4JjZ
pkANJ/jTKQUQ9ufWkO/12k/zplICMJEDQERZu/SP+t8ILDZhgA99eaJDA8bEL9Mf0bSQIb7WAA1L9d5
UDNDMD0b7Yn0G+WAZP9d0ChdTpQnxNE9m9nPVATA+D9W1sPDL6MIf5LTYEt2V1hwN92CmwfAv7losDa
TEUwOJaUBCa7bwF/SyuCBAW2jmQkgR04/8VRAnovuPMGdf9CKcH23ph08iP3L0QaYDgwGO1RKT8m/2J
ygIICDvp/wVldWA4MjioJwc7uG/i+BfUD7/s+Cba2v4hdSWqys/8WmX+BY8Uhweu+rwTHJTmws7P/T4
D/MjKBIwf9fzlScHw4Loi9k/YB/gtLBQ4NXv/w298d730YZ2d8F3oHe5R7LzUZ/GPt3+v/+b/f/+n/f
/xxPB4zuP95d99DHtAjEAgEAoFAIBAIBAKBQCxgaIoXRtXnMRQl8bM68zMndOd1uybvL+J8y39ebf6r
+NH1DjQVRQt+Uv0fw2BGl5MA/omplQBd/4XNHIjVeQRQxQgQQD5DgPD/pSOApnGBkEyA75XglVOPMEP
zCNARIoARQq3ruvOA81VnR5eOAKb/5msmgO69sFaWearYH5ogAEv4WEQopsHCEqBThgAGIQHMsrqb+7
xGMQJ44qPOEoAdXVYCaD3HlgVoug/4qKq9yKy5c8Q/l8yoO6J8nyaAFpzQbiSxM4zR46Pjl+B7gOgAZ
8j9Xps9gPnLwneRSYBY9lkCsKNLRwDTe+taaMtM/2waamI0OklKJzHaUzjnTo3NpeHxQJ01mkoME/sS
XALEB6gJ06rOwKfMPhmPAM7fo3VTBEiMLhsBjJ6LkRnbMmew5//gjaoqY5Q9OJlR96uq8wigu19Wgrn
fTQlsVH4mXoJHAPeAV1r0akrwvGkCBH9Z2lx2lZg4zoNG5D00hgmGtoQmQIsLIc07gZ42eyndQdCXAs
M01UgpfJlkRn1ku2kC6OHk1124OOnVZY2efDI+AbQwh5jx8xppAqjxsfoMlCwBuj6t1SQBEqNLSACTS
ZNqxxdwxR3Rw1Nn+KfZiLALRn1k1TQBIm/hPqXBazWYPvMSL8EjgJ4Y8iUrTQAjmezVGQJomqZ7f4RL
uyhjRQRIjC5bH0AzZnSyq3TDCcPOnXgaM6P+WEYV4D/oJxOTXwsYyZfgEaA7O8QlAK/lw/EAJtv8iQi
QHF0uAnBOkhbPyURunznNzJiaRwDnh97sqY1dxYx8yCZA9N5M9v0nR5eTAHoIlupZLdcTO6dDnymg2G
PNOAPoXAkwg1qP02xmCJB4iWwJ8FRkLgGYd8EnQDdsfans+0+OLicB/HffC+pAzxMrjI9L9IuYUS38h
VQVEGT4QDFMTvL3D0u8BI8A/gHdwATmE4B5F3wCKIm2RPj+FW6zYrkI4FXmvUi1TSP6tttNnR9m1HeL
vD5A+Gu8VrMnuj3fu7Evwe0DxH/ZXAKw74JHAJNLgMToUlkBlgBmrILhiUiOMqeZGVXzGkEqMy1TzWN
FSb8ElwAm2wiaQwAzvxH0SlFexa3o8P0nRpfLC7LvdUUP+7MMAbxRzZw5zeyoC6WZNnlmvM9ADxo3yR
ygRwcwT8ZfDYwPCGeqkTxg5i/T+Tx3niRsNwSF4ndBY5sdXc5lQYkRp/dFmRMgAGWYJFuOEIusMsu41
wqBQCAQCMSCx38BAxFKs5qg5n4AAAAldEVYdGRhdGU6Y3JlYXRlADIwMjUtMDItMDZUMDA6Mjg6MDIr
MDA6MDC8FBY7AAAAJXRFWHRkYXRlOm1vZGlmeQAyMDI1LTAyLTA2VDAwOjI4OjAyKzAwOjAwzUmuhwA
AAABJRU5ErkJggg==
'''

HALF_FULL_TABLE = [
    [' ', '\u3000'],
    ['!', '！'], ['"', '”'], ['#', '＃'], ['$', '＄'], ['%', '％'],
    ['&', '＆'], ["'", '’'], ['(', '（'], [')', '）'], ['*', '＊'],
    ['+', '＋'], [',', '，'], ['-', '－'], ['.', '．'], ['/', '／'],
    ['0', '０'], ['1', '１'], ['2', '２'], ['3', '３'], ['4', '４'],
    ['5', '５'], ['6', '６'], ['7', '７'], ['8', '８'], ['9', '９'],
    [':', '：'], [';', '；'], ['<', '＜'], ['=', '＝'], ['>', '＞'],
    ['?', '？'], ['@', '＠'],
    ['A', 'Ａ'], ['B', 'Ｂ'], ['C', 'Ｃ'], ['D', 'Ｄ'], ['E', 'Ｅ'],
    ['F', 'Ｆ'], ['G', 'Ｇ'], ['H', 'Ｈ'], ['I', 'Ｉ'], ['J', 'Ｊ'],
    ['K', 'Ｋ'], ['L', 'Ｌ'], ['M', 'Ｍ'], ['N', 'Ｎ'], ['O', 'Ｏ'],
    ['P', 'Ｐ'], ['Q', 'Ｑ'], ['R', 'Ｒ'], ['S', 'Ｓ'], ['T', 'Ｔ'],
    ['U', 'Ｕ'], ['V', 'Ｖ'], ['W', 'Ｗ'], ['X', 'Ｘ'], ['Y', 'Ｙ'],
    ['Z', 'Ｚ'],
    ['[', '［'], ['\\', '＼'], [']', '］'], ['^', '＾'], ['_', '＿'],
    ['`', '｀'],
    ['a', 'ａ'], ['b', 'ｂ'], ['c', 'ｃ'], ['d', 'ｄ'], ['e', 'ｅ'],
    ['f', 'ｆ'], ['g', 'ｇ'], ['h', 'ｈ'], ['i', 'ｉ'], ['j', 'ｊ'],
    ['k', 'ｋ'], ['l', 'ｌ'], ['m', 'ｍ'], ['n', 'ｎ'], ['o', 'ｏ'],
    ['p', 'ｐ'], ['q', 'ｑ'], ['r', 'ｒ'], ['s', 'ｓ'], ['t', 'ｔ'],
    ['u', 'ｕ'], ['v', 'ｖ'], ['w', 'ｗ'], ['x', 'ｘ'], ['y', 'ｙ'],
    ['z', 'ｚ'],
    ['{', '｛'], ['|', '｜'], ['}', '｝'], ['~', '〜'],
    #
    ['ｳﾞ', 'ヴ'],
    ['ｶﾞ', 'ガ'], ['ｷﾞ', 'ギ'], ['ｸﾞ', 'グ'], ['ｹﾞ', 'ゲ'], ['ｺﾞ', 'ゴ'],
    ['ｻﾞ', 'ザ'], ['ｼﾞ', 'ジ'], ['ｽﾞ', 'ズ'], ['ｾﾞ', 'ゼ'], ['ｿﾞ', 'ゾ'],
    ['ﾀﾞ', 'ダ'], ['ﾁﾞ', 'ヂ'], ['ﾂﾞ', 'ヅ'], ['ﾃﾞ', 'デ'], ['ﾄﾞ', 'ド'],
    ['ﾊﾞ', 'バ'], ['ﾋﾞ', 'ビ'], ['ﾌﾞ', 'ブ'], ['ﾍﾞ', 'ベ'], ['ﾎﾞ', 'ボ'],
    ['ﾊﾟ', 'パ'], ['ﾋﾟ', 'ピ'], ['ﾌﾟ', 'プ'], ['ﾍﾟ', 'ペ'], ['ﾎﾟ', 'ポ'],
    ['ﾜﾞ', 'ヷ'], ['ｦﾞ', 'ヺ'],
    ['｡', '。'], ['｢', '「'], ['｣', '」'], ['､', '、'], ['･', '・'],
    ['ｦ', 'ヲ'],
    ['ｧ', 'ァ'], ['ｨ', 'ィ'], ['ｩ', 'ゥ'], ['ｪ', 'ェ'], ['ｫ', 'ォ'],
    ['ｬ', 'ャ'], ['ｭ', 'ュ'], ['ｮ', 'ョ'], ['ｯ', 'ッ'], ['ｰ', 'ー'],
    ['ｱ', 'ア'], ['ｲ', 'イ'], ['ｳ', 'ウ'], ['ｴ', 'エ'], ['ｵ', 'オ'],
    ['ｶ', 'カ'], ['ｷ', 'キ'], ['ｸ', 'ク'], ['ｹ', 'ケ'], ['ｺ', 'コ'],
    ['ｻ', 'サ'], ['ｼ', 'シ'], ['ｽ', 'ス'], ['ｾ', 'セ'], ['ｿ', 'ソ'],
    ['ﾀ', 'タ'], ['ﾁ', 'チ'], ['ﾂ', 'ツ'], ['ﾃ', 'テ'], ['ﾄ', 'ト'],
    ['ﾅ', 'ナ'], ['ﾆ', 'ニ'], ['ﾇ', 'ヌ'], ['ﾈ', 'ネ'], ['ﾉ', 'ノ'],
    ['ﾊ', 'ハ'], ['ﾋ', 'ヒ'], ['ﾌ', 'フ'], ['ﾍ', 'ヘ'], ['ﾎ', 'ホ'],
    ['ﾏ', 'マ'], ['ﾐ', 'ミ'], ['ﾑ', 'ム'], ['ﾒ', 'メ'], ['ﾓ', 'モ'],
    ['ﾔ', 'ヤ'], ['ﾕ', 'ユ'], ['ﾖ', 'ヨ'],
    ['ﾗ', 'ラ'], ['ﾘ', 'リ'], ['ﾙ', 'ル'], ['ﾚ', 'レ'], ['ﾛ', 'ロ'],
    ['ﾜ', 'ワ'], ['ﾝ', 'ン'],
    ['ﾞ', '゛'], ['ﾟ', '゜']]

# 平成22年内閣告示第2号
JOYOKANJI = (
    ('0001', '亜亞', 'ア', ''),
    ('0002', '哀', 'アイ、あわ-れ、あわ-れむ', ''),
    ('0003', '挨', 'アイ', ''),
    ('0004', '愛', 'アイ', '愛媛'),
    ('0005', '曖', 'アイ', ''),
    ('0006', '悪惡', 'アク、オ、わる-い', ''),
    ('0007', '握', 'アク、にぎ-る', ''),
    ('0008', '圧壓', 'アツ', ''),
    ('0009', '扱', 'あつか-う', ''),
    ('0010', '宛', 'あ-てる', ''),
    ('0011', '嵐', 'あらし', ''),
    ('0012', '安', 'アン、やす-い', ''),
    ('0013', '案', 'アン', ''),
    ('0014', '暗', 'アン、くら-い', ''),
    ('0015', '以', 'イ', ''),
    ('0016', '衣', 'イ、ころも', '浴衣'),
    ('0017', '位', 'イ、くらい', '三位一体、従三位'),
    ('0018', '囲圍', 'イ、かこ-む、かこ-う', ''),
    ('0019', '医醫', 'イ', ''),
    ('0020', '依', 'イ、（エ）', ''),
    ('0021', '委', 'イ、ゆだ-ねる', ''),
    ('0022', '威', 'イ', ''),
    ('0023', '為爲', 'イ', '為替'),
    ('0024', '畏', 'イ、おそ-れる', ''),
    ('0025', '胃', 'イ', ''),
    ('0026', '尉', 'イ', ''),
    ('0027', '異', 'イ、こと', ''),
    ('0028', '移', 'イ、うつ-る、うつ-す', ''),
    ('0029', '萎', 'イ、な-える', ''),
    ('0030', '偉', 'イ、えら-い', ''),
    ('0031', '椅', 'イ', ''),
    ('0032', '彙', 'イ', ''),
    ('0033', '意', 'イ', '意気地'),
    ('0034', '違', 'イ、ちが-う、ちが-える', ''),
    ('0035', '維', 'イ', ''),
    ('0036', '慰', 'イ、なぐさ-める、なぐさ-む', ''),
    ('0037', '遺', 'イ、（ユイ）', ''),
    ('0038', '緯', 'イ', ''),
    ('0039', '域', 'イキ', ''),
    ('0040', '育', 'イク、そだ-つ、そだ-てる、はぐく-む', ''),
    ('0041', '一', 'イチ、イツ、ひと、ひと-つ', '一日、一人'),
    ('0042', '壱壹', 'イチ', ''),
    ('0043', '逸逸', 'イツ', ''),
    ('0044', '茨', '（いばら）', '茨城'),
    ('0045', '芋', 'いも', ''),
    ('0046', '引', 'イン、ひ-く、ひ-ける', ''),
    ('0047', '印', 'イン、しるし', ''),
    ('0048', '因', 'イン、よ-る', ''),
    ('0049', '咽', 'イン', ''),
    ('0050', '姻', 'イン', ''),
    ('0051', '員', 'イン', ''),
    ('0052', '院', 'イン', ''),
    ('0053', '淫', 'イン、みだ-ら', ''),
    ('0054', '陰', 'イン、かげ、かげ-る', ''),
    ('0055', '飲飮', 'イン、の-む', ''),
    ('0056', '隠隱', 'イン、かく-す、かく-れる', ''),
    ('0057', '韻', 'イン', ''),
    ('0058', '右', 'ウ、ユウ、みぎ', ''),
    ('0059', '宇', 'ウ', ''),
    ('0060', '羽羽', 'ウ、は、はね', ''),
    ('0061', '雨', 'ウ、あめ、（あま）', '五月雨、時雨、梅雨、春雨、小雨、霧雨'),
    ('0062', '唄', '（うた）', ''),
    ('0063', '鬱', 'ウツ', ''),
    ('0064', '畝', 'うね', ''),
    ('0065', '浦', 'うら', ''),
    ('0066', '運', 'ウン、はこ-ぶ', ''),
    ('0067', '雲', 'ウン、くも', ''),
    ('0068', '永', 'エイ、なが-い', ''),
    ('0069', '泳', 'エイ、およ-ぐ', ''),
    ('0070', '英', 'エイ', ''),
    ('0071', '映', 'エイ、うつ-る、うつ-す、は-える', ''),
    ('0072', '栄榮', 'エイ、さか-える、は-え、は-える', ''),
    ('0073', '営營', 'エイ、いとな-む', ''),
    ('0074', '詠', 'エイ、よ-む', ''),
    ('0075', '影', 'エイ、かげ', ''),
    ('0076', '鋭銳', 'エイ、するど-い', ''),
    ('0077', '衛衞', 'エイ', ''),
    ('0078', '易', 'エキ、イ、やさ-しい', ''),
    ('0079', '疫', 'エキ、（ヤク）', ''),
    ('0080', '益益', 'エキ、（ヤク）', ''),
    ('0081', '液', 'エキ', ''),
    ('0082', '駅驛', 'エキ', ''),
    ('0083', '悦悅', 'エツ', ''),
    ('0084', '越', 'エツ、こ-す、こ-える', ''),
    ('0085', '謁謁', 'エツ', ''),
    ('0086', '閲閱', 'エツ', ''),
    ('0087', '円圓', 'エン、まる-い', ''),
    ('0088', '延', 'エン、の-びる、の-べる、の-ばす', ''),
    ('0089', '沿', 'エン、そ-う', ''),
    ('0090', '炎', 'エン、ほのお', ''),
    ('0091', '怨', 'エン、オン', ''),
    ('0092', '宴', 'エン', ''),
    ('0093', '媛', 'エン', '愛媛'),
    ('0094', '援', 'エン', ''),
    ('0095', '園', 'エン、その', ''),
    ('0096', '煙', 'エン、けむ-る、けむり、けむ-い', ''),
    ('0097', '猿', 'エン、さる', ''),
    ('0098', '遠', 'エン、（オン）、とお-い', ''),
    ('0099', '鉛', 'エン、なまり', ''),
    ('0100', '塩鹽', 'エン、しお', ''),
    ('0101', '演', 'エン', ''),
    ('0102', '縁緣', 'エン、ふち', '因縁'),
    ('0103', '艶艷', 'エン、つや', ''),
    ('0104', '汚', 'オ、けが-す、けが-れる、けが-らわしい、よご-す、よご-れる、きたな-い', ''),
    ('0105', '王', 'オウ', '親王、勤王'),
    ('0106', '凹', 'オウ', '凸凹'),
    ('0107', '央', 'オウ', ''),
    ('0108', '応應', 'オウ、こた-える', '反応、順応'),
    ('0109', '往', 'オウ', ''),
    ('0110', '押', 'オウ、お-す、お-さえる', ''),
    ('0111', '旺', 'オウ', ''),
    ('0112', '欧歐', 'オウ', ''),
    ('0113', '殴毆', 'オウ、なぐ-る', ''),
    ('0114', '桜櫻', 'オウ、さくら', ''),
    ('0115', '翁', 'オウ', ''),
    ('0116', '奥奧', 'オウ、おく', ''),
    ('0117', '横橫', 'オウ、よこ', ''),
    ('0118', '岡', '（おか）', ''),
    ('0119', '屋', 'オク、や', '母屋、数寄屋、数奇屋、部屋、八百屋、紺屋'),
    ('0120', '億', 'オク', ''),
    ('0121', '憶', 'オク', ''),
    ('0122', '臆', 'オク', ''),
    ('0123', '虞', 'おそれ', ''),
    ('0124', '乙', 'オツ', '乙女、早乙女'),
    ('0125', '俺', 'おれ', ''),
    ('0126', '卸', 'おろ-す、おろし', ''),
    ('0127', '音', 'オン、イン、おと、ね', '観音'),
    ('0128', '恩', 'オン', ''),
    ('0129', '温溫', 'オン、あたた-か、あたた-かい、あたた-まる、あたた-める', ''),
    ('0130', '穏穩', 'オン、おだ-やか', '安穏'),
    ('0131', '下', 'カ、ゲ、した、しも、もと、さ-げる、さ-がる、くだ-る、くだ-す、くだ-さる、お-ろす、お-りる', '下手'),
    ('0132', '化', 'カ、ケ、ば-ける、ば-かす', ''),
    ('0133', '火', 'カ、ひ、（ほ）', ''),
    ('0134', '加', 'カ、くわ-える、くわ-わる', ''),
    ('0135', '可', 'カ', ''),
    ('0136', '仮假', 'カ、（ケ）、かり', '仮名'),
    ('0137', '何', 'カ、なに、（なん）', ''),
    ('0138', '花', 'カ、はな', ''),
    ('0139', '佳', 'カ', ''),
    ('0140', '価價', 'カ、あたい', ''),
    ('0141', '果', 'カ、は-たす、は-てる、は-て', '果物'),
    ('0142', '河', 'カ、かわ', '河岸、河原'),
    ('0143', '苛', 'カ', ''),
    ('0144', '科', 'カ', ''),
    ('0145', '架', 'カ、か-ける、か-かる', ''),
    ('0146', '夏', 'カ、（ゲ）、なつ', ''),
    ('0147', '家', 'カ、ケ、いえ、や', '母家'),
    ('0148', '荷', 'カ、に', ''),
    ('0149', '華', 'カ、（ケ）、はな', ''),
    ('0150', '菓', 'カ', ''),
    ('0151', '貨', 'カ', ''),
    ('0152', '渦', 'カ、うず', ''),
    ('0153', '過', 'カ、す-ぎる、す-ごす、あやま-つ、あやま-ち', ''),
    ('0154', '嫁', 'カ、よめ、とつ-ぐ', ''),
    ('0155', '暇', 'カ、ひま', ''),
    ('0156', '禍禍', 'カ', ''),
    ('0157', '靴', 'カ、くつ', ''),
    ('0158', '寡', 'カ', ''),
    ('0159', '歌', 'カ、うた、うた-う', '詩歌'),
    ('0160', '箇', 'カ', ''),
    ('0161', '稼', 'カ、かせ-ぐ', ''),
    ('0162', '課', 'カ', ''),
    ('0163', '蚊', 'か', '蚊帳'),
    ('0164', '牙', 'ガ、（ゲ）、きば', ''),
    ('0165', '瓦', 'ガ、かわら', ''),
    ('0166', '我', 'ガ、われ、わ', ''),
    ('0167', '画畫', 'ガ、カク', ''),
    ('0168', '芽', 'ガ、め', ''),
    ('0169', '賀', 'ガ', '滋賀'),
    ('0170', '雅', 'ガ', ''),
    ('0171', '餓', 'ガ', ''),
    ('0172', '介', 'カイ', ''),
    ('0173', '回', 'カイ、（エ）、まわ-る、まわ-す', ''),
    ('0174', '灰', 'カイ、はい', ''),
    ('0175', '会會', 'カイ、エ、あ-う', ''),
    ('0176', '快', 'カイ、こころよ-い', ''),
    ('0177', '戒', 'カイ、いまし-める', ''),
    ('0178', '改', 'カイ、あらた-める、あらた-まる', ''),
    ('0179', '怪', 'カイ、あや-しい、あや-しむ', ''),
    ('0180', '拐', 'カイ', ''),
    ('0181', '悔悔', 'カイ、く-いる、く-やむ、くや-しい', ''),
    ('0182', '海海', 'カイ、うみ', '海女、海士、海原'),
    ('0183', '界', 'カイ', ''),
    ('0184', '皆', 'カイ、みな', ''),
    ('0185', '械', 'カイ', ''),
    ('0186', '絵繪', 'カイ、エ', ''),
    ('0187', '開', 'カイ、ひら-く、ひら-ける、あ-く、あ-ける', ''),
    ('0188', '階', 'カイ', ''),
    ('0189', '塊', 'カイ、かたまり', ''),
    ('0190', '楷', 'カイ', ''),
    ('0191', '解', 'カイ、ゲ、と-く、と-かす、と-ける', ''),
    ('0192', '潰', 'カイ、つぶ-す、つぶ-れる', ''),
    ('0193', '壊壞', 'カイ、こわ-す、こわ-れる', ''),
    ('0194', '懐懷', 'カイ、ふところ、なつ-かしい、なつ-かしむ、なつ-く、なつ-ける', ''),
    ('0195', '諧', 'カイ', ''),
    ('0196', '貝', 'かい', ''),
    ('0197', '外', 'ガイ、ゲ、そと、ほか、はず-す、はず-れる', ''),
    ('0198', '劾', 'ガイ', ''),
    ('0199', '害', 'ガイ', ''),
    ('0200', '崖', 'ガイ、がけ', ''),
    ('0201', '涯', 'ガイ', ''),
    ('0202', '街', 'ガイ、（カイ）、まち', ''),
    ('0203', '慨慨', 'ガイ', ''),
    ('0204', '蓋', 'ガイ、ふた', ''),
    ('0205', '該', 'ガイ', ''),
    ('0206', '概槪', 'ガイ', ''),
    ('0207', '骸', 'ガイ', ''),
    ('0208', '垣', 'かき', ''),
    ('0209', '柿', 'かき', ''),
    ('0210', '各', 'カク、おのおの', ''),
    ('0211', '角', 'カク、かど、つの', ''),
    ('0212', '拡擴', 'カク', ''),
    ('0213', '革', 'カク、かわ', ''),
    ('0214', '格', 'カク、（コウ）', ''),
    ('0215', '核', 'カク', ''),
    ('0216', '殻殼', 'カク、から', ''),
    ('0217', '郭', 'カク', ''),
    ('0218', '覚覺', 'カク、おぼ-える、さ-ます、さ-める', ''),
    ('0219', '較', 'カク', ''),
    ('0220', '隔', 'カク、へだ-てる、へだ-たる', ''),
    ('0221', '閣', 'カク', ''),
    ('0222', '確', 'カク、たし-か、たし-かめる', ''),
    ('0223', '獲', 'カク、え-る', ''),
    ('0224', '嚇', 'カク', ''),
    ('0225', '穫', 'カク', ''),
    ('0226', '学學', 'ガク、まな-ぶ', ''),
    ('0227', '岳嶽', 'ガク、たけ', ''),
    ('0228', '楽樂', 'ガク、ラク、たの-しい、たの-しむ', '神楽'),
    ('0229', '額', 'ガク、ひたい', ''),
    ('0230', '顎', 'ガク、あご', ''),
    ('0231', '掛', 'か-ける、か-かる、かかり', ''),
    ('0232', '潟', 'かた', ''),
    ('0233', '括', 'カツ', ''),
    ('0234', '活', 'カツ', ''),
    ('0235', '喝喝', 'カツ', ''),
    ('0236', '渇渴', 'カツ、かわ-く', ''),
    ('0237', '割', 'カツ、わ-る、わり、わ-れる、さ-く', ''),
    ('0238', '葛', 'カツ、くず', ''),
    ('0239', '滑', 'カツ、コツ、すべ-る、なめ-らか', ''),
    ('0240', '褐褐', 'カツ', ''),
    ('0241', '轄', 'カツ', ''),
    ('0242', '且', 'か-つ', ''),
    ('0243', '株', 'かぶ', ''),
    ('0244', '釜', 'かま', ''),
    ('0245', '鎌', 'かま', ''),
    ('0246', '刈', 'か-る', ''),
    ('0247', '干', 'カン、ほ-す、ひ-る', ''),
    ('0248', '刊', 'カン', ''),
    ('0249', '甘', 'カン、あま-い、あま-える、あま-やかす', ''),
    ('0250', '汗', 'カン、あせ', ''),
    ('0251', '缶罐', 'カン', ''),
    ('0252', '完', 'カン', ''),
    ('0253', '肝', 'カン、きも', ''),
    ('0254', '官', 'カン', ''),
    ('0255', '冠', 'カン、かんむり', ''),
    ('0256', '巻卷', 'カン、ま-く、まき', ''),
    ('0257', '看', 'カン', ''),
    ('0258', '陥陷', 'カン、おちい-る、おとしい-れる', ''),
    ('0259', '乾', 'カン、かわ-く、かわ-かす', ''),
    ('0260', '勘', 'カン', ''),
    ('0261', '患', 'カン、わずら-う', ''),
    ('0262', '貫', 'カン、つらぬ-く', ''),
    ('0263', '寒', 'カン、さむ-い', ''),
    ('0264', '喚', 'カン', ''),
    ('0265', '堪', 'カン、た-える', '堪能'),
    ('0266', '換', 'カン、か-える、か-わる', ''),
    ('0267', '敢', 'カン', ''),
    ('0268', '棺', 'カン', ''),
    ('0269', '款', 'カン', ''),
    ('0270', '間', 'カン、ケン、あいだ、ま', ''),
    ('0271', '閑', 'カン', ''),
    ('0272', '勧勸', 'カン、すす-める', ''),
    ('0273', '寛寬', 'カン', ''),
    ('0274', '幹', 'カン、みき', ''),
    ('0275', '感', 'カン', ''),
    ('0276', '漢漢', 'カン', ''),
    ('0277', '慣', 'カン、な-れる、な-らす', ''),
    ('0278', '管', 'カン、くだ', ''),
    ('0279', '関關', 'カン、せき、かか-わる', ''),
    ('0280', '歓歡', 'カン', ''),
    ('0281', '監', 'カン', ''),
    ('0282', '緩', 'カン、ゆる-い、ゆる-やか、ゆる-む、ゆる-める', ''),
    ('0283', '憾', 'カン', ''),
    ('0284', '還', 'カン', ''),
    ('0285', '館館', 'カン、やかた', ''),
    ('0286', '環', 'カン', ''),
    ('0287', '簡', 'カン', ''),
    ('0288', '観觀', 'カン', ''),
    ('0289', '韓', 'カン', ''),
    ('0290', '艦', 'カン', ''),
    ('0291', '鑑', 'カン、かんが-みる', ''),
    ('0292', '丸', 'ガン、まる、まる-い、まる-める', ''),
    ('0293', '含', 'ガン、ふく-む、ふく-める', ''),
    ('0294', '岸', 'ガン、きし', '河岸'),
    ('0295', '岩', 'ガン、いわ', ''),
    ('0296', '玩', 'ガン', ''),
    ('0297', '眼', 'ガン、（ゲン）、まなこ', '眼鏡'),
    ('0298', '頑', 'ガン', ''),
    ('0299', '顔顏', 'ガン、かお', '笑顔'),
    ('0300', '願', 'ガン、ねが-う', ''),
    ('0301', '企', 'キ、くわだ-てる', ''),
    ('0302', '伎', 'キ', ''),
    ('0303', '危', 'キ、あぶ-ない、あや-うい、あや-ぶむ', ''),
    ('0304', '机', 'キ、つくえ', ''),
    ('0305', '気氣', 'キ、ケ', '意気地、浮気'),
    ('0306', '岐', 'キ', '岐阜'),
    ('0307', '希', 'キ', ''),
    ('0308', '忌', 'キ、い-む、い-まわしい', ''),
    ('0309', '汽', 'キ', ''),
    ('0310', '奇', 'キ', '数奇屋'),
    ('0311', '祈祈', 'キ、いの-る', ''),
    ('0312', '季', 'キ', ''),
    ('0313', '紀', 'キ', ''),
    ('0314', '軌', 'キ', ''),
    ('0315', '既旣', 'キ、すで-に', ''),  # "旣"は康熙字典体
    ('0316', '記', 'キ、しる-す', ''),
    ('0317', '起', 'キ、お-きる、お-こる、お-こす', ''),
    ('0318', '飢', 'キ、う-える', ''),
    ('0319', '鬼', 'キ、おに', ''),
    ('0320', '帰歸', 'キ、かえ-る、かえ-す', ''),
    ('0321', '基', 'キ、もと、もとい', ''),
    ('0322', '寄', 'キ、よ-る、よ-せる', '数寄屋、最寄り、寄席'),
    ('0323', '規', 'キ', ''),
    ('0324', '亀龜', 'キ、かめ', ''),
    ('0325', '喜', 'キ、よろこ-ぶ', ''),
    ('0326', '幾', 'キ、いく', ''),
    ('0327', '揮', 'キ', ''),
    ('0328', '期', 'キ、（ゴ）', ''),
    ('0329', '棋', 'キ', ''),
    ('0330', '貴', 'キ、たっと-い、とうと-い、たっと-ぶ、とうと-ぶ', '富貴'),
    ('0331', '棄', 'キ', ''),
    ('0332', '毀', 'キ', ''),
    ('0333', '旗', 'キ、はた', ''),
    ('0334', '器器', 'キ、うつわ', ''),
    ('0335', '畿', 'キ', ''),
    ('0336', '輝', 'キ、かがや-く', ''),
    ('0337', '機', 'キ、はた', ''),
    ('0338', '騎', 'キ', ''),
    ('0339', '技', 'ギ、わざ', ''),
    ('0340', '宜', 'ギ', ''),
    ('0341', '偽僞', 'ギ、いつわ-る、にせ', ''),
    ('0342', '欺', 'ギ、あざむ-く', ''),
    ('0343', '義', 'ギ', ''),
    ('0344', '疑', 'ギ、うたが-う', ''),
    ('0345', '儀', 'ギ', ''),
    ('0346', '戯戲', 'ギ、たわむ-れる', ''),
    ('0347', '擬', 'ギ', ''),
    ('0348', '犠犧', 'ギ', ''),
    ('0349', '議', 'ギ', ''),
    ('0350', '菊', 'キク', ''),
    ('0351', '吉', 'キチ、キツ', ''),
    ('0352', '喫', 'キツ', ''),
    ('0353', '詰', 'キツ、つ-める、つ-まる、つ-む', ''),
    ('0354', '却', 'キャク', ''),
    ('0355', '客', 'キャク、カク', ''),
    ('0356', '脚', 'キャク、（キャ）、あし', ''),
    ('0357', '逆', 'ギャク、さか、さか-らう', ''),
    ('0358', '虐', 'ギャク、しいた-げる', ''),
    ('0359', '九', 'キュウ、ク、ここの、ここの-つ', ''),
    ('0360', '久', 'キュウ、（ク）、ひさ-しい', ''),
    ('0361', '及', 'キュウ、およ-ぶ、およ-び、およ-ぼす', ''),
    ('0362', '弓', 'キュウ、ゆみ', ''),
    ('0363', '丘', 'キュウ、おか', ''),
    ('0364', '旧舊', 'キュウ', ''),
    ('0365', '休', 'キュウ、やす-む、やす-まる、やす-める', ''),
    ('0366', '吸', 'キュウ、す-う', ''),
    ('0367', '朽', 'キュウ、く-ちる', ''),
    ('0368', '臼', 'キュウ、うす', ''),
    ('0369', '求', 'キュウ、もと-める', ''),
    ('0370', '究', 'キュウ、きわ-める', ''),
    ('0371', '泣', 'キュウ、な-く', ''),
    ('0372', '急', 'キュウ、いそ-ぐ', ''),
    ('0373', '級', 'キュウ', ''),
    ('0374', '糾', 'キュウ', ''),
    ('0375', '宮', 'キュウ、グウ、（ク）、みや', '宮城、宮内庁'),
    ('0376', '救', 'キュウ、すく-う', ''),
    ('0377', '球', 'キュウ、たま', ''),
    ('0378', '給', 'キュウ', ''),
    ('0379', '嗅', 'キュウ、か-ぐ', ''),
    ('0380', '窮', 'キュウ、きわ-める、きわ-まる', ''),
    ('0381', '牛', 'ギュウ、うし', ''),
    ('0382', '去', 'キョ、コ、さ-る', ''),
    ('0383', '巨', 'キョ', ''),
    ('0384', '居', 'キョ、い-る', '居士'),
    ('0385', '拒', 'キョ、こば-む', ''),
    ('0386', '拠據', 'キョ、コ', ''),
    ('0387', '挙擧', 'キョ、あ-げる、あ-がる', ''),
    ('0388', '虚虛', 'キョ、（コ）', ''),
    ('0389', '許', 'キョ、ゆる-す', ''),
    ('0390', '距', 'キョ', ''),
    ('0391', '魚', 'ギョ、うお、さかな', '雑魚'),
    ('0392', '御', 'ギョ、ゴ、おん', ''),
    ('0393', '漁', 'ギョ、リョウ', ''),
    ('0394', '凶', 'キョウ', ''),
    ('0395', '共', 'キョウ、とも', ''),
    ('0396', '叫', 'キョウ、さけ-ぶ', ''),
    ('0397', '狂', 'キョウ、くる-う、くる-おしい', ''),
    ('0398', '京', 'キョウ、ケイ', '京浜、京阪'),
    ('0399', '享', 'キョウ', ''),
    ('0400', '供', 'キョウ、（ク）、そな-える、とも', ''),
    ('0401', '協', 'キョウ', ''),
    ('0402', '況', 'キョウ', ''),
    ('0403', '峡峽', 'キョウ', ''),
    ('0404', '挟挾', 'キョウ、はさ-む、はさ-まる', ''),
    ('0405', '狭狹', 'キョウ、せま-い、せば-める、せば-まる', ''),
    ('0406', '恐', 'キョウ、おそ-れる、おそ-ろしい', ''),
    ('0407', '恭', 'キョウ、うやうや-しい', ''),
    ('0408', '胸', 'キョウ、むね、（むな）', ''),
    ('0409', '脅', 'キョウ、おびや-かす、おど-す、おど-かす', ''),
    ('0410', '強', 'キョウ、ゴウ、つよ-い、つよ-まる、つよ-める、し-いる', ''),
    ('0411', '教敎', 'キョウ、おし-える、おそ-わる', ''),
    ('0412', '郷鄕', 'キョウ、ゴウ', ''),
    ('0413', '境', 'キョウ、（ケイ）、さかい', ''),
    ('0414', '橋', 'キョウ、はし', ''),
    ('0415', '矯', 'キョウ、た-める', ''),
    ('0416', '鏡', 'キョウ、かがみ', '眼鏡'),
    ('0417', '競', 'キョウ、ケイ、きそ-う、せ-る', ''),
    ('0418', '響響', 'キョウ、ひび-く', ''),
    ('0419', '驚', 'キョウ、おどろ-く、おどろ-かす', ''),
    ('0420', '仰', 'ギョウ、（コウ）、あお-ぐ、おお-せ', ''),
    ('0421', '暁曉', 'ギョウ、あかつき', ''),
    ('0422', '業', 'ギョウ、ゴウ、わざ', ''),
    ('0423', '凝', 'ギョウ、こ-る、こ-らす', ''),
    ('0424', '曲', 'キョク、ま-がる、ま-げる', ''),
    ('0425', '局', 'キョク', ''),
    ('0426', '極', 'キョク、ゴク、きわ-める、きわ-まる、きわ-み', ''),
    ('0427', '玉', 'ギョク、たま', ''),
    ('0428', '巾', 'キン', ''),
    ('0429', '斤', 'キン', ''),
    ('0430', '均', 'キン', ''),
    ('0431', '近', 'キン、ちか-い', ''),
    ('0432', '金', 'キン、コン、かね、（かな）', ''),
    ('0433', '菌', 'キン', ''),
    ('0434', '勤勤', 'キン、（ゴン）、つと-める、つと-まる', ''),
    ('0435', '琴', 'キン、こと', ''),
    ('0436', '筋', 'キン、すじ', ''),
    ('0437', '僅', 'キン、わず-か', ''),
    ('0438', '禁', 'キン', ''),
    ('0439', '緊', 'キン', ''),
    ('0440', '錦', 'キン、にしき', ''),
    ('0441', '謹謹', 'キン、つつし-む', ''),
    ('0442', '襟', 'キン、えり', ''),
    ('0443', '吟', 'ギン', ''),
    ('0444', '銀', 'ギン', ''),
    ('0445', '区區', 'ク', ''),
    ('0446', '句', 'ク', ''),
    ('0447', '苦', 'ク、くる-しい、くる-しむ、くる-しめる、にが-い、にが-る', ''),
    ('0448', '駆驅', 'ク、か-ける、か-る', ''),
    ('0449', '具', 'グ', ''),
    ('0450', '惧', 'グ', ''),
    ('0451', '愚', 'グ、おろ-か', ''),
    ('0452', '空', 'クウ、そら、あ-く、あ-ける、から', ''),
    ('0453', '偶', 'グウ', ''),
    ('0454', '遇', 'グウ', ''),
    ('0455', '隅', 'グウ、すみ', ''),
    ('0456', '串', 'くし', ''),
    ('0457', '屈', 'クツ', ''),
    ('0458', '掘', 'クツ、ほ-る', ''),
    ('0459', '窟', 'クツ', ''),
    ('0460', '熊', 'くま', ''),
    ('0461', '繰', 'く-る', ''),
    ('0462', '君', 'クン、きみ', ''),
    ('0463', '訓', 'クン', ''),
    ('0464', '勲勳', 'クン', ''),
    ('0465', '薫薰', 'クン、かお-る', ''),
    ('0466', '軍', 'グン', ''),
    ('0467', '郡', 'グン', ''),
    ('0468', '群', 'グン、む-れる、む-れ、（むら）', ''),
    ('0469', '兄', 'ケイ、（キョウ）、あに', '兄さん'),
    ('0470', '刑', 'ケイ', ''),
    ('0471', '形', 'ケイ、ギョウ、かた、かたち', ''),
    ('0472', '系', 'ケイ', ''),
    ('0473', '径徑', 'ケイ', ''),
    ('0474', '茎莖', 'ケイ、くき', ''),
    ('0475', '係', 'ケイ、かか-る、かかり', ''),
    ('0476', '型', 'ケイ、かた', ''),
    ('0477', '契', 'ケイ、ちぎ-る', ''),
    ('0478', '計', 'ケイ、はか-る、はか-らう', '時計'),
    ('0479', '恵惠', 'ケイ、エ、めぐ-む', ''),
    ('0480', '啓', 'ケイ', ''),
    ('0481', '掲揭', 'ケイ、かか-げる', ''),
    ('0482', '渓溪', 'ケイ', ''),
    ('0483', '経經', 'ケイ、キョウ、へ-る', '読経'),
    ('0484', '蛍螢', 'ケイ、ほたる', ''),
    ('0485', '敬', 'ケイ、うやま-う', ''),
    ('0486', '景', 'ケイ', '景色'),
    ('0487', '軽輕', 'ケイ、かる-い、かろ-やか', ''),
    ('0488', '傾', 'ケイ、かたむ-く、かたむ-ける', ''),
    ('0489', '携', 'ケイ、たずさ-える、たずさ-わる', ''),
    ('0490', '継繼', 'ケイ、つ-ぐ', ''),
    ('0491', '詣', 'ケイ、もう-でる', ''),
    ('0492', '慶', 'ケイ', ''),
    ('0493', '憬', 'ケイ', '憧憬'),
    ('0494', '稽', 'ケイ', ''),
    ('0495', '憩', 'ケイ、いこ-い、いこ-う', ''),
    ('0496', '警', 'ケイ', ''),
    ('0497', '鶏鷄', 'ケイ、にわとり', ''),
    ('0498', '芸藝', 'ゲイ', ''),
    ('0499', '迎', 'ゲイ、むか-える', ''),
    ('0500', '鯨', 'ゲイ、くじら', ''),
    ('0501', '隙', 'ゲキ、すき', ''),
    ('0502', '劇', 'ゲキ', ''),
    ('0503', '撃擊', 'ゲキ、う-つ', ''),
    ('0504', '激', 'ゲキ、はげ-しい', ''),
    ('0505', '桁', 'けた', ''),
    ('0506', '欠缺', 'ケツ、か-ける、か-く', ''),
    ('0507', '穴', 'ケツ、あな', ''),
    ('0508', '血', 'ケツ、ち', ''),
    ('0509', '決', 'ケツ、き-める、き-まる', ''),
    ('0510', '結', 'ケツ、むす-ぶ、ゆ-う、ゆ-わえる', ''),
    ('0511', '傑', 'ケツ', ''),
    ('0512', '潔', 'ケツ、いさぎよ-い', ''),
    ('0513', '月', 'ゲツ、ガツ、つき', '五月、五月雨'),
    ('0514', '犬', 'ケン、いぬ', ''),
    ('0515', '件', 'ケン', ''),
    ('0516', '見', 'ケン、み-る、み-える、み-せる', ''),
    ('0517', '券', 'ケン', ''),
    ('0518', '肩', 'ケン、かた', ''),
    ('0519', '建', 'ケン、（コン）、た-てる、た-つ', ''),
    ('0520', '研硏', 'ケン、と-ぐ', ''),
    ('0521', '県縣', 'ケン', ''),
    ('0522', '倹儉', 'ケン', ''),
    ('0523', '兼', 'ケン、か-ねる', ''),
    ('0524', '剣劍', 'ケン、つるぎ', ''),
    ('0525', '拳', 'ケン、こぶし', ''),
    ('0526', '軒', 'ケン、のき', ''),
    ('0527', '健', 'ケン、すこ-やか', ''),
    ('0528', '険險', 'ケン、けわ-しい', ''),
    ('0529', '圏圈', 'ケン', ''),
    ('0530', '堅', 'ケン、かた-い', ''),
    ('0531', '検檢', 'ケン', ''),
    ('0532', '嫌', 'ケン、（ゲン）、きら-う、いや', ''),
    ('0533', '献獻', 'ケン、（コン）', ''),
    ('0534', '絹', 'ケン、きぬ', ''),
    ('0535', '遣', 'ケン、つか-う、つか-わす', ''),
    ('0536', '権權', 'ケン、（ゴン）', ''),
    ('0537', '憲', 'ケン', ''),
    ('0538', '賢', 'ケン、かしこ-い', ''),
    ('0539', '謙', 'ケン', ''),
    ('0540', '鍵', 'ケン、かぎ', ''),
    ('0541', '繭', 'ケン、まゆ', ''),
    ('0542', '顕顯', 'ケン', ''),
    ('0543', '験驗', 'ケン、（ゲン）', ''),
    ('0544', '懸', 'ケン、（ケ）、か-ける、か-かる', ''),
    ('0545', '元', 'ゲン、ガン、もと', ''),
    ('0546', '幻', 'ゲン、まぼろし', ''),
    ('0547', '玄', 'ゲン', '玄人'),
    ('0548', '言', 'ゲン、ゴン、い-う、こと', ''),
    ('0549', '弦', 'ゲン、つる', ''),
    ('0550', '限', 'ゲン、かぎ-る', ''),
    ('0551', '原', 'ゲン、はら', '海原、河原、川原'),
    ('0552', '現', 'ゲン、あらわ-れる、あらわ-す', ''),
    ('0553', '舷', 'ゲン', ''),
    ('0554', '減', 'ゲン、へ-る、へ-らす', ''),
    ('0555', '源', 'ゲン、みなもと', ''),
    ('0556', '厳嚴', 'ゲン、（ゴン）、おごそ-か、きび-しい', ''),
    ('0557', '己', 'コ、キ、おのれ', ''),
    ('0558', '戸戶', 'コ、と', ''),
    ('0559', '古', 'コ、ふる-い、ふる-す', ''),
    ('0560', '呼', 'コ、よ-ぶ', ''),
    ('0561', '固', 'コ、かた-める、かた-まる、かた-い', '固唾'),
    ('0562', '股', 'コ、また', ''),
    ('0563', '虎', 'コ、とら', ''),
    ('0564', '孤', 'コ', ''),
    ('0565', '弧', 'コ', ''),
    ('0566', '故', 'コ、ゆえ', ''),
    ('0567', '枯', 'コ、か-れる、か-らす', ''),
    ('0568', '個', 'コ', ''),
    ('0569', '庫', 'コ、（ク）', ''),
    ('0570', '湖', 'コ、みずうみ', ''),
    ('0571', '雇', 'コ、やと-う', ''),
    ('0572', '誇', 'コ、ほこ-る', ''),
    ('0573', '鼓', 'コ、つづみ', ''),
    ('0574', '錮', 'コ', ''),
    ('0575', '顧', 'コ、かえり-みる', ''),
    ('0576', '五', 'ゴ、いつ、いつ-つ', '五月、五月雨'),
    ('0577', '互', 'ゴ、たが-い', ''),
    ('0578', '午', 'ゴ', ''),
    ('0579', '呉吳', 'ゴ', ''),
    ('0580', '後', 'ゴ、コウ、のち、うし-ろ、あと、おく-れる', ''),
    ('0581', '娯娛', 'ゴ', ''),
    ('0582', '悟', 'ゴ、さと-る', ''),
    ('0583', '碁', 'ゴ', ''),
    ('0584', '語', 'ゴ、かた-る、かた-らう', ''),
    ('0585', '誤', 'ゴ、あやま-る', ''),
    ('0586', '護', 'ゴ', ''),
    ('0587', '口', 'コウ、ク、くち', ''),
    ('0588', '工', 'コウ、ク', ''),
    ('0589', '公', 'コウ、おおやけ', ''),
    ('0590', '勾', 'コウ', ''),
    ('0591', '孔', 'コウ', ''),
    ('0592', '功', 'コウ、（ク）', ''),
    ('0593', '巧', 'コウ、たく-み', ''),
    ('0594', '広廣', 'コウ、ひろ-い、ひろ-まる、ひろ-める、ひろ-がる、ひろ-げる', ''),
    ('0595', '甲', 'コウ、カン', ''),
    ('0596', '交', 'コウ、まじ-わる、まじ-える、ま-じる、ま-ざる、ま-ぜる、か-う、か-わす', ''),
    ('0597', '光', 'コウ、ひか-る、ひかり', ''),
    ('0598', '向', 'コウ、む-く、む-ける、む-かう、む-こう', ''),
    ('0599', '后', 'コウ', ''),
    ('0600', '好', 'コウ、この-む、す-く', ''),
    ('0601', '江', 'コウ、え', ''),
    ('0602', '考', 'コウ、かんが-える', ''),
    ('0603', '行', 'コウ、ギョウ、（アン）、い-く、ゆ-く、おこな-う', '行方'),
    ('0604', '坑', 'コウ', ''),
    ('0605', '孝', 'コウ', ''),
    ('0606', '抗', 'コウ', ''),
    ('0607', '攻', 'コウ、せ-める', ''),
    ('0608', '更', 'コウ、さら、ふ-ける、ふ-かす', ''),
    ('0609', '効效', 'コウ、き-く', ''),  # "效"は康熙字典体
    ('0610', '幸', 'コウ、さいわ-い、さち、しあわ-せ', ''),
    ('0611', '拘', 'コウ', ''),
    ('0612', '肯', 'コウ', ''),
    ('0613', '侯', 'コウ', ''),
    ('0614', '厚', 'コウ、あつ-い', ''),
    ('0615', '恒恆', 'コウ', ''),
    ('0616', '洪', 'コウ', ''),
    ('0617', '皇', 'コウ、オウ', '天皇'),
    ('0618', '紅', 'コウ、（ク）、べに、くれない', '紅葉'),
    ('0619', '荒', 'コウ、あら-い、あ-れる、あ-らす', ''),
    ('0620', '郊', 'コウ', ''),
    ('0621', '香', 'コウ、（キョウ）、か、かお-り、かお-る', ''),
    ('0622', '候', 'コウ、そうろう', ''),
    ('0623', '校', 'コウ', ''),
    ('0624', '耕', 'コウ、たがや-す', ''),
    ('0625', '航', 'コウ', ''),
    ('0626', '貢', 'コウ、（ク）、みつ-ぐ', ''),
    ('0627', '降', 'コウ、お-りる、お-ろす、ふ-る', ''),
    ('0628', '高', 'コウ、たか-い、たか、たか-まる、たか-める', ''),
    ('0629', '康', 'コウ', ''),
    ('0630', '控', 'コウ、ひか-える', ''),
    ('0631', '梗', 'コウ', ''),
    ('0632', '黄黃', 'コウ、オウ、き、（こ）', '硫黄'),
    ('0633', '喉', 'コウ、のど', ''),
    ('0634', '慌', 'コウ、あわ-てる、あわ-ただしい', ''),
    ('0635', '港', 'コウ、みなと', ''),
    ('0636', '硬', 'コウ、かた-い', ''),
    ('0637', '絞', 'コウ、しぼ-る、し-める、し-まる', ''),
    ('0638', '項', 'コウ', ''),
    ('0639', '溝', 'コウ、みぞ', ''),
    ('0640', '鉱鑛', 'コウ', ''),
    ('0641', '構', 'コウ、かま-える、かま-う', ''),
    ('0642', '綱', 'コウ、つな', ''),
    ('0643', '酵', 'コウ', ''),
    ('0644', '稿', 'コウ', ''),
    ('0645', '興', 'コウ、キョウ、おこ-る、おこ-す', ''),
    ('0646', '衡', 'コウ', ''),
    ('0647', '鋼', 'コウ、はがね', ''),
    ('0648', '講', 'コウ', ''),
    ('0649', '購', 'コウ', ''),
    ('0650', '乞', 'こ-う', ''),
    ('0651', '号號', 'ゴウ', ''),
    ('0652', '合', 'ゴウ、ガッ、（カッ）、あ-う、あ-わす、あ-わせる', '合点'),
    ('0653', '拷', 'ゴウ', ''),
    ('0654', '剛', 'ゴウ', ''),
    ('0655', '傲', 'ゴウ', ''),
    ('0656', '豪', 'ゴウ', ''),
    ('0657', '克', 'コク', ''),
    ('0658', '告吿', 'コク、つ-げる', ''),
    ('0659', '谷', 'コク、たに', ''),
    ('0660', '刻', 'コク、きざ-む', ''),
    ('0661', '国國', 'コク、くに', ''),
    ('0662', '黒黑', 'コク、くろ、くろ-い', ''),
    ('0663', '穀穀', 'コク', ''),
    ('0664', '酷', 'コク', ''),
    ('0665', '獄', 'ゴク', ''),
    ('0666', '骨', 'コツ、ほね', ''),
    ('0667', '駒', 'こま', ''),
    ('0668', '込', 'こ-む、こ-める', ''),
    ('0669', '頃', 'ころ', ''),
    ('0670', '今', 'コン、キン、いま', '今日、今朝、今年'),
    ('0671', '困', 'コン、こま-る', ''),
    ('0672', '昆', 'コン', '昆布'),
    ('0673', '恨', 'コン、うら-む、うら-めしい', ''),
    ('0674', '根', 'コン、ね', ''),
    ('0675', '婚', 'コン', ''),
    ('0676', '混', 'コン、ま-じる、ま-ざる、ま-ぜる、こ-む', ''),
    ('0677', '痕', 'コン、あと', ''),
    ('0678', '紺', 'コン', '紺屋'),
    ('0679', '魂', 'コン、たましい', ''),
    ('0680', '墾', 'コン', ''),
    ('0681', '懇', 'コン、ねんご-ろ', ''),
    ('0682', '左', 'サ、ひだり', ''),
    ('0683', '佐', 'サ', ''),
    ('0684', '沙', 'サ', ''),
    ('0685', '査', 'サ', ''),
    ('0686', '砂', 'サ、シャ、すな', '砂利'),
    ('0687', '唆', 'サ、そそのか-す', ''),
    ('0688', '差', 'サ、さ-す', '差し支える'),
    ('0689', '詐', 'サ', ''),
    ('0690', '鎖', 'サ、くさり', ''),
    ('0691', '座', 'ザ、すわ-る', ''),
    ('0692', '挫', 'ザ', ''),
    ('0693', '才', 'サイ', ''),
    ('0694', '再', 'サイ、（サ）、ふたた-び', ''),
    ('0695', '災', 'サイ、わざわ-い', ''),
    ('0696', '妻', 'サイ、つま', ''),
    ('0697', '采', 'サイ', ''),
    ('0698', '砕碎', 'サイ、くだ-く、くだ-ける', ''),
    ('0699', '宰', 'サイ', ''),
    ('0700', '栽', 'サイ', ''),
    ('0701', '彩', 'サイ、いろど-る', ''),
    ('0702', '採', 'サイ、と-る', ''),
    ('0703', '済濟', 'サイ、す-む、す-ます', ''),
    ('0704', '祭', 'サイ、まつ-る、まつ-り', ''),
    ('0705', '斎齋', 'サイ', ''),
    ('0706', '細', 'サイ、ほそ-い、ほそ-る、こま-か、こま-かい', ''),
    ('0707', '菜', 'サイ、な', ''),
    ('0708', '最', 'サイ、もっと-も', '最寄り'),
    ('0709', '裁', 'サイ、た-つ、さば-く', ''),
    ('0710', '債', 'サイ', ''),
    ('0711', '催', 'サイ、もよお-す', ''),
    ('0712', '塞', 'サイ、ソク、ふさ-ぐ、ふさ-がる', ''),
    ('0713', '歳歲', 'サイ、（セイ）', '二十歳'),
    ('0714', '載', 'サイ、の-せる、の-る', ''),
    ('0715', '際', 'サイ、きわ', ''),
    ('0716', '埼', '（さい）', '埼玉'),
    ('0717', '在', 'ザイ、あ-る', ''),
    ('0718', '材', 'ザイ', ''),
    ('0719', '剤劑', 'ザイ', ''),
    ('0720', '財', 'ザイ、（サイ）', ''),
    ('0721', '罪', 'ザイ、つみ', ''),
    ('0722', '崎', 'さき', ''),
    ('0723', '作', 'サク、サ、つく-る', ''),
    ('0724', '削', 'サク、けず-る', ''),
    ('0725', '昨', 'サク', '昨日'),
    ('0726', '柵', 'サク', ''),
    ('0727', '索', 'サク', ''),
    ('0728', '策', 'サク', ''),
    ('0729', '酢', 'サク、す', ''),
    ('0730', '搾', 'サク、しぼ-る', ''),
    ('0731', '錯', 'サク', ''),
    ('0732', '咲', 'さ-く', ''),
    ('0733', '冊册', 'サツ、サク', ''),
    ('0734', '札', 'サツ、ふだ', ''),
    ('0735', '刷', 'サツ、す-る', ''),
    ('0736', '刹', 'サツ、セツ', ''),
    ('0737', '拶', 'サツ', ''),
    ('0738', '殺殺', 'サツ、（サイ）、（セツ）、ころ-す', ''),
    ('0739', '察', 'サツ', ''),
    ('0740', '撮', 'サツ、と-る', ''),
    ('0741', '擦', 'サツ、す-る、す-れる', ''),
    ('0742', '雑雜', 'ザツ、ゾウ', '雑魚'),
    ('0743', '皿', 'さら', ''),
    ('0744', '三', 'サン、み、み-つ、みっ-つ', '三味線'),
    ('0745', '山', 'サン、やま', '山車、築山、富山'),
    ('0746', '参參', 'サン、まい-る', ''),
    ('0747', '桟棧', 'サン', '桟敷'),
    ('0748', '蚕蠶', 'サン、かいこ', ''),
    ('0749', '惨慘', 'サン、ザン、みじ-め', ''),
    ('0750', '産產', 'サン、う-む、う-まれる、うぶ', '土産'),
    ('0751', '傘', 'サン、かさ', ''),
    ('0752', '散', 'サン、ち-る、ち-らす、ち-らかす、ち-らかる', ''),
    ('0753', '算', 'サン', ''),
    ('0754', '酸', 'サン、す-い', ''),
    ('0755', '賛贊', 'サン', ''),
    ('0756', '残殘', 'ザン、のこ-る、のこ-す', '名残'),
    ('0757', '斬', 'ザン、き-る', ''),
    ('0758', '暫', 'ザン', ''),
    ('0759', '士', 'シ', '海士、居士、博士'),
    ('0760', '子', 'シ、ス、こ', '迷子、息子'),
    ('0761', '支', 'シ、ささ-える', '差し支える'),
    ('0762', '止', 'シ、と-まる、と-める', '波止場'),
    ('0763', '氏', 'シ、うじ', ''),
    ('0764', '仕', 'シ、（ジ）、つか-える', ''),
    ('0765', '史', 'シ', ''),
    ('0766', '司', 'シ', ''),
    ('0767', '四', 'シ、よ、よ-つ、よっ-つ、よん', ''),
    ('0768', '市', 'シ、いち', ''),
    ('0769', '矢', 'シ、や', ''),
    ('0770', '旨', 'シ、むね', ''),
    ('0771', '死', 'シ、し-ぬ', ''),
    ('0772', '糸絲', 'シ、いと', ''),
    ('0773', '至', 'シ、いた-る', ''),
    ('0774', '伺', 'シ、うかが-う', ''),
    ('0775', '志', 'シ、こころざ-す、こころざし', ''),
    ('0776', '私', 'シ、わたくし、わたし', ''),
    ('0777', '使', 'シ、つか-う', ''),
    ('0778', '刺', 'シ、さ-す、さ-さる', ''),
    ('0779', '始', 'シ、はじ-める、はじ-まる', ''),
    ('0780', '姉', 'シ、あね', '姉さん'),
    ('0781', '枝', 'シ、えだ', ''),
    ('0782', '祉祉', 'シ', ''),
    ('0783', '肢', 'シ', ''),
    ('0784', '姿', 'シ、すがた', ''),
    ('0785', '思', 'シ、おも-う', ''),
    ('0786', '指', 'シ、ゆび、さ-す', ''),
    ('0787', '施', 'シ、セ、ほどこ-す', ''),
    ('0788', '師', 'シ', '師走'),
    ('0789', '恣', 'シ', ''),
    ('0790', '紙', 'シ、かみ', ''),
    ('0791', '脂', 'シ、あぶら', ''),
    ('0792', '視視', 'シ', ''),
    ('0793', '紫', 'シ、むらさき', ''),
    ('0794', '詞', 'シ', '祝詞'),
    ('0795', '歯齒', 'シ、は', ''),
    ('0796', '嗣', 'シ', ''),
    ('0797', '試', 'シ、こころ-みる、ため-す', ''),
    ('0798', '詩', 'シ', '詩歌'),
    ('0799', '資', 'シ', ''),
    ('0800', '飼飼', 'シ、か-う', ''),
    ('0801', '誌', 'シ', ''),
    ('0802', '雌', 'シ、め、めす', ''),
    ('0803', '摯', 'シ', ''),
    ('0804', '賜', 'シ、たまわ-る', ''),
    ('0805', '諮', 'シ、はか-る', ''),
    ('0806', '示', 'ジ、シ、しめ-す', ''),
    ('0807', '字', 'ジ、あざ', '文字'),
    ('0808', '寺', 'ジ、てら', ''),
    ('0809', '次', 'ジ、シ、つ-ぐ、つぎ', ''),
    ('0810', '耳', 'ジ、みみ', ''),
    ('0811', '自', 'ジ、シ、みずか-ら', ''),
    ('0812', '似', 'ジ、に-る', ''),
    ('0813', '児兒', 'ジ、（ニ）', '稚児、鹿児島'),
    ('0814', '事', 'ジ、（ズ）、こと', ''),
    ('0815', '侍', 'ジ、さむらい', ''),
    ('0816', '治', 'ジ、チ、おさ-める、おさ-まる、なお-る、なお-す', ''),
    ('0817', '持', 'ジ、も-つ', ''),
    ('0818', '時', 'ジ、とき', '時雨、時計'),
    ('0819', '滋', 'ジ', '滋賀'),
    ('0820', '慈', 'ジ、いつく-しむ', ''),
    ('0821', '辞辭', 'ジ、や-める', ''),
    ('0822', '磁', 'ジ', ''),
    ('0823', '餌', 'ジ、えさ、え', ''),
    ('0824', '璽', 'ジ', ''),
    ('0825', '鹿', 'しか、（か）', '鹿児島'),
    ('0826', '式', 'シキ', ''),
    ('0827', '識', 'シキ', ''),
    ('0828', '軸', 'ジク', ''),
    ('0829', '七', 'シチ、なな、なな-つ、（なの）', '七夕、七日'),
    ('0830', '𠮟(叱)', 'シツ、しか-る', ''),
    ('0831', '失', 'シツ、うしな-う', ''),
    ('0832', '室', 'シツ、むろ', ''),
    ('0833', '疾', 'シツ', ''),
    ('0834', '執', 'シツ、シュウ、と-る', ''),
    ('0835', '湿濕', 'シツ、しめ-る、しめ-す', ''),
    ('0836', '嫉', 'シツ', ''),
    ('0837', '漆', 'シツ、うるし', ''),
    ('0838', '質', 'シツ、シチ、（チ）', ''),
    ('0839', '実實', 'ジツ、み、みの-る', ''),
    ('0840', '芝', 'しば', '芝生'),
    ('0841', '写寫', 'シャ、うつ-す、うつ-る', ''),
    ('0842', '社社', 'シャ、やしろ', ''),
    ('0843', '車', 'シャ、くるま', '山車'),
    ('0844', '舎舍', 'シャ', '田舎'),
    ('0845', '者者', 'シャ、もの', '猛者'),
    ('0846', '射', 'シャ、い-る', ''),
    ('0847', '捨', 'シャ、す-てる', ''),
    ('0848', '赦', 'シャ', ''),
    ('0849', '斜', 'シャ、なな-め', ''),
    ('0850', '煮煮', 'シャ、に-る、に-える、に-やす', ''),
    ('0851', '遮', 'シャ、さえぎ-る', ''),
    ('0852', '謝', 'シャ、あやま-る', ''),
    ('0853', '邪', 'ジャ', '風邪'),
    ('0854', '蛇', 'ジャ、ダ、へび', ''),
    ('0855', '尺', 'シャク', ''),
    ('0856', '借', 'シャク、か-りる', ''),
    ('0857', '酌', 'シャク、く-む', ''),
    ('0858', '釈釋', 'シャク', ''),
    ('0859', '爵', 'シャク', ''),
    ('0860', '若', 'ジャク、（ニャク）、わか-い、も-しくは', '若人'),
    ('0861', '弱', 'ジャク、よわ-い、よわ-る、よわ-まる、よわ-める', ''),
    ('0862', '寂', 'ジャク、（セキ）、さび、さび-しい、さび-れる', ''),
    ('0863', '手', 'シュ、て、（た）', '上手、手伝う、下手'),
    ('0864', '主', 'シュ、（ス）、ぬし、おも', ''),
    ('0865', '守', 'シュ、（ス）、まも-る、も-り', ''),
    ('0866', '朱', 'シュ', ''),
    ('0867', '取', 'シュ、と-る', '鳥取'),
    ('0868', '狩', 'シュ、か-る、か-り', ''),
    ('0869', '首', 'シュ、くび', ''),
    ('0870', '殊', 'シュ、こと', ''),
    ('0871', '珠', 'シュ', '数珠'),
    ('0872', '酒', 'シュ、さけ、（さか）', 'お神酒'),
    ('0873', '腫', 'シュ、は-れる、は-らす', ''),
    ('0874', '種', 'シュ、たね', ''),
    ('0875', '趣', 'シュ、おもむき', ''),
    ('0876', '寿壽', 'ジュ、ことぶき', ''),
    ('0877', '受', 'ジュ、う-ける、う-かる', ''),
    ('0878', '呪', 'ジュ、のろ-う', ''),
    ('0879', '授', 'ジュ、さず-ける、さず-かる', ''),
    ('0880', '需', 'ジュ', ''),
    ('0881', '儒', 'ジュ', ''),
    ('0882', '樹', 'ジュ', ''),
    ('0883', '収收', 'シュウ、おさ-める、おさ-まる', ''),
    ('0884', '囚', 'シュウ', ''),
    ('0885', '州', 'シュウ、す', ''),
    ('0886', '舟', 'シュウ、ふね、（ふな）', ''),
    ('0887', '秀', 'シュウ、ひい-でる', ''),
    ('0888', '周', 'シュウ、まわ-り', ''),
    ('0889', '宗', 'シュウ、ソウ', ''),
    ('0890', '拾', 'シュウ、ジュウ、ひろ-う', ''),
    ('0891', '秋', 'シュウ、あき', ''),
    ('0892', '臭臭', 'シュウ、くさ-い、にお-う', ''),
    ('0893', '修', 'シュウ、（シュ）、おさ-める、おさ-まる', ''),
    ('0894', '袖', 'シュウ、そで', ''),
    ('0895', '終', 'シュウ、お-わる、お-える', ''),
    ('0896', '羞', 'シュウ', ''),
    ('0897', '習', 'シュウ、なら-う', ''),
    ('0898', '週', 'シュウ', ''),
    ('0899', '就', 'シュウ、（ジュ）、つ-く、つ-ける', ''),
    ('0900', '衆', 'シュウ、（シュ）', ''),
    ('0901', '集', 'シュウ、あつ-まる、あつ-める、つど-う', ''),
    ('0902', '愁', 'シュウ、うれ-える、うれ-い', ''),
    ('0903', '酬', 'シュウ', ''),
    ('0904', '醜', 'シュウ、みにく-い', ''),
    ('0905', '蹴', 'シュウ、け-る', ''),
    ('0906', '襲', 'シュウ、おそ-う', ''),
    ('0907', '十', 'ジュウ、ジッ、とお、と', '十重二十重、二十、二十歳、二十日、十'),
    ('0908', '汁', 'ジュウ、しる', ''),
    ('0909', '充', 'ジュウ、あ-てる', ''),
    ('0910', '住', 'ジュウ、す-む、す-まう', ''),
    ('0911', '柔', 'ジュウ、ニュウ、やわ-らか、やわ-らかい', ''),
    ('0912', '重', 'ジュウ、チョウ、え、おも-い、かさ-ねる、かさ-なる', '十重二十重'),
    ('0913', '従從', 'ジュウ、（ショウ）、（ジュ）、したが-う、したが-える', ''),
    ('0914', '渋澁', 'ジュウ、しぶ、しぶ-い、しぶ-る', ''),
    ('0915', '銃', 'ジュウ', ''),
    ('0916', '獣獸', 'ジュウ、けもの', ''),
    ('0917', '縦縱', 'ジュウ、たて', ''),
    ('0918', '叔', 'シュク', '叔父、叔母'),
    ('0919', '祝祝', 'シュク、（シュウ）、いわ-う', '祝詞'),
    ('0920', '宿', 'シュク、やど、やど-る、やど-す', ''),
    ('0921', '淑', 'シュク', ''),
    ('0922', '粛肅', 'シュク', ''),
    ('0923', '縮', 'シュク、ちぢ-む、ちぢ-まる、ちぢ-める、ちぢ-れる、ちぢ-らす', ''),
    ('0924', '塾', 'ジュク', ''),
    ('0925', '熟', 'ジュク、う-れる', ''),
    ('0926', '出', 'シュツ、（スイ）、で-る、だ-す', ''),
    ('0927', '述', 'ジュツ、の-べる', ''),
    ('0928', '術', 'ジュツ', ''),
    ('0929', '俊', 'シュン', ''),
    ('0930', '春', 'シュン、はる', ''),
    ('0931', '瞬', 'シュン、またた-く', ''),
    ('0932', '旬', 'ジュン、（シュン）', ''),
    ('0933', '巡', 'ジュン、めぐ-る', 'お巡りさん'),
    ('0934', '盾', 'ジュン、たて', ''),
    ('0935', '准', 'ジュン', ''),
    ('0936', '殉', 'ジュン', ''),
    ('0937', '純', 'ジュン', ''),
    ('0938', '循', 'ジュン', ''),
    ('0939', '順', 'ジュン', ''),
    ('0940', '準', 'ジュン', ''),
    ('0941', '潤', 'ジュン、うるお-う、うるお-す、うる-む', ''),
    ('0942', '遵', 'ジュン', ''),
    ('0943', '処處', 'ショ', ''),
    ('0944', '初', 'ショ、はじ-め、はじ-めて、はつ、うい、そ-める', ''),
    ('0945', '所', 'ショ、ところ', ''),
    ('0946', '書', 'ショ、か-く', ''),
    ('0947', '庶', 'ショ', ''),
    ('0948', '暑暑', 'ショ、あつ-い', ''),
    ('0949', '署署', 'ショ', ''),
    ('0950', '緒緖', 'ショ、（チョ）、お', ''),
    ('0951', '諸諸', 'ショ', ''),
    ('0952', '女', 'ジョ、ニョ、（ニョウ）、おんな、め', '海女、乙女、早乙女'),
    ('0953', '如', 'ジョ、ニョ', ''),
    ('0954', '助', 'ジョ、たす-ける、たす-かる、すけ', ''),
    ('0955', '序', 'ジョ', ''),
    ('0956', '叙敍', 'ジョ', ''),
    ('0957', '徐', 'ジョ', ''),
    ('0958', '除', 'ジョ、（ジ）、のぞ-く', ''),
    ('0959', '小', 'ショウ、ちい-さい、こ、お', '小豆'),
    ('0960', '升', 'ショウ、ます', ''),
    ('0961', '少', 'ショウ、すく-ない、すこ-し', ''),
    ('0962', '召', 'ショウ、め-す', ''),
    ('0963', '匠', 'ショウ', ''),
    ('0964', '床', 'ショウ、とこ、ゆか', ''),
    ('0965', '抄', 'ショウ', ''),
    ('0966', '肖', 'ショウ', ''),
    ('0967', '尚尙', 'ショウ', ''),
    ('0968', '招', 'ショウ、まね-く', ''),
    ('0969', '承', 'ショウ、うけたまわ-る', ''),
    ('0970', '昇', 'ショウ、のぼ-る', ''),
    ('0971', '松', 'ショウ、まつ', ''),
    ('0972', '沼', 'ショウ、ぬま', ''),
    ('0973', '昭', 'ショウ', ''),
    ('0974', '宵', 'ショウ、よい', ''),
    ('0975', '将將', 'ショウ', ''),
    ('0976', '消', 'ショウ、き-える、け-す', ''),
    ('0977', '症', 'ショウ', ''),
    ('0978', '祥祥', 'ショウ', ''),
    ('0979', '称稱', 'ショウ', ''),
    ('0980', '笑', 'ショウ、わら-う、え-む', '笑顔'),
    ('0981', '唱', 'ショウ、とな-える', ''),
    ('0982', '商', 'ショウ、あきな-う', ''),
    ('0983', '渉涉', 'ショウ', ''),
    ('0984', '章', 'ショウ', ''),
    ('0985', '紹', 'ショウ', ''),
    ('0986', '訟', 'ショウ', ''),
    ('0987', '勝', 'ショウ、か-つ、まさ-る', ''),
    ('0988', '掌', 'ショウ', ''),
    ('0989', '晶', 'ショウ', ''),
    ('0990', '焼燒', 'ショウ、や-く、や-ける', ''),
    ('0991', '焦', 'ショウ、こ-げる、こ-がす、こ-がれる、あせ-る', ''),
    ('0992', '硝', 'ショウ', ''),
    ('0993', '粧', 'ショウ', ''),
    ('0994', '詔', 'ショウ、みことのり', ''),
    ('0995', '証證', 'ショウ', ''),
    ('0996', '象', 'ショウ、ゾウ', ''),
    ('0997', '傷', 'ショウ、きず、いた-む、いた-める', ''),
    ('0998', '奨奬', 'ショウ', ''),
    ('0999', '照', 'ショウ、て-る、て-らす、て-れる', ''),
    ('1000', '詳', 'ショウ、くわ-しい', ''),
    ('1001', '彰', 'ショウ', ''),
    ('1002', '障', 'ショウ、さわ-る', ''),
    ('1003', '憧', 'ショウ、あこが-れる', '憧憬'),
    ('1004', '衝', 'ショウ', ''),
    ('1005', '賞', 'ショウ', ''),
    ('1006', '償', 'ショウ、つぐな-う', ''),
    ('1007', '礁', 'ショウ', ''),
    ('1008', '鐘', 'ショウ、かね', ''),
    ('1009', '上', 'ジョウ、（ショウ）、うえ、（うわ）、かみ、あ-げる、あ-がる、のぼ-る、のぼ-せる、のぼ-す', '上手'),
    ('1010', '丈', 'ジョウ、たけ', ''),
    ('1011', '冗', 'ジョウ', ''),
    ('1012', '条條', 'ジョウ', ''),
    ('1013', '状狀', 'ジョウ', ''),
    ('1014', '乗乘', 'ジョウ、の-る、の-せる', ''),
    ('1015', '城', 'ジョウ、しろ', '茨城、宮城'),
    ('1016', '浄淨', 'ジョウ', ''),
    ('1017', '剰剩', 'ジョウ', ''),
    ('1018', '常', 'ジョウ、つね、とこ', ''),
    ('1019', '情', 'ジョウ、（セイ）、なさ-け', ''),
    ('1020', '場', 'ジョウ、ば', '波止場'),
    ('1021', '畳疊', 'ジョウ、たた-む、たたみ', ''),
    ('1022', '蒸', 'ジョウ、む-す、む-れる、む-らす', ''),
    ('1023', '縄繩', 'ジョウ、なわ', ''),
    ('1024', '壌壤', 'ジョウ', ''),
    ('1025', '嬢孃', 'ジョウ', ''),
    ('1026', '錠', 'ジョウ', ''),
    ('1027', '譲讓', 'ジョウ、ゆず-る', ''),
    ('1028', '醸釀', 'ジョウ、かも-す', ''),
    ('1029', '色', 'ショク、シキ、いろ', '景色'),
    ('1030', '拭', 'ショク、ふ-く、ぬぐ-う', ''),
    ('1031', '食', 'ショク、（ジキ）、く-う、く-らう、た-べる', ''),
    ('1032', '植', 'ショク、う-える、う-わる', ''),
    ('1033', '殖', 'ショク、ふ-える、ふ-やす', ''),
    ('1034', '飾', 'ショク、かざ-る', ''),
    ('1035', '触觸', 'ショク、ふ-れる、さわ-る', ''),
    ('1036', '嘱囑', 'ショク', ''),
    ('1037', '織', 'ショク、シキ、お-る', ''),
    ('1038', '職', 'ショク', ''),
    ('1039', '辱', 'ジョク、はずかし-める', ''),
    ('1040', '尻', 'しり', '尻尾'),
    ('1041', '心', 'シン、こころ', '心地'),
    ('1042', '申', 'シン、もう-す', ''),
    ('1043', '伸', 'シン、の-びる、の-ばす、の-べる', ''),
    ('1044', '臣', 'シン、ジン', ''),
    ('1045', '芯', 'シン', ''),
    ('1046', '身', 'シン、み', ''),
    ('1047', '辛', 'シン、から-い', ''),
    ('1048', '侵', 'シン、おか-す', ''),
    ('1049', '信', 'シン', ''),
    ('1050', '津', 'シン、つ', ''),
    ('1051', '神神', 'シン、ジン、かみ、（かん）、（こう）', 'お神酒、神楽、神奈川'),
    ('1052', '唇', 'シン、くちびる', ''),
    ('1053', '娠', 'シン', ''),
    ('1054', '振', 'シン、ふ-る、ふ-るう、ふ-れる', ''),
    ('1055', '浸', 'シン、ひた-す、ひた-る', ''),
    ('1056', '真眞', 'シン、ま', '真面目、真っ赤、真っ青'),
    ('1057', '針', 'シン、はり', ''),
    ('1058', '深', 'シン、ふか-い、ふか-まる、ふか-める', ''),
    ('1059', '紳', 'シン', ''),
    ('1060', '進', 'シン、すす-む、すす-める', ''),
    ('1061', '森', 'シン、もり', ''),
    ('1062', '診', 'シン、み-る', ''),
    ('1063', '寝寢', 'シン、ね-る、ね-かす', ''),
    ('1064', '慎愼', 'シン、つつし-む', ''),
    ('1065', '新', 'シン、あたら-しい、あら-た、にい', ''),
    ('1066', '審', 'シン', ''),
    ('1067', '震', 'シン、ふる-う、ふる-える', ''),
    ('1068', '薪', 'シン、たきぎ', ''),
    ('1069', '親', 'シン、おや、した-しい、した-しむ', ''),
    ('1070', '人', 'ジン、ニン、ひと', '大人、玄人、素人、仲人、一人、二人、若人'),
    ('1071', '刃', 'ジン、は', ''),
    ('1072', '仁', 'ジン、（ニ）', ''),
    ('1073', '尽盡', 'ジン、つ-くす、つ-きる、つ-かす', ''),
    ('1074', '迅', 'ジン', ''),
    ('1075', '甚', 'ジン、はなは-だ、はなは-だしい', ''),
    ('1076', '陣', 'ジン', ''),
    ('1077', '尋', 'ジン、たず-ねる', ''),
    ('1078', '腎', 'ジン', ''),
    ('1079', '須', 'ス', ''),
    ('1080', '図圖', 'ズ、ト、はか-る', ''),
    ('1081', '水', 'スイ、みず', '清水'),
    ('1082', '吹', 'スイ、ふ-く', '息吹、吹雪'),
    ('1083', '垂', 'スイ、た-れる、た-らす', ''),
    ('1084', '炊', 'スイ、た-く', ''),
    ('1085', '帥', 'スイ', ''),
    ('1086', '粋粹', 'スイ、いき', ''),
    ('1087', '衰', 'スイ、おとろ-える', ''),
    ('1088', '推', 'スイ、お-す', ''),
    ('1089', '酔醉', 'スイ、よ-う', ''),
    ('1090', '遂', 'スイ、と-げる', ''),
    ('1091', '睡', 'スイ', ''),
    ('1092', '穂穗', 'スイ、ほ', ''),
    ('1093', '随隨', 'ズイ', ''),
    ('1094', '髄髓', 'ズイ', ''),
    ('1095', '枢樞', 'スウ', ''),
    ('1096', '崇', 'スウ', ''),
    ('1097', '数數', 'スウ、（ス）、かず、かぞ-える', '数珠、数寄屋、数奇屋'),
    ('1098', '据', 'す-える、す-わる', ''),
    ('1099', '杉', 'すぎ', ''),
    ('1100', '裾', 'すそ', ''),
    ('1101', '寸', 'スン', ''),
    ('1102', '瀬瀨', 'せ', ''),
    ('1103', '是', 'ゼ', ''),
    ('1104', '井', 'セイ、（ショウ）、い', ''),
    ('1105', '世', 'セイ、セ、よ', ''),
    ('1106', '正', 'セイ、ショウ、ただ-しい、ただ-す、まさ', ''),
    ('1107', '生',
     'セイ、ショウ、い-きる、い-かす、い-ける、う-まれる、う-む、お-う、は-える、は-やす、き、なま',
     '芝生、弥生'),
    ('1108', '成', 'セイ、（ジョウ）、な-る、な-す', ''),
    ('1109', '西', 'セイ、サイ、にし', ''),
    ('1110', '声聲', 'セイ、（ショウ）、こえ、（こわ）', ''),
    ('1111', '制', 'セイ', ''),
    ('1112', '姓', 'セイ、ショウ', ''),
    ('1113', '征', 'セイ', ''),
    ('1114', '性', 'セイ、ショウ', ''),
    ('1115', '青靑', 'セイ、（ショウ）、あお、あお-い', '真っ青'),
    ('1116', '斉齊', 'セイ', ''),
    ('1117', '政', 'セイ、（ショウ）、まつりごと', ''),
    ('1118', '星', 'セイ、（ショウ）、ほし', ''),
    ('1119', '牲', 'セイ', ''),
    ('1120', '省', 'セイ、ショウ、かえり-みる、はぶ-く', ''),
    ('1121', '凄', 'セイ', ''),
    ('1122', '逝', 'セイ、ゆ-く、い-く', ''),
    ('1123', '清淸', 'セイ、（ショウ）、きよ-い、きよ-まる、きよ-める', '清水'),
    ('1124', '盛', 'セイ、（ジョウ）、も-る、さか-る、さか-ん', ''),
    ('1125', '婿', 'セイ、むこ', ''),
    ('1126', '晴晴', 'セイ、は-れる、は-らす', ''),
    ('1127', '勢', 'セイ、いきお-い', ''),
    ('1128', '聖', 'セイ', ''),
    ('1129', '誠', 'セイ、まこと', ''),
    ('1130', '精精', 'セイ、（ショウ）', ''),
    ('1131', '製', 'セイ', ''),
    ('1132', '誓', 'セイ、ちか-う', ''),
    ('1133', '静靜', 'セイ、（ジョウ）、しず、しず-か、しず-まる、しず-める', ''),
    ('1134', '請', 'セイ、（シン）、こ-う、う-ける', ''),
    ('1135', '整', 'セイ、ととの-える、ととの-う', ''),
    ('1136', '醒', 'セイ', ''),
    ('1137', '税稅', 'ゼイ', ''),
    ('1138', '夕', 'セキ、ゆう', '七夕'),
    ('1139', '斥', 'セキ', ''),
    ('1140', '石', 'セキ、（シャク）、（コク）、いし', ''),
    ('1141', '赤', 'セキ、（シャク）、あか、あか-い、あか-らむ、あか-らめる', '真っ赤'),
    ('1142', '昔', 'セキ、（シャク）、むかし', ''),
    ('1143', '析', 'セキ', ''),
    ('1144', '席', 'セキ', '寄席'),
    ('1145', '脊', 'セキ', ''),
    ('1146', '隻', 'セキ', ''),
    ('1147', '惜', 'セキ、お-しい、お-しむ', ''),
    ('1148', '戚', 'セキ', ''),
    ('1149', '責', 'セキ、せ-める', ''),
    ('1150', '跡', 'セキ、あと', ''),
    ('1151', '積', 'セキ、つ-む、つ-もる', ''),
    ('1152', '績', 'セキ', ''),
    ('1153', '籍', 'セキ', ''),
    ('1154', '切', 'セツ、（サイ）、き-る、き-れる', ''),
    ('1155', '折', 'セツ、お-る、おり、お-れる', ''),
    ('1156', '拙', 'セツ、つたな-い', ''),
    ('1157', '窃竊', 'セツ', ''),
    ('1158', '接', 'セツ、つ-ぐ', ''),
    ('1159', '設', 'セツ、もう-ける', ''),
    ('1160', '雪', 'セツ、ゆき', '雪崩、吹雪'),
    ('1161', '摂攝', 'セツ', ''),
    ('1162', '節節', 'セツ、（セチ）、ふし', ''),
    ('1163', '説說', 'セツ、（ゼイ）、と-く', ''),
    ('1164', '舌', 'ゼツ、した', ''),
    ('1165', '絶絕', 'ゼツ、た-える、た-やす、た-つ', ''),
    ('1166', '千', 'セン、ち', ''),
    ('1167', '川', 'セン、かわ', '川原、神奈川'),
    ('1168', '仙', 'セン', ''),
    ('1169', '占', 'セン、し-める、うらな-う', ''),
    ('1170', '先', 'セン、さき', ''),
    ('1171', '宣', 'セン', ''),
    ('1172', '専專', 'セン、もっぱ-ら', ''),
    ('1173', '泉', 'セン、いずみ', ''),
    ('1174', '浅淺', 'セン、あさ-い', ''),
    ('1175', '洗', 'セン、あら-う', ''),
    ('1176', '染', 'セン、そ-める、そ-まる、し-みる、し-み', ''),
    ('1177', '扇', 'セン、おうぎ', ''),
    ('1178', '栓', 'セン', ''),
    ('1179', '旋', 'セン', ''),
    ('1180', '船', 'セン、ふね、（ふな）', '伝馬船'),
    ('1181', '戦戰', 'セン、いくさ、たたか-う', ''),
    ('1182', '煎', 'セン、い-る', ''),
    ('1183', '羨', 'セン、うらや-む、うらや-ましい', ''),
    ('1184', '腺', 'セン', ''),
    ('1185', '詮', 'セン', ''),
    ('1186', '践踐', 'セン', ''),
    ('1187', '箋', 'セン', ''),
    ('1188', '銭錢', 'セン、ぜに', ''),
    ('1189', '潜潛', 'セン、ひそ-む、もぐ-る', ''),
    ('1190', '線', 'セン', '三味線'),
    ('1191', '遷', 'セン', ''),
    ('1192', '選', 'セン、えら-ぶ', ''),
    ('1193', '薦', 'セン、すす-める', ''),
    ('1194', '繊纖', 'セン', ''),
    ('1195', '鮮', 'セン、あざ-やか', ''),
    ('1196', '全', 'ゼン、まった-く、すべ-て', ''),
    ('1197', '前', 'ゼン、まえ', ''),
    ('1198', '善', 'ゼン、よ-い', ''),
    ('1199', '然', 'ゼン、ネン', ''),
    ('1200', '禅禪', 'ゼン', ''),
    ('1201', '漸', 'ゼン', ''),
    ('1202', '膳', 'ゼン', ''),
    ('1203', '繕', 'ゼン、つくろ-う', ''),
    ('1204', '狙', 'ソ、ねら-う', ''),
    ('1205', '阻', 'ソ、はば-む', ''),
    ('1206', '祖祖', 'ソ', ''),
    ('1207', '租', 'ソ', ''),
    ('1208', '素', 'ソ、ス', '素人'),
    ('1209', '措', 'ソ', ''),
    ('1210', '粗', 'ソ、あら-い', ''),
    ('1211', '組', 'ソ、く-む、くみ', ''),
    ('1212', '疎', 'ソ、うと-い、うと-む', ''),
    ('1213', '訴', 'ソ、うった-える', ''),
    ('1214', '塑', 'ソ', ''),
    ('1215', '遡', 'ソ、さかのぼ-る', ''),
    ('1216', '礎', 'ソ、いしずえ', ''),
    ('1217', '双雙', 'ソウ、ふた', ''),
    ('1218', '壮壯', 'ソウ', ''),
    ('1219', '早', 'ソウ、（サッ）、はや-い、はや-まる、はや-める', '早乙女、早苗'),
    ('1220', '争爭', 'ソウ、あらそ-う', ''),
    ('1221', '走', 'ソウ、はし-る', '師走'),
    ('1222', '奏', 'ソウ、かな-でる', ''),
    ('1223', '相', 'ソウ、ショウ、あい', '相撲'),
    ('1224', '荘莊', 'ソウ', ''),
    ('1225', '草', 'ソウ、くさ', '草履'),
    ('1226', '送', 'ソウ、おく-る', ''),
    ('1227', '倉', 'ソウ、くら', ''),
    ('1228', '捜搜', 'ソウ、さが-す', ''),
    ('1229', '挿揷插', 'ソウ、さ-す', ''),  # "插"は康熙字典体
    ('1230', '桑', 'ソウ、くわ', ''),
    ('1231', '巣巢', 'ソウ、す', ''),
    ('1232', '掃', 'ソウ、は-く', ''),
    ('1233', '曹', 'ソウ', ''),
    ('1234', '曽曾', 'ソウ、（ゾ）', ''),
    ('1235', '爽', 'ソウ、さわ-やか', ''),
    ('1236', '窓', 'ソウ、まど', ''),
    ('1237', '創', 'ソウ、つく-る', ''),
    ('1238', '喪', 'ソウ、も', ''),
    ('1239', '痩瘦', 'ソウ、や-せる', ''),
    ('1240', '葬', 'ソウ、ほうむ-る', ''),
    ('1241', '装裝', 'ソウ、ショウ、よそお-う', ''),
    ('1242', '僧僧', 'ソウ', ''),
    ('1243', '想', 'ソウ、（ソ）', ''),
    ('1244', '層層', 'ソウ', ''),
    ('1245', '総總', 'ソウ', ''),
    ('1246', '遭', 'ソウ、あ-う', ''),
    ('1247', '槽', 'ソウ', ''),
    ('1248', '踪', 'ソウ', ''),
    ('1249', '操', 'ソウ、みさお、あやつ-る', ''),
    ('1250', '燥', 'ソウ', ''),
    ('1251', '霜', 'ソウ、しも', ''),
    ('1252', '騒騷', 'ソウ、さわ-ぐ', ''),
    ('1253', '藻', 'ソウ、も', ''),
    ('1254', '造', 'ゾウ、つく-る', ''),
    ('1255', '像', 'ゾウ', ''),
    ('1256', '増增', 'ゾウ、ま-す、ふ-える、ふ-やす', ''),
    ('1257', '憎憎', 'ゾウ、にく-む、にく-い、にく-らしい、にく-しみ', ''),
    ('1258', '蔵藏', 'ゾウ、くら', ''),
    ('1259', '贈贈', 'ゾウ、（ソウ）、おく-る', ''),
    ('1260', '臓臟', 'ゾウ', ''),
    ('1261', '即卽', 'ソク', ''),
    ('1262', '束', 'ソク、たば', ''),
    ('1263', '足', 'ソク、あし、た-りる、た-る、た-す', '足袋'),
    ('1264', '促', 'ソク、うなが-す', ''),
    ('1265', '則', 'ソク', ''),
    ('1266', '息', 'ソク、いき', '息吹、息子'),
    ('1267', '捉', 'ソク、とら-える', ''),
    ('1268', '速', 'ソク、はや-い、はや-める、はや-まる、すみ-やか', ''),
    ('1269', '側', 'ソク、がわ', '側'),
    ('1270', '測', 'ソク、はか-る', ''),
    ('1271', '俗', 'ゾク', ''),
    ('1272', '族', 'ゾク', ''),
    ('1273', '属屬', 'ゾク', ''),
    ('1274', '賊', 'ゾク', ''),
    ('1275', '続續', 'ゾク、つづ-く、つづ-ける', ''),
    ('1276', '卒', 'ソツ', ''),
    ('1277', '率', 'ソツ、リツ、ひき-いる', ''),
    ('1278', '存', 'ソン、ゾン', ''),
    ('1279', '村', 'ソン、むら', ''),
    ('1280', '孫', 'ソン、まご', ''),
    ('1281', '尊', 'ソン、たっと-い、とうと-い、たっと-ぶ、とうと-ぶ', ''),
    ('1282', '損', 'ソン、そこ-なう、そこ-ねる', ''),
    ('1283', '遜', 'ソン', ''),
    ('1284', '他', 'タ、ほか', ''),
    ('1285', '多', 'タ、おお-い', ''),
    ('1286', '汰', 'タ', ''),
    ('1287', '打', 'ダ、う-つ', ''),
    ('1288', '妥', 'ダ', ''),
    ('1289', '唾', 'ダ、つば', '固唾、唾'),
    ('1290', '堕墮', 'ダ', ''),
    ('1291', '惰', 'ダ', ''),
    ('1292', '駄', 'ダ', ''),
    ('1293', '太', 'タイ、タ、ふと-い、ふと-る', '太刀'),
    ('1294', '対對', 'タイ、ツイ', ''),
    ('1295', '体體', 'タイ、テイ、からだ', ''),
    ('1296', '耐', 'タイ、た-える', ''),
    ('1297', '待', 'タイ、ま-つ', ''),
    ('1298', '怠', 'タイ、おこた-る、なま-ける', ''),
    ('1299', '胎', 'タイ', ''),
    ('1300', '退', 'タイ、しりぞ-く、しりぞ-ける', '立ち退く'),
    ('1301', '帯帶', 'タイ、お-びる、おび', ''),
    ('1302', '泰', 'タイ', ''),
    ('1303', '堆', 'タイ', ''),
    ('1304', '袋', 'タイ、ふくろ', '足袋'),
    ('1305', '逮', 'タイ', ''),
    ('1306', '替', 'タイ、か-える、か-わる', '為替'),
    ('1307', '貸', 'タイ、か-す', ''),
    ('1308', '隊', 'タイ', ''),
    ('1309', '滞滯', 'タイ、とどこお-る', ''),
    ('1310', '態', 'タイ', ''),
    ('1311', '戴', 'タイ', ''),
    ('1312', '大', 'ダイ、タイ、おお、おお-きい、おお-いに', '大人、大和、大阪、大分'),
    ('1313', '代', 'ダイ、タイ、か-わる、か-える、よ、しろ', ''),
    ('1314', '台臺', 'ダイ、タイ', ''),
    ('1315', '第', 'ダイ', ''),
    ('1316', '題', 'ダイ', ''),
    ('1317', '滝瀧', 'たき', ''),
    ('1318', '宅', 'タク', ''),
    ('1319', '択擇', 'タク', ''),
    ('1320', '沢澤', 'タク、さわ', ''),
    ('1321', '卓', 'タク', ''),
    ('1322', '拓', 'タク', ''),
    ('1323', '託', 'タク', ''),
    ('1324', '濯', 'タク', ''),
    ('1325', '諾', 'ダク', ''),
    ('1326', '濁', 'ダク、にご-る、にご-す', ''),
    ('1327', '但', 'ただ-し', ''),
    ('1328', '達', 'タツ', '友達'),
    ('1329', '脱脫', 'ダツ、ぬ-ぐ、ぬ-げる', ''),
    ('1330', '奪', 'ダツ、うば-う', ''),
    ('1331', '棚', 'たな', ''),
    ('1332', '誰', 'だれ', ''),
    ('1333', '丹', 'タン', ''),
    ('1334', '旦', 'タン、ダン', ''),
    ('1335', '担擔', 'タン、かつ-ぐ、にな-う', ''),
    ('1336', '単單', 'タン', ''),
    ('1337', '炭', 'タン、すみ', ''),
    ('1338', '胆膽', 'タン', ''),
    ('1339', '探', 'タン、さぐ-る、さが-す', ''),
    ('1340', '淡', 'タン、あわ-い', ''),
    ('1341', '短', 'タン、みじか-い', ''),
    ('1342', '嘆嘆', 'タン、なげ-く、なげ-かわしい', ''),
    ('1343', '端', 'タン、はし、は、はた', ''),
    ('1344', '綻', 'タン、ほころ-びる', ''),
    ('1345', '誕', 'タン', ''),
    ('1346', '鍛', 'タン、きた-える', '鍛冶'),
    ('1347', '団團', 'ダン、（トン）', ''),
    ('1348', '男', 'ダン、ナン、おとこ', ''),
    ('1349', '段', 'ダン', ''),
    ('1350', '断斷', 'ダン、た-つ、ことわ-る', ''),
    ('1351', '弾彈', 'ダン、ひ-く、はず-む、たま', ''),
    ('1352', '暖', 'ダン、あたた-か、あたた-かい、あたた-まる、あたた-める', ''),
    ('1353', '談', 'ダン', ''),
    ('1354', '壇', 'ダン、（タン）', ''),
    ('1355', '地', 'チ、ジ', '意気地、心地'),
    ('1356', '池', 'チ、いけ', ''),
    ('1357', '知', 'チ、し-る', ''),
    ('1358', '値', 'チ、ね、あたい', ''),
    ('1359', '恥', 'チ、は-じる、はじ、は-じらう、は-ずかしい', ''),
    ('1360', '致', 'チ、いた-す', ''),
    ('1361', '遅遲', 'チ、おく-れる、おく-らす、おそ-い', ''),
    ('1362', '痴癡', 'チ', ''),  # "癡"は康熙字典体
    ('1363', '稚', 'チ', '稚児'),
    ('1364', '置', 'チ、お-く', ''),
    ('1365', '緻', 'チ', ''),
    ('1366', '竹', 'チク、たけ', '竹刀'),
    ('1367', '畜', 'チク', ''),
    ('1368', '逐', 'チク', ''),
    ('1369', '蓄', 'チク、たくわ-える', ''),
    ('1370', '築', 'チク、きず-く', '築山'),
    ('1371', '秩', 'チツ', ''),
    ('1372', '窒', 'チツ', ''),
    ('1373', '茶', 'チャ、サ', ''),
    ('1374', '着', 'チャク、（ジャク）、き-る、き-せる、つ-く、つ-ける', ''),
    ('1375', '嫡', 'チャク', ''),
    ('1376', '中', 'チュウ、（ジュウ）、なか', ''),
    ('1377', '仲', 'チュウ、なか', '仲人'),
    ('1378', '虫蟲', 'チュウ、むし', ''),
    ('1379', '沖', 'チュウ、おき', ''),
    ('1380', '宙', 'チュウ', ''),
    ('1381', '忠', 'チュウ', ''),
    ('1382', '抽', 'チュウ', ''),
    ('1383', '注', 'チュウ、そそ-ぐ', ''),
    ('1384', '昼晝', 'チュウ、ひる', ''),
    ('1385', '柱', 'チュウ、はしら', ''),
    ('1386', '衷', 'チュウ', ''),
    ('1387', '酎', 'チュウ', ''),
    ('1388', '鋳鑄', 'チュウ、い-る', ''),
    ('1389', '駐', 'チュウ', ''),
    ('1390', '著著', 'チョ、あらわ-す、いちじる-しい', ''),
    ('1391', '貯', 'チョ', ''),
    ('1392', '丁', 'チョウ、テイ', ''),
    ('1393', '弔', 'チョウ、とむら-う', ''),
    ('1394', '庁廳', 'チョウ', ''),
    ('1395', '兆', 'チョウ、きざ-す、きざ-し', ''),
    ('1396', '町', 'チョウ、まち', ''),
    ('1397', '長', 'チョウ、なが-い', '八百長'),
    ('1398', '挑', 'チョウ、いど-む', ''),
    ('1399', '帳', 'チョウ', '蚊帳'),
    ('1400', '張', 'チョウ、は-る', ''),
    ('1401', '彫', 'チョウ、ほ-る', ''),
    ('1402', '眺', 'チョウ、なが-める', ''),
    ('1403', '釣', 'チョウ、つ-る', ''),
    ('1404', '頂', 'チョウ、いただ-く、いただき', ''),
    ('1405', '鳥', 'チョウ、とり', '鳥取'),
    ('1406', '朝', 'チョウ、あさ', '今朝'),
    ('1407', '貼', 'チョウ、は-る', '貼付'),
    ('1408', '超', 'チョウ、こ-える、こ-す', ''),
    ('1409', '腸', 'チョウ', ''),
    ('1410', '跳', 'チョウ、は-ねる、と-ぶ', ''),
    ('1411', '徴徵', 'チョウ', ''),
    ('1412', '嘲', 'チョウ、あざけ-る', ''),
    ('1413', '潮', 'チョウ、しお', ''),
    ('1414', '澄', 'チョウ、す-む、す-ます', ''),
    ('1415', '調', 'チョウ、しら-べる、ととの-う、ととの-える', ''),
    ('1416', '聴聽', 'チョウ、き-く', ''),
    ('1417', '懲懲', 'チョウ、こ-りる、こ-らす、こ-らしめる', ''),
    ('1418', '直', 'チョク、ジキ、ただ-ちに、なお-す、なお-る', ''),
    ('1419', '勅敕', 'チョク', ''),  # "敕"は康熙字典体
    ('1420', '捗', 'チョク', ''),
    ('1421', '沈', 'チン、しず-む、しず-める', ''),
    ('1422', '珍', 'チン、めずら-しい', ''),
    ('1423', '朕', 'チン', ''),
    ('1424', '陳', 'チン', ''),
    ('1425', '賃', 'チン', ''),
    ('1426', '鎮鎭', 'チン、しず-める、しず-まる', ''),
    ('1427', '追', 'ツイ、お-う', ''),
    ('1428', '椎', 'ツイ', ''),
    ('1429', '墜', 'ツイ', ''),
    ('1430', '通', 'ツウ、（ツ）、とお-る、とお-す、かよ-う', ''),
    ('1431', '痛', 'ツウ、いた-い、いた-む、いた-める', ''),
    ('1432', '塚塚', 'つか', ''),
    ('1433', '漬', 'つ-ける、つ-かる', ''),
    ('1434', '坪', 'つぼ', ''),
    ('1435', '爪', 'つめ、（つま）', ''),
    ('1436', '鶴', 'つる', ''),
    ('1437', '低', 'テイ、ひく-い、ひく-める、ひく-まる', ''),
    ('1438', '呈', 'テイ', ''),
    ('1439', '廷', 'テイ', ''),
    ('1440', '弟', 'テイ、（ダイ）、（デ）、おとうと', ''),
    ('1441', '定', 'テイ、ジョウ、さだ-める、さだ-まる、さだ-か', ''),
    ('1442', '底', 'テイ、そこ', ''),
    ('1443', '抵', 'テイ', ''),
    ('1444', '邸', 'テイ', ''),
    ('1445', '亭', 'テイ', ''),
    ('1446', '貞', 'テイ', ''),
    ('1447', '帝', 'テイ', ''),
    ('1448', '訂', 'テイ', ''),
    ('1449', '庭', 'テイ、にわ', ''),
    ('1450', '逓遞', 'テイ', ''),
    ('1451', '停', 'テイ', ''),
    ('1452', '偵', 'テイ', ''),
    ('1453', '堤', 'テイ、つつみ', ''),
    ('1454', '提', 'テイ、さ-げる', ''),
    ('1455', '程', 'テイ、ほど', ''),
    ('1456', '艇', 'テイ', ''),
    ('1457', '締', 'テイ、し-まる、し-める', ''),
    ('1458', '諦', 'テイ、あきら-める', ''),
    ('1459', '泥', 'デイ、どろ', ''),
    ('1460', '的', 'テキ、まと', ''),
    ('1461', '笛', 'テキ、ふえ', ''),
    ('1462', '摘', 'テキ、つ-む', ''),
    ('1463', '滴', 'テキ、しずく、したた-る', ''),
    ('1464', '適', 'テキ', ''),
    ('1465', '敵', 'テキ、かたき', ''),
    ('1466', '溺', 'デキ、おぼ-れる', ''),
    ('1467', '迭', 'テツ', ''),
    ('1468', '哲', 'テツ', ''),
    ('1469', '鉄鐵', 'テツ', ''),
    ('1470', '徹', 'テツ', ''),
    ('1471', '撤', 'テツ', ''),
    ('1472', '天', 'テン、あめ、（あま）', ''),
    ('1473', '典', 'テン', ''),
    ('1474', '店', 'テン、みせ', ''),
    ('1475', '点點', 'テン', '合点'),
    ('1476', '展', 'テン', ''),
    ('1477', '添', 'テン、そ-える、そ-う', ''),
    ('1478', '転轉', 'テン、ころ-がる、ころ-げる、ころ-がす、ころ-ぶ', ''),
    ('1479', '塡(填)', 'テン', ''),
    ('1480', '田', 'デン、た', '田舎'),
    ('1481', '伝傳', 'デン、つた-わる、つた-える、つた-う', '手伝う、伝馬船'),
    ('1482', '殿', 'デン、テン、との、どの', ''),
    ('1483', '電', 'デン', ''),
    ('1484', '斗', 'ト', ''),
    ('1485', '吐', 'ト、は-く', ''),
    ('1486', '妬', 'ト、ねた-む', ''),
    ('1487', '徒', 'ト', ''),
    ('1488', '途', 'ト', ''),
    ('1489', '都都', 'ト、ツ、みやこ', ''),
    ('1490', '渡', 'ト、わた-る、わた-す', ''),
    ('1491', '塗', 'ト、ぬ-る', ''),
    ('1492', '賭', 'ト、か-ける', ''),
    ('1493', '土', 'ド、ト、つち', '土産'),
    ('1494', '奴', 'ド', ''),
    ('1495', '努', 'ド、つと-める', ''),
    ('1496', '度', 'ド、（ト）、（タク）、たび', ''),
    ('1497', '怒', 'ド、いか-る、おこ-る', ''),
    ('1498', '刀', 'トウ、かたな', '竹刀、太刀'),
    ('1499', '冬', 'トウ、ふゆ', ''),
    ('1500', '灯燈', 'トウ、ひ', ''),
    ('1501', '当當', 'トウ、あ-たる、あ-てる', ''),
    ('1502', '投', 'トウ、な-げる', '投網'),
    ('1503', '豆', 'トウ、（ズ）、まめ', '小豆'),
    ('1504', '東', 'トウ、ひがし', ''),
    ('1505', '到', 'トウ', ''),
    ('1506', '逃', 'トウ、に-げる、に-がす、のが-す、のが-れる', ''),
    ('1507', '倒', 'トウ、たお-れる、たお-す', ''),
    ('1508', '凍', 'トウ、こお-る、こご-える', ''),
    ('1509', '唐', 'トウ、から', ''),
    ('1510', '島', 'トウ、しま', '鹿児島'),
    ('1511', '桃', 'トウ、もも', ''),
    ('1512', '討', 'トウ、う-つ', ''),
    ('1513', '透', 'トウ、す-く、す-かす、す-ける', ''),
    ('1514', '党黨', 'トウ', ''),
    ('1515', '悼', 'トウ、いた-む', ''),
    ('1516', '盗盜', 'トウ、ぬす-む', ''),
    ('1517', '陶', 'トウ', ''),
    ('1518', '塔', 'トウ', ''),
    ('1519', '搭', 'トウ', ''),
    ('1520', '棟', 'トウ、むね、（むな）', ''),
    ('1521', '湯', 'トウ、ゆ', ''),
    ('1522', '痘', 'トウ', ''),
    ('1523', '登', 'トウ、ト、のぼ-る', ''),
    ('1524', '答', 'トウ、こた-える、こた-え', ''),
    ('1525', '等', 'トウ、ひと-しい', ''),
    ('1526', '筒', 'トウ、つつ', ''),
    ('1527', '統', 'トウ、す-べる', ''),
    ('1528', '稲稻', 'トウ、いね、（いな）', ''),
    ('1529', '踏', 'トウ、ふ-む、ふ-まえる', ''),
    ('1530', '糖', 'トウ', ''),
    ('1531', '頭', 'トウ、ズ、（ト）、あたま、かしら', ''),
    ('1532', '謄', 'トウ', ''),
    ('1533', '藤', 'トウ、ふじ', ''),
    ('1534', '闘鬪鬭', 'トウ、たたか-う', ''),  # "鬭"は康熙字典体
    ('1535', '騰', 'トウ', ''),
    ('1536', '同', 'ドウ、おな-じ', ''),
    ('1537', '洞', 'ドウ、ほら', ''),
    ('1538', '胴', 'ドウ', ''),
    ('1539', '動', 'ドウ、うご-く、うご-かす', ''),
    ('1540', '堂', 'ドウ', ''),
    ('1541', '童', 'ドウ、わらべ', ''),
    ('1542', '道', 'ドウ、（トウ）、みち', ''),
    ('1543', '働', 'ドウ、はたら-く', ''),
    ('1544', '銅', 'ドウ', ''),
    ('1545', '導', 'ドウ、みちび-く', ''),
    ('1546', '瞳', 'ドウ、ひとみ', ''),
    ('1547', '峠', 'とうげ', ''),
    ('1548', '匿', 'トク', ''),
    ('1549', '特', 'トク', ''),
    ('1550', '得', 'トク、え-る、う-る', ''),
    ('1551', '督', 'トク', ''),
    ('1552', '徳德', 'トク', ''),
    ('1553', '篤', 'トク', ''),
    ('1554', '毒', 'ドク', ''),
    ('1555', '独獨', 'ドク、ひと-り', ''),
    ('1556', '読讀', 'ドク、トク、（トウ）、よ-む', '読経'),
    ('1557', '栃', '（とち）', '栃木'),
    ('1558', '凸', 'トツ', '凸凹'),
    ('1559', '突突', 'トツ、つ-く', ''),
    ('1560', '届屆', 'とど-ける、とど-く', ''),
    ('1561', '屯', 'トン', ''),
    ('1562', '豚', 'トン、ぶた', ''),
    ('1563', '頓', 'トン', ''),
    ('1564', '貪', 'ドン、むさぼ-る', ''),
    ('1565', '鈍', 'ドン、にぶ-い、にぶ-る', ''),
    ('1566', '曇', 'ドン、くも-る', ''),
    ('1567', '丼', 'どんぶり、（どん）', ''),
    ('1568', '那', 'ナ', ''),
    ('1569', '奈', 'ナ', '神奈川、奈良'),
    ('1570', '内內', 'ナイ、（ダイ）、うち', ''),
    ('1571', '梨', 'なし', ''),
    ('1572', '謎', 'なぞ', ''),
    ('1573', '鍋', 'なべ', ''),
    ('1574', '南', 'ナン、（ナ）、みなみ', ''),
    ('1575', '軟', 'ナン、やわ-らか、やわ-らかい', ''),
    ('1576', '難難', 'ナン、かた-い、むずか-しい', '難しい'),
    ('1577', '二', 'ニ、ふた、ふた-つ', '十重二十重、二十、二十歳、二十日、二人、二日'),
    ('1578', '尼', 'ニ、あま', ''),
    ('1579', '弐貳', 'ニ', ''),
    ('1580', '匂', 'にお-う', ''),
    ('1581', '肉', 'ニク', ''),
    ('1582', '虹', 'にじ', ''),
    ('1583', '日', 'ニチ、ジツ、ひ、か', '明日、昨日、今日、一日、二十日、日和、二日、七日'),
    ('1584', '入', 'ニュウ、い-る、い-れる、はい-る', ''),
    ('1585', '乳', 'ニュウ、ちち、ち', '乳母'),
    ('1586', '尿', 'ニョウ', ''),
    ('1587', '任', 'ニン、まか-せる、まか-す', ''),
    ('1588', '妊', 'ニン', ''),
    ('1589', '忍', 'ニン、しの-ぶ、しの-ばせる', ''),
    ('1590', '認', 'ニン、みと-める', ''),
    ('1591', '寧', 'ネイ', ''),
    ('1592', '熱', 'ネツ、あつ-い', ''),
    ('1593', '年', 'ネン、とし', '今年'),
    ('1594', '念', 'ネン', ''),
    ('1595', '捻', 'ネン', ''),
    ('1596', '粘', 'ネン、ねば-る', ''),
    ('1597', '燃', 'ネン、も-える、も-やす、も-す', ''),
    ('1598', '悩惱', 'ノウ、なや-む、なや-ます', ''),
    ('1599', '納', 'ノウ、（ナッ）、（ナ）、（ナン）、（トウ）、おさ-める、おさ-まる', ''),
    ('1600', '能', 'ノウ', '堪能'),
    ('1601', '脳腦', 'ノウ', ''),
    ('1602', '農', 'ノウ', ''),
    ('1603', '濃', 'ノウ、こ-い', ''),
    ('1604', '把', 'ハ', ''),
    ('1605', '波', 'ハ、なみ', '波止場'),
    ('1606', '派', 'ハ', ''),
    ('1607', '破', 'ハ、やぶ-る、やぶ-れる', ''),
    ('1608', '覇霸', 'ハ', ''),  # "霸"は康熙字典体
    ('1609', '馬', 'バ、うま、（ま）', '伝馬船'),
    ('1610', '婆', 'バ', ''),
    ('1611', '罵', 'バ、ののし-る', ''),
    ('1612', '拝拜', 'ハイ、おが-む', ''),
    ('1613', '杯', 'ハイ、さかずき', ''),
    ('1614', '背', 'ハイ、せ、せい、そむ-く、そむ-ける', ''),
    ('1615', '肺', 'ハイ', ''),
    ('1616', '俳', 'ハイ', ''),
    ('1617', '配', 'ハイ、くば-る', ''),
    ('1618', '排', 'ハイ', ''),
    ('1619', '敗', 'ハイ、やぶ-れる', ''),
    ('1620', '廃廢', 'ハイ、すた-れる、すた-る', ''),
    ('1621', '輩', 'ハイ', ''),
    ('1622', '売賣', 'バイ、う-る、う-れる', ''),
    ('1623', '倍', 'バイ', ''),
    ('1624', '梅梅', 'バイ、うめ', '梅雨'),
    ('1625', '培', 'バイ、つちか-う', ''),
    ('1626', '陪', 'バイ', ''),
    ('1627', '媒', 'バイ', ''),
    ('1628', '買', 'バイ、か-う', ''),
    ('1629', '賠', 'バイ', ''),
    ('1630', '白', 'ハク、ビャク、しろ、（しら）、しろ-い', '白髪'),
    ('1631', '伯', 'ハク', '伯父、伯母'),
    ('1632', '拍', 'ハク、（ヒョウ）', ''),
    ('1633', '泊', 'ハク、と-まる、と-める', ''),
    ('1634', '迫', 'ハク、せま-る', ''),
    ('1635', '剝(剥)', 'ハク、は-がす、は-ぐ、は-がれる、は-げる', ''),
    ('1636', '舶', 'ハク', ''),
    ('1637', '博', 'ハク、（バク）', '博士'),
    ('1638', '薄', 'ハク、うす-い、うす-める、うす-まる、うす-らぐ、うす-れる', ''),
    ('1639', '麦麥', 'バク、むぎ', ''),
    ('1640', '漠', 'バク', ''),
    ('1641', '縛', 'バク、しば-る', ''),
    ('1642', '爆', 'バク', ''),
    ('1643', '箱', 'はこ', ''),
    ('1644', '箸', 'はし', ''),
    ('1645', '畑', 'はた、はたけ', ''),
    ('1646', '肌', 'はだ', ''),
    ('1647', '八', 'ハチ、や、や-つ、やっ-つ、（よう）', '八百長、八百屋'),
    ('1648', '鉢', 'ハチ、（ハツ）', ''),
    ('1649', '発發', 'ハツ、ホツ', ''),
    ('1650', '髪髮', 'ハツ、かみ', '白髪'),
    ('1651', '伐', 'バツ', ''),
    ('1652', '抜拔', 'バツ、ぬ-く、ぬ-ける、ぬ-かす、ぬ-かる', ''),
    ('1653', '罰', 'バツ、バチ', ''),
    ('1654', '閥', 'バツ', ''),
    ('1655', '反', 'ハン、（ホン）、（タン）、そ-る、そ-らす', ''),
    ('1656', '半', 'ハン、なか-ば', ''),
    ('1657', '氾', 'ハン', ''),
    ('1658', '犯', 'ハン、おか-す', ''),
    ('1659', '帆', 'ハン、ほ', ''),
    ('1660', '汎', 'ハン', ''),
    ('1661', '伴', 'ハン、バン、ともな-う', ''),
    ('1662', '判', 'ハン、バン', ''),
    ('1663', '坂', 'ハン、さか', ''),
    ('1664', '阪', 'ハン', '大阪'),
    ('1665', '板', 'ハン、バン、いた', ''),
    ('1666', '版', 'ハン', ''),
    ('1667', '班', 'ハン', ''),
    ('1668', '畔', 'ハン', ''),
    ('1669', '般', 'ハン', ''),
    ('1670', '販', 'ハン', ''),
    ('1671', '斑', 'ハン', ''),
    ('1672', '飯飯', 'ハン、めし', ''),
    ('1673', '搬', 'ハン', ''),
    ('1674', '煩', 'ハン、（ボン）、わずら-う、わずら-わす', ''),
    ('1675', '頒', 'ハン', ''),
    ('1676', '範', 'ハン', ''),
    ('1677', '繁繁', 'ハン', ''),
    ('1678', '藩', 'ハン', ''),
    ('1679', '晩晚', 'バン', ''),
    ('1680', '番', 'バン', ''),
    ('1681', '蛮蠻', 'バン', ''),
    ('1682', '盤', 'バン', ''),
    ('1683', '比', 'ヒ、くら-べる', ''),
    ('1684', '皮', 'ヒ、かわ', ''),
    ('1685', '妃', 'ヒ', ''),
    ('1686', '否', 'ヒ、いな', ''),
    ('1687', '批', 'ヒ', ''),
    ('1688', '彼', 'ヒ、かれ、（かの）', ''),
    ('1689', '披', 'ヒ', ''),
    ('1690', '肥', 'ヒ、こ-える、こえ、こ-やす、こ-やし', ''),
    ('1691', '非', 'ヒ', ''),
    ('1692', '卑卑', 'ヒ、いや-しい、いや-しむ、いや-しめる', ''),
    ('1693', '飛', 'ヒ、と-ぶ、と-ばす', ''),
    ('1694', '疲', 'ヒ、つか-れる', ''),
    ('1695', '秘祕', 'ヒ、ひ-める', ''),
    ('1696', '被', 'ヒ、こうむ-る', ''),
    ('1697', '悲', 'ヒ、かな-しい、かな-しむ', ''),
    ('1698', '扉', 'ヒ、とびら', ''),
    ('1699', '費', 'ヒ、つい-やす、つい-える', ''),
    ('1700', '碑碑', 'ヒ', ''),
    ('1701', '罷', 'ヒ', ''),
    ('1702', '避', 'ヒ、さ-ける', ''),
    ('1703', '尾', 'ビ、お', '尻尾'),
    ('1704', '眉', 'ビ、（ミ）、まゆ', ''),
    ('1705', '美', 'ビ、うつく-しい', ''),
    ('1706', '備', 'ビ、そな-える、そな-わる', ''),
    ('1707', '微', 'ビ', ''),
    ('1708', '鼻', 'ビ、はな', ''),
    ('1709', '膝', 'ひざ', ''),
    ('1710', '肘', 'ひじ', ''),
    ('1711', '匹', 'ヒツ、ひき', ''),
    ('1712', '必', 'ヒツ、かなら-ず', ''),
    ('1713', '泌', 'ヒツ、ヒ', '分泌'),
    ('1714', '筆', 'ヒツ、ふで', ''),
    ('1715', '姫姬', 'ひめ', ''),
    ('1716', '百', 'ヒャク', '八百長、八百屋'),
    ('1717', '氷', 'ヒョウ、こおり、ひ', ''),
    ('1718', '表', 'ヒョウ、おもて、あらわ-す、あらわ-れる', ''),
    ('1719', '俵', 'ヒョウ、たわら', ''),
    ('1720', '票', 'ヒョウ', ''),
    ('1721', '評', 'ヒョウ', ''),
    ('1722', '漂', 'ヒョウ、ただよ-う', ''),
    ('1723', '標', 'ヒョウ', ''),
    ('1724', '苗', 'ビョウ、なえ、（なわ）', '早苗'),
    ('1725', '秒', 'ビョウ', ''),
    ('1726', '病', 'ビョウ、（ヘイ）、や-む、やまい', ''),
    ('1727', '描', 'ビョウ、えが-く、か-く', ''),
    ('1728', '猫', 'ビョウ、ねこ', ''),
    ('1729', '品', 'ヒン、しな', ''),
    ('1730', '浜濱', 'ヒン、はま', ''),
    ('1731', '貧', 'ヒン、ビン、まず-しい', ''),
    ('1732', '賓賓', 'ヒン', ''),
    ('1733', '頻頻', 'ヒン', ''),
    ('1734', '敏敏', 'ビン', ''),
    ('1735', '瓶甁', 'ビン', ''),
    ('1736', '不', 'フ、ブ', ''),
    ('1737', '夫', 'フ、（フウ）、おっと', ''),
    ('1738', '父', 'フ、ちち', '叔父、伯父、父さん'),
    ('1739', '付', 'フ、つ-ける、つ-く', '貼付'),
    ('1740', '布', 'フ、ぬの', '昆布'),
    ('1741', '扶', 'フ', ''),
    ('1742', '府', 'フ', ''),
    ('1743', '怖', 'フ、こわ-い', ''),
    ('1744', '阜', '（フ）', '岐阜'),
    ('1745', '附', 'フ', ''),
    ('1746', '訃', 'フ', ''),
    ('1747', '負', 'フ、ま-ける、ま-かす、お-う', ''),
    ('1748', '赴', 'フ、おもむ-く', ''),
    ('1749', '浮', 'フ、う-く、う-かれる、う-かぶ、う-かべる', '浮気、浮つく'),
    ('1750', '婦', 'フ', ''),
    ('1751', '符', 'フ', ''),
    ('1752', '富', 'フ、（フウ）、と-む、とみ', '富山、富貴'),
    ('1753', '普', 'フ', ''),
    ('1754', '腐', 'フ、くさ-る、くさ-れる、くさ-らす', ''),
    ('1755', '敷', 'フ、し-く', '桟敷'),
    ('1756', '膚', 'フ', ''),
    ('1757', '賦', 'フ', ''),
    ('1758', '譜', 'フ', ''),
    ('1759', '侮侮', 'ブ、あなど-る', ''),
    ('1760', '武', 'ブ、ム', ''),
    ('1761', '部', 'ブ', '部屋'),
    ('1762', '舞', 'ブ、ま-う、まい', ''),
    ('1763', '封', 'フウ、ホウ', ''),
    ('1764', '風', 'フウ、（フ）、かぜ、（かざ）', '風邪'),
    ('1765', '伏', 'フク、ふ-せる、ふ-す', ''),
    ('1766', '服', 'フク', ''),
    ('1767', '副', 'フク', ''),
    ('1768', '幅', 'フク、はば', ''),
    ('1769', '復', 'フク', ''),
    ('1770', '福福', 'フク', ''),
    ('1771', '腹', 'フク、はら', ''),
    ('1772', '複', 'フク', ''),
    ('1773', '覆', 'フク、おお-う、くつがえ-す、くつがえ-る', ''),
    ('1774', '払拂', 'フツ、はら-う', ''),
    ('1775', '沸', 'フツ、わ-く、わ-かす', ''),
    ('1776', '仏佛', 'ブツ、ほとけ', ''),
    ('1777', '物', 'ブツ、モツ、もの', '果物'),
    ('1778', '粉', 'フン、こ、こな', ''),
    ('1779', '紛', 'フン、まぎ-れる、まぎ-らす、まぎ-らわす、まぎ-らわしい', ''),
    ('1780', '雰', 'フン', ''),
    ('1781', '噴', 'フン、ふ-く', ''),
    ('1782', '墳', 'フン', ''),
    ('1783', '憤', 'フン、いきどお-る', ''),
    ('1784', '奮', 'フン、ふる-う', ''),
    ('1785', '分', 'ブン、フン、ブ、わ-ける、わ-かれる、わ-かる、わ-かつ', '大分'),
    ('1786', '文', 'ブン、モン、ふみ', '文字'),
    ('1787', '聞', 'ブン、モン、き-く、き-こえる', ''),
    ('1788', '丙', 'ヘイ', ''),
    ('1789', '平', 'ヘイ、ビョウ、たい-ら、ひら', ''),
    ('1790', '兵', 'ヘイ、ヒョウ', ''),
    ('1791', '併倂', 'ヘイ、あわ-せる', ''),
    ('1792', '並竝', 'ヘイ、なみ、なら-べる、なら-ぶ、なら-びに', ''),  # "竝"は康熙字典体
    ('1793', '柄', 'ヘイ、がら、え', ''),
    ('1794', '陛', 'ヘイ', ''),
    ('1795', '閉', 'ヘイ、と-じる、と-ざす、し-める、し-まる', ''),
    ('1796', '塀塀', 'ヘイ', ''),
    ('1797', '幣', 'ヘイ', ''),
    ('1798', '弊', 'ヘイ', ''),
    ('1799', '蔽', 'ヘイ', ''),
    ('1800', '餅餠', 'ヘイ、もち', ''),  # "餠"は康熙字典体
    ('1801', '米', 'ベイ、マイ、こめ', ''),
    ('1802', '壁', 'ヘキ、かべ', ''),
    ('1803', '璧', 'ヘキ', ''),
    ('1804', '癖', 'ヘキ、くせ', ''),
    ('1805', '別', 'ベツ、わか-れる', ''),
    ('1806', '蔑', 'ベツ、さげす-む', ''),
    ('1807', '片', 'ヘン、かた', ''),
    ('1808', '辺邊', 'ヘン、あた-り、べ', ''),
    ('1809', '返', 'ヘン、かえ-す、かえ-る', ''),
    ('1810', '変變', 'ヘン、か-わる、か-える', ''),
    ('1811', '偏', 'ヘン、かたよ-る', ''),
    ('1812', '遍', 'ヘン', ''),
    ('1813', '編', 'ヘン、あ-む', ''),
    ('1814', '弁辨瓣辯', 'ベン', ''),
    ('1815', '便', 'ベン、ビン、たよ-り', ''),
    ('1816', '勉勉', 'ベン', ''),
    ('1817', '歩步', 'ホ、ブ、（フ）、ある-く、あゆ-む', ''),
    ('1818', '保', 'ホ、たも-つ', ''),
    ('1819', '哺', 'ホ', ''),
    ('1820', '捕', 'ホ、と-らえる、と-らわれる、と-る、つか-まえる、つか-まる', ''),
    ('1821', '補', 'ホ、おぎな-う', ''),
    ('1822', '舗舖', 'ホ', '老舗'),
    ('1823', '母', 'ボ、はは', '乳母、叔母、伯母、母屋、母家、母さん'),
    ('1824', '募', 'ボ、つの-る', ''),
    ('1825', '墓', 'ボ、はか', ''),
    ('1826', '慕', 'ボ、した-う', ''),
    ('1827', '暮', 'ボ、く-れる、く-らす', ''),
    ('1828', '簿', 'ボ', ''),
    ('1829', '方', 'ホウ、かた', '行方'),
    ('1830', '包', 'ホウ、つつ-む', ''),
    ('1831', '芳', 'ホウ、かんば-しい', ''),
    ('1832', '邦', 'ホウ', ''),
    ('1833', '奉', 'ホウ、（ブ）、たてまつ-る', ''),
    ('1834', '宝寶', 'ホウ、たから', ''),
    ('1835', '抱', 'ホウ、だ-く、いだ-く、かか-える', ''),
    ('1836', '放', 'ホウ、はな-す、はな-つ、はな-れる、ほう-る', ''),
    ('1837', '法', 'ホウ、（ハッ）、（ホッ）', ''),
    ('1838', '泡', 'ホウ、あわ', ''),
    ('1839', '胞', 'ホウ', ''),
    ('1840', '俸', 'ホウ', ''),
    ('1841', '倣', 'ホウ、なら-う', ''),
    ('1842', '峰', 'ホウ、みね', ''),
    ('1843', '砲', 'ホウ', ''),
    ('1844', '崩', 'ホウ、くず-れる、くず-す', '雪崩'),
    ('1845', '訪', 'ホウ、おとず-れる、たず-ねる', ''),
    ('1846', '報', 'ホウ、むく-いる', ''),
    ('1847', '蜂', 'ホウ、はち', ''),
    ('1848', '豊豐', 'ホウ、ゆた-か', ''),
    ('1849', '飽', 'ホウ、あ-きる、あ-かす', ''),
    ('1850', '褒襃', 'ホウ、ほ-める', ''),  # "襃"は康熙字典体
    ('1851', '縫', 'ホウ、ぬ-う', ''),
    ('1852', '亡', 'ボウ、（モウ）、な-い', ''),
    ('1853', '乏', 'ボウ、とぼ-しい', ''),
    ('1854', '忙', 'ボウ、いそが-しい', ''),
    ('1855', '坊', 'ボウ、（ボッ）', ''),
    ('1856', '妨', 'ボウ、さまた-げる', ''),
    ('1857', '忘', 'ボウ、わす-れる', ''),
    ('1858', '防', 'ボウ、ふせ-ぐ', ''),
    ('1859', '房', 'ボウ、ふさ', ''),
    ('1860', '肪', 'ボウ', ''),
    ('1861', '某', 'ボウ', ''),
    ('1862', '冒', 'ボウ、おか-す', ''),
    ('1863', '剖', 'ボウ', ''),
    ('1864', '紡', 'ボウ、つむ-ぐ', ''),
    ('1865', '望', 'ボウ、モウ、のぞ-む', ''),
    ('1866', '傍', 'ボウ、かたわ-ら', ''),
    ('1867', '帽', 'ボウ', ''),
    ('1868', '棒', 'ボウ', ''),
    ('1869', '貿', 'ボウ', ''),
    ('1870', '貌', 'ボウ', ''),
    ('1871', '暴', 'ボウ、（バク）、あば-く、あば-れる', ''),
    ('1872', '膨', 'ボウ、ふく-らむ、ふく-れる', ''),
    ('1873', '謀', 'ボウ、（ム）、はか-る', ''),
    ('1874', '頰(頬)', 'ほお', ''),
    ('1875', '北', 'ホク、きた', ''),
    ('1876', '木', 'ボク、モク、き、（こ）', '木綿'),
    ('1877', '朴', 'ボク', ''),
    ('1878', '牧', 'ボク、まき', ''),
    ('1879', '睦', 'ボク', ''),
    ('1880', '僕', 'ボク', ''),
    ('1881', '墨墨', 'ボク、すみ', ''),
    ('1882', '撲', 'ボク', '相撲'),
    ('1883', '没沒', 'ボツ', ''),
    ('1884', '勃', 'ボツ', ''),
    ('1885', '堀', 'ほり', ''),
    ('1886', '本', 'ホン、もと', ''),
    ('1887', '奔', 'ホン', ''),
    ('1888', '翻飜', 'ホン、ひるがえ-る、ひるがえ-す', ''),  # "飜"は康熙字典体
    ('1889', '凡', 'ボン、（ハン）', ''),
    ('1890', '盆', 'ボン', ''),
    ('1891', '麻', 'マ、あさ', ''),
    ('1892', '摩', 'マ', ''),
    ('1893', '磨', 'マ、みが-く', ''),
    ('1894', '魔', 'マ', ''),
    ('1895', '毎每', 'マイ', ''),
    ('1896', '妹', 'マイ、いもうと', ''),
    ('1897', '枚', 'マイ', ''),
    ('1898', '昧', 'マイ', ''),
    ('1899', '埋', 'マイ、う-める、う-まる、う-もれる', ''),
    ('1900', '幕', 'マク、バク', ''),
    ('1901', '膜', 'マク', ''),
    ('1902', '枕', 'まくら', ''),
    ('1903', '又', 'また', ''),
    ('1904', '末', 'マツ、バツ、すえ', ''),
    ('1905', '抹', 'マツ', ''),
    ('1906', '万萬', 'マン、バン', ''),
    ('1907', '満滿', 'マン、み-ちる、み-たす', ''),
    ('1908', '慢', 'マン', ''),
    ('1909', '漫', 'マン', ''),
    ('1910', '未', 'ミ', ''),
    ('1911', '味', 'ミ、あじ、あじ-わう', '三味線'),
    ('1912', '魅', 'ミ', ''),
    ('1913', '岬', 'みさき', ''),
    ('1914', '密', 'ミツ', ''),
    ('1915', '蜜', 'ミツ', ''),
    ('1916', '脈', 'ミャク', ''),
    ('1917', '妙', 'ミョウ', ''),
    ('1918', '民', 'ミン、たみ', ''),
    ('1919', '眠', 'ミン、ねむ-る、ねむ-い', ''),
    ('1920', '矛', 'ム、ほこ', ''),
    ('1921', '務', 'ム、つと-める、つと-まる', ''),
    ('1922', '無', 'ム、ブ、な-い', ''),
    ('1923', '夢', 'ム、ゆめ', ''),
    ('1924', '霧', 'ム、きり', ''),
    ('1925', '娘', 'むすめ', ''),
    ('1926', '名', 'メイ、ミョウ、な', '仮名、名残'),
    ('1927', '命', 'メイ、ミョウ、いのち', ''),
    ('1928', '明',
     'メイ、ミョウ、あ-かり、あか-るい、あか-るむ、あか-らむ、あき-らか、あ-ける、あ-く、あ-くる、あ-かす',
     '明日'),
    ('1929', '迷', 'メイ、まよ-う', '迷子'),
    ('1930', '冥', 'メイ、ミョウ', ''),
    ('1931', '盟', 'メイ', ''),
    ('1932', '銘', 'メイ', ''),
    ('1933', '鳴', 'メイ、な-く、な-る、な-らす', ''),
    ('1934', '滅', 'メツ、ほろ-びる、ほろ-ぼす', ''),
    ('1935', '免免', 'メン、まぬか-れる', '免れる'),
    ('1936', '面', 'メン、おも、おもて、つら', '真面目'),
    ('1937', '綿', 'メン、わた', '木綿'),
    ('1938', '麺麵', 'メン', ''),
    ('1939', '茂', 'モ、しげ-る', ''),
    ('1940', '模', 'モ、ボ', ''),
    ('1941', '毛', 'モウ、け', ''),
    ('1942', '妄', 'モウ、ボウ', ''),
    ('1943', '盲', 'モウ', ''),
    ('1944', '耗', 'モウ、（コウ）', ''),
    ('1945', '猛', 'モウ', '猛者'),
    ('1946', '網', 'モウ、あみ', '投網'),
    ('1947', '目', 'モク、（ボク）、め、（ま）', '真面目'),
    ('1948', '黙默', 'モク、だま-る', ''),
    ('1949', '門', 'モン、かど', ''),
    ('1950', '紋', 'モン', ''),
    ('1951', '問', 'モン、と-う、と-い、（とん）', '問屋'),
    ('1952', '冶', 'ヤ', '鍛冶'),
    ('1953', '夜', 'ヤ、よ、よる', ''),
    ('1954', '野', 'ヤ、の', '野良'),
    ('1955', '弥彌', 'や', '弥生'),
    ('1956', '厄', 'ヤク', ''),
    ('1957', '役', 'ヤク、エキ', ''),
    ('1958', '約', 'ヤク', ''),
    ('1959', '訳譯', 'ヤク、わけ', ''),
    ('1960', '薬藥', 'ヤク、くすり', ''),
    ('1961', '躍', 'ヤク、おど-る', ''),
    ('1962', '闇', 'やみ', ''),
    ('1963', '由', 'ユ、ユウ、（ユイ）、よし', ''),
    ('1964', '油', 'ユ、あぶら', ''),
    ('1965', '喩', 'ユ', ''),
    ('1966', '愉', 'ユ', ''),
    ('1967', '諭', 'ユ、さと-す', ''),
    ('1968', '輸', 'ユ', ''),
    ('1969', '癒', 'ユ、い-える、い-やす', ''),
    ('1970', '唯', 'ユイ、（イ）', ''),
    ('1971', '友', 'ユウ、とも', '友達'),
    ('1972', '有', 'ユウ、ウ、あ-る', ''),
    ('1973', '勇', 'ユウ、いさ-む', ''),
    ('1974', '幽', 'ユウ', ''),
    ('1975', '悠', 'ユウ', ''),
    ('1976', '郵', 'ユウ', ''),
    ('1977', '湧', 'ユウ、わ-く', ''),
    ('1978', '猶', 'ユウ', ''),
    ('1979', '裕', 'ユウ', ''),
    ('1980', '遊', 'ユウ、（ユ）、あそ-ぶ', ''),
    ('1981', '雄', 'ユウ、お、おす', ''),
    ('1982', '誘', 'ユウ、さそ-う', ''),
    ('1983', '憂', 'ユウ、うれ-える、うれ-い、う-い', ''),
    ('1984', '融', 'ユウ', ''),
    ('1985', '優', 'ユウ、やさ-しい、すぐ-れる', ''),
    ('1986', '与與', 'ヨ、あた-える', ''),
    ('1987', '予豫', 'ヨ', ''),
    ('1988', '余餘', 'ヨ、あま-る、あま-す', ''),
    ('1989', '誉譽', 'ヨ、ほま-れ', ''),
    ('1990', '預', 'ヨ、あず-ける、あず-かる', ''),
    ('1991', '幼', 'ヨウ、おさな-い', ''),
    ('1992', '用', 'ヨウ、もち-いる', ''),
    ('1993', '羊', 'ヨウ、ひつじ', ''),
    ('1994', '妖', 'ヨウ、あや-しい', ''),
    ('1995', '洋', 'ヨウ', ''),
    ('1996', '要', 'ヨウ、かなめ、い-る', ''),
    ('1997', '容', 'ヨウ', ''),
    ('1998', '庸', 'ヨウ', ''),
    ('1999', '揚', 'ヨウ、あ-げる、あ-がる', ''),
    ('2000', '揺搖', 'ヨウ、ゆ-れる、ゆ-る、ゆ-らぐ、ゆ-るぐ、ゆ-する、ゆ-さぶる、ゆ-すぶる', ''),
    ('2001', '葉', 'ヨウ、は', '紅葉'),
    ('2002', '陽', 'ヨウ', ''),
    ('2003', '溶', 'ヨウ、と-ける、と-かす、と-く', ''),
    ('2004', '腰', 'ヨウ、こし', ''),
    ('2005', '様樣', 'ヨウ、さま', ''),
    ('2006', '瘍', 'ヨウ', ''),
    ('2007', '踊', 'ヨウ、おど-る、おど-り', ''),
    ('2008', '窯', 'ヨウ、かま', ''),
    ('2009', '養', 'ヨウ、やしな-う', ''),
    ('2010', '擁', 'ヨウ', ''),
    ('2011', '謡謠', 'ヨウ、うたい、うた-う', ''),
    ('2012', '曜', 'ヨウ', ''),
    ('2013', '抑', 'ヨク、おさ-える', ''),
    ('2014', '沃', 'ヨク', ''),
    ('2015', '浴', 'ヨク、あ-びる、あ-びせる', '浴衣'),
    ('2016', '欲', 'ヨク、ほっ-する、ほ-しい', ''),
    ('2017', '翌', 'ヨク', ''),
    ('2018', '翼', 'ヨク、つばさ', ''),
    ('2019', '拉', 'ラ', ''),
    ('2020', '裸', 'ラ、はだか', ''),
    ('2021', '羅', 'ラ', ''),
    ('2022', '来來', 'ライ、く-る、きた-る、きた-す', ''),
    ('2023', '雷', 'ライ、かみなり', ''),
    ('2024', '頼賴', 'ライ、たの-む、たの-もしい、たよ-る', ''),
    ('2025', '絡', 'ラク、から-む、から-まる、から-める', ''),
    ('2026', '落', 'ラク、お-ちる、お-とす', ''),
    ('2027', '酪', 'ラク', ''),
    ('2028', '辣', 'ラツ', ''),
    ('2029', '乱亂', 'ラン、みだ-れる、みだ-す', ''),
    ('2030', '卵', 'ラン、たまご', ''),
    ('2031', '覧覽', 'ラン', ''),
    ('2032', '濫', 'ラン', ''),
    ('2033', '藍', 'ラン、あい', ''),
    ('2034', '欄欄', 'ラン', ''),
    ('2035', '吏', 'リ', ''),
    ('2036', '利', 'リ、き-く', '砂利'),
    ('2037', '里', 'リ、さと', ''),
    ('2038', '理', 'リ', ''),
    ('2039', '痢', 'リ', ''),
    ('2040', '裏', 'リ、うら', ''),
    ('2041', '履', 'リ、は-く', '草履'),
    ('2042', '璃', 'リ', ''),
    ('2043', '離', 'リ、はな-れる、はな-す', ''),
    ('2044', '陸', 'リク', ''),
    ('2045', '立', 'リツ、（リュウ）、た-つ、た-てる', '立ち退く'),
    ('2046', '律', 'リツ、（リチ）', ''),
    ('2047', '慄', 'リツ', ''),
    ('2048', '略', 'リャク', ''),
    ('2049', '柳', 'リュウ、やなぎ', ''),
    ('2050', '流', 'リュウ、（ル）、なが-れる、なが-す', ''),
    ('2051', '留', 'リュウ、（ル）、と-める、と-まる', ''),
    ('2052', '竜龍', 'リュウ、たつ', ''),
    ('2053', '粒', 'リュウ、つぶ', ''),
    ('2054', '隆隆', 'リュウ', ''),
    ('2055', '硫', 'リュウ', '硫黄'),
    ('2056', '侶', 'リョ', ''),
    ('2057', '旅', 'リョ、たび', ''),
    ('2058', '虜虜', 'リョ', ''),
    ('2059', '慮', 'リョ', ''),
    ('2060', '了', 'リョウ', ''),
    ('2061', '両兩', 'リョウ', ''),
    ('2062', '良', 'リョウ、よ-い', '野良、奈良'),
    ('2063', '料', 'リョウ', ''),
    ('2064', '涼', 'リョウ、すず-しい、すず-む', ''),
    ('2065', '猟獵', 'リョウ', ''),
    ('2066', '陵', 'リョウ、みささぎ', ''),
    ('2067', '量', 'リョウ、はか-る', ''),
    ('2068', '僚', 'リョウ', ''),
    ('2069', '領', 'リョウ', ''),
    ('2070', '寮', 'リョウ', ''),
    ('2071', '療', 'リョウ', ''),
    ('2072', '瞭', 'リョウ', ''),
    ('2073', '糧', 'リョウ、（ロウ）、かて', ''),
    ('2074', '力', 'リョク、リキ、ちから', ''),
    ('2075', '緑綠', 'リョク、（ロク）、みどり', ''),
    ('2076', '林', 'リン、はやし', ''),
    ('2077', '厘', 'リン', ''),
    ('2078', '倫', 'リン', ''),
    ('2079', '輪', 'リン、わ', ''),
    ('2080', '隣', 'リン、とな-る、となり', ''),
    ('2081', '臨', 'リン、のぞ-む', ''),
    ('2082', '瑠', 'ル', ''),
    ('2083', '涙淚', 'ルイ、なみだ', ''),
    ('2084', '累', 'ルイ', ''),
    ('2085', '塁壘', 'ルイ', ''),
    ('2086', '類類', 'ルイ、たぐ-い', ''),
    ('2087', '令', 'レイ', ''),
    ('2088', '礼禮', 'レイ、ライ', ''),
    ('2089', '冷', 'レイ、つめ-たい、ひ-える、ひ-や、ひ-やす、ひ-やかす、さ-める、さ-ます', ''),
    ('2090', '励勵', 'レイ、はげ-む、はげ-ます', ''),
    ('2091', '戻戾', 'レイ、もど-す、もど-る', ''),
    ('2092', '例', 'レイ、たと-える', ''),
    ('2093', '鈴', 'レイ、リン、すず', ''),
    ('2094', '零', 'レイ', ''),
    ('2095', '霊靈', 'レイ、リョウ、たま', ''),
    ('2096', '隷', 'レイ', ''),
    ('2097', '齢齡', 'レイ', ''),
    ('2098', '麗', 'レイ、うるわ-しい', ''),
    ('2099', '暦曆', 'レキ、こよみ', ''),
    ('2100', '歴歷', 'レキ', ''),
    ('2101', '列', 'レツ', ''),
    ('2102', '劣', 'レツ、おと-る', ''),
    ('2103', '烈', 'レツ', ''),
    ('2104', '裂', 'レツ、さ-く、さ-ける', ''),
    ('2105', '恋戀', 'レン、こ-う、こい、こい-しい', ''),
    ('2106', '連', 'レン、つら-なる、つら-ねる、つ-れる', ''),
    ('2107', '廉', 'レン', ''),
    ('2108', '練練', 'レン、ね-る', ''),
    ('2109', '錬鍊', 'レン', ''),
    ('2110', '呂', 'ロ', ''),
    ('2111', '炉爐', 'ロ', ''),
    ('2112', '賂', 'ロ', ''),
    ('2113', '路', 'ロ、じ', ''),
    ('2114', '露', 'ロ、（ロウ）、つゆ', ''),
    ('2115', '老', 'ロウ、お-いる、ふ-ける', '老舗'),
    ('2116', '労勞', 'ロウ', ''),
    ('2117', '弄', 'ロウ、もてあそ-ぶ', ''),
    ('2118', '郎郞', 'ロウ', ''),
    ('2119', '朗朗', 'ロウ、ほが-らか', ''),
    ('2120', '浪', 'ロウ', ''),
    ('2121', '廊廊', 'ロウ', ''),
    ('2122', '楼樓', 'ロウ', ''),
    ('2123', '漏', 'ロウ、も-る、も-れる、も-らす', ''),
    ('2124', '籠', 'ロウ、かご、こ-もる', ''),
    ('2125', '六', 'ロク、む、む-つ、むっ-つ、（むい）', ''),
    ('2126', '録錄', 'ロク', ''),
    ('2127', '麓', 'ロク、ふもと', ''),
    ('2128', '論', 'ロン', ''),
    ('2129', '和', 'ワ、（オ）、やわ-らぐ、やわ-らげる、なご-む、なご-やか', '日和、大和'),
    ('2130', '話', 'ワ、はな-す、はなし', ''),
    ('2131', '賄', 'ワイ、まかな-う', ''),
    ('2132', '脇', 'わき', ''),
    ('2133', '惑', 'ワク、まど-う', ''),
    ('2134', '枠', 'わく', ''),
    ('2135', '湾灣', 'ワン', ''),
    ('2136', '腕', 'ワン、うで', ''),
)

TYPEFACES = (
    'アｱ', 'イｲ', 'ウｳ', 'エｴ', 'オｵ', 'カｶ', 'キｷ', 'クｸ', 'ケｹ', 'コｺ',
    'サｻ', 'シｼ', 'スｽ', 'セｾ', 'ソｿ', 'タﾀ', 'チﾁ', 'ツﾂ', 'テﾃ', 'トﾄ',
    'ナﾅ', 'ニﾆ', 'ヌﾇ', 'ネﾈ', 'ノﾉ', 'ハﾊ', 'ヒﾋ', 'フﾌ', 'ヘﾍ', 'ホﾎ',
    'マﾏ', 'ミﾐ', 'ムﾑ', 'メﾒ', 'モﾓ', 'ヤﾔ', 'ユﾕ', 'ヨﾖ',
    'ラﾗ', 'リﾘ', 'ルﾙ', 'レﾚ', 'ロﾛ', 'ワﾜ', 'ヲｦ',
    'ンﾝ', '、､', '。｡',
    'ァｧ', 'ィｨ', 'ゥｩ', 'ェｪ', 'ォｫ', 'ッｯ',
    'ャｬ', 'ュｭ', 'ョｮ',
    '+⁺₊', '-⁻₋', '=⁼₌', '(⁽₍', ')⁾₎',
    '0⁰₀', '1⑴①', '2⑵②²₂', '3⑶③³₃', '4⑷④⁴₄',
    '5⑸⑤⁵₅', '6⑹⑥⁶₆', '7⑺⑦⁷₇', '8⑻⑧⁸₈', '9⑼⑨⁹₉',
    'nⁿₙ',
    '印㊞', '有㈲', '株㈱', '社㈳', '財㈶', '学㈻',
    '吉𠮷', '崎﨑嵜', '高髙',
    '頬頰', '侠俠', '巌巖', '桑桒', '桧檜', '槙槇', '祐祐', '祷禱', '禄祿',
    '秦䅈', '穣穰', '第㐧', '蝉蟬', '脇𦚰', '鴎鷗', '鴬鶯', '今𫝆',
    # 常用漢字
    '亜亞', '悪惡', '圧壓', '囲圍', '医醫', '為爲', '壱壹', '逸逸', '飲飮',
    '隠隱', '羽羽', '栄榮', '営營', '鋭銳', '衛衞', '益益', '駅驛', '悦悅',
    '謁謁', '閲閱', '円圓', '塩鹽', '縁緣', '艶艷', '応應', '欧歐', '殴毆',
    '桜櫻', '奥奧', '横橫', '温溫', '穏穩', '仮假', '価價', '禍禍', '画畫',
    '会會', '悔悔', '海海', '絵繪', '壊壞', '懐懷', '慨慨', '概槪', '拡擴',
    '殻殼', '覚覺', '学學', '岳嶽', '楽樂', '喝喝', '渇渴', '褐褐', '缶罐',
    '巻卷', '陥陷', '勧勸', '寛寬', '漢漢', '関關', '歓歡', '館館', '観觀',
    '顔顏', '気氣', '祈祈', '既旣', '帰歸', '亀龜', '器器', '偽僞', '戯戲',
    '犠犧', '旧舊', '拠據', '挙擧', '虚虛', '峡峽', '挟挾', '狭狹', '教敎',
    '郷鄕', '響響', '暁曉', '勤勤', '謹謹', '区區', '駆驅駈',  # "駈"を追加
    '勲勳', '薫薰', '径徑', '茎莖', '恵惠', '掲揭', '渓溪', '経經', '蛍螢',
    '軽輕', '継繼', '鶏鷄', '芸藝', '撃擊', '欠缺', '研硏', '県縣', '倹儉',
    '剣劍', '険險', '圏圈', '検檢', '献獻', '権權', '顕顯', '験驗', '厳嚴',
    '戸戶', '呉吳', '娯娛', '広廣', '効效', '恒恆', '黄黃', '鉱鑛', '号號',
    '告吿', '国國', '黒黑', '穀穀', '砕碎', '済濟', '斎齋', '歳歲', '剤劑',
    '冊册', '殺殺', '雑雜', '参參', '桟棧', '蚕蠶', '惨慘', '産產', '賛贊',
    '残殘', '糸絲', '祉祉', '視視', '歯齒', '飼飼', '児兒', '辞辭', '𠮟叱',
    '湿濕', '実實', '写寫', '社社', '舎舍', '者者', '煮煮', '釈釋', '寿壽',
    '収收', '臭臭', '従從', '渋澁', '獣獸', '縦縱', '祝祝', '粛肅', '処處',
    '暑暑', '署署', '緒緖', '諸諸', '叙敍敘',  # "敘"を追加
    '尚尙', '将將', '祥祥', '称稱', '渉涉', '焼燒', '証證', '奨奬', '条條',
    '状狀', '乗乘', '浄淨', '剰剩', '畳疊', '縄繩', '壌壤', '嬢孃', '譲讓',
    '醸釀', '触觸', '嘱囑', '神神', '真眞', '寝寢', '慎愼', '尽盡', '図圖',
    '粋粹', '酔醉', '穂穗', '随隨', '髄髓', '枢樞', '数數', '瀬瀨', '声聲',
    '青靑', '斉齊', '清淸', '晴晴', '精精', '静靜', '税稅', '窃竊', '摂攝',
    '節節', '説說', '絶絕', '専專', '浅淺', '戦戰', '践踐', '銭錢', '潜潛',
    '繊纖', '禅禪', '祖祖', '双雙', '壮壯', '争爭', '荘莊', '捜搜', '挿揷插',
    '巣巢', '曽曾', '痩瘦', '装裝', '僧僧', '層層', '総總', '騒騷', '増增',
    '憎憎', '蔵藏', '贈贈', '臓臟', '即卽', '属屬', '続續', '堕墮', '対對',
    '体體', '帯帶', '滞滯', '台臺', '滝瀧', '択擇', '沢澤', '脱脫', '担擔',
    '単單', '胆膽', '嘆嘆', '団團', '断斷', '弾彈', '遅遲', '痴癡', '虫蟲',
    '昼晝', '鋳鑄', '著著', '庁廳', '徴徵', '聴聽', '懲懲', '勅敕', '鎮鎭',
    '塚塚', '逓遞', '鉄鐵', '点點', '転轉', '塡填', '伝傳', '都都', '灯燈',
    '当當', '党黨', '盗盜', '稲稻', '闘鬪鬭', '徳德', '独獨', '読讀', '突突',
    '届屆', '内內', '難難', '弐貳', '悩惱', '脳腦', '覇霸', '拝拜', '廃廢',
    '売賣', '梅梅', '剝剥', '麦麥', '発發', '髪髮', '抜拔', '飯飯', '繁繁',
    '晩晚', '蛮蠻', '卑卑', '秘祕', '碑碑', '姫姬', '浜濱濵',  # "濵"を追加
    '賓賓', '頻頻', '敏敏', '瓶甁', '侮侮', '福福', '払拂', '仏佛', '併倂',
    '並竝', '塀塀', '餅餠', '辺邊邉',  # "邉"を追加
    '変變', '弁辨瓣辯', '勉勉', '歩步', '舗舖', '宝寶', '豊豐', '褒襃', '頰頬',
    '墨墨', '没沒', '翻飜', '毎每', '万萬', '満滿', '免免', '麺麵', '黙默',
    '弥彌', '訳譯', '薬藥', '与與', '予豫', '余餘', '誉譽', '揺搖', '様樣',
    '謡謠', '来來', '頼賴', '乱亂', '覧覽', '欄欄', '竜龍', '隆隆', '虜虜',
    '両兩', '猟獵', '緑綠', '涙淚', '塁壘', '類類', '礼禮礼',  # "礼"を追加
    '励勵', '戻戾', '霊靈', '齢齡', '暦曆', '歴歷', '恋戀', '練練', '錬鍊',
    '炉爐', '労勞', '郎郞', '朗朗', '廊廊', '楼樓', '録錄', '湾灣',
)

SAMPLE_BASIS = '''
# ★（タイトル）

v=+1.0
### ★（第1項）

★

### ★（第2項）

★
'''

SAMPLE_LAW = '''
v=+0.5 V=+0.5
$ 総則

v=+0.5 V=+0.5
$$ 通則

v=+0.5
: （基本原則）

##
私権は、公共の福祉に適合しなければならない。

###
権利の行使及び義務の履行は、信義に従い誠実に行わなければならない。

###
権利の濫用は、これを許さない。

v=+0.5
: （解釈の基準）

##
この法律は、個人の尊厳と両性の本質的平等を旨として、解釈しなければならない。

v=+0.5 V=+0.5
$$ 人

v=+0.5 V=+0.5
$$$ 権利能力

v=+0.5
##
私権の享有は、出生に始まる。

###
外国人は、法令又は条約の規定により禁止される場合を除き、私権を享有する。

v=+0.5 V=+0.5
$$$ 意思能力

v=+0.5
##-#
法律行為の当事者が意思表示をした時に意思能力を有しなかったときは、
その法律行為は、無効とする。

v=+0.5 V=+0.5
$$$ 行為能力

v=+0.5
: （成年）

##
年齢十八歳をもって、成年とする。

v=+0.5
: （未成年者の法律行為）

##
未成年者が法律行為をするには、その法定代理人の同意を得なければならない。
ただし、単に権利を得、又は義務を免れる法律行為については、この限りでない。

###
前項の規定に反する法律行為は、取り消すことができる。

###
第一項の規定にかかわらず、法定代理人が目的を定めて処分を許した財産は、
その目的の範囲内において、未成年者が自由に処分することができる。
目的を定めないで処分を許した財産を処分するときも、同様とする。

v=+0.5
: （未成年者の営業の許可）

##
一種又は数種の営業を許された未成年者は、その営業に関しては、
成年者と同一の行為能力を有する。

###
前項の場合において、未成年者がその営業に堪えることができない事由があるときは、
その法定代理人は、第四編（親族）の規定に従い、その許可を取り消し、
又はこれを制限することができる。

v=+0.5
: （後見開始の審判）

##
精神上の障害により事理を弁識する能力を欠く常況にある者については、
家庭裁判所は、本人、配偶者、四親等内の親族、未成年後見人、未成年後見監督人、
保佐人、保佐監督人、補助人、補助監督人又は検察官の請求により、
後見開始の審判をすることができる。

v=+0.5
: （成年被後見人及び成年後見人）

##
後見開始の審判を受けた者は、成年被後見人とし、これに成年後見人を付する。

v=+0.5
: （成年被後見人の法律行為）

##
成年被後見人の法律行為は、取り消すことができる。
ただし、日用品の購入その他日常生活に関する行為については、この限りでない。

v=+0.5
: （後見開始の審判の取消し）

##
第七条に規定する原因が消滅したときは、家庭裁判所は、本人、配偶者、
四親等内の親族、後見人（未成年後見人及び成年後見人をいう。以下同じ。）、
後見監督人（未成年後見監督人及び成年後見監督人をいう。以下同じ。）又は
検察官の請求により、後見開始の審判を取り消さなければならない。

v=+0.5
: （保佐開始の審判）

##
精神上の障害により事理を弁識する能力が著しく不十分である者については、
家庭裁判所は、本人、配偶者、四親等内の親族、後見人、後見監督人、補助人、
補助監督人又は検察官の請求により、保佐開始の審判をすることができる。
ただし、第七条に規定する原因がある者については、この限りでない。

v=+0.5
: （被保佐人及び保佐人）

##
保佐開始の審判を受けた者は、被保佐人とし、これに保佐人を付する。

v=+0.5
: （保佐人の同意を要する行為等）

##
被保佐人が次に掲げる行為をするには、その保佐人の同意を得なければならない。
ただし、第九条ただし書に規定する行為については、この限りでない。

####
元本を領収し、又は利用すること。

####
借財又は保証をすること。

####
不動産その他重要な財産に関する権利の得喪を目的とする行為をすること。

####
訴訟行為をすること。

####
贈与、和解又は仲裁合意
（仲裁法（平成十五年法律第百三十八号）第二条第一項に規定する仲裁合意をいう。）
をすること。

####
相続の承認若しくは放棄又は遺産の分割をすること。

####
贈与の申込みを拒絶し、遺贈を放棄し、負担付贈与の申込みを承諾し、
又は負担付遺贈を承認すること。

####
新築、改築、増築又は大修繕をすること。

####
第六百二条に定める期間を超える賃貸借をすること。

####
前各号に掲げる行為を制限行為能力者
（未成年者、成年被後見人、被保佐人及び第十七条第一項の審判を受けた被補助人を
いう。以下同じ。）の法定代理人としてすること。

###
家庭裁判所は、
第十一条本文に規定する者又は保佐人若しくは保佐監督人の請求により、
被保佐人が前項各号に掲げる行為以外の行為をする場合であっても
その保佐人の同意を得なければならない旨の審判をすることができる。
ただし、第九条ただし書に規定する行為については、この限りでない。

###
保佐人の同意を得なければならない行為について、
保佐人が被保佐人の利益を害するおそれがないにもかかわらず同意をしないときは、
家庭裁判所は、被保佐人の請求により、保佐人の同意に代わる許可を与えることができる。

###
保佐人の同意を得なければならない行為であって、
その同意又はこれに代わる許可を得ないでしたものは、取り消すことができる。
'''

SAMPLE_SETTLEMENT = '''
# 和解契約書

v=+1.0
★（以下「甲」という。）と
★（以下「乙」という。）は、
★に関し、次のとおり和解した。

## （★）

★

###
★

###
★

## （債務）

乙は、甲に対し、
★万★円の債務を負っていることを認める。

## （支払）

乙は、甲に対し、
令和★年★月★日限り、
前条の★万★円を下記の口座に振り込んで支払う。
ただし、振込手数料は乙の負担とする。

<=-1.0 v=+0.5
金融機関__　　　　　　　　　　　__　本支店名__　　　　　　　　　　　__

<=-1.0 v=+0.5
普通・当座等__　　　　　　　__　口座番号__　　　　　　　　　　　　　__

<=-1.0 v=+0.5
<名義/フリガナ>__　　　　　　　　　　　　　　　　　　　　　　　　　　　　　__

## （清算条項）

甲と乙は、甲と乙の間には、
★本件に関し、
上記各条項に定めるほか、何らの債権債務のないことを相互に確認する。

v=+1.0
本和解の成立を証するため、本書を★通作成し、各自1通を所持するものとする。

# ##=1 ###=1

v=+1.0
: 令和★年★月★日

v=+1.0
: 甲　　　　★
: \\　　　　　　　　　　　★　　　　　　　　　　　　　　　　　　　　㊞

: ★代理人　★
: \\　　　　　　　弁護士　★　　　　　　　　　　　　　　　　　　　　㊞

v=+0.5
: ★　住所　^DDD^__　　　　　　　　　　　　　　　　　　　　　　　　　　　　__^DDD^

v=+0.5
: \\　　氏名　__　　　　　　　　　　　　　　　　　　　　　　　　　　　㊞__
'''

SAMPLE_PETITION = '''
# 訴状

v=+0.5
令和★年★月★日 :

v=+0.5
: ★裁判所　御中

v=+0.5
<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　原告★訴訟代理人弁護士　　　　★
: \\　　　　　　　　　　　　　　　同　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　　　同（担当）　　　　　　　　★

v=+0.5
: 〒★－★　広島市★
<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　原告　　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　上記代表者代表取締役　　　　　★

: 〒★－★　広島市★
<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　原告　　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　上記代表者代表取締役　　　　　★

: 〒★－★　広島市★
: \\　　　　　　　★法律事務所（送達場所）
<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　原告★訴訟代理人弁護士　　　　★
: \\　　　　　　　　　　　　　　　同　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　　　同（担当）　　　　　　　　★
: \\　　　　　　　TEL ★－★－★　　FAX ★－★－★

: 〒★－★　広島市★
<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　被告　　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　上記代表者代表取締役　　　　　★

: 〒★－★　広島市★
<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　被告　　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　上記代表者代表取締役　　　　　★

v=+1.0
: ★請求事件
: 訴訟物の価額　　★★★★万★★★★円
: 貼用印紙額　　　　　　★万★★★★円

## 請求の趣旨

###
被告★は、原告に対し、★連帯して、
★円及びこれに対する令和★年★月★日から支払済みまで年３分
の割合による金員を支払え。

###
訴訟費用は被告★の負担とする。

<<=1 <=1
との判決並びに仮執行の宣言を求める。

# 請求の原因

### ★について

★

### ★について

★

### まとめ

よって、原告は、被告★に対し、不法行為に基づき、
損害金★円及びこれに対する本件事故日である
令和★年★月★日から支払済みまで民法所定年３分
の割合による遅延損害金の支払を求める。

v=+1.0
# ##=1 ###=1

: 証拠方法 :

### 甲第１号証　　　　　★
### 甲第２号証　　　　　★
### 甲第３号証　　　　　★
### 甲第４号証　　　　　★
### 甲第５号証　　　　　★
### 甲第６号証　　　　　★
### 甲第７号証　　　　　★
### 甲第８号証　　　　　★
### 甲第９号証　　　　　★
### 甲第１０号証　　　　★
### 甲第１１号証　　　　★

v=+1.0
# ##=1 ###=1

: 附属書類 :

### 訴状副本　　　　　　　　　　　　　　★通<!--[被告の数]-->
### 資格証明書　　　　　　　　　　　　　★通<!--[法人当事者の数]-->
### 訴訟委任状　　　　　　　　　　　　　★通<!--[原告の数]-->
### 甲号証の写し　　　　　　　　　　　各★通<!--[被告の数＋1]-->
'''

SAMPLE_EVIDENCE = '''
: 令和★年（★）第★号　★請求事件
: 原告　★
: 被告　★

v=+0.5
# 証拠説明書

v=+0.5
令和★年★月★日 :

v=+0.5
: ★裁判所　御中

v=+0.5

<!-------q1--------q2--------q3--------q4--------q5--------q6--------q7-->
: \\　　　　　　　　　　　　　★★★訴訟代理人弁護士　　　　★
: \\　　　　　　　　　　　　　　　同　　　　　　　　　　　　★
: \\　　　　　　　　　　　　　　　同（担当）　　　　　　　　★

v=+1.0
--
| 号証|   標目   |原写| 作成日 |   作成者   |        立証趣旨        |   備考   |
=
|:----|:---------|:--:|:-------|:-----------|:-----------------------|:---------|
<!-- ↓ 改行しない書き方 -->
|★1  |★書      |写し|R★.★.★|★         |①★であったこと<br>②★であったこと||
|★2  |★書      |写し|R★.★.★|★         |①★であったこと<br>②★であったこと||
<!-- ↓ 改行する書き方 -->
|★3  |★書      |写し|R★.★.★
                               |★          |①★であったこと<br>
                                             ②★であったこと        |          |
|★4  |★書      |写し|R★.★.★
                               |★          |①★であったこと<br>
                                             ②★であったこと        |          |
--
'''

SAMPLE_INDICTMENT = '''
令和★年検第★号 :

v=+0.5
# 起訴状

v=+0.5
: ★地方裁判所　殿

v=+0.5 <=-20.0
: ★地方検察庁
: \\　　検察官検事　★★　★★

v=+1.0
: 下記被告事件につき公訴を提起する。

: 記 :

: 本籍　★
: 住居　★
: 職業　★
: \\　　　　　　　　　　　　　　　勾留中（★拘置所）
: \\　　　　　　　　　　　　　　　　　　　　　　　　　★★　★★
: \\　　　　　　　　　　　　　　　　　　　　　　　　　平成★年★月★日生

v=+1.0
: 公訴事実 :

被告人は、

## ★し

## ★し

#

<<=1.0
たものである。

: 罪名及び罰条 :

<=-8.0
: 第1　★　　　　　　　　　　　★法第★条
: 第2　★　　　　　　　　　　　★法第★条
'''

SAMPLE_CIVIL_JUDGEMENT = '''
: 令和★年★月★日判決言渡　同日原本領収　裁判所書記官
: 令和★年（ワ）第★号　★請求事件
: 口頭弁論終結の日　令和★年★月★日

: 判決 :

<=-4.0
: ★県★市★
: \\　　　　　　　　原告　　　　　　　　　　　　　★★　★★
: \\　　　　　　　　同代表者代表取締役　　　　　　★★　★★
: \\　　　　　　　　同訴訟代理人弁護士　　　　　　★★　★★
: ★県★市★
: \\　　　　　　　　被告　　　　　　　　　　　　　★★　★★
: \\　　　　　　　　同代表者代表取締役　　　　　　★★　★★
: \\　　　　　　　　同訴訟代理人弁護士　　　　　　★★　★★

: 主文 :

<=-4.0
###
被告は、原告に対し、
★万円及びこれに対する令和★年★月★日から支払済みまで年3分の割合による
金員を支払え。

<=-4.0
###
訴訟費用は被告の負担とする。

<=-4.0
###
この判決は第1項に限り仮に執行することができる。

## 請求の趣旨

★

## 事案の概要

★

## 当裁判所の判断

★

## 結論

★

よって、主文のとおり判決する。

: \\　　　　★裁判所
: \\　　　　　　裁判官
'''

SAMPLE_CRIMINAL_JUDGEMENT = '''
: 令和★年★月★日宣告　裁判所書記官
: 令和★年（わ）第★号

: 判決 :

<=-1.0
: 本籍　★
: 住居　★
: 職業　★
: \\　　　　　　　　　　　　　　　　　　　　　　　　　★★　★★
: \\　　　　　　　　　　　　　　　　　　　　　　　　（★★　★★）
: \\　　　　　　　　　　　　　　　　　　　　　　　　　平成★年★月★日生

上記の者に対する★被告事件につき、当裁判所は、
検察官★、
弁護人★
出席の上審理し、次のとおり判決する。

: 主文 :

<=-5.0
: 被告を懲役★年★月に処する。
: この裁判が確定した日から★年間その刑の執行を猶予する。

: 理由 :

: （罪となるべき事実）

## 被告人は、★したものである。

## 被告人は、★したものである。

: （証拠の標目）

第★回公判調書中の被告人の供述部分

被害届（甲1）、★

: （事実認定の補足説明）

### 争点

★

### 前提となる事実

★

### 当裁判所の判断

★

### 結論

: （法令の適用）

: 罰<7.0>条　　刑法★条
: 刑<1.0>種<1.0>の<1.0>選<1.0>択　　懲役刑を選択
: 刑の全部の執行猶予　　刑法25条1項
: 訴<.143>訟<.143>費<.143>用<.143>の<.143>不<.143>負<.143>担　　刑事訴訟法181条1項ただし書
<!--
: 罰条　　　　　　　　　刑法★条
: 刑種の選択　　　　　　懲役刑を選択
: 刑の全部の執行猶予　　刑法25条1項
: 訴訟費用の不負担　　　刑事訴訟法181条1項ただし書
-->

: （量刑の理由）

★

: （検察官の求刑）

: 令和★年★月★日

: ★地方裁判所
: \\　　裁判官
'''

DONT_EDIT_MESSAGE = '<!--【以下は必要なデータですので編集しないでください】-->'


######################################################################
# FUNCTION


def get_real_width(s: str) -> int:
    wid = 0
    for c in s:
        if c == '\t':
            wid += (int(wid / TAB_WIDTH) + 1) * TAB_WIDTH
            continue
        w = unicodedata.east_asian_width(c)
        if c == '':
            wid += 0
        elif re.match('^[☐☑]$', c):
            wid += 2
        elif re.match('^[´¨―‐∥…‥‘’“”±×÷≠≦≧∞∴♂♀°′″℃§]$', c):
            wid += 2
        elif re.match('^[☆★○●◎◇◆□■△▲▽▼※→←↑↓]$', c):
            wid += 2
        elif re.match('^[∈∋⊆⊇⊂⊃∪∩∧∨⇒⇔∀∃∠⊥⌒∂∇≡≒≪≫√∽∝∵]$', c):
            wid += 2
        elif re.match('^[∫∬Å‰♯♭♪†‡¶◯]$', c):
            wid += 2
        elif re.match('^[ΑΒΓΔΕΖΗΘΙΚΛΜΝΞΟΠΡΣΤΥΦΧΨΩ]$', c):
            wid += 2
        elif re.match('^[αβγδεζηθικλμνξοπρστυφχψω]$', c):
            wid += 2
        elif re.match('^[АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ]$', c):
            wid += 2
        elif re.match('^[абвгдеёжзийклмнопрстуфхцчшщъыьэюя]$', c):
            wid += 2
        elif re.match('^[─│┌┐┘└├┬┤┴┼━┃┏┓┛┗┣┳┫┻╋┠┯┨┷┿┝┰┥┸╂]$', c):
            wid += 2
        elif re.match('^[№℡≒≡∫∮∑√⊥∠∟⊿∵∩∪]$', c):
            wid += 2
        elif re.match('^[⑴⑵⑶⑷⑸⑹⑺⑻⑼⑽⑾⑿⒀⒁⒂⒃⒄⒅⒆⒇]$', c):
            wid += 2
        elif re.match('^[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳]$', c):
            wid += 2
        elif re.match('^[⒈⒉⒊⒋⒌⒍⒎⒏⒐⒑⒒⒓⒔⒕⒖⒗⒘⒙⒚⒛]$', c):
            wid += 2
        elif re.match('^[ⅰⅱⅲⅳⅴⅵⅶⅷⅸⅹⅺⅻ]$', c):
            wid += 2
        elif re.match('^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩⅪⅫ]$', c):
            wid += 2
        elif re.match('^[⒜⒝⒞⒟⒠⒡⒢⒣⒤⒥⒦⒧⒨⒩⒪⒫⒬⒭⒮⒯⒰⒱⒲⒳⒴⒵]$', c):
            wid += 2
        elif re.match('^[ⓐⓑⓒⓓⓔⓕⓖⓗⓘⓙⓚⓛⓜⓝⓞⓟⓠⓡⓢⓣⓤⓥⓦⓧⓨⓩ]$', c):
            wid += 2
        elif re.match('^[🄐🄑🄒🄓🄔🄕🄖🄗🄘🄙🄚🄛🄜🄝🄞🄟🄠🄡🄢🄣🄤🄥🄦🄧🄨🄩]$', c):
            wid += 2
        elif re.match('^[ⒶⒷⒸⒹⒺⒻⒼⒽⒾⒿⓀⓁⓂⓃⓄⓅⓆⓇⓈⓉⓊⓋⓌⓍⓎⓏ]$', c):
            wid += 2
        elif re.match('^[㉑㉒㉓㉔㉕㉖㉗㉘㉙㉚㉛㉜㉝㉞㉟㊱㊲㊳㊴㊵㊶㊷㊸㊹㊺㊻㊼㊽㊾㊿]$', c):
            wid += 2
        elif re.match('^[🄋➀➁➂➃➄➅➆➇➈➉]$', c):
            wid += 2
        elif re.match('^[㋐㋑㋒㋓㋔㋕㋖㋗㋘㋙㋚㋛㋜㋝㋞㋟㋠㋡㋢㋣㋤㋥㋦㋧㋨]$', c):
            wid += 2
        elif re.match('^[㋩㋪㋫㋬㋭㋮㋯㋰㋱㋲㋳㋴㋵㋶㋷㋸㋹㋺㋻㋼㋽㋾]$', c):
            wid += 2
        elif re.match('^[㊀㊁㊂㊃㊄㊅㊆㊇㊈㊉]$', c):
            wid += 2
        elif (w == 'F'):  # Full alphabet ...
            wid += 2
        elif(w == 'H'):   # Half katakana ...
            wid += 1
        elif(w == 'W'):   # Chinese character ...
            wid += 2
        elif(w == 'Na'):  # Half alphabet ...
            wid += 1
        elif(w == 'A'):   # Greek character ...
            wid += 1
        elif(w == 'N'):   # Arabic character ...
            wid += 1
    return wid


def c2n_n_arab(s: str) -> int:
    n = 0
    for c in s:
        n *= 10
        if re.match('^[0-9]$', c):
            n += int(c)
        elif re.match('^[０-９]$', c):
            n += ord(c) - 65296
        else:
            return -1
    return n


def c2n_n_kata(s: str) -> int:
    i = -1
    if len(s) == 1:
        i = ord(s)
    n = 65392
    if i >= n + 1 and i <= n + 44:
        # ｱｲｳｴｵｶｷｸｹｺｻｼｽｾｿﾀﾁﾂﾃﾄﾅﾆﾇﾈﾉﾊﾋﾌﾍﾎﾏﾐﾑﾒﾓﾔﾕﾖﾗﾘﾙﾚﾛﾜ
        return i - n
    n = 65337
    if i == n + 45:
        # ｦ
        return i - n
    n = 65391
    if i == n + 46:
        # ﾝ
        return i - n
    n = 12448
    if i >= n + 2 * 1 and i <= n + 2 * 5:
        # アイウエオ
        return int((i - n) / 2)
    n = 12447
    if i >= n + 2 * 6 and i <= n + 2 * 17:
        # カキクケコサシスセソタチ
        return int((i - n) / 2)
    n = 12448
    if i >= n + 2 * 18 and i <= n + 2 * 20:
        # ツテト
        return int((i - n) / 2)
    n = 12469
    if i >= n + 1 * 21 and i <= n + 1 * 25:
        # ナニヌネノ
        return int((i - n) / 1)
    n = 12417
    if i >= n + 3 * 26 and i <= n + 3 * 30:
        # ハヒフヘホ
        return int((i - n) / 3)
    n = 12479
    if i >= n + 1 * 31 and i <= n + 1 * 35:
        # マミムメモ
        return int((i - n) / 1)
    n = 12444
    if i >= n + 2 * 36 and i <= n + 2 * 38:
        # ヤユヨ
        return int((i - n) / 2)
    n = 12482
    if i >= n + 1 * 39 and i <= n + 1 * 43:
        # ラリルレロ
        return int((i - n) / 1)
    n = 12483
    if i >= n + 1 * 44 and i <= n + 1 * 49:
        # ワヰヱヲン
        return int((i - n) / 1)
    return -1


def c2n_n_alph(s: str) -> int:
    i = -1
    if len(s) == 1:
        i = ord(s)
    n = 96
    if i >= n + 1 and i <= n + 26:
        # a...z
        return i - n
    n = 65344
    if i >= n + 1 and i <= n + 26:
        # ａ...ｚ
        return i - n
    return -1


def c2n_n_kanj(s: str) -> int:
    i = s
    i = re.sub('[０〇零]', '0', i)
    i = re.sub('[１一壱]', '1', i)
    i = re.sub('[２二弐]', '2', i)
    i = re.sub('[３三参]', '3', i)
    i = re.sub('[４四]', '4', i)
    i = re.sub('[５五伍]', '5', i)
    i = re.sub('[６六]', '6', i)
    i = re.sub('[７七]', '7', i)
    i = re.sub('[８八]', '8', i)
    i = re.sub('[９九]', '9', i)
    #
    i = re.sub('[拾]', '十', i)
    i = re.sub('[佰陌]', '百', i)
    i = re.sub('[仟阡]', '千', i)
    i = re.sub('[萬]', '万', i)
    #
    i = re.sub('^([千百十])', '1\\1', i)
    i = re.sub('([^0-9])([千百十])', '\\1 1\\2', i)
    #
    i = re.sub('(万)([^千]*)$', '\\1 0千\\2', i)
    i = re.sub('(千)([^百]*)$', '\\1 0百\\2', i)
    i = re.sub('(百)([^十]*)$', '\\1 0十\\2', i)
    i = re.sub('(十)$', '\\1 0', i)
    #
    i = re.sub('[万千百十 ]', '', i)
    #
    if re.match('^[0-9]+$', i):
        return int(i)
    return -1


def adjust_line(old_doc: str) -> str:
    old_doc = re.sub('\n\n+', '\n\n', old_doc)
    new_doc = ''
    for old_lin in old_doc.split('\n'):
        old_lin = re.sub('([,，、.．。]+)', '\\1\n', old_lin)
        old_lin = re.sub('([0-9][,.])\n([0-9])', '\\1\\2', old_lin)
        old_lin = re.sub('([０-９][，．])\n([０-９])', '\\1\\2', old_lin)
        # old_lin = re.sub('(を)', '\\1\n', old_lin)
        old_lin = re.sub('([「『（\\(]+)', '\n\\1', old_lin)
        old_lin = re.sub('([\\)）』」]+)', '\\1\n', old_lin)
        old_lin = re.sub('([ \t\u3000]+)\n([「『（\\(])', '\\1\\2', old_lin)
        old_lin = re.sub('([.．。]+)\n([\\)）』」])', '\\1\\2', old_lin)
        old_lin = re.sub('^\n+', '', old_lin)
        old_lin = re.sub('\n+$', '', old_lin)
        old_lin = re.sub('\n+', '\n', old_lin)
        new_lin = ''
        for phr in old_lin.split('\n'):
            if get_real_width(new_lin + phr) > MD_TEXT_WIDTH:
                new_doc += new_lin + '\n'
                new_lin = ''
            new_lin += phr
        new_doc += new_lin + '\n'
    new_doc = re.sub('\n+$', '', new_doc)
    return new_doc


def count_days(date: str) -> int:
    res = '([MTSHR]?)([0-9]+)-([0-9]+)-([0-9]+)'
    era = re.sub(res, '\\1', date)
    yea = re.sub(res, '\\2', date)
    mon = re.sub(res, '\\3', date)
    day = re.sub(res, '\\4', date)
    if era == 'M':
        yea = str(int(yea) + 1867)
    elif era == 'T':
        yea = str(int(yea) + 1911)
    elif era == 'S':
        yea = str(int(yea) + 1925)
    elif era == 'H':
        yea = str(int(yea) + 1988)
    elif era == 'R':
        yea = str(int(yea) + 2018)
    if int(yea) < 100:
        yea = str(int(yea) + 2000)
    ymd_hms = yea + '-' + mon + '-' + day + ' 09:00:00 UTC'
    date = datetime.datetime.strptime(ymd_hms, '%Y-%m-%d %H:%M:%S %Z')
    unix_time = date.timestamp()
    days: int = round(unix_time / 86400)
    return days


######################################################################
# CLASS


############################################################
# SIMPLE DAILOG


class OneWordDialog(tkinter.simpledialog.Dialog):

    def __init__(self, pane, mother, title, prompt, head, tail,
                 init='', cand=[]):
        self.pane = pane
        self.mother = mother
        self.prompt = prompt
        self.head = head
        self.tail = tail
        self.init = init
        self.cand = cand
        self.value = None
        self.cand.append(init)
        self.cnum = len(cand) - 1
        super().__init__(pane, title=title)

    def body(self, pane):
        fon = self.mother.gothic_font
        prompt = tkinter.Label(pane, text=self.prompt + '\n', justify='left')
        prompt.pack(side='top', anchor='w')
        frm = tkinter.Frame(pane)
        frm.pack()
        tkinter.Label(frm, text=self.head).pack(side='left')
        self.entry = tkinter.Entry(frm, width=25, font=fon)
        self.entry.pack(side='left')
        self.entry.insert(0, self.init)
        tkinter.Label(frm, text=self.tail).pack(side='left')
        self.entry.bind('<Key>', self.entry_key)
        self.entry.bind('<Button-1>', self.entry_button0)
        self.entry.bind('<Button-2>', self.entry_button0)
        self.entry.bind('<Button-3>', self.entry_button3)
        super().body(pane)
        return self.entry

    def apply(self):
        self.value = self.entry.get()
        if (len(self.cand) > 1) and (self.cand[-2] == self.value):
            self.cand.pop(-1)
        else:
            self.cand[-1] = self.value

    def get_value(self):
        return self.value

    def entry_key(self, key):
        k = Makdo._get_key(key)
        if k is None:
            return 'break'  # shift, control, alt, mode_switch
        elif self.mother._is_key(k, 'Up', 'C-r', 'C-o'):
            if self.cnum == len(self.cand) - 1:
                self.cand[-1] = self.entry.get()
            if self.cnum > 0:
                self.cnum -= 1
                self.entry.delete(0, 'end')
                self.entry.insert(0, self.cand[self.cnum])
            return 'break'
        elif self.mother._is_key(k, 'Down', 'C-n', 'C-l'):
            # if self.cnum == len(self.cand) - 1:
            #     self.cand[-1] = self.entry.get()
            if self.cnum < len(self.cand) - 1:
                self.cnum += 1
                self.entry.delete(0, 'end')
                self.entry.insert(0, self.cand[self.cnum])
            return 'break'
        elif self.mother._is_key(k, 'F15', 'C-g', 'C-u', 'C-v'):
            self.entry_paste_word()
            return 'break'

    def entry_button0(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry.focus_force()

    def entry_button3(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry.focus_force()
        self.mother.bt3 = tkinter.Menu(self, tearoff=False)
        e = self.entry.get()
        if e != '':
            self.mother.bt3.add_command(label='コピー',
                                        command=self.entry_copy_word)
            self.mother.bt3.add_separator()
        self.mother.bt3.add_command(label='貼り付け',
                                    command=self.entry_paste_word)
        self.mother.bt3.post(click.x_root, click.y_root)

    def entry_copy_word(self):
        e = self.entry.get()
        if e != '':
            self.mother._clipboard_append(e)

    def entry_paste_word(self):
        try:
            cb = self.mother.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb != '':
            self.entry.insert('insert', cb)


class TwoWordsDialog(tkinter.simpledialog.Dialog):

    def __init__(self, pane, mother, title, prompt, head1, head2, tail1, tail2,
                 init1='', init2='', cand1=[], cand2=[]):
        self.pane = pane
        self.mother = mother
        self.prompt = prompt
        self.head1 = head1
        self.tail1 = tail1
        self.head2 = head2
        self.tail2 = tail2
        self.init1 = init1
        self.init2 = init2
        self.cand1 = cand1
        self.cand2 = cand2
        self.value1 = None
        self.value2 = None
        self.cand1.append(init1)
        self.cnum1 = len(cand1) - 1
        self.cand2.append(init2)
        self.cnum2 = len(cand2) - 1
        super().__init__(pane, title=title)

    def body(self, pane):
        fon = self.mother.gothic_font
        prompt = tkinter.Label(pane, text=self.prompt + '\n', justify='left')
        prompt.pack(side='top', anchor='w')
        frm = tkinter.Frame(pane)
        frm.pack()
        tkinter.Label(frm, text=self.head1).pack(side='left')
        self.entry1 = tkinter.Entry(frm, width=25, font=fon)
        self.entry1.pack(side='top')
        self.entry1.insert(0, self.init1)
        tkinter.Label(frm, text=self.tail1).pack(side='left')
        frm = tkinter.Frame(pane)
        frm.pack()
        tkinter.Label(frm, text=self.head2).pack(side='left')
        self.entry2 = tkinter.Entry(frm, width=25, font=fon)
        self.entry2.pack(side='top')
        self.entry2.insert(0, self.init2)
        tkinter.Label(frm, text=self.tail2).pack(side='left')
        self.entry1.bind('<Key>', self.entry1_key)
        self.entry1.bind('<Button-1>', self.entry1_button0)
        self.entry1.bind('<Button-2>', self.entry1_button0)
        self.entry1.bind('<Button-3>', self.entry1_button3)
        self.entry2.bind('<Key>', self.entry2_key)
        self.entry2.bind('<Button-1>', self.entry2_button0)
        self.entry2.bind('<Button-2>', self.entry2_button0)
        self.entry2.bind('<Button-3>', self.entry2_button3)
        super().body(pane)
        return self.entry1

    def apply(self):
        self.value1 = self.entry1.get()
        self.value2 = self.entry2.get()
        if (len(self.cand1) > 1) and (self.cand1[-2] == self.value1):
            self.cand1.pop(-1)
        else:
            self.cand1[-1] = self.value1
        if (len(self.cand2) > 1) and (self.cand2[-2] == self.value2):
            self.cand2.pop(-1)
        else:
            self.cand2[-1] = self.value2

    def get_value(self):
        return self.value1, self.value2

    def entry1_key(self, key):
        k = Makdo._get_key(key)
        if k is None:
            return 'break'  # shift, control, alt, mode_switch
        if self.mother._is_key(k, 'Up', 'C-r', 'C-o'):
            if self.cnum1 == len(self.cand1) - 1:
                self.cand1[-1] = self.entry1.get()
            if self.cnum1 > 0:
                self.cnum1 -= 1
                self.entry1.delete(0, 'end')
                self.entry1.insert(0, self.cand1[self.cnum1])
            return 'break'
        elif self.mother._is_key(k, 'Down', 'C-n', 'C-l'):
            # if self.cnum1 == len(self.cand1) - 1:
            #     self.cand1[-1] = self.entry1.get()
            if self.cnum1 < len(self.cand1) - 1:
                self.cnum1 += 1
                self.entry1.delete(0, 'end')
                self.entry1.insert(0, self.cand1[self.cnum1])
            return 'break'
        elif self.mother._is_key(k, 'F15', 'C-g', 'C-u', 'C-v'):
            self.entry1_paste_word()
            return 'break'

    def entry2_key(self, key):
        k = Makdo._get_key(key)
        if k is None:
            return 'break'  # shift, control, alt, mode_switch
        if self.mother._is_key(k, 'Up', 'C-r', 'C-o'):
            if self.cnum2 == len(self.cand2) - 1:
                self.cand2[-1] = self.entry2.get()
            if self.cnum2 > 0:
                self.cnum2 -= 1
                self.entry2.delete(0, 'end')
                self.entry2.insert(0, self.cand2[self.cnum2])
            return 'break'
        elif self.mother._is_key(k, 'Down', 'C-n', 'C-l'):
            # if self.cnum2 == len(self.cand2) - 1:
            #     self.cand2[-1] = self.entry2.get()
            if self.cnum2 < len(self.cand2) - 1:
                self.cnum2 += 1
                self.entry2.delete(0, 'end')
                self.entry2.insert(0, self.cand2[self.cnum2])
            return 'break'
        elif self.mother._is_key(k, 'F15', 'C-g', 'C-u', 'C-v'):
            self.entry2_paste_word()
            return 'break'

    def entry1_button0(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry1.focus_force()

    def entry2_button0(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry2.focus_force()

    def entry1_button3(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry1.focus_force()
        self.mother.bt3 = tkinter.Menu(self, tearoff=False)
        e = self.entry1.get()
        if e != '':
            self.mother.bt3.add_command(label='コピー',
                                        command=self.entry1_copy_word)
            self.mother.bt3.add_separator()
        self.mother.bt3.add_command(label='貼り付け',
                                    command=self.entry1_paste_word)
        self.mother.bt3.post(click.x_root, click.y_root)

    def entry2_button3(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry2.focus_force()
        self.mother.bt3 = tkinter.Menu(self, tearoff=False)
        e = self.entry2.get()
        if e != '':
            self.mother.bt3.add_command(label='コピー',
                                        command=self.entry2_copy_word)
            self.mother.bt3.add_separator()
        self.mother.bt3.add_command(label='貼り付け',
                                    command=self.entry2_paste_word)
        self.mother.bt3.post(click.x_root, click.y_root)

    def entry1_copy_word(self):
        e = self.entry1.get()
        if e != '':
            self.mother._clipboard_append(e)

    def entry1_paste_word(self):
        try:
            cb = self.mother.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb != '':
            self.entry1.insert('insert', cb)

    def entry2_copy_word(self):
        e = self.entry2.get()
        if e != '':
            self.mother._clipboard_append(e)

    def entry2_paste_word(self):
        try:
            cb = self.mother.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb != '':
            self.entry2.insert('insert', cb)


class RadiobuttonDialog(tkinter.simpledialog.Dialog):

    def __init__(self, pane, mother, title, prompt, cand, init=-1):
        self.pane = pane
        self.mother = mother
        self.prompt = prompt
        self.cand = cand
        self.init = init
        self.value = None
        self.has_pressed_ok = False
        super().__init__(pane, title=title)

    def body(self, pane):
        prompt = tkinter.Label(pane, text=self.prompt + '\n', justify='left')
        prompt.pack(side='top', anchor='w')
        m = len(self.cand) - 1
        self.value = tkinter.IntVar()
        self.value.set(self.init)
        focused = None
        for i in range(len(self.cand)):
            head = re.sub('\n', ' ', self.cand[i])
            if head == '':
                head = '（空）'
            if len(head) > 15:
                head = head[:14] + '…'
            rb = tkinter.Radiobutton(pane, text=head,
                                     variable=self.value, value=i)
            rb.pack(side='top', anchor='w')
            if i == self.init:
                focused = rb
        self.bind('<Key-Up>', self.process_Up)
        self.bind('<Key-Down>', self.process_Down)
        super().body(pane)
        return focused

    def process_Up(self, event):
        event.widget.tk_focusPrev().focus()
        return 'break'

    def process_Down(self, event):
        event.widget.tk_focusNext().focus()
        return 'break'

    def apply(self):
        self.has_pressed_ok = True

    def get_value(self):
        if self.has_pressed_ok:
            return self.cand[self.value.get()]
        else:
            return None


class NumberRevisersDialog(tkinter.simpledialog.Dialog):

    def __init__(self, pane, mother, title, prompt, confs):
        self.pane = pane
        self.mother = mother
        self.prompt = prompt
        self.confs = confs
        self.values = []
        self.entries = []
        super().__init__(pane, title=title)

    def body(self, pane):
        frm1 = tkinter.Frame(pane)
        frm1.pack()
        prompt = tkinter.Label(frm1, text=self.prompt + '\n', justify='left')
        prompt.pack(side='top', anchor='w')
        frm2 = tkinter.Frame(pane)
        frm2.pack()
        for n, conf in enumerate(self.confs):
            entry = self._body(frm2, n, conf)
            self.entries.append(entry)
        return self.entries[0]

    def _body(self, frm, n, conf):
        fon = self.mother.gothic_font
        txt = tkinter.Label(frm, text=conf[0] + conf[1] + conf[2])
        txt.grid(row=n, column=0)
        txt = tkinter.Label(frm, text='　→　')
        txt.grid(row=n, column=1)
        txt = tkinter.Label(frm, text=conf[3])
        txt.grid(row=n, column=2)
        entry = tkinter.Entry(frm, width=3, justify='center', font=fon)
        entry.grid(row=n, column=3)
        entry.insert(0, conf[4])
        txt = tkinter.Label(frm, text=conf[5])
        txt.grid(row=n, column=4)
        return entry

    def apply(self):
        for e in self.entries:
            v = e.get()
            self.values.append(v)

    def get_values(self):
        return self.values


class PasswordDialog(tkinter.simpledialog.Dialog):

    def __init__(self, pane, mother, title, prompt):
        self.pane = pane
        self.mother = mother
        self.prompt = prompt
        self.value = None
        super().__init__(pane, title=title)

    def body(self, pane):
        fon = self.mother.gothic_font
        prompt = tkinter.Label(pane, text=self.prompt + '\n', justify='left')
        prompt.pack(side='top', anchor='w')
        self.entry = tkinter.Entry(pane, width=25, font=fon, show='*')
        self.entry.pack(side='left')
        self.entry.bind('<Key>', self.entry_key)
        self.entry.bind('<Button-1>', self.entry_button0)
        self.entry.bind('<Button-2>', self.entry_button0)
        self.entry.bind('<Button-3>', self.entry_button3)
        super().body(pane)
        return self.entry

    def apply(self):
        self.value = self.entry.get()

    def get_value(self):
        return self.value

    def entry_key(self, key):
        k = Makdo._get_key(key)
        if k is None:
            return 'break'  # shift, control, alt, mode_switch
        if self.mother._is_key(k, 'F15', 'C-g', 'C-u', 'C-v'):
            self.entry_paste_word()
            return 'break'

    def entry_button0(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry.focus_force()

    def entry_button3(self, click):
        self.mother.close_mouse_menu()  # close mouse menu
        self.entry.focus_force()
        self.mother.bt3 = tkinter.Menu(self, tearoff=False)
        self.mother.bt3.add_command(label='貼り付け',
                                    command=self.entry_paste_word)
        self.mother.bt3.post(click.x_root, click.y_root)

    def entry_paste_word(self):
        try:
            cb = self.mother.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb != '':
            self.entry.insert('insert', cb)


############################################################
# WITCH


class Witch:

    # 00-94 (32-126)
    pepper = [
        42, 84, 62, 18, 76,  4, 68, 37, 53, 26, 59, 71, 23, 27, 33, 29,
        51, 20, 25, 76, 89, 30, 90, 86, 45, 74,  6, 42, 14,  7, 34, 51,
        31, 31, 13, 74, 68, 32, 41, 44, 17, 39, 34,  4, 41, 25, 79, 94,
        56, 61, 23, 42, 58, 44, 79, 91, 38,  7, 42, 14,  7, 11, 12, 75,
        43, 71,  5,  1,  4, 42, 45, 32, 68, 83, 42,  5, 52, 13, 32, 47,
        39,  7, 48, 90,  1,  1, 53, 80, 42, 57, 64, 56,  5, 82, 30, 15,
        12, 82, 51, 48, 43, 63, 12, 14, 20, 62, 93, 55, 13, 24, 68, 63,
        71, 30, 79, 20, 22, 42, 29, 81, 56, 61, 70, 37, 35, 37, 27, 37,
        57, 82, 58, 71, 83,  4, 57, 62,  3, 31, 40, 48, 21, 51, 87, 49,
        38, 27, 48,  7, 54, 35, 45, 58, 85, 35, 39, 11, 88, 37, 18, 90,
        90, 21, 66, 56, 18, 91, 36, 71, 63, 48, 46, 75, 52, 65, 12, 33,
        42, 72, 41, 31, 86, 59, 24, 56, 27, 94, 23, 47, 92, 42, 15, 15,
        40, 27, 62, 53, 65, 59, 36, 38, 93, 21, 37, 32, 43, 55, 77, 64,
        17, 67, 48, 88, 74, 75, 67,  9, 94, 84,  4,  0, 90, 48, 24, 50,
        22,  6, 27, 39, 38, 10, 68, 46, 90,  5, 66, 34,  4, 40, 50, 31,
        93,  5, 54, 89, 43, 44, 54, 57, 90, 26, 60, 61, 33, 33, 45, 28,
    ]

    @staticmethod
    def enchant(dechant_word):
        m = len(dechant_word)
        ns = []
        for i in range(m):
            j = i - m // 2
            if j < 0:
                j += m
            # j = i - 1
            # if j == -1:
            #     j = -1
            c_i = dechant_word[i]
            c_j = dechant_word[j]
            n_i = (ord(c_i) - 32) // 5  # 0-18
            n_j = (ord(c_j) - 32) % 5   # 0-4
            n = (n_j * 19) + n_i        # (4 * 19) + 18 = 94
            # n = (n_i * 5) + n_j        # (18 * 5) + 4 = 94
            ns.append(n)
        enchant_word = ''
        for i in range(m):
            n = ns[i]
            n += Witch.pepper[i % len(Witch.pepper)]
            if n >= 95:
                n -= 95
            e = chr(n + 32)
            enchant_word += e
        return enchant_word

    @staticmethod
    def dechant(enchant_word):
        m = len(enchant_word)
        ns = []
        for i in range(m):
            e = enchant_word[i]
            n = ord(e) - 32
            n -= Witch.pepper[i % len(Witch.pepper)]
            if n < 0:
                n += 95
            ns.append(n)
        dechant_word = ''
        for i in range(m):
            j = i + m // 2
            if j >= m:
                j -= m
            # j = i + 1
            # if j == m:
            #     j = 0
            n_i = ns[i] % 19     # 0 -> 18
            n_j = ns[j] // 19    # 0 -> 4
            # n_i = ns[i] // 5    # 0 -> 18
            # n_j = ns[j] % 5     # 0 -> 4
            n = (n_i * 5) + n_j  # (18 * 5) + 4 = 94
            d = chr(n + 32)
            dechant_word += d
        return dechant_word


############################################################
# CHARS STATE


class CharsState:

    def __init__(self):
        self.del_or_ins = ''
        self.is_in_comment = False
        self.parentheses = []
        self.has_underline = False
        self.has_specific_font = False
        self.has_frame = False
        self.has_ruby = False
        self.standard_size = ''
        self.is_resized = ''
        self.is_stretched = ''
        self.is_in_preformatted = False
        self.is_in_italic = False
        self.is_in_bold = False
        self.script_parenthesis = ''
        self.is_length_reviser = False
        self.chapter_depth = 0
        self.section_depth = 0

    def __eq__(self, other):
        if self.del_or_ins != other.del_or_ins:
            return False
        if self.is_in_comment != other.is_in_comment:
            return False
        if self.parentheses != other.parentheses:
            return False
        if self.has_underline != other.has_underline:
            return False
        if self.has_specific_font != other.has_specific_font:
            return False
        if self.has_frame != other.has_frame:
            return False
        if self.has_ruby != other.has_ruby:
            return False
        if self.standard_size != other.standard_size:
            return False
        if self.is_resized != other.is_resized:
            return False
        if self.is_stretched != other.is_stretched:
            return False
        if self.is_in_preformatted != other.is_in_preformatted:
            return False
        if self.is_in_italic != other.is_in_italic:
            return False
        if self.is_in_bold != other.is_in_bold:
            return False
        if self.script_parenthesis != other.script_parenthesis:
            return False
        return True

    def copy(self):
        copy = CharsState()
        copy.del_or_ins = self.del_or_ins
        copy.is_in_comment = self.is_in_comment
        for p in self.parentheses:
            copy.parentheses.append(p)
        copy.has_underline = self.has_underline
        copy.has_specific_font = self.has_specific_font
        copy.has_frame = self.has_frame
        copy.has_ruby = self.has_ruby
        copy.standard_size = self.standard_size
        copy.is_resized = self.is_resized
        copy.is_stretched = self.is_stretched
        copy.is_in_preformatted = self.is_in_preformatted
        copy.is_in_italic = self.is_in_italic
        copy.is_in_bold = self.is_in_bold
        copy.script_parenthesis = self.script_parenthesis
        copy.is_length_reviser = self.is_length_reviser
        copy.chapter_depth = self.chapter_depth
        copy.section_depth = self.section_depth
        # for v in vars(copy):
        #     vars(copy)[v] = vars(self)[v]
        return copy

    def reset_partially(self):
        self.is_length_reviser = False
        self.chapter_depth = 0
        self.section_depth = 0

    def toggle_is_in_comment(self):
        self.is_in_comment = not self.is_in_comment

    def set_del_or_ins(self, del_or_ins):
        if del_or_ins == 'del':
            if self.del_or_ins == 'del':
                self.del_or_ins = ''
            else:
                self.del_or_ins = 'del'
        if del_or_ins == 'ins':
            if self.del_or_ins == 'ins':
                self.del_or_ins = ''
            else:
                self.del_or_ins = 'ins'

    def toggle_has_underline(self):
        self.has_underline = not self.has_underline

    def toggle_has_specific_font(self):
        self.has_specific_font = not self.has_specific_font

    def attach_or_remove_frame(self, fd):
        if fd == '[|':
            self.has_frame = True
        elif fd == '|]':
            self.has_frame = False

    def set_or_unset_has_ruby(self, has_ruby):
        self.has_ruby = has_ruby

    def set_is_resized(self, fd):
        if fd == '---':
            if self.is_resized == '---':
                self.is_resized = ''
            else:
                self.is_resized = '---'
        elif fd == '--':
            if self.is_resized == '--':
                self.is_resized = ''
            else:
                self.is_resized = '--'
        elif fd == '++':
            if self.is_resized == '++':
                self.is_resized = ''
            else:
                self.is_resized = '++'
        elif fd == '+++':
            if self.is_resized == '+++':
                self.is_resized = ''
            else:
                self.is_resized = '+++'

    def set_is_stretched(self, fd):
        if fd == '>>>':
            if self.is_stretched == '<<<':
                self.is_stretched = ''
            else:
                self.is_stretched = '>>>'
        elif fd == '>>':
            if self.is_stretched == '<<':
                self.is_stretched = ''
            else:
                self.is_stretched = '>>'
        elif fd == '<<':
            if self.is_stretched == '>>':
                self.is_stretched = ''
            else:
                self.is_stretched = '<<'
        elif fd == '<<<':
            if self.is_stretched == '>>>':
                self.is_stretched = ''
            else:
                self.is_stretched = '<<<'

    def toggle_is_in_preformatted(self):
        self.is_in_preformatted = not self.is_in_preformatted

    def toggle_is_in_italic(self):
        self.is_in_italic = not self.is_in_italic

    def toggle_is_in_bold(self):
        self.is_in_bold = not self.is_in_bold

    def set_or_unset_script_parenthesis(self, par):
        if self.script_parenthesis == '':
            if re.match('^{[0-9]?{$', par):
                self.script_parenthesis = par
        else:
            n = self.script_parenthesis.replace('{', '')
            if par == '}' + n + '}':
                self.script_parenthesis = ''

    def apply_parenthesis(self, parenthesis):
        ps = self.parentheses
        p = parenthesis
        if p == '「' or p == '『' or p == '[' or \
           p == '｛' or p == '{' or \
           p == '（' or p == '(':
            ps.append(p)
            return
        if p == ')' or p == '）' or \
           p == '}' or p == '｝' or \
           p == ']' or p == '』' or p == '」':
            if len(ps) > 0:
                if ps[-1] == '(' and p == ')' or \
                   ps[-1] == '（' and p == '）' or \
                   ps[-1] == '{' and p == '}' or \
                   ps[-1] == '｛' and p == '｝' or \
                   ps[-1] == '[' and p == ']' or \
                   ps[-1] == '『' and p == '』' or \
                   ps[-1] == '「' and p == '」':
                    ps.pop(-1)
            return

    def set_chapter_depth(self, depth):
        self.chapter_depth = depth

    def set_section_depth(self, depth):
        self.section_depth = depth

    def get_key(self, chars):
        key = 'c'
        # ANGLE
        if False:
            pass
        elif chars == 'substitute phrase A':  # substitute phrase
            key += '-230'                     # substitute phrase
        elif chars == 'substitute phrase B':  # substitute phrase
            key += '-290'                     # substitute phrase
        elif chars == 'substitute phrase C':  # substitute phrase
            key += '-270'                     # substitute phrase
        elif chars == ' ':
            return 'hsp_tag'
        elif chars == '\u3000':
            return 'fsp_tag'
        elif chars == '\t':
            return 'tab_tag'
        elif self.is_in_comment:
            key += '-0'
        elif chars == 'escape':
            key += '-340'
        elif chars == 'font decorator':
            key += '-120'
        elif chars == 'table':
            key += '-190'
        elif self.is_length_reviser:  # Should be before "half number"
            key += '-150'
        elif chars == 'half number':
            key += '-30'
        elif chars == 'full number':
            key += '-330'
        elif chars == 'list':
            key += '-330'
        elif chars == 'alignment':
            key += '-180'
        elif chars == 'half katakana':
            key += '-050'
        elif re.match('^horizontalline[0-9]{3}$', chars):
            key += '-' + re.sub('^horizontalline0?0?', '', chars)
        elif chars == 'image':
            if len(self.parentheses) == 0:
                key += '-120'
            elif len(self.parentheses) == 1:
                key += '-160'
            elif len(self.parentheses) == 2:
                key += '-180'
            elif len(self.parentheses) == 3:
                key += '-200'
            elif len(self.parentheses) == 4:
                key += '-220'
            elif len(self.parentheses) == 5:
                key += '-240'
            elif len(self.parentheses) == 6:
                key += '-260'
            elif len(self.parentheses) == 7:
                key += '-280'
            elif len(self.parentheses) >= 8:
                key += '-290'
        elif len(self.parentheses) == 1:
            key += '-120'
        elif len(self.parentheses) == 2:
            key += '-160'
        elif len(self.parentheses) == 3:
            key += '-180'
        elif len(self.parentheses) == 4:
            key += '-200'
        elif len(self.parentheses) == 5:
            key += '-220'
        elif len(self.parentheses) == 6:
            key += '-240'
        elif len(self.parentheses) == 7:
            key += '-260'
        elif len(self.parentheses) == 8:
            key += '-280'
        elif len(self.parentheses) >= 9:
            key += '-290'
        elif chars == '<sp>':
            key += '-140'
        elif chars == '<br>' or chars == '<pgbr>':
            key += '-310'
        elif chars == 'hline':
            key += '-310'
        elif chars == 'ruby':
            key += '-110'
        elif chars == 'R' or chars == 'red' or chars == 'DR':
            key += '-0'
        elif chars == 'Y' or chars == 'yellow' or chars == 'DY':
            key += '-60'
        elif chars == 'G' or chars == 'green' or chars == 'DG':
            key += '-120'
        elif chars == 'C' or chars == 'cyan' or chars == 'DC':
            key += '-180'
        elif chars == 'B' or chars == 'blue' or chars == 'DB':
            key += '-240'
        elif chars == 'M' or chars == 'magenta' or chars == 'DM':
            key += '-300'
        elif chars == 'gray':
            key += '-360'  # gray
        elif chars == 'fold':
            key += '-170'
        elif self.chapter_depth > 0:
            key += '-' + str(210 + ((self.chapter_depth - 1) * 10))
        elif self.section_depth > 0:
            key += '-' + str(30 + ((self.section_depth - 1) * 10))
        else:
            key += '-XXX'
        # LIGHTNESS
        if self.del_or_ins == 'del':
            key += '-0'
        elif self.del_or_ins == 'ins':
            key += '-2'
        else:
            key += '-1'
        # FONT
        if chars == 'mincho':
            key += '-m'  # mincho
        else:
            key += '-g'  # gothic
        # UNDERLINE
        if chars == 'font decorator':
            key += '-x'  # no underline
        elif chars == 'table':
            key += '-x'  # no underline
        elif chars == ' ' or chars == '\t' or chars == '\u3000':
            # if not self.is_in_comment:
            key += '-u'  # underline
        elif not self.is_in_comment and self.has_underline:
            key += '-u'  # comment
        elif not self.is_in_comment and self.has_specific_font:
            key += '-u'  # specific font
        elif not self.is_in_comment and self.has_frame:
            key += '-u'  # frame
        elif not self.is_in_comment and self.has_ruby:
            key += '-u'  # ruby
        elif not self.is_in_comment and self.is_resized != self.standard_size:
            key += '-u'  # resized
        elif not self.is_in_comment and self.is_stretched != '':
            key += '-u'  # stretched
        elif not self.is_in_comment and self.is_in_preformatted:
            key += '-u'  # preformatted
        elif not self.is_in_comment and self.is_in_italic:
            key += '-u'  # italic
        elif not self.is_in_comment and self.is_in_bold:
            key += '-u'  # bold
        else:
            key += '-x'  # no underline
        # RETURN
        return key  # c-XXX-1-g-x, ...

############################################################
# LINE DATUM


class LineDatum:

    def __init__(self):
        self.line_number = 0
        self.line_text = ''
        self.beg_chars_state = CharsState()
        self.end_chars_state = CharsState()
        self.paint_keywords = False

    def paint_line(self, pane, paint_keywords=False):
        # PREPARE
        i = self.line_number
        line_text = self.line_text
        chars_state = self.beg_chars_state.copy()
        self.paint_keywords = paint_keywords
        # EMPTY LINE
        if line_text == '\n':
            chars_state.standard_size = ''  # for table
            self.end_chars_state = chars_state.copy()
            return
        # RESET TAG
        for tag in pane.tag_names():
            if tag == 'IMEmarkedtext':  # macos ime
                continue
            if tag != 'search_tag':
                pane.tag_remove(tag, str(i + 1) + '.0', str(i + 1) + '.end')
        # LINE
        if not chars_state.is_in_comment:
            # PAGE BREAK
            if line_text == '<pgbr>\n' or line_text == '<Pgbr>\n':
                beg, end = str(i + 1) + '.0', str(i + 1) + '.end'
                key = chars_state.get_key('<pgbr>')                     # 1.key
                #                                                       # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                #                                                       # 5.tmp
                #                                                       # 6.beg
                self.end_chars_state = chars_state.copy()
                return
            # HORIZONTAL LINE
            res_color = '(?:|R|red|Y|yellow|G|green|C|cyan|B|blue|M|magenta)'
            res = '^' \
                + '((?:\\^' + res_color + '\\^)?)' \
                + '(-{5,})' \
                + '((?:\\^' + res_color + '\\^)?)' \
                + '\n$'
            if (line_text[0] == '^' or line_text[0] == '-') and \
               re.match(res, line_text):
                hfre = re.sub(res, '\\1', line_text)
                line = re.sub(res, '\\2', line_text)
                tfre = re.sub(res, '\\3', line_text)
                hlen = len(hfre)
                llen = len(line)
                hcol = 'gray' if hfre == '^^' else hfre.replace('^', '')
                tcol = 'gray' if tfre == '^^' else tfre.replace('^', '')
                beg = str(i + 1) + '.0'
                if hfre != '':
                    key = chars_state.get_key(hcol)                     # 1.key
                    end = str(i + 1) + '.' + str(hlen)                  # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    #                                                   # 5.tmp
                    beg = end                                           # 6.beg
                key = chars_state.get_key('hline')                      # 1.key
                end = str(i + 1) + '.' + str(hlen + llen)               # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                #                                                       # 5.tmp
                beg = end                                               # 6.beg
                if tfre != '':
                    key = chars_state.get_key(tcol)                     # 1.key
                    end = str(i + 1) + '.end'                           # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    #                                                   # 5.tmp
                    beg = end                                           # 6.beg
                self.end_chars_state = chars_state.copy()
                return
            # LENGTH REVISERS
            if line_text[0] == '<' or line_text[0] == '>' or \
               line_text[0] == 'v' or line_text[0] == 'V' or \
               line_text[0] == 'x' or line_text[0] == 'X':
                res = '^((<<|<|>|v|V|x|X)=(\\+|\\-)?[\\.0-9]+\\s+)+'
                if re.match(res, line_text):
                    chars_state.is_length_reviser = True
            # CHAPTER
            if line_text[0] == '$':
                res = '^(\\${,5})(?:-\\$+)*(=[\\.0-9]+)?(?:\\s.*)?\n?$'
                if re.match(res, line_text):
                    dep = len(re.sub(res, '\\1', line_text))
                    chars_state.set_chapter_depth(dep)
            # SECTION
            if line_text[0] == '#':
                res = '^(#{,8})(?:-#+)*(=[\\.0-9]+)?(?:\\s.*)?\n?$'
                if re.match(res, line_text):
                    dep = len(re.sub(res, '\\1', line_text))
                    chars_state.set_section_depth(dep)
            # CHAPTER AND SECTION
            if line_text[0] == '`' or line_text[0] == '*' or \
               line_text[0] == '-' or line_text[0] == '+':
                # CHAPTER
                res = '^((?:`|\\*|\\-{2,}|\\+{2,})+)' \
                    + '(\\${,5})(?:-\\$+)*(=[\\.0-9]+)?(?:\\s.*)?' \
                    + '((?:`|\\*|\\-{2,}|\\+{2,})*)\n?$'
                if re.match(res, line_text):
                    dep = len(re.sub(res, '\\2', line_text))
                    chars_state.set_chapter_depth(dep)
                # SECTION
                res = '^((?:`|\\*|\\-{2,}|\\+{2,})+)' \
                    + '(#{,8})(?:-#+)*(=[\\.0-9]+)?(?:\\s.*)?' \
                    + '((?:`|\\*|\\-{2,}|\\+{2,})*)\n?$'
                if re.match(res, line_text):
                    dep = len(re.sub(res, '\\2', line_text))
                    chars_state.set_section_depth(dep)
            if line_text[0] == '.':
                res = '^(?:\\.\\.\\.\\[[0-9]+\\])' \
                    + '(#{,8})(?:-#+)*(=[\\.0-9]+)?(?:\\s.*)?\n?$'
                if re.match(res, line_text):
                    dep = len(re.sub(res, '\\1', line_text))
                    chars_state.set_section_depth(dep)
            # TABLE
            if line_text[0] == '|':
                if chars_state.standard_size == '':
                    chars_state.standard_size = chars_state.is_resized
            res = '^(:\\s+)?\\s*(\\|:?-*:?[\\^=]?)+(\\|(\\s+:)?|\\\\)?$'
            if re.match(res, line_text):
                for j, c in enumerate(line_text + '\0'):
                    if c == '|':
                        key = chars_state.get_key('table')              # 1.key
                    elif c == ':':
                        key = chars_state.get_key('alignment')          # 1.key
                    elif c == '-':
                        key = chars_state.get_key('font decorator')     # 1.key
                    elif c == '^' or c == '=':
                        key = chars_state.get_key('hline')              # 1.key
                    elif c == ' ' or c == '\t' or c == '\u3000':
                        key = chars_state.get_key(c)                    # 1.key
                    elif c == '\\':
                        key = chars_state.get_key('escape')             # 1.key
                    elif c != '\0':
                        key = chars_state.get_key('')                   # 1.key
                    else:  # c == '\0'
                        break
                    beg = str(i + 1) + '.' + str(j)                     # 2.beg
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                self.end_chars_state = chars_state.copy()
                return
            if re.match('^\\^+$', line_text) or re.match('^=+$', line_text):
                key = chars_state.get_key('hline')                      # 1.key
                beg = str(i + 1) + '.0'                                 # 2.end
                end = str(i + 1) + '.end'                               # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                self.end_chars_state = chars_state.copy()
                return
        # PARTS
        beg, tmp = str(i + 1) + '.0', ''
        is_in_substitute_phrase = False  # substitute phrase
        for j, c in enumerate(line_text):
            tmp += c
            s1 = line_text[j - 0:j + 1] if True else ''
            s2 = line_text[j - 1:j + 1] if j > 0 else ''
            s3 = line_text[j - 2:j + 1] if j > 1 else ''
            s4 = line_text[j - 3:j + 1] if j > 2 else ''
            c0 = line_text[j + 1] if j < len(line_text) - 1 else ''
            c1 = c
            c2 = line_text[j - 1] if j > 0 else ''
            c3 = line_text[j - 2] if j > 1 else ''
            c4 = line_text[j - 3] if j > 2 else ''
            c5 = line_text[j - 4] if j > 3 else ''
            s_lft = line_text[:j + 1]
            s_rgt = line_text[j + 1:]
            # SUBSTITUTE PHRASE
            if c2 == '%' and c1 == '[' and not is_in_substitute_phrase and \
               re.match(NOT_ESCAPED + '%\\[$', s_lft) and \
               re.match('^.*\\]%.*$', s_rgt):
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j - 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                chars_state.toggle_is_in_comment()                      # 4.set
                tmp = '%['                                              # 5.tmp
                beg = end                                               # 6.beg
                is_in_substitute_phrase = True
                continue
            if c2 == ']' and c1 == '%' and is_in_substitute_phrase:
                if re.match('%\\[A[0-9A-Za-z]?:.*\\]%', tmp):
                    key = chars_state.get_key('substitute phrase A')    # 1.key
                elif re.match('%\\[B[0-9A-Za-z]?:.*\\]%', tmp):
                    key = chars_state.get_key('substitute phrase B')    # 1.key
                else:
                    key = chars_state.get_key('substitute phrase C')    # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                chars_state.toggle_is_in_comment()                      # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                is_in_substitute_phrase = False
                continue
            if is_in_substitute_phrase:
                continue
            # END OF THE LINE "\n"
            if c1 == '\n':
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                #                                                       # 5.tmp
                #                                                       # 6.beg
                break
            # COMMENT
            if s4 == '<!--' and not chars_state.is_in_comment and \
               (c5 != '\\' or re.match(NOT_ESCAPED + '<!--$', tmp)):
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j - 3)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                chars_state.toggle_is_in_comment()                      # 4.set
                tmp = '<!--'                                            # 5.tmp
                beg = end                                               # 6.beg
                continue
            if s3 == '-->' and chars_state.is_in_comment and \
               (c4 != '\\' or re.match(NOT_ESCAPED + '-->$', tmp)):
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                chars_state.toggle_is_in_comment()                      # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            if chars_state.is_in_comment and c1 != '\n':
                continue
            # LENGTH REVISERS
            if chars_state.is_length_reviser:
                res_c = '^[<>vVxX=\\-\\+0-9\\.\\s]$'
                res_s = '^' \
                    + '((<<|<|>|v|V|x|X)=(\\+|\\-)?[\\.0-9]+\\s+)*' \
                    + '(((((<<|<|>|v|V|x|X)=?)(\\+|\\-)?)[\\.0-9]*)\\s*)' \
                    + '$'
                if not re.match(res_c, c):
                    chars_state.is_length_reviser = False
                elif not re.match(res_s, s_lft + ' '):
                    chars_state.is_length_reviser = False
            # ASCII
            if c.isascii() and not c.isalnum():
                # ESCAPE SYMBOL
                if c == '\\':
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '\\'                                        # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('escape')                 # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # RELAX
                if c2 == '<' and c1 == '>':
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '<>'                                        # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('font decorator')         # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # LIST
                if j == 0 and c == '-' and c0 != '\n' and re.match('\\s', c0):
                    key = chars_state.get_key('list')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                if c == '.' and re.match('^[0-9]+\\.$', s_lft) and \
                   re.match('\\s', c0):
                    key = chars_state.get_key('half number')
                    pane.tag_remove(key, str(i + 1) + '.0', str(i + 1) + '.1')
                    beg, end = str(i + 1) + '.0', str(i + 1) + '.' + str(j + 1)
                    key = chars_state.get_key('list')                   # 1.key
                    #                                                   # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # ALIGNMENT
                if j == 0 and c == ':' and re.match('\\s', c0):
                    key = chars_state.get_key('alignment')              # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                if j >= 2 and re.match('\\s', c2) and c == ':' and c0 == '\n':
                    key = chars_state.get_key('alignment')              # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # DEL ("->", "<-")
                if len(tmp) >= 2 and \
                   ((chars_state.del_or_ins == '' and s2 == '->' and
                     (c3 != '\\' or re.match(NOT_ESCAPED + '\\->$', tmp))) or
                    (chars_state.del_or_ins == 'del' and s2 == '<-' and
                     (c3 != '\\' or re.match(NOT_ESCAPED + '<\\-$', tmp)))):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.set_del_or_ins('del')                   # 4.set
                    # tmp = '->' or '<-'                                # 5.tmp
                    beg = end                                           # 6.beg
                    key = 'c-20-1-g-x'                                  # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # INS ("+>", "<+")
                if len(tmp) >= 2 and \
                   ((chars_state.del_or_ins == '' and s2 == '+>' and
                     (c3 != '\\' or re.match(NOT_ESCAPED + '\\+>$', tmp))) or
                    (chars_state.del_or_ins == 'ins' and s2 == '<+' and
                     (c3 != '\\' or re.match(NOT_ESCAPED + '<\\+$', tmp)))):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.set_del_or_ins('ins')                   # 4.set
                    # tmp = '+>' or '<+'                                # 5.tmp
                    beg = end                                           # 6.beg
                    key = 'c-200-1-g-x'                                 # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # LINE BREAK
                if len(tmp) >= 4 and (s4 == '<br>' or s4 == '<Br>'):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 3)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = <br>                                        # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('<br>')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # COLOR
                res_color = '(' \
                    + 'R|red|Y|yellow|G|green|C|cyan|B|blue|M|magenta' \
                    + '|' \
                    + 'DR|DY|DG|DC|DB|DM' \
                    + ')'
                if (c == '_' and re.match('^.*_' + res_color + '_$', tmp)) or \
                   (c == '^' and re.match('^.*\\^' + res_color + '\\^$', tmp)):
                    res = '^(.*)[_\\^]' + res_color + '[_\\^]$'
                    mdt = re.sub(res, '\\1', tmp)
                    col = re.sub(res, '\\2', tmp)
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - len(col) - 1)      # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '_.+_' or '^.+^'                            # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key(col)                      # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                if len(tmp) >= 2 and s2 == '^^':
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '^^'                                        # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('gray')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # FONT DECORATOR ("---", "+++", ">>>", "<<<")
                if len(tmp) >= 3 and \
                   (s3 == '---' or s3 == '+++' or s3 == '>>>' or s3 == '<<<') \
                   and (c4 != '\\' or re.match(NOT_ESCAPED + '...$', tmp)):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 2)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = re.sub('^(.*)(...)$', '\\2', tmp)             # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('font decorator')         # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    res1, res2 = '^.*:-+$', '^-*:.*$'
                    if not re.match(res1, s_lft) and not re.match(res2, s_rgt):
                        if tmp == '---' or tmp == '+++':
                            chars_state.set_is_resized(tmp)             # 4.set
                        else:
                            chars_state.set_is_stretched(tmp)           # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # FONT DECORATOR ("--", "++", ">>", "<<")
                if len(tmp) >= 2 and \
                   (s2 == '--' or s2 == '++' or s2 == '>>' or s2 == '<<') and \
                   (c3 != '\\' or re.match(NOT_ESCAPED + '..$', tmp)):
                    if (c1 != c0) or \
                       (chars_state.is_stretched == '<<' and s2 == '>>') or \
                       (chars_state.is_stretched == '>>' and s2 == '<<'):
                        key = chars_state.get_key('')                   # 1.key
                        end = str(i + 1) + '.' + str(j - 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = re.sub('^(.*)(..)$', '\\2', tmp)          # 5.tmp
                        beg = end                                       # 6.beg
                        key = chars_state.get_key('font decorator')     # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        res1, res2 = '^.*:-+$', '^-*:.*$'
                        if not re.match(res1, s_lft) and \
                           not re.match(res2, s_rgt):
                            res = '^=[-\\+]?[0-9]*(\\.?[0-9]+)(\\s.*)?$'
                            if s2 != '<<' or not re.match(res, s_rgt):
                                if tmp == '--' or tmp == '++':
                                    chars_state.set_is_resized(tmp)     # 4.set
                                else:
                                    chars_state.set_is_stretched(tmp)   # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
                        continue
                # UNDERLINE ("_.*_")
                res = NOT_ESCAPED + '(_[\\$=\\.#\\-~\\+]{,4}_)$'
                if c == '_' and re.match(res, tmp):
                    mdt = re.sub(res, '\\2', tmp)
                    hul = chars_state.has_underline
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - len(mdt) + 1)      # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if hul:
                        chars_state.toggle_has_underline()              # 4.set
                    tmp = mdt                                           # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('font decorator')         # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if not hul:
                        chars_state.toggle_has_underline()              # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # FONT DECORATOR ("@.+@")
                res = NOT_ESCAPED + '(@[^@]{1,66}@)$'
                if c == '@' and re.match(res, tmp):
                    mdt = re.sub(res, '\\2', tmp)
                    hsf = chars_state.has_specific_font
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - len(mdt) + 1)      # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if hsf:
                        chars_state.toggle_has_specific_font()          # 4.set
                    tmp = mdt                                           # 5.tmp
                    beg = end                                           # 6.beg
                    for k, tmp_c in enumerate(mdt):
                        key = chars_state.get_key('font decorator')     # 1.key
                        if tmp_c == ' ' or tmp_c == '\t' or tmp_c == '\u3000':
                            key = chars_state.get_key(tmp_c)            # 1.key
                        end = str(i + 1) + '.' \
                            + str(j - len(mdt) + 1 + (k + 1))           # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        beg = end                                       # 6.beg
                    if not hsf:
                        chars_state.toggle_has_specific_font()          # 4.set
                    tmp = ''                                            # 5.tmp
                    continue
                # FRAME (Should be before TABLE and IMAGE)
                if (c1 == '[' and c0 == '|') or (c1 == '|' and c0 == ']'):
                    continue
                if (c2 == '[' and c1 == '|') or (c2 == '|' and c1 == ']'):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j - 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '[|' or '|]'                                # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('font decorator')         # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.attach_or_remove_frame(c2 + c1)         # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # TABLE
                if c == '|':
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '|'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('table')                  # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                if c == ':':
                    res1, res2 = '^.*\\|', '^-*:?[\\^|=]?$'
                    res3, res4 = '^.*\\|:?-*:$', '^[\\^|=]?$'
                    res5, res6 = '^.*\\|:$', '^\\s.*$'
                    res7, res8 = '^.*\\s:$', '^(@([0-9]*x)?[0-9]+)?\\|.*$'
                    if (re.match(res1, s_lft) and re.match(res2, s_rgt)) or \
                       (re.match(res3, s_lft) and re.match(res4, s_rgt)) or \
                       (re.match(res5, s_lft) and re.match(res6, s_rgt)) or \
                       (re.match(res7, s_lft) and re.match(res8, s_rgt)):
                        key = chars_state.get_key('alignment')          # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
                        continue
                if c == '-' and tmp == '-':
                    res1, res2 = '^.*\\|:?-{4,}$', '^-*:?[\\^|=]?$'
                    if re.match(res1, s_lft) and re.match(res2, s_rgt):
                        key = chars_state.get_key('font decorator')     # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
                        continue
                if c == '^' or c == '=':
                    res1, res2 = '^.*\\|:?-*:?.$', '^$'
                    if re.match(res1, s_lft) and re.match(res2, s_rgt):
                        key = chars_state.get_key('hline')              # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
                        continue
                # IMAGE
                if c == '!' and re.match('^\\[.*\\]\\(.*\\)', line_text[j+1:]):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '!'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('image')                  # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                # SPACE (< n >)
                if c == '<' and re.match('^\\s*[\\.0-9]+\\s*>.*$', s_rgt):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '<'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('<sp>')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                if c == '>' and re.match('^.*<\\s*[\\.0-9]+\\s*>$', s_lft):
                    key = chars_state.get_key('<sp>')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # RUBY (<xxx/yyy>)
                if c == '<' and re.match('^[^</>]*/[^</>]*>.*$', s_rgt):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    # tmp = '<'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('ruby')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.set_or_unset_has_ruby(True)             # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                if c == '/' and \
                   re.match('^.*<[^</>]*/$', s_lft) and \
                   re.match('^[^</>]*>.*$', s_rgt):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.set_or_unset_has_ruby(False)            # 4.set
                    # tmp = '/'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('ruby')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.set_or_unset_has_ruby(True)             # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                if c == '>' and re.match('^.*<[^</>]*/[^</>]*>$', s_lft):
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    chars_state.set_or_unset_has_ruby(False)            # 4.set
                    # tmp = '>'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('ruby')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # PREFORMATTED
                if c == '`' and re.match(NOT_ESCAPED + c + '$', s_lft):
                    iip = chars_state.is_in_preformatted
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j)                     # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if iip and not re.match(NOT_ESCAPED + '```$', s_lft):
                        chars_state.toggle_is_in_preformatted()         # 4.set
                    # tmp = '`'                                         # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('font decorator')         # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if not iip and not re.match(NOT_ESCAPED + '```$', s_lft):
                        chars_state.toggle_is_in_preformatted()         # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # SCRIPT PARENTHESIS (DISABLE ITALIC)
                if c == '{':
                    if chars_state.script_parenthesis == '':
                        if re.match(NOT_ESCAPED + '{[0-9]?{$', s_lft):
                            par = re.sub('^(.*)({[0-9]?{)$', '\\2', s_lft)
                            chars_state.set_or_unset_script_parenthesis(par)
                elif c == '}':
                    if chars_state.script_parenthesis != '':
                        if re.match(NOT_ESCAPED + '}[0-9]?}$', s_lft):
                            par = re.sub('^(.*)(}[0-9]?})$', '\\2', s_lft)
                            chars_state.set_or_unset_script_parenthesis(par)
                # SCRIPT PARENTHESIS (DISABLE ITALIC AND PARENTHESIS)
                if chars_state.script_parenthesis != '':
                    if c == '*' or c == '(' or c == ')':
                        continue
                # ITALIC AND BOLD
                if c == '*' and re.match(NOT_ESCAPED + '\\*$', s_lft) and \
                   (c0 != '*' or re.match(NOT_ESCAPED + '\\*\\*\\*$', s_lft)):
                    # if chars_state.script_parenthesis == '':
                    iii = chars_state.is_in_italic
                    iib = chars_state.is_in_bold
                    if re.match(NOT_ESCAPED + '\\*\\*\\*$', s_lft):
                        n = 3
                    elif re.match(NOT_ESCAPED + '\\*\\*$', s_lft):
                        n = 2
                    elif re.match(NOT_ESCAPED + '\\*$', s_lft):
                        n = 1
                    key = chars_state.get_key('')                       # 1.key
                    end = str(i + 1) + '.' + str(j + 1 - n)             # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if iii and (n == 1 or n == 3):
                        chars_state.toggle_is_in_italic()               # 4.set
                    if iib and (n == 2 or n == 3):
                        chars_state.toggle_is_in_bold()                 # 4.set
                    # tmp = '*{1,3}'                                    # 5.tmp
                    beg = end                                           # 6.beg
                    key = chars_state.get_key('font decorator')         # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    if not iii and (n == 1 or n == 3):
                        chars_state.toggle_is_in_italic()               # 4.set
                    if not iib and (n == 2 or n == 3):
                        chars_state.toggle_is_in_bold()                 # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # FOLDING
                if line_text[0] == '#':
                    if re.match('^#+(-#+)*(\\s.*)?\\.{3}\\[$', s_lft) and \
                       re.match(NOT_ESCAPED + '\\.{3}\\[$', s_lft) and \
                       re.match('^[0-9]+\\]$', s_rgt):
                        continue  # #+ xxx...[ / n]
                    if re.match('^#+(-#+)*(\\s.*)?\\.{3}\\[[0-9]+$', s_lft) \
                       and re.match(NOT_ESCAPED + '\\.{3}\\[[0-9]+$', s_lft) \
                       and re.match('^[0-9]*\\]$', s_rgt):
                        continue  # #+ xxx...[n / ]
                    res_l = '^(#+(?:-#+)*(?:\\s.*)?)(\\.{3}\\[[0-9]+\\])$'
                    res_r = '^\n$'
                    if re.match('^(.*)(\\.{3}\\[[0-9]+\\])$', tmp) and \
                       re.match(res_l, s_lft) and re.match(res_r, s_rgt):
                        sym = re.sub('^(.*)(\\.{3}\\[[0-9]+\\])$', '\\2', tmp)
                        key = chars_state.get_key('')                   # 1.key
                        end = str(i + 1) + '.' + str(j + 1 - len(sym))  # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        # tmp = '...[n]'                                # 5.tmp
                        beg = end                                       # 6.beg
                        key = chars_state.get_key('fold')               # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
                        continue  # #+ xxx...[n] /
                if line_text[0] == '.':
                    if re.match('^\\.{3}\\[$', s_lft) and \
                       re.match('^[0-9]+\\]#+(-#+)*(\\s.*)?$', s_rgt):
                        continue  # ...[ / n]#+ xxx
                    if re.match('^\\.{3}\\[[0-9]+$', s_lft) and \
                       re.match('^[0-9]*\\]#+(-#+)*(\\s.*)?$', s_rgt):
                        continue  # ...[n / ]#+ xxx
                    res_l = '^\\.{3}\\[[0-9]+\\]$'
                    res_r = '^(#+(?:-#+)*(?:\\s.*)?)$'
                    if re.match('^\\.{3}\\[[0-9]+\\]$', tmp) and \
                       re.match(res_l, s_lft) and re.match(res_r, s_rgt):
                        key = chars_state.get_key('fold')               # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
                        continue  # ...[n] / #+
            if re.match('^[0-9]$', c):
                # SAPCE (< n >)
                if ((re.match('^.*<\\s*[0-9]+$', s_lft) and
                     re.match('^[0-9]*\\s*>.*$', s_rgt)) or
                    (re.match('^.*<\\s*[0-9]+$', s_lft) and
                     re.match('^[0-9]*\\.[0-9]+\\s*>.*$', s_rgt)) or
                    (re.match('^.*<\\s*[0-9]*\\.$', s_lft) and
                     re.match('^[0-9]+\\s*>.*$', s_rgt)) or
                    (re.match('^.*<\\s*[0-9]*\\.[0-9]+$', s_lft) and
                     re.match('^[0-9]*\\s*>.*$', s_rgt))):
                    key = chars_state.get_key('<sp>')                   # 1.key
                    end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                    pane.tag_add(key, beg, end)                         # 3.tag
                    #                                                   # 4.set
                    tmp = ''                                            # 5.tmp
                    beg = end                                           # 6.beg
                    continue
                # FONT SIZE (@n@)
                if ((re.match('^.*@[0-9]+$', s_lft) and
                     re.match('^[0-9]*@.*$', s_rgt)) or
                    (re.match('^.*@[0-9]+$', s_lft) and
                     re.match('^[0-9]*\\.[0-9]+@.*$', s_rgt)) or
                    (re.match('^.*@[0-9]*\\.$', s_lft) and
                     re.match('^[0-9]+@.*$', s_rgt)) or
                    (re.match('^.*@[0-9]*\\.[0-9]+$', s_lft) and
                     re.match('^[0-9]*@.*$', s_rgt))):
                    continue
            # NUMBER
            if re.match('^[0-9]$', c):
                if re.match('^#+(-#+)*(\\s.*)?\\.\\.\\.\\[[0-9]+$', s_lft) \
                   and re.match(NOT_ESCAPED + '\\.\\.\\.\\[[0-9]+$', s_lft) \
                   and re.match('^[0-9]*\\]$', s_rgt):
                    continue  # folded section (#+ xxx...[n / ])
                if re.match('^\\.\\.\\.\\[[0-9]+$', s_lft) and \
                   re.match('^[0-9]*\\]#+(-#+)*(\\s.*)?$', s_rgt):
                    continue  # folded section (...[n / ]#+ xxx)
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                # tmp = '[0-9]'                                         # 5.tmp
                beg = end                                               # 6.beg
                key = chars_state.get_key('half number')                # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            if re.match('[' +
                        '０-９' +
                        '零一二三四五六七八九十' +
                        '⑴⑵⑶⑷⑸⑹⑺⑻⑼⑽⑾⑿⒀⒁⒂⒃⒄⒅⒆⒇' +
                        '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳' +
                        ']', c):
                if c2 == '第' and c1 == '三':
                    if re.match('^(?:債務)?者', s_rgt):
                        continue
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                # tmp = '[０-９...]'                                    # 5.tmp
                beg = end                                               # 6.beg
                key = chars_state.get_key('full number')                # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            # SPACE (" ", "\t", "\u3000")
            if c == ' ' or c == '\t' or c == '\u3000':
                if re.match(NOT_ESCAPED + '@[^@]{1,66}.$', tmp) and \
                   re.match('[^@]{1,66}@.*$', s_rgt):
                    continue
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                # tmp = ' ' or '\t' or '\u3000'                         # 5.tmp
                beg = end                                               # 6.beg
                key = chars_state.get_key(c)                            # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            # PARENTHESES
            if c == '「' or c == '『' or c == '[' or \
               c == '｛' or c == '{' or \
               c == '（' or c == '(':
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                chars_state.apply_parenthesis(c)                        # 4.set
                tmp = c                                                 # 5.tmp
                beg = end                                               # 6.beg
                continue
            if c == ')' or c == '）' or \
               c == '}' or c == '｝' or \
               c == ']' or c == '』' or c == '」':
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                chars_state.apply_parenthesis(c)                        # 4.set
                # tmp = ''                                              # 5.tmp
                beg = end                                               # 6.beg
                continue
            # MINCHO
            if c == '\u30FC':  # 長音記号
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                # tmp = c                                               # 5.tmp
                beg = end                                               # 6.beg
                key = chars_state.get_key('mincho')                     # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            # HALF KATAKANA
            if re.match('^[ｦ-ﾟ]$', c):
                key = chars_state.get_key('')                       # 1.key
                end = str(i + 1) + '.' + str(j)                     # 2.end
                pane.tag_add(key, beg, end)                         # 3.tag
                chars_state.set_or_unset_has_ruby(False)            # 4.set
                # tmp = '[ｦ-ﾟ]'                                     # 5.tmp
                beg = end                                           # 6.beg
                key = chars_state.get_key('half katakana')          # 1.key
                end = str(i + 1) + '.' + str(j + 1)                 # 2.end
                pane.tag_add(key, beg, end)                         # 3.tag
                #                                                   # 4.set
                tmp = ''                                            # 5.tmp
                beg = end                                           # 6.beg
                continue
            # HORIZONTAL LINES
            if not c.isascii() and \
               ((c == '\u00AD' or c == '\u058A' or c == '\u05BE' or
                 c == '\u1806' or c == '\u180A' or c == '\u2010' or
                 c == '\u2011' or c == '\u2012' or c == '\u2013' or
                 c == '\u2014' or c == '\u2015' or c == '\u2043' or
                 c == '\u207B' or c == '\u208B' or c == '\u2212' or
                 c == '\u2500' or c == '\u2501' or c == '\u2796' or
                 c == '\u2E3A' or c == '\u2E3B' or c == '\u3127' or
                 c == '\u3161' or c == '\uFE58' or c == '\uFE63' or
                 c == '\uFF0D' or c == '\uFF70')):
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                # tmp = '-'                                             # 5.tmp
                beg = end                                               # 6.beg
                if False:
                    pass
                elif c == '\u2010':  # ハイフン（EUC:A1BE）
                    key = chars_state.get_key('horizontalline000')      # 1.key
                elif c == '\u2015':  # 水平線（EUC:A1BD）
                    key = chars_state.get_key('horizontalline010')      # 1.key
                elif c == '\u2212':  # マイナス記号（EUC:A1DD）
                    key = chars_state.get_key('horizontalline020')      # 1.key
                elif c == '\u2500':  # 罫線（EUC:A8A1）
                    key = chars_state.get_key('horizontalline030')      # 1.key
                elif c == '\u2501':  # 太字罫線（EUC:A8AC）
                    key = chars_state.get_key('horizontalline040')      # 1.key
                elif c == '\uFF70':  # 半角カナの長音記号（EUC:8EB0）
                    key = chars_state.get_key('horizontalline050')      # 1.key
                elif c == '\u00AD':  # 改行時だけに表示されるハイフン
                    key = chars_state.get_key('horizontalline060')      # 1.key
                elif c == '\u058A':  # アメリカンハイフン
                    key = chars_state.get_key('horizontalline070')      # 1.key
                elif c == '\u05BE':  # ヘブライ語のマカフ
                    key = chars_state.get_key('horizontalline080')      # 1.key
                elif c == '\u1806':  # モンゴル語のソフトハイフン
                    key = chars_state.get_key('horizontalline090')      # 1.key
                elif c == '\u180A':  # モンゴル語のニルグ
                    key = chars_state.get_key('horizontalline100')      # 1.key
                elif c == '\u2011':  # 改行しないハイフン
                    key = chars_state.get_key('horizontalline110')      # 1.key
                elif c == '\u2012':  # 数字幅のダッシュ
                    key = chars_state.get_key('horizontalline120')      # 1.key
                elif c == '\u2013':  # Ｎ幅ダッシュ
                    key = chars_state.get_key('horizontalline130')      # 1.key
                elif c == '\u2014':  # Ｍ幅ダッシュ
                    key = chars_state.get_key('horizontalline140')      # 1.key
                elif c == '\u2043':  # 箇条書きの記号
                    key = chars_state.get_key('horizontalline150')      # 1.key
                elif c == '\u207B':  # 上付きマイナス
                    key = chars_state.get_key('horizontalline160')      # 1.key
                elif c == '\u208B':  # 下付きマイナス
                    key = chars_state.get_key('horizontalline170')      # 1.key
                elif c == '\u2796':  # 太字マイナス記号
                    key = chars_state.get_key('horizontalline180')      # 1.key
                elif c == '\u2E3A':  # 2倍幅のＭ幅ダッシュ
                    key = chars_state.get_key('horizontalline190')      # 1.key
                elif c == '\u2E3B':  # 3倍幅のＭ幅ダッシュ
                    key = chars_state.get_key('horizontalline200')      # 1.key
                elif c == '\u3127':  # 注音符号の「Ｉ」の発
                    key = chars_state.get_key('horizontalline210')      # 1.key
                elif c == '\u3161':  # ハングルの「ウ」
                    key = chars_state.get_key('horizontalline220')      # 1.key
                elif c == '\uFE58':  # 小さいＭ幅ダッシュ
                    key = chars_state.get_key('horizontalline230')      # 1.key
                elif c == '\uFE63':  # 小さいハイフンマイナス
                    key = chars_state.get_key('horizontalline240')      # 1.key
                elif c == '\uFF0D':  # 全角ハイフンマイナス
                    key = chars_state.get_key('horizontalline250')      # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            # ERROR ("★")
            if c == '★' or \
               re.match('^[⺟⺠⻁⻄⻑⻘⻝⻤⻨⻩⻫⻭⻯⻲戶黑]$', c):  # bushu
                key = chars_state.get_key('')                           # 1.key
                end = str(i + 1) + '.' + str(j)                         # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                # tmp = '★'                                            # 5.tmp
                beg = end                                               # 6.beg
                key = 'error_tag'                                       # 1.key
                end = str(i + 1) + '.' + str(j + 1)                     # 2.end
                pane.tag_add(key, beg, end)                             # 3.tag
                #                                                       # 4.set
                tmp = ''                                                # 5.tmp
                beg = end                                               # 6.beg
                continue
            # KEYWORD
            if Makdo.keywords_to_paint is not None and \
               Makdo.keywords_to_paint != '':
                kws = []
                kw = ''
                for c in Makdo.keywords_to_paint + '|':
                    if re.match(NOT_ESCAPED + '\\|$', kw + c):
                        kws.append(kw)
                        kw = ''
                    else:
                        kw += c
                for kw in kws:
                    if re.match('^(.*?)' + kw + '$', tmp):
                        key = chars_state.get_key('')                   # 1.key
                        end = str(i + 1) + '.' + str(j + 1 - len(kw))   # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        # tmp = kw                                      # 5.tmp
                        beg = end                                       # 6.beg
                        key = chars_state.get_key('red')                # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
            if self.paint_keywords:
                for kw in KEYWORDS:
                    if re.match('^(.*?)' + kw[0] + '$', tmp):
                        t1 = re.sub('^(.*?)' + kw[0] + '$', '\\1', tmp)
                        t2 = re.sub('^(.*?)' + kw[0] + '$', '\\2', tmp)
                        if t2 == '本訴' or t2 == '反訴' or t2 == '別訴':
                            if re.match('^(原|被)告', s_rgt):
                                continue  # 本訴/原告
                        if t2 == '原告' or t2 == '被告':
                            if re.match('^(?:.|\n)*(本|反|別)訴$', t1):
                                continue  # 本訴原告/
                        if t2 == '被告' and c0 == '人':
                            continue  # 被告/人
                        if t2 == '債務者' and re.match('^.*第三$', t1):
                            continue  # 第三/債務者
                        key = chars_state.get_key('')                   # 1.key
                        end = str(i + 1) + '.' + str(j + 1 - len(t2))   # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        # tmp = t2                                      # 5.tmp
                        beg = end                                       # 6.beg
                        key = chars_state.get_key(kw[1])                # 1.key
                        end = str(i + 1) + '.' + str(j + 1)             # 2.end
                        pane.tag_add(key, beg, end)                     # 3.tag
                        #                                               # 4.set
                        tmp = ''                                        # 5.tmp
                        beg = end                                       # 6.beg
        self.end_chars_state = chars_state.copy()
        return


############################################################
# MAKDO

class Makdo:

    args_dont_show_help = None         # True|+False
    file_dont_show_help = None
    args_make_backup_file = False      # True|+False
    file_make_backup_file = False
    # args_is_toc_display_mode = True  # True|+False
    file_is_toc_display_mode = True
    args_read_only = None              # True|+False
    # file_read_only = None
    args_background_color = None       # +W|B|G
    file_background_color = None
    args_font_size = None              # 3|6|9|12|15|+18|21|24|27|30|33|36|...
    file_font_size = None
    args_paint_keywords = None         # True|+False
    file_paint_keywords = None
    args_keywords_to_paint = None      # 'foo|bar|baz'
    file_keywords_to_paint = None
    args_use_regexps = None            # True|+False
    file_use_regexps = None
    args_digit_separator = None        # +0|3|4
    file_digit_separator = None
    # args_shortcut_key_layout = None  # +normal|qwerty|akauni|
    file_shortcut_key_layout = None

    args_input_file = None

    search_word = ''

    ##############################################
    # INIT

    def __init__(self):
        self.win = None
        self.temp_dir = tempfile.TemporaryDirectory()
        self.file_path = self.args_input_file
        self.init_text = ''
        self.saved_text = ''
        self.file_lines = []
        self.has_made_backup_file = False
        self.line_data = []
        self.clipboard_list = ['']
        self.key_history = ['' for i in range(21)]
        self.key_pressed_time = [0 for i in range(21)]
        self.current_pane = 'txt'
        self.formula_number = -1
        self.memo_pad_memory = None
        self.rectangle_text_list = []
        #
        self.onedrive_directory = None
        #
        self.must_show_folding_help_message = True
        self.must_show_keyboard_macro_help_message = True
        self.must_show_config_help_message = True
        # GET CONFIGURATION
        self.get_and_set_configurations()
        # WINDOW
        # mac doesn't support "tkinterdnd2" (drag and drop)
        if sys.platform != 'darwin':
            self.win = tkinterdnd2.TkinterDnD.Tk()  # need to do first
        else:
            self.win = tkinter.Tk()
        self.win.title('MAKDO')
        self.win.geometry(WINDOW_SIZE)
        self.win.protocol("WM_DELETE_WINDOW", self.quit_makdo)
        # SPLASH SCREEN
        if getattr(sys, 'frozen', False):
            import _socket
            _socket.setdefaulttimeout(5)
            import pyi_splash
            pyi_splash.close()
        else:
            if sys.platform != 'darwin':
                # mac doesn't support splash screen
                self.show_splash_screen(SPLASH_IMG)
        # TITLE BAR ICON
        if sys.platform != 'darwin' or getattr(sys, 'frozen', False):
            # mac doesn't support iconphoto
            icon8_img = tkinter.PhotoImage(data=ICON8_IMG, master=self.win)
            self.win.iconphoto(False, icon8_img)
        # FRAME
        # self.frm = tkinter.Frame()
        # self.frm.pack(expand=True, fill=tkinter.BOTH)
        # MENU BAR
        self.mnb = tkinter.Menu(self.win)
        self._make_menu()
        # STATUS BAR
        stb = tkinter.Frame(self.win)
        stb.pack(fill='x', side='bottom', anchor='s')
        self.stb_l = tkinter.Frame(stb)  # left
        self.stb_l.pack(side='left', anchor='w')
        self.stb_r = tkinter.Frame(stb)  # right
        self.stb_r.pack(side='right', anchor='e')
        self._make_status_bar()
        # PANED WINDOW
        self.pnd = tkinter.PanedWindow(self.win, bd=0, sashwidth=3,
                                       orient='horizontal')
        self.pnd.pack(expand=True, fill='both', side='left', anchor='n')
        self.pnd_l = tkinter.PanedWindow(self.win, bd=0, sashwidth=3,
                                         orient='vertical')
        # self.pnd.add(self.pnd_l, minsize=100)  # -> table of contents
        self.pnd_r = tkinter.PanedWindow(self.win, bd=0, sashwidth=3,
                                         orient='vertical')
        self.pnd.add(self.pnd_r, minsize=100)  # -> table of contents
        self.pnd1 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#FF5D5D')  # 000
        self.pnd2 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#BC7A00')  # 040
        self.pnd3 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#758F00')  # 070
        self.pnd4 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#00A586')  # 170
        self.pnd5 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#7676FF')  # 240
        self.pnd6 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#C75DFF')  # 280
        self.pnd_r.add(self.pnd1, minsize=100)
        # MAIN TEXT
        self.txt = tkinter.Text(self.pnd1, undo=True)
        scb = tkinter.Scrollbar(self.pnd1, orient='vertical',
                                command=self.txt.yview)
        scb.pack(side='right', fill='y')
        self.txt['yscrollcommand'] = scb.set
        self.txt.pack(expand=True, fill='both')
        self.txt.config(insertbackground='#FF7777', blockcursor=True)  # cursor
        self._make_txt_key_configuration()
        # mac doesn't support "tkinterdnd2" (drag and drop)
        if sys.platform != 'darwin':
            self.txt.drop_target_register(tkinterdnd2.DND_FILES)
            self.txt.dnd_bind('<<Drop>>', self.open_dropped_file)
        # SUB TEXT
        self.sub = tkinter.Text(self.pnd2, undo=True)
        self.sub.config(insertbackground='#FF7777', blockcursor=True)  # cursor
        self._make_sub_key_configuration()
        self.sub_scb = tkinter.Scrollbar(self.pnd2, orient='vertical',
                                         command=self.sub.yview)
        self.sub_frm = tkinter.Frame(self.pnd2)
        # FONT
        families = tkinter.font.families()
        self.gothic_font = None
        self.mincho_font = None
        if self.gothic_font is None:
            for f in families:
                for gf in BIZUD_GOTHIC_FONT:
                    if re.match('^' + gf, f):
                        self.gothic_font \
                            = tkinter.font.Font(self.win, family=f)
                        break
                else:
                    continue
                break
        if self.gothic_font is None:
            self.show_font_help_message()
        if self.mincho_font is None:
            for f in families:
                for mf in BIZUD_MINCHO_FONT:
                    if re.match('^' + mf, f):
                        self.mincho_font \
                            = tkinter.font.Font(self.win, family=f)
                        break
                else:
                    continue
                break
        if self.mincho_font is None:
            if sys.platform == 'darwin':
                mfs = HIRAGINO_MINCHO_FONT + YU_MINCHO_FONT + MS_MINCHO_FONT
            else:
                mfs = YU_MINCHO_FONT + HIRAGINO_MINCHO_FONT + MS_MINCHO_FONT
            for f in families:
                for mf in mfs:
                    if re.match('^' + mf, f):
                        self.mincho_font \
                            = tkinter.font.Font(self.win, family=f)
                        break
                else:
                    continue
                break
        if self.gothic_font is None:
            self.gothic_font = tkinter.font.nametofont("TkFixedFont").copy()
        if self.mincho_font is None:
            self.mincho_font = tkinter.font.nametofont("TkFixedFont").copy()
        self.set_font()
        # OPEN FILE
        if self.args_input_file is not None:
            if os.path.exists(self.args_input_file):
                self.just_open_file(self.args_input_file)
            else:
                self.set_message_on_status_bar('新しいファイルです')
                self.file_path = self.args_input_file
                self._set_file_name(self.file_path)
        else:
            self.show_first_help_message()
        # TABLE OF CONTENTS
        is_toc_display_mode = self.is_toc_display_mode.get()
        if self.init_text != '' and is_toc_display_mode:
            self.set_message_on_status_bar('目次を作成しています', True)
            self.settle_or_remove_toc()
            self.set_message_on_status_bar('', True)
        # CURRENT PANE
        if self.current_pane == 'sub':
            self.sub.focus_set()
        elif self.current_pane == 'toc':
            self.toc_cvs_frm.focus_set()
        else:
            self.txt.focus_set()
        # RUN PERIODICALLY
        self.run_periodically()
        # LOOP
        self.win.mainloop()

    ####################################
    # SPLASH SCREEN

    def show_splash_screen(self, image):
        # mac doesn't support splash screen
        if sys.platform == 'darwin':
            return
        self.splash_screen = tkinter.Tk()
        sw = self.splash_screen.winfo_screenwidth()
        sh = self.splash_screen.winfo_screenheight()
        self.splash_image \
            = tkinter.PhotoImage(data=image, master=self.splash_screen)
        iw = self.splash_image.width()
        ih = self.splash_image.height()
        size = str(iw - 1) + 'x' + str(ih - 1)
        position = str(int((sw - iw) / 2)) + '+' + str(int((sh - ih) / 2))
        self.splash_screen.geometry(size + '+' + position)
        self.splash_screen.overrideredirect(1)  # no title bar
        canvas \
            = tkinter.Canvas(self.splash_screen, bg=None, width=iw, height=ih)
        canvas.place(x=-1, y=-1)
        canvas.create_image(0, 0, image=self.splash_image, anchor='nw')
        self.win.after(5000, self.destroy_splash_screen)
        # self.splash_screen.after(5000, self.destroy_splash_screen)

    def destroy_splash_screen(self):
        # mac doesn't support splash screen
        if sys.platform == 'darwin':
            return
        if ('splash_image' in vars(self)) and (self.splash_image is not None):
            self.splash_image = None
            self.splash_screen.destroy()

    ####################################
    # TOOLS

    def _set_file_name(self, file_path):
        file_name = os.path.basename(file_path)
        self.win.title(file_name + ' - MAKDO')
        self.set_file_name_on_status_bar(file_name)

    @staticmethod
    def _get_v_position_of_insert(pane):
        insert_position = pane.index('insert')
        insert_v_position = int(re.sub('\\.[0-9]+$', '', insert_position))
        return insert_v_position

    @staticmethod
    def _get_h_position_of_insert(pane):
        insert_position = pane.index('insert')
        insert_h_position = int(re.sub('^[0-9]+\\.', '', insert_position))
        return insert_h_position

    @staticmethod
    def _get_max_v_position(pane):
        max_position = pane.index('end-1c')
        max_v_position = int(re.sub('\\.[0-9]+$', '', max_position))
        return max_v_position

    @staticmethod
    def _get_max_h_position(pane):
        line_end_position = pane.txt.index('insert lineend')
        max_h_position = int(re.sub('^[0-9]+\\.', '', line_end_position))
        return max_h_position

    def _execute_when_delete_is_pressed(self, pane):
        if pane.tag_ranges('sel'):
            if self._is_read_only_pane(pane):
                self.copy_region()
            else:
                self.cut_region()
        elif 'akauni' in pane.mark_names():
            akn = pane.index('akauni')
            ins = pane.index('insert')
            beg = re.sub('\\..*$', '.0', ins)
            if akn == ins and akn != beg:
                c = pane.get(beg, akn)
                self.win.clipboard_clear()
                self.win.clipboard_append(c)
                if self.clipboard_list[-1] != '':
                    self.clipboard_list.append('')
                self.clipboard_list[-1] += c
                if not self._is_read_only_pane(pane):
                    pane.edit_separator()
                    pane.delete(beg, akn)
                    pane.edit_separator()
                self.cancel_region(pane)
            else:
                if self._is_read_only_pane(pane):
                    self.copy_region()
                else:
                    self.cut_region()
        else:
            ins = pane.index('insert')
            end = re.sub('\\..*$', '.end', ins)
            c = pane.get(ins, end)
            if self._is_read_only_pane(pane):
                self.win.clipboard_clear()
                self.win.clipboard_append(c)
                if self.clipboard_list[-1] != '':
                    self.clipboard_list.append('')
                self.clipboard_list[-1] += c
            else:
                if c == '':
                    self.win.clipboard_append('\n')
                    self.clipboard_list[-1] += '\n'
                    # pane.edit_separator()
                    pane.delete(ins, end + '+1c')
                    pane.edit_separator()
                else:
                    self.win.clipboard_append(c)
                    self.clipboard_list[-1] += c
                    pane.edit_separator()
                    pane.delete(ins, end)
                    # pane.edit_separator()

    def paint_out_line(self, line_number, pane=None):
        if pane is None:
            pane = self.txt
        ln = line_number
        # REGION IS SET
        if pane.tag_ranges('sel'):
            return
        if 'akauni' in pane.mark_names():
            return
        # UPDATE TEXT
        file_text = pane.get('1.0', 'end-1c')
        self.file_lines = file_text.split('\n')
        m = len(self.file_lines) - 1
        while len(self.line_data) < m + 1:
            self.line_data.append(LineDatum())
            self.line_data[-1].line_number = len(self.line_data) - 1
        while len(self.line_data) > m + 1:
            self.line_data.pop(-1)
        if m < 0:
            return
        # BAD LINE ID
        if ln < 0 or ln >= len(self.line_data):
            return
        # PREPARE
        line_text = self.file_lines[ln] + '\n'
        if ln == 0:
            chars_state = CharsState()
        else:
            chars_state \
                = self.line_data[ln - 1].end_chars_state.copy()
            chars_state.reset_partially()
        paint_keywords = self.paint_keywords.get()
        # EXCLUSION
        # if self.line_data[ln].line_text == line_text and \
        #    self.line_data[ln].beg_chars_state == chars_state and \
        #    self.line_data[ln].paint_keywords == paint_keywords:
        #     return
        # PAINT
        self.line_data[ln].line_text = line_text
        self.line_data[ln].beg_chars_state = chars_state
        self.line_data[ln].end_chars_state = CharsState()
        self.line_data[ln].paint_line(pane, paint_keywords)

    def paint_all_lines(self, pane):
        document = pane.get('1.0', 'end-1c')
        if document != '':
            paint_keywords = self.paint_keywords.get()
            self.line_data = [LineDatum() for line in self.file_lines]
            for i, line in enumerate(self.file_lines):
                self.line_data[i].line_number = i
                self.line_data[i].line_text = line + '\n'
                if i > 0:
                    self.line_data[i].beg_chars_state \
                        = self.line_data[i - 1].end_chars_state.copy()
                    self.line_data[i].beg_chars_state.reset_partially()
                n = i + 1
                if (n % 1000) == 0:
                    t = '行を色付けしています（' + str(n) + '行目）'
                    self.set_message_on_status_bar(t, True)
                self.line_data[i].paint_line(pane, paint_keywords)
            self.set_message_on_status_bar('', True)

    @staticmethod
    def _get_now():
        now = datetime.datetime.now(datetime.UTC) \
            + datetime.timedelta(hours=+9)
        jst = datetime.timezone(datetime.timedelta(hours=+9))
        now = now.replace(tzinfo=jst)
        return now

    @staticmethod
    def _convert_half_to_full(half):
        full = half
        full = re.sub('0', '０', full)
        full = re.sub('1', '１', full)
        full = re.sub('2', '２', full)
        full = re.sub('3', '３', full)
        full = re.sub('4', '４', full)
        full = re.sub('5', '５', full)
        full = re.sub('6', '６', full)
        full = re.sub('7', '７', full)
        full = re.sub('8', '８', full)
        full = re.sub('9', '９', full)
        return full

    @staticmethod
    def _get_encoding(raw_data):
        encoding = 'SHIFT_JIS'
        if raw_data != '':
            encoding = chardet.detect(raw_data)['encoding']
        if encoding is None:
            encoding = 'SHIFT_JIS'
        elif (re.match('^utf[-_]?.*$', encoding, re.I)) or \
             (re.match('^shift[-_]?jis.*$', encoding, re.I)) or \
             (re.match('^cp932.*$', encoding, re.I)) or \
             (re.match('^euc[-_]?(jp|jis).*$', encoding, re.I)) or \
             (re.match('^iso[-_]?2022[-_]?jp.*$', encoding, re.I)) or \
             (re.match('^ascii.*$', encoding, re.I)):
            pass
        else:
            # Windows-1252 (Western Europe)
            # MacCyrillic (Macintosh Cyrillic)
            # ...
            encoding = 'SHIFT_JIS'
            n = '警告'
            m = '文字コードを「SHIFT_JIS」に修正しました．'
            tkinter.messagebox.showwarning(n, m)
        return encoding

    @staticmethod
    def _decode_data(encoding, raw_data):
        try:
            decoded_data = raw_data.decode(encoding)
        except BaseException:
            try:
                decoded_data = raw_data.decode('utf-8')
            except BaseException:
                n = 'エラー'
                m = 'データを読みません（テキストでないかも？）'
                tkinter.messagebox.showwarning(n, m)
                raise BaseException('failed to read data')
                return None
        return decoded_data

    def _get_tmp_md(self):
        md_path = self.temp_dir.name + '/doc.md'
        file_text = self.txt.get('1.0', 'end-1c')
        # No warning here. Warning will be given during conversion.
        file_text = self.get_fully_unfolded_document(file_text, False)
        with open(md_path, 'w') as f:
            f.write(file_text)
        return md_path

    def _get_tmp_docx(self):
        md_path = self._get_tmp_md()
        docx_path = re.sub('md$', 'docx', md_path)
        stderr = sys.stderr
        sys.stderr = tempfile.TemporaryFile(mode='w+')
        importlib.reload(makdo.makdo_md2docx)
        try:
            m2d = makdo.makdo_md2docx.Md2Docx(md_path)
            m2d.save(docx_path)
        except BaseException:
            pass
        sys.stderr.seek(0)
        msg = sys.stderr.read()
        sys.stderr = stderr
        if msg != '':
            n = 'エラー'
            tkinter.messagebox.showerror(n, msg)
        return docx_path

    def _read_file(self, file_path):
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read()
        except BaseException:
            return None
        encoding = self._get_encoding(raw_data)
        try:
            document = self._decode_data(encoding, raw_data)
        except BaseException:
            return None
        return document

    def _read_docx_file(self, file_path):
        md_path = self.temp_dir.name + '/doc.md'
        stderr = sys.stderr
        sys.stderr = tempfile.TemporaryFile(mode='w+')
        importlib.reload(makdo.makdo_docx2md)
        try:
            d2m = makdo.makdo_docx2md.Docx2Md(file_path)
            d2m.save(md_path)
        except BaseException:
            pass
        sys.stderr.seek(0)
        msg = sys.stderr.read()
        sys.stderr = stderr
        if msg != '':
            n = 'エラー'
            tkinter.messagebox.showerror(n, msg)
            return None
        document = self._read_md_file(md_path)
        return document

    def _read_md_file(self, file_path):
        document = self._read_file(file_path)
        if document is None:
            return None
        document = self.get_fully_unfolded_document(document)
        return document

    def _read_txt_file(self, file_path):
        document = self._read_file(file_path)
        if document is None:
            return None
        return document

    def _read_csv_file(self, file_path):
        document = self._read_file(file_path)
        if document is None:
            return None
        is_in_cell = False
        table = '|'
        for i, c in enumerate(document):
            if c == '\n':
                if not is_in_cell:
                    table += '|\n|'
                else:
                    table += '<br>'
            elif c == '\r':
                continue
            elif c == ',':
                if not is_in_cell:
                    table += '|'
                else:
                    table += ','
            elif c == '"':
                is_in_cell = not is_in_cell
                if i > 0 and document[i - 1] == '"':
                    if is_in_cell:
                        table += '"'
            else:
                table += c
        if not re.match('^(.|\n)*\\|$', table):
            table += '|'
        if re.match('^(.|\n)*\n\\|$', table):
            table = re.sub('\n\\|$', '', table)
        table += '\n'
        return table

    def _read_xlsx_file(self, file_path):
        document = ''
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            document += '\n<!-- ' + sheet_name + ' -->\n'
            ws = wb[sheet_name]
            table = ''
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is None:
                        table += '|'
                    else:
                        table += '|' + str(cell.value)
                table += '|\n'
            document += table
        document = re.sub('^\n+', '', document)
        return document

    def _is_read_only_pane(self, pane):
        if pane == self.sub:
            if self.sub_pane_is_read_only:
                return True
            else:
                return False
        else:
            if self.is_read_only.get():
                return True
            else:
                return False

    def _is_region_specified(self, pane):
        if pane.tag_ranges('sel'):
            return True
        elif 'akauni' in pane.mark_names():
            return True
        return False

    def _get_region(self, pane):
        if pane.tag_ranges('sel'):
            beg, end = pane.index('sel.first'), pane.index('sel.last')
        elif 'akauni' in pane.mark_names():
            beg, end = self._get_indices_in_order(pane, 'insert', 'akauni')
        else:
            beg, end = '', ''
        return beg, end

    def cancel_region(self, pane):
        if pane.tag_ranges('sel'):
            pane.tag_remove('sel', "1.0", "end")
        if 'akauni' in pane.mark_names():
            pane.tag_remove('akauni_tag', '1.0', 'end')
            pane.mark_unset('akauni')

    def _show_no_region_error(self):
        n = 'エラー'
        m = '範囲が選択されていません．'
        tkinter.messagebox.showerror(n, m)

    def _get_indices_in_order(self, pane, index1, index2):
        position1 = pane.index(index1)
        position2 = pane.index(index2)
        p1_v = int(re.sub('\\..+$', '', position1))
        p1_h = int(re.sub('^.+\\.', '', position1))
        p2_v = int(re.sub('\\..+$', '', position2))
        p2_h = int(re.sub('^.+\\.', '', position2))
        if (p1_v < p2_v) or (p1_v == p2_v and p1_h < p2_h):
            return position1, position2
        if (p2_v < p1_v) or (p2_v == p1_v and p2_h < p1_h):
            return position2, position1
        return position1, position2

    def _open_sub_pane(self, document, is_read_only, exec_button='') -> bool:
        self.sub_pane_is_read_only = is_read_only
        if len(self.pnd_r.panes()) > 1:
            n = 'エラー'
            m = '別のサブウィンドウが開いています．'
            tkinter.messagebox.showerror(n, m)
            return False
        # self.quit_editing_formula()
        # self.close_memo_pad()
        half_height = int(self.pnd_r.winfo_height() / 2) - 5
        self.pnd_r.remove(self.pnd1)
        self.pnd_r.remove(self.pnd2)
        self.pnd_r.remove(self.pnd3)
        self.pnd_r.remove(self.pnd4)
        self.pnd_r.remove(self.pnd5)
        self.pnd_r.remove(self.pnd6)
        self.pnd_r.add(self.pnd1, height=half_height, minsize=100)
        self.pnd_r.add(self.pnd2, height=half_height, minsize=100)
        self.pnd_r.update()
        self._put_back_cursor_to_pane(self.txt)
        self.txt.focus_force()
        self.pnd_r.remove(self.pnd1)
        self.pnd_r.remove(self.pnd2)
        self.pnd_r.add(self.pnd1, height=half_height, minsize=100)
        self.pnd_r.add(self.pnd2, height=half_height, minsize=100)
        self.sub_frm.pack(side='bottom')
        try:
            self.sub_btn1.destroy()
            self.sub_btn2.destroy()
            self.sub_btn3.destroy()
        except BaseException:
            pass
        if exec_button == '質問':
            self.sub_btn1 = tkinter.Button(self.sub_frm, text='質問',
                                           command=self._execute_sub_pane)
            self.sub_btn1.pack(side='left', anchor='e')
            self.sub_btn2 = tkinter.Label(self.sub_frm, text='\u3000',
                                          bg='#BC7A00')
            self.sub_btn2.pack(side='left', anchor='e', fill='both')
            self.sub_btn3 = tkinter.Button(self.sub_frm, text='終了',
                                           command=self._close_sub_pane)
            self.sub_btn3.pack(side='right', anchor='w')
        elif exec_button == '入替':
            self.sub_btn1 = tkinter.Button(self.sub_frm, text='入替',
                                           command=self.swap_windows)
            self.sub_btn1.pack(side='left', anchor='e')
            self.sub_btn2 = tkinter.Label(self.sub_frm, text='\u3000',
                                          bg='#BC7A00')
            self.sub_btn2.pack(side='left', anchor='e', fill='both')
            self.sub_btn3 = tkinter.Button(self.sub_frm, text='終了',
                                           command=self._close_sub_pane)
            self.sub_btn3.pack(side='right', anchor='w')
        else:
            self.sub_btn1 = tkinter.Button(self.sub_frm, text='終了',
                                           command=self._close_sub_pane)
            self.sub_btn1.pack(side='top')
        self.sub_scb.pack(side='right', fill='y')
        self.sub.pack(expand=True, fill='both')
        for key in self.txt.configure():
            self.sub.configure({key: self.txt.cget(key)})
        self.sub['yscrollcommand'] = self.sub_scb.set
        #
        self.sub.delete('1.0', 'end')
        self.sub.insert('1.0', document)
        self.sub.mark_set('insert', '1.0')
        # self.sub.configure(state='disabled')
        self.sub.focus_force()
        self.current_pane = 'sub'
        return True

    def _execute_sub_pane(self) -> bool:
        return True

    def _close_sub_pane(self) -> bool:
        if len(self.pnd_r.panes()) == 1:
            return False
        # SUB CURSOR
        if 'x0sixteenth' in self.sub.mark_names():
            sub_ins = self.sub.get('1.0', 'insert')
            for i in range(16):
                j = i + 1
                ind_min = 'x' + str(i) + 'sixteenth'
                ind_max = 'x' + str(j) + 'sixteenth'
                sub_min = self.sub.get('1.0', ind_min)
                sub_max = self.sub.get('1.0', ind_max)
                n_up = len(sub_ins) - len(sub_min)
                n_dn = len(sub_max) - len(sub_ins)
                if n_up >= 0 and n_dn >= 0:
                    sub_up = self.sub.get(ind_min, 'insert')
                    sub_dn = self.sub.get('insert', ind_max)
                    txt_up = self.txt.get(ind_min,
                                          ind_min + '+' + str(n_up) + 'c')
                    txt_dn = self.txt.get(ind_max + '-' + str(n_dn) + 'c',
                                          ind_max)
                    if n_up + n_dn == 0:
                        n_rt = len(txt_up + txt_dn) / 2
                    else:
                        n_rt = (len(txt_up + txt_dn) * n_up) / (n_up + n_dn)
                    n_rt = round(n_rt)
                    if sub_up == txt_up:
                        self.txt.mark_set('sub_insert',
                                          ind_min + '+' + str(n_up) + 'c')
                    elif sub_dn == txt_dn:
                        self.txt.mark_set('sub_insert',
                                          ind_max + '-' + str(n_dn) + 'c')
                    else:
                        self.txt.mark_set('sub_insert',
                                          ind_min + '+' + str(n_rt) + 'c')
                    break
            # for i in range(17):
            #     self.txt.mark_unset('x' + str(i) + 'sixteenth')
            #     self.sub.mark_unset('x' + str(i) + 'sixteenth')
        #
        self.quit_editing_formula()
        self.update_memo_pad()
        # RAG (Retrieval-Augmented Generation)
        if hasattr(self, 'quit_editing_llama_rag_data'):
            if self.is_editing_llama_rag_data:
                self.quit_editing_llama_rag_data()
        self.memo_pad_memory = None
        self.close_mouse_menu()  # close mouse menu
        self.pnd_r.remove(self.pnd2)
        #
        self.txt.focus_force()
        self.current_pane = 'txt'
        #
        return True

    @staticmethod
    def _put_back_cursor_to_pane(pane, must_place_in_center=False):
        pane.update()
        p = pane.index('@0,0')  # "x,y" not "y,x"
        h_min = int(re.sub('\\.[0-9]+$', '', p))
        p = pane.index('@1000000,1000000')  # "x,y" not "y,x"
        h_max = int(re.sub('\\.[0-9]+$', '', p)) - 1
        p = pane.index('insert')
        h_cur = int(re.sub('\\.[0-9]+$', '', p))
        if must_place_in_center:
            half = (h_max - h_min) // 2
            pane.yview('insert-' + str(half) + 'l')
        elif h_cur <= h_min + 2:
            pane.yview('insert-2l')
        elif h_cur >= h_max - 2:
            pane.yview('insert-' + str(h_max - h_min - 2) + 'l')

    @staticmethod
    def _get_lines_of_pane(pane):
        pane.update()
        p = pane.index('@0,0')  # "x,y" not "y,x"
        h_min = int(re.sub('\\.[0-9]+$', '', p))
        p = pane.index('@1000000,1000000')  # "x,y" not "y,x"
        h_max = int(re.sub('\\.[0-9]+$', '', p)) - 1
        lines = h_max - h_min + 1
        return lines

    @staticmethod
    def _get_width_of_pane(pane):
        pane_width = pane.winfo_width()
        char_width = tkinter.font.Font(font=pane['font']).measure('X')
        return pane_width // char_width

    def _move_vertically(self, pane, ideal_h_position, height_to_move):
        i = self._get_v_position_of_insert(pane) + height_to_move
        j = ideal_h_position
        m = self._get_ideal_position_index_in_line(pane, i, j)
        pane.mark_set('insert', m)
        self._put_back_cursor_to_pane(pane)

    def _move_horizontally(self, pane, width_to_move):
        i = self._get_v_position_of_insert(pane)
        j = self._get_ideal_h_position_of_insert(pane) + width_to_move
        m = self._get_ideal_position_index_in_line(pane, i, j)
        pane.mark_set('insert', m)

    @staticmethod
    def _get_ideal_h_position_of_insert(pane):
        s = pane.get('insert linestart', 'insert')
        return get_real_width(s)

    @staticmethod
    def _get_ideal_position_index_in_line(pane, v_position, ideal_width):
        i = v_position
        line = pane.get(str(i) + '.0', str(i) + '.end')
        line_pre, line_pos, iw = '', '', 0
        for c in line:
            if c == '\t':
                iw += (int(iw / TAB_WIDTH) + 1) * TAB_WIDTH
            else:
                iw += get_real_width(c)
            if iw > ideal_width:
                break
            line_pre += c
        j = len(line_pre)
        return str(i) + '.' + str(j)

    def _jump_to_prev_pane(self):
        if self.current_pane == 'txt':
            if 'toc_cvs_frm' in vars(self):
                self.toc_cvs_frm.focus_set()
                self.current_pane = 'toc'
                self.set_message_on_status_bar('目次ウィンドウです')
            elif len(self.pnd_r.panes()) > 1:
                self.sub.focus_set()
                self.current_pane = 'sub'
                self.set_message_on_status_bar('サブウィンドウです')
        elif self.current_pane == 'sub':
            self.txt.focus_set()
            self.current_pane = 'txt'
            self.set_message_on_status_bar('メインウィンドウです')
        else:
            if len(self.pnd_r.panes()) > 1:
                self.sub.focus_set()
                self.current_pane = 'sub'
                self.set_message_on_status_bar('サブウィンドウです')
            else:
                self.txt.focus_set()
                self.current_pane = 'txt'
                self.set_message_on_status_bar('メインウィンドウです')

    def _jump_to_next_pane(self):
        if self.current_pane == 'txt':
            if len(self.pnd_r.panes()) > 1:
                self.sub.focus_set()
                self.current_pane = 'sub'
                self.set_message_on_status_bar('サブウィンドウです')
            elif 'toc_cvs_frm' in vars(self):
                self.toc_cvs_frm.focus_set()
                self.current_pane = 'toc'
                self.set_message_on_status_bar('目次ウィンドウです')
        elif self.current_pane == 'sub':
            if 'toc_cvs_frm' in vars(self):
                self.toc_cvs_frm.focus_set()
                self.current_pane = 'toc'
                self.set_message_on_status_bar('目次ウィンドウです')
            else:
                self.txt.focus_set()
                self.current_pane = 'txt'
                self.set_message_on_status_bar('メインウィンドウです')
        else:
            self.txt.focus_set()
            self.current_pane = 'txt'
            self.set_message_on_status_bar('メインウィンドウです')

    @staticmethod
    def _save_config_file(file_path, contents):
        if os.path.exists(file_path + '~'):
            os.chmod(file_path + '~', 0o600)
            os.remove(file_path + '~')
        if os.path.exists(file_path):
            os.rename(file_path, file_path + '~')
        with open(file_path, 'w') as f:
            f.write(contents)
        os.chmod(file_path, 0o400)

    def _execute_external_command(self, command: list) -> bool:
        self.set_message_on_status_bar('外部アプリを起動します', True)
        try:
            ret = subprocess.run(command,
                                 check=True, shell=False,
                                 stdout=subprocess.PIPE,
                                 encoding='utf-8')
            if ret.returncode == 0:
                self.set_message_on_status_bar('')
                return True
        except BaseException:
            self.set_message_on_status_bar('外部アプリの起動に失敗しました')
            return False
        self.set_message_on_status_bar('外部アプリの起動に失敗しました')
        return False

    # It is unclear why this process is necessary.
    def scroll_by_mouse_wheel(self, event):
        ex, ey = event.x_root, event.y_root
        if ('cmp_cvs' in vars(self)) and self.cmp_cvs.winfo_exists():
            ox, oy = self.cmp_cvs.winfo_rootx(), self.cmp_cvs.winfo_rooty()
            sx, sy = self.cmp_cvs.winfo_width(), self.cmp_cvs.winfo_height()
            if (ox <= ex and ex <= ox + sx) and (oy <= ey and ey <= oy + sy):
                self._scroll_by_mouse_wheel(self.cmp_cvs, event)
        if ('toc_cvs' in vars(self)) and self.toc_cvs.winfo_exists():
            ox, oy = self.toc_cvs.winfo_rootx(), self.toc_cvs.winfo_rooty()
            sx, sy = self.toc_cvs.winfo_width(), self.toc_cvs.winfo_height()
            if (ox <= ex and ex <= ox + sx) and (oy <= ey and ey <= oy + sy):
                self._scroll_by_mouse_wheel(self.toc_cvs, event)

    def _scroll_by_mouse_wheel(self, canvas, event):
        if sys.platform == 'win32':
            canvas.yview_scroll(- int(event.delta / 120), 'units')
        elif sys.platform == 'darwin':
            canvas.yview_scroll(- int(event.delta / 120), 'units')
        elif sys.platform == 'linux':
            if event.num == 4:
                canvas.yview_scroll(-1, 'units')
            elif event.num == 5:
                canvas.yview_scroll(+1, 'units')

    @staticmethod
    def _escape_search_word(search_word):
        search_word = search_word.replace('\\', '\\\\')
        search_word = search_word.replace('.', '\\.')
        search_word = search_word.replace('(', '\\(').replace(')', '\\)')
        search_word = search_word.replace('{', '\\{').replace('}', '\\}')
        search_word = search_word.replace('[', '\\[').replace(']', '\\]')
        search_word = search_word.replace('-', '\\-').replace('+', '\\+')
        search_word = search_word.replace('*', '\\*').replace('?', '\\?')
        search_word = search_word.replace('^', '\\^').replace('$', '\\$')
        search_word = search_word.replace('|', '\\|')
        return search_word

    def _clipboard_append(self, string):
        if string != '':
            self.win.clipboard_clear()
            self.win.clipboard_append(string)
            if self.clipboard_list[-1] != '':
                self.clipboard_list.append('')
            self.clipboard_list[-1] += string

    @staticmethod
    def _remove_head_and_tail_fds(doc):
        if doc[0] == '`' or doc[0] == '*' or \
           doc[0] == '-' or doc[0] == '+' or \
           doc[0] == '>' or doc[0] == '<' or \
           doc[0] == '^' or doc[0] == '_' or \
           doc[0] == '@' or doc[0] == '|':
            fds = ['`',                        # preformatted
                   '\\*{1,3}',                 # italic and bold
                   '\\-{2,3}', '\\+{2,3}',     # font scale
                   '\\+>', '<\\+',             # insert
                   '<{2,3}', '>{2,3}',         # font width
                   '\\^[0-9A-Za-z]{0,11}\\^',  # font color
                   '_[0-9A-Za-z]{1,11}_',      # highlight color
                   '@[^@]{1,66}@',             # font / font scale
                   '\\[\\|', '\\|\\]',         # frame
                   ]
            tmp = ''
            while doc != tmp:
                tmp = doc
                for fd in fds:
                    doc = re.sub('^' + fd + '((?:.|\n)*?)$', '\\1', doc)
                    doc = re.sub('^((?:.|\n)*?)' + fd + '$', '\\1', doc)
        return doc

    def _insert_inline_text(self, inline_text, step=0):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return
        pane.edit_separator()
        pane.insert('insert', inline_text)
        s = str(step) + 'c'
        if step >= 0:
            s = '+' + str(step) + 'c'
        pane.mark_set('insert', 'insert' + s)
        self.paint_out_line(self._get_v_position_of_insert(pane) - 1)
        pane.edit_separator()

    def _insert_paragraph_text(self, paragraph_text):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return
        pane.edit_separator()
        self._insert_line_break_as_necessary()
        pane.insert('insert', paragraph_text)
        pane.edit_separator()

    def _insert_line_break_as_necessary(self):
        t = self.txt.get('1.0', 'insert')
        if len(t) == 0:
            pass
        elif len(t) == 1:
            if t[-1] == '\n':
                pass
            else:
                self.txt.insert('insert', '\n\n')
        elif len(t) >= 2:
            if t[-2] == '\n' and t[-1] == '\n':
                pass
            elif t[-1] == '\n':
                self.txt.insert('insert', '\n')
            else:
                self.txt.insert('insert', '\n\n')
        p = self.txt.index('insert')
        t = self.txt.get('insert', 'end-1c')
        if len(t) == 0:
            self.txt.insert('insert', '\n')
        elif len(t) == 1:
            if t[0] == '\n':
                pass
            else:
                self.txt.insert('insert', '\n\n')
        elif len(t) >= 2:
            if t[0] == '\n' and t[1] == '\n':
                pass
            elif t[0] == '\n':
                self.txt.insert('insert', '\n')
            else:
                self.txt.insert('insert', '\n\n')
        self.txt.mark_set('insert', p)

    @staticmethod
    def _get_key(event):
        sym, stt = event.keysym, event.state
        if sym == 'Shift_L' or sym == 'Shift_R' or \
           sym == 'Control_L' or sym == 'Control_R' or \
           sym == 'Alt_L' or sym == 'Alt_R' or \
           sym == 'Mode_switch':
            return None
        if sym in ASCII_SYMBOLS:
            sym = ASCII_SYMBOLS[sym]
        # |           |windows| macOS | linux |
        # |Shift      |      1|      1|      1|
        # |Lock       |       |      2|       |
        # |Control    |      4|      4|      4|
        # |Alt        |       |     16|      8|
        # |Mode_switch|       |       |   8192|
        if stt & 4:
            return 'C-' + sym
        else:
            return sym

    def _is_key(self, key, code1=None, code2=None, code3=None, code4=None):
        if key is None:
            return False
        layout = self.shortcut_key_layout.get()
        if code1 is not None and key == code1:
            return True
        elif layout == 'akauni':
            if sys.platform == 'win32':
                # AUTOHOTKEY
                akey = {'C-c': 'C-f', 'C-f': 'C-c',
                        'C-g': 'C-v', 'C-v': 'C-g',
                        'C-w': 'C-z', 'C-z': 'C-w'}
                if code2 in akey:
                    code2 = akey[code2]
            if code2 is not None and key == code2:
                return True
        elif layout == 'qwerty':
            if code3 is not None and key == code3:
                return True
        elif layout == 'normal':
            if code4 is not None and key == code4:
                return True
        return False

    ####################################
    # MENU

    def _make_menu(self):
        self._make_menu_file()
        self._make_menu_edit()
        self._make_menu_insert()
        self._make_menu_paragraph()
        self._make_menu_move()
        self._make_menu_tool()
        self._make_menu_configuration()
        self._make_menu_internet()
        self._make_menu_special()
        self._make_menu_help()
        self.win['menu'] = self.mnb

    ##########################
    # MENU FILE

    def _make_menu_file(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='ファイル(F)', menu=menu, underline=5)
        #
        menu.add_command(label='ファイルを開く(O)', underline=8,
                         command=self.open_file)
        menu.add_command(label='ファイルを閉じる(C)', underline=9,
                         command=self.close_file)
        menu.add_separator()
        #
        menu.add_command(label='ファイルを保存(S)', underline=8,
                         command=self.save_file, accelerator='Ctrl+S')
        menu.add_command(label='Markdown形式で名前を付けて保存(M)', underline=20,
                         command=self.name_and_save_by_md)
        menu.add_command(label='MS Word形式で名前を付けて保存(D)', underline=19,
                         command=self.name_and_save_by_docx)
        # menu.add_command(label='名前を付けて保存(A)', underline=9,
        #                  command=self.name_and_save)
        menu.add_separator()
        #
        menu.add_command(label='ファイル形式を相互に直接変換',
                         command=self.convert_directly)
        menu.add_separator()
        #
        menu.add_command(label='PDFに変換',
                         command=self.convert_to_pdf)
        menu.add_command(label='見た目の確認・印刷(P)', underline=18,
                         command=self.start_writer, accelerator='Ctrl+P')
        menu.add_separator()
        #
        menu.add_command(label='OneDriveフォルダにコピーをアップロード',
                         command=self.upload_to_onedrive)
        menu.add_separator()
        #
        menu.add_command(label='終了(Q)', underline=3,
                         command=self.quit_makdo, accelerator='Ctrl+Q')
        # menu.add_separator()

    ################
    # COMMAND

    # OPEN FILE

    def open_file(self):
        _d, _f = None, None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
            _f = os.path.basename(self.file_path)
        ans = self.close_file()
        if ans is None:
            return False
        ti = 'ファイルを開く'
        ty = [('可能な形式', '.md .docx'),
              ('Markdown', '.md'), ('MS Word', '.docx'),
              ('全てのファイル', '*')]
        file_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d, initialfile=_f)
        if file_path == () or file_path == '':
            return False
        self.just_open_file(file_path)
        return True

    def just_open_file(self, file_path):
        if self.exists_auto_file(file_path):
            self.file_path = ''
            self.init_text = ''
            self.saved_text = ''
            self.file_lines = []
            return
        if re.match('^(?:.|\n)+.docx$', file_path):
            document = self._read_docx_file(file_path)
        else:
            document = self._read_md_file(file_path)
        if document is None:
            self.file_path = None
            return
        self.file_path = file_path
        self.init_text = document
        self.saved_text = document
        self.file_lines = document.split('\n')
        # self.txt.delete('1.0', 'end')
        self.txt.insert('1.0', document)
        self.txt.focus_set()
        self.current_pane = 'txt'
        self.txt.mark_set('insert', '1.0')
        self._set_file_name(file_path)
        if document == '':
            self.set_message_on_status_bar('空のファイルを開きました')
        # PAINT ALL LINES
        self.paint_all_lines(self.txt)
        # TABLE OF CONTENTS
        self.update_toc()
        # CLEAR THE UNDO STACK
        self.txt.edit_reset()

    def open_dropped_file(self, event):
        res_doc = '^(.|\n)+\\.(md|docx)$'
        res_xls = '^(.|\n)+\\.(xlsx)$'
        res_img = '^(.|\n)+\\.(jpg|jpeg|png|gif|tif|tiff|bmp)$'
        file_path = event.data
        file_path = re.sub('^{(.*)}$', '\\1', file_path)
        if re.match(res_doc, file_path, re.I):
            ans = self.close_file()
            if ans is None:
                return None
            self.just_open_file(file_path)
        elif re.match(res_xls, file_path, re.I):
            self.insert_table_from_excel(file_path)
        elif re.match(res_img, file_path, re.I):
            image_md_text = '![代替テキスト:縦x横](' + file_path + ' "説明")'
            self.txt.edit_separator()
            self.txt.insert('insert', image_md_text)
            self.paint_out_line(self._get_v_position_of_insert(pane) - 1)
            self.txt.edit_separator()

    # CLOSE FILE

    def close_file(self):
        # SAVE FILE
        if self._has_edited(True):  # must warn
            ans = self._ask_to_save('保存しますか？')
            if ans is None:
                return None
            elif ans is True:
                if not self.save_file():
                    return None
        if self._has_edited(False):  # must not warn
            ans = self._ask_to_save('データが消えますが、保存しますか？')
            if ans is None:
                return None
            elif ans is True:
                if not self.save_file():
                    return None
        # REMOVE AUTO SAVE FILE
        self.remove_auto_file(self.file_path)
        self.file_path = None
        self.init_text = ''
        self.saved_text = ''
        self.txt.delete('1.0', 'end')
        self.win.title('MAKDO')
        self.set_file_name_on_status_bar('')
        # TABLE OF CONTENTS
        is_toc_display_mode = self.is_toc_display_mode.get()
        if is_toc_display_mode:
            if 'toc_cvs' in vars(self):
                self.remove_toc()
            if 'toc_cvs' not in vars(self):
                self.settle_toc()
        # RETURN
        return True

    # SAVE FILE

    def _has_edited(self, must_warn=True):
        file_text = self.txt.get('1.0', 'end-1c')
        file_text = self.get_fully_unfolded_document(file_text, must_warn)
        # REMOVED 24.11.13 >
        # if file_text != '':
        #     if self.saved_text != file_text:
        #         return True
        # <
        if file_text == self.saved_text:
            return False
        return True

    def _ask_to_save(self, message):
        tkinter.Tk().withdraw()
        n, m, d = '確認', message, 'yes'
        return tkinter.messagebox.askyesnocancel(n, m, default=d)

    def save_file(self) -> bool:
        # FILE PATH
        if (self.file_path is None) or (self.file_path == ''):
            ti = 'ファイルを保存'
            ty = [('可能な形式', '.md .docx'),
                  ('Markdown', '.md'), ('MS Word', '.docx'),
                  ('全てのファイル', '*')]
            file_path \
                = tkinter.filedialog.asksaveasfilename(title=ti, filetypes=ty)
            if file_path == () or file_path == '':
                return False
            # if not re.match('^(?:.|\n)\\.md$', file_path):
            #     file_path += '.md'
            self.file_path = file_path
            self._set_file_name(file_path)
        # FILE
        file_text = self.txt.get('1.0', 'end-1c')
        if file_text != '' and file_text[-1] != '\n':
            file_text += '\n'
            self.txt.insert('end', '\n')
            self._put_back_cursor_to_pane(self.txt)
        must_warn = True
        if re.match('^(.|\n)+.docx$', self.file_path):
            must_warn = False
        if not self._has_edited(must_warn):
            self.set_message_on_status_bar('保存済みです')
            return False
        self._stamp_config(file_text)
        file_text = self.txt.get('1.0', 'end-1c')
        file_text = self.get_fully_unfolded_document(file_text,
                                                     False)  # must not warn
        # BACKUP FILE
        if self.make_backup_file.get() and not self.has_made_backup_file:
            if os.path.exists(self.file_path) and \
               not os.path.islink(self.file_path):
                try:
                    os.rename(self.file_path, self.file_path + '~')
                    self.has_made_backup_file = True
                except BaseException:
                    n = 'エラー'
                    m = 'バックアップに失敗しました．\n\n' \
                        + 'ファイルを上書きして保存しますか？'
                    d = 'no'
                    r = tkinter.messagebox.askyesnocancel(n, m, default=d)
                    if (r is None) or (not r):
                        return False
        # DOCX OR MD
        if re.match('^(?:.|\n)+.docx$', self.file_path):
            md_path = self.temp_dir.name + '/doc.md'
        else:
            md_path = self.file_path
        # SAVE MD FILE
        try:
            with open(md_path, 'w') as f:
                f.write(file_text)
        except BaseException:
            n, m = 'エラー', 'ファイルの保存に失敗しました．'
            tkinter.messagebox.showerror(n, m)
            return False
        # SAVE DOCX FILE
        if re.match('^(?:.|\n)+\\.docx$', self.file_path):
            stderr = sys.stderr
            sys.stderr = tempfile.TemporaryFile(mode='w+')
            importlib.reload(makdo.makdo_md2docx)
            try:
                m2d = makdo.makdo_md2docx.Md2Docx(md_path)
                m2d.save(self.file_path)
            except BaseException:
                pass
            sys.stderr.seek(0)
            msg = sys.stderr.read()
            sys.stderr = stderr
            if msg != '':
                n = '警告'
                tkinter.messagebox.showwarning(n, msg)
                # return
        self.txt.edit_separator()
        self.set_message_on_status_bar('保存しました')
        self.saved_text = file_text
        # RETURN
        return True

    def _stamp_config(self, file_text):
        if not re.match('^\\s*<!--', file_text):
            return
        file_text = re.sub('-->(.|\n)*$', '', file_text)
        now = datetime.datetime.now(datetime.UTC) \
            + datetime.timedelta(hours=+9)
        jst = datetime.timezone(datetime.timedelta(hours=+9))
        now = now.replace(tzinfo=jst)
        tit_res = '^((?:書題名|document_title):\\s*)(.*)$'
        cre_res = '^((?:作成時|created_time):\\s*)(\\S+)?(\\s.*)?$'
        mod_res = '^((?:更新時|modified_time):\\s*)(\\S+)?(\\s.*)?$'
        for i, line in enumerate(file_text.split('\n')):
            # DOCUMENT TITLE
            if re.match(tit_res, line):
                cfg = re.sub(tit_res, '\\1', line)
                tit = re.sub(tit_res, '\\2', line)
                if tit == '':
                    beg = str(i + 1) + '.' + str(len(cfg))
                    unix_time = datetime.datetime.timestamp(now)
                    self.txt.insert(beg, hex(int(unix_time * 1000000)))
                    if not re.match('^.*\\s$', cfg):
                        self.txt.insert(beg, ' ')
            # CREATED TIME
            if re.match(cre_res, line):
                cfg = re.sub(cre_res, '\\1', line)
                tim = re.sub(cre_res, '\\2', line)
                usr = re.sub(cre_res, '\\3', line)
                j, k = len(cfg), len(tim)
                beg = str(i + 1) + '.' + str(j)
                end = str(i + 1) + '.' + str(j + k)
                res_jst = '^' + '[0-9]{4}-[0-9]{2}-[0-9]{2}' + \
                    'T[0-9]{2}:[0-9]{2}:[0-9]{2}\\+09:00' + '(\\s.*)?$'
                if not re.match(res_jst, tim):
                    tim = ''
                try:
                    dt = datetime.datetime.fromisoformat(tim)
                except BaseException:
                    self.txt.delete(beg, end)
                    ts = now.isoformat(timespec='seconds')
                    self.txt.insert(beg, ts)
                    if not re.match('^.*\\s$', cfg):
                        self.txt.insert(beg, ' ')
            # MODIFIED TIME
            if re.match(mod_res, line):
                cfg = re.sub(mod_res, '\\1', line)
                tim = re.sub(mod_res, '\\2', line)
                usr = re.sub(mod_res, '\\3', line)
                j, k = len(cfg), len(tim)
                beg = str(i + 1) + '.' + str(j)
                end = str(i + 1) + '.' + str(j + k)
                self.txt.delete(beg, end)
                ts = now.isoformat(timespec='seconds')
                self.txt.insert(beg, ts)
                if not re.match('^.*\\s$', cfg):
                    self.txt.insert(beg, ' ')

    # NAME AND SAVE

    def name_and_save_by_md(self):
        ti = 'Markdown形式で名前をつけて保存'
        ty = [('Markdown', '.md')]
        file_path \
            = tkinter.filedialog.asksaveasfilename(title=ti, filetypes=ty)
        if file_path == () or file_path == '':
            return False
        if not re.match('^(?:.|\n)+\\.md$', file_path):
            file_path += '.md'
        self.remove_auto_file(self.file_path)
        self.file_path = file_path
        self.saved_text = ''
        self._set_file_name(file_path)
        self.save_file()
        return True

    def name_and_save_by_docx(self):
        ti = 'MS Word形式で名前をつけて保存'
        ty = [('MS Word', '.docx')]
        file_path \
            = tkinter.filedialog.asksaveasfilename(title=ti, filetypes=ty)
        if file_path == () or file_path == '':
            return False
        if not re.match('^(?:.|\n)+\\.docx$', file_path):
            file_path += '.docx'
        self.remove_auto_file(self.file_path)
        self.file_path = file_path
        self.saved_text = ''
        self._set_file_name(file_path)
        self.save_file()
        return True

    # AUTO SAVE

    def get_auto_path(self, file_path):
        if file_path is None or file_path == '':
            return None
        if '/' in file_path or '\\' in file_path:
            d = re.sub('^((?:.|\n)*[/\\\\])(.*)$', '\\1', file_path)
            f = re.sub('^((?:.|\n)*[/\\\\])(.*)$', '\\2', file_path)
        else:
            d = ''
            f = file_path
        if '.' in f:
            n = re.sub('^((?:.|\n)*)(\\..*)$', '\\1', f)
            e = re.sub('^((?:.|\n)*)(\\..*)$', '\\2', f)
        else:
            n = f
            e = ''
        n = re.sub('^((?:.|\n){,240})(.*)$', '\\1', n)
        return d + '~$' + n + e + '.zip'

    def exists_auto_file(self, file_path):
        auto_path = self.get_auto_path(file_path)
        if os.path.exists(auto_path):
            # auto_file = re.sub('^(.|\n)*[/\\\\]', '', auto_path)
            n = 'エラー'
            m = '自動保存ファイルが存在します．\n' + \
                '"' + auto_path + '"\n\n' + \
                '①現在、ファイルを編集中\n' + \
                '②過去の編集中のファイルが残存\n' + \
                'の2つの可能性が考えられます．\n\n' + \
                '①現在、ファイルを編集中\n' + \
                'の場合は、「No」を選択してください．\n\n' + \
                '②過去の編集中のファイルが残存\n' + \
                'の場合、異常終了したものと思われます．\n' + \
                '「No」を選択して、' + \
                '自動保存ファイルの中身を確認してから、' + \
                '削除することをおすすめします．\n\n' + \
                '自動保存ファイルを削除しますか？'
            ans = tkinter.messagebox.askyesno(n, m, default='no')
            if ans:
                try:
                    self.remove_auto_file(file_path)
                except BaseException:
                    n, m = 'エラー', '自動保存ファイルの削除に失敗しました．'
                    tkinter.messagebox.showerror(n, m)
        if os.path.exists(auto_path):
            return True
        else:
            return False

    def save_auto_file(self, file_path):
        if file_path is not None and file_path != '':
            new_text = self.txt.get('1.0', 'end-1c')
            auto_path = self.get_auto_path(file_path)
            if os.path.exists(auto_path):
                with zipfile.ZipFile(auto_path, 'r') as old_zip:
                    with old_zip.open('doc.md', 'r') as f:
                        old_text = f.read()
                        if new_text == old_text.decode():
                            return
            try:
                with zipfile.ZipFile(auto_path, 'w',
                                     compression=zipfile.ZIP_DEFLATED,
                                     compresslevel=9) as new_zip:
                    new_zip.writestr('doc.md', new_text)
            except BaseException:
                if 'must_show_auto_file_save_failed_message' not in vars(self):
                    n = 'エラー'
                    m = '自動保存ファイルの作成に\n' \
                        + '失敗しました．\n\n' \
                        + '異常終了してしまった場合に、\n' \
                        + '編集中のデータが失われてしまう\n' \
                        + '可能性があります．\n\n' \
                        + 'フォルダの書込み権限の有無を\n' \
                        + 'ご確認ください．'
                    tkinter.messagebox.showerror(n, m)
                    self.must_show_auto_file_save_failed_message = False

    def remove_auto_file(self, file_path):
        if file_path is not None and file_path != '':
            auto_path = self.get_auto_path(file_path)
            if re.match('(^|(.|\n)*[/\\\\])~\\$(.|\n)+\\.zip$', auto_path):
                if os.path.exists(auto_path):
                    os.remove(auto_path)

    # CONVERT DIRECTLY

    def convert_directly(self):
        # mac doesn't support "tkinterdnd2" (drag and drop)
        if sys.platform != 'darwin':
            self.convert_directly_on_non_mac()
        else:
            self.convert_directly_on_mac()

    def convert_directly_on_non_mac(self):
        if len(self.pnd_r.panes()) > 1:
            return False
        # self.quit_editing_formula()
        # self.close_memo_pad()
        self.pnd_r.update()
        half_height = int(self.pnd_r.winfo_height() / 2) - 5
        self.pnd_r.remove(self.pnd1)
        self.pnd_r.remove(self.pnd2)
        self.pnd_r.remove(self.pnd3)
        self.pnd_r.remove(self.pnd4)
        self.pnd_r.remove(self.pnd5)
        self.pnd_r.remove(self.pnd6)
        self.pnd_r.add(self.pnd1, height=half_height, minsize=100)
        self.pnd_r.add(self.pnd4, height=half_height, minsize=100)
        # self.pnd_r.update()
        #
        btn = tkinter.Button(self.pnd4, text='終了',
                             command=self._quit_converting_directly)
        btn.pack(side='bottom')
        #
        self.pool = tkinter.Text(self.pnd4)
        self.pool.drop_target_register(tkinterdnd2.DND_FILES)
        self.pool.insert('end', 'ここにmdファイル又はdocxファイルをドロップしてください\n')
        self.pool.dnd_bind('<<Drop>>', self._convert_dropped_file)
        self.pool.pack(expand=True, side='top', fill='both')
        self.pool.config(bg='#00A586', fg='white')
        self.pool['font'] = self.gothic_font

    def _convert_dropped_file(self, event):
        filename = event.data
        filename = re.sub('^{(.*)}$', '\\1', filename)
        basename = os.path.basename(filename)
        self.pool.delete('1.0', 'end')
        self.pool.insert('end', '"' + basename + '"を受け取りました\n')
        stderr = sys.stderr
        sys.stderr = tempfile.TemporaryFile(mode='w+')
        if re.match('^.*\\.(m|M)(d|D)$', filename):
            self.pool.insert('end', 'docxファイルを作成します\n')
            try:
                importlib.reload(makdo.makdo_md2docx)
                m2d = makdo.makdo_md2docx.Md2Docx(filename)
                m2d.save('')
                self.pool.insert('end', 'docxファイルを作成しました\n')
            except BaseException:
                sys.stderr.seek(0)
                self.pool.insert('end', sys.stderr.read())
                self.pool.insert('end', 'docxファイルを作成できませんでした\n')
        elif re.match('^.*\\.(d|D)(o|O)(c|C)(x|X)$', filename):
            self.pool.insert('end', 'mdファイルを作成します\n')
            try:
                importlib.reload(makdo.makdo_docx2md)
                d2m = makdo.makdo_docx2md.Docx2Md(filename)
                d2m.save('')
                self.pool.insert('end', 'mdファイルを作成しました\n')
            except BaseException:
                sys.stderr.seek(0)
                self.pool.insert('end', sys.stderr.read())
                self.pool.insert('end', 'mdファイルを作成できませんでした\n')
        else:
            self.pool.insert('end', '不適切なファイルです\n')
        sys.stderr = stderr
        self.pool.insert('end', '\nここにmdファイル又はdocxファイルをドロップしてください\n')

    def _quit_converting_directly(self):
        self.pnd_r.remove(self.pnd4)
        self.txt.focus_set()
        self.current_pane = 'txt'

    def convert_directly_on_mac(self):
        ti = '相互に直接変換'
        ty = [('可能な形式', '.md .docx'),
              ('Markdown', '.md'), ('MS Word', '.docx'),
              ('全てのファイル', '*')]
        _d, _f = None, None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
            _f = os.path.basename(self.file_path)
        file_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d, initialfile=_f)
        if file_path == () or file_path == '':
            return
        stderr = sys.stderr
        sys.stderr = tempfile.TemporaryFile(mode='w+')
        msg = ''
        if re.match('^.*\\.(m|M)(d|D)$', file_path):
            try:
                importlib.reload(makdo.makdo_md2docx)
                m2d = makdo.makdo_md2docx.Md2Docx(file_path)
                m2d.save('')
            except BaseException:
                pass
        elif re.match('^.*\\.(d|D)(o|O)(c|C)(x|X)$', file_path):
            try:
                importlib.reload(makdo.makdo_docx2md)
                d2m = makdo.makdo_docx2md.Docx2Md(file_path)
                d2m.save('')
            except BaseException:
                pass
        else:
            n, m = 'エラー', '変換できないファイル形式です．'
            tkinter.messagebox.showerror(n, m)
        sys.stderr.seek(0)
        msg = sys.stderr.read()
        if msg != '':
            n = '警告'
            tkinter.messagebox.showwarning(n, msg)
        sys.stderr = stderr

    # CONVERT TO PDF

    def convert_to_pdf(self) -> bool:
        ti, ty = 'PDFに変換', [('PDF', '.pdf')]
        _d = '.'
        _f = ''
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
            _d = '.' if _d == '' else _d
            _f = os.path.basename(self.file_path)
            _f = re.sub('\\.(md|docx)$', '', _f) + '.pdf'
        pdf_path = tkinter.filedialog.asksaveasfilename(
            title=ti, filetypes=ty, initialdir=_d, initialfile=_f)
        if pdf_path == () or pdf_path == '':
            return False
        if not re.match('^(?:.|\n)+\\.pdf$', pdf_path):
            pdf_path += '.pdf'
        tmp_docx = self._get_tmp_docx()
        if sys.platform == 'win32':
            # MS Word
            try:
                self.set_message_on_status_bar('PDFに変換します', True)
                app = win32com.client.Dispatch("Word.Application")
                app.Visible = False
                doc = app.Documents.Open(FileName=tmp_docx,
                                         ConfirmConversions=False,
                                         ReadOnly=True)
                doc.SaveAs(pdf_path, FileFormat=17)  # 17=PDF
                doc.Close()
                app.Quit()
                self.set_message_on_status_bar('PDFに変換しました')
                return True
            except BaseException:
                pass
            # LibreOffice
            lo = 'C:/Program Files/LibreOffice/program/soffice.exe'
            if self._convert_to_pdf_by_libreoffice(lo, tmp_docx, pdf_path):
                return True
        elif sys.platform == 'darwin':
            # LibreOffice
            lo = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
            if self._convert_to_pdf_by_libreoffice(lo, tmp_docx, pdf_path):
                return True
            # TextEdit / Pages
            if 'has_showed_help_message_of_converting_to_pdf' not in locals():
                n = 'お知らせ'
                m = 'LibreOfficeの起動に失敗しました．\n\n' \
                    + 'mac環境では、標準で、\n' \
                    + '直接PDFを作成する方法が\n' \
                    + 'ありません．\n\n' \
                    + 'ワープロアプリを起動しますので、\n' \
                    + '印刷メニュー等から\n' \
                    + 'PDFに変換してください．'
                tkinter.messagebox.showinfo(n, m)
                self.has_showed_help_message_of_converting_to_pdf = True
            com = ['open', tmp_docx]
            if self._execute_external_command(com):
                return True
            return False
        elif sys.platform == 'linux':
            # LibreOffice
            lo = '/usr/bin/libreoffice'
            if self._convert_to_pdf_by_libreoffice(lo, tmp_docx, pdf_path):
                return True
        n = '警告'
        m = '外部アプリ（MS Word等）の\n' \
            + '起動に失敗しました．\n\n' \
            + '下記をインストールしてください．\n' \
            + '- MS Word\n' \
            + '- LibreOffice（無料）'
        tkinter.messagebox.showwarning(n, m)
        return False

    def _convert_to_pdf_by_libreoffice(self, libreoffice, tmp_docx, pdf_path):
        dir_path = re.sub('((?:.|\n)*)/(?:.|\n)+$', '\\1', tmp_docx)
        com = [libreoffice,
               '--headless',
               '--convert-to', 'pdf',
               '--outdir', dir_path,
               tmp_docx]
        if self._execute_external_command(com):
            tmp_pdf = re.sub('docx$', 'pdf', tmp_docx)
            if not os.path.exists(tmp_pdf):
                self.set_message_on_status_bar('PDFの変換に失敗しました')
                return True
            shutil.move(tmp_pdf, pdf_path)
            if not os.path.exists(pdf_path):
                self.set_message_on_status_bar('PDFの変換に失敗しました')
                return True
            self.set_message_on_status_bar('PDFに変換しました')
            return True
        return False

    # START WRITER

    def start_writer(self) -> bool:
        docx_path = self._get_tmp_docx()
        if sys.platform == 'win32':
            # MS Word
            try:
                self.set_message_on_status_bar('MS Wordを起動します', True)
                app = win32com.client.Dispatch("Word.Application")
                app.Visible = True
                doc = app.Documents.Open(FileName=docx_path,
                                         ConfirmConversions=False,
                                         ReadOnly=True)
                self.set_message_on_status_bar('')
                return True
            except BaseException:
                self.set_message_on_status_bar('MS Wordの起動に失敗しました')
            # LibreOffice
            com = ['C:/Program Files/LibreOffice/program/soffice.exe',
                   docx_path]
            if self._execute_external_command(com):
                return True
            # WordPad
            com = ['C:/Program Files/Windows NT/Accessories/wordpad.exe',
                   docx_path]
            if self._execute_external_command(com):
                return True
        elif sys.platform == 'darwin':
            # MS Word
            com = ['/Applications/Microsoft Word.app'
                   + '/Contents/MacOS/Microsoft Word', docx_path]
            if self._execute_external_command(com):
                return True
            # LibreOffice
            com = ['/Applications/LibreOffice.app/Contents/MacOS/soffice',
                   docx_path]
            if self._execute_external_command(com):
                return True
            # TextEdit / Pages
            com = ['open', docx_path]
            if self._execute_external_command(com):
                return True
        elif sys.platform == 'linux':
            # LibreOffice
            com = ['/usr/bin/libreoffice', docx_path]
            if self._execute_external_command(com):
                return True
        n = '警告'
        m = '外部アプリ（MS Word等）の\n' \
            + '起動に失敗しました．\n\n' \
            + '下記をインストールしてください．\n' \
            + '- MS Word\n' \
            + '- LibreOffice（無料）'
        tkinter.messagebox.showwarning(n, m)
        return False

    # UPLOAD TO ONEDRIVE

    def upload_to_onedrive(self):
        if self.onedrive_directory is None:
            self.set_onedrive_directory()
        if self.onedrive_directory is None:
            return False
        if self.file_path is None:
            ti = 'OneDriveフォルダにコピーをアップロード'
            ty = [('MS Word', '.docx')]
            file_path = tkinter.filedialog.asksaveasfilename(
                title=ti, filetypes=ty, initialdir=self.onedrive_directory)
            if file_path == () or file_path == '':
                return False
            if not re.match('^(.|\n)*\\.docx$', file_path):
                file_path += '.docx'
        elif re.match('^(.|\n)*\\.md$', self.file_path):
            file_name = re.sub('md$', 'docx', os.path.basename(self.file_path))
            file_path = self.onedrive_directory + '/' + file_name
        elif re.match('^(.|\n)*\\.docx$', self.file_path):
            file_name = os.path.basename(filepath)
            file_path = self.onedrive_directory + '/' + file_name
        file_text = self.txt.get('1.0', 'end-1c')
        md_path = self.temp_dir.name + '/doc.md'
        try:
            with open(md_path, 'w') as f:
                f.write(file_text)
        except BaseException:
            n, m = 'エラー', 'ファイルの保存に失敗しました．'
            tkinter.messagebox.showerror(n, m)
            return False
        stderr = sys.stderr
        sys.stderr = tempfile.TemporaryFile(mode='w+')
        importlib.reload(makdo.makdo_md2docx)
        try:
            m2d = makdo.makdo_md2docx.Md2Docx(md_path)
            m2d.save(file_path)
        except BaseException:
            pass
        sys.stderr.seek(0)
        msg = sys.stderr.read()
        sys.stderr = stderr
        if msg != '':
            n = '警告'
            tkinter.messagebox.showwarning(n, msg)
        # return
        self.set_message_on_status_bar('アップロードしました')
        return True

    # QUIT

    def quit_makdo(self):
        ans = self.close_file()
        if ans is None:
            return None
        self.win.quit()
        self.win.destroy()
        sys.exit(0)

    ##########################
    # MENU EDIT

    def _make_menu_edit(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='編集(E)', menu=menu, underline=3)
        #
        menu.add_command(label='元に戻す(U)', underline=5,
                         command=self.edit_modified_undo, accelerator='Ctrl+Z')
        menu.add_command(label='やり直す(R)', underline=5,
                         command=self.edit_modified_redo, accelerator='Ctrl+Y')
        menu.add_separator()
        #
        menu.add_command(label='切り取り(C)', underline=5,
                         command=self.cut_region, accelerator='Ctrl+X')
        menu.add_command(label='コピー(Y)', underline=4,
                         command=self.copy_region, accelerator='Ctrl+C')
        menu.add_command(label='貼り付け(P)', underline=5,
                         command=self.paste_region, accelerator='Ctrl+V')
        menu.add_command(label='リストから貼り付け',
                         command=self.paste_region_from_list)
        menu.add_separator()
        #
        menu.add_command(label='矩形（四角形）を切り取り',
                         command=self.cut_rectangle)
        menu.add_command(label='矩形（四角形）をコピー',
                         command=self.copy_rectangle)
        menu.add_command(label='矩形（四角形）を貼り付け',
                         command=self.paste_rectangle)
        menu.add_separator()
        #
        menu.add_command(label='全て選択(A)', underline=5,
                         command=self.select_all, accelerator='Ctrl+A')
        menu.add_separator()
        #
        menu.add_command(label='前を置換',
                         command=self.replace_backward)
        menu.add_command(label='後を置換',
                         command=self.replace_forward, accelerator='Ctrl+L')
        menu.add_command(label='全て置換',
                         command=self.replace_all)
        menu.add_separator()
        #
        menu.add_command(label='選択範囲を大文字に変換',
                         command=self.replace_lower_case_with_upper_case)
        menu.add_command(label='選択範囲を小文字に変換',
                         command=self.replace_upper_case_with_lower_case)
        menu.add_command(label='選択範囲を全角文字に変換',
                         command=self.replace_half_width_with_full_width)
        menu.add_command(label='選択範囲を半角文字に変換',
                         command=self.replace_full_width_with_half_width)
        menu.add_command(label='選択範囲の半角全角を推奨の形に変換',
                         command=self.adjust_character_width)
        menu.add_separator()
        #
        menu.add_command(label='選択範囲の行を正順にソート（並替え）',
                         command=self.sort_lines)
        menu.add_command(label='選択範囲の行を逆順にソート（並替え）',
                         command=self.sort_lines_in_reverse_order)
        menu.add_separator()
        #
        menu.add_command(label='数式を計算',
                         command=self.calculate)
        menu.add_separator()
        #
        menu.add_command(label='字体を変える',
                         command=self.change_typeface)
        menu.add_separator()
        #
        menu.add_command(label='コメントアウトにする',
                         command=self.comment_out_region)
        menu.add_command(label='コメントアウトを取り消す',
                         command=self.uncomment_in_region)
        # menu.add_separator()

    ######
    # COMMAND

    def edit_modified_undo(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        try:
            pane.edit_undo()
        except BaseException:
            pass
        word1 = self.stb_sor1.get()
        if Makdo.search_word == word1:
            self._highlight_search_word()
        self.set_message_on_status_bar('元に戻しました（undo）')

    def edit_modified_redo(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        try:
            pane.edit_redo()
        except BaseException:
            pass
        word1 = self.stb_sor1.get()
        if Makdo.search_word == word1:
            self._highlight_search_word()
        self.set_message_on_status_bar('やり直しました（redo）')

    def cut_region(self):
        self._cut_or_copy_region(True)

    def copy_region(self):
        self._cut_or_copy_region(False)

    def _cut_or_copy_region(self, must_cut=False):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if must_cut:
            if self._is_read_only_pane(pane):
                return False
        beg, end = self._get_region(pane)
        if beg == '' or end == '':
            self._show_no_region_error()
            return False
        c = pane.get(beg, end)
        self.win.clipboard_clear()
        self.win.clipboard_append(c)
        if self.clipboard_list[-1] != '':
            self.clipboard_list.append('')
        self.clipboard_list[-1] += c
        self.cancel_region(pane)
        if must_cut:
            pane.edit_separator()
            pane.delete(beg, end)
            if self.current_pane == 'txt':
                # PAINT LINES
                vp = int(re.sub('\\.[0-9]+$', '', beg))
                n = c.count('\n')
                for i in range(n):
                    self.line_data.pop(vp)
                for i, ld in enumerate(self.line_data):
                    ld.line_number = i
                self.paint_out_line(vp - 1)
                # UPDATE TOC
                self.update_toc()
            pane.edit_separator()
        return True

    def paste_region(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return False
        if self.current_pane == 'txt':
            beg_v = self._get_v_position_of_insert(self.txt)
        try:
            cb = self.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb == '':
            return True
        pane.edit_separator()
        pane.insert('insert', cb)
        if self.current_pane == 'txt':
            # PAINT LINES
            end_v = self._get_v_position_of_insert(self.txt)
            n = end_v - beg_v
            for i in range(n):
                self.line_data.insert(beg_v, LineDatum())
            for i, ld in enumerate(self.line_data):
                ld.line_number = i
            for i in range(beg_v - 1, end_v):
                self.paint_out_line(i)
            # UPDATE TOC
            self.update_toc()
        self._put_back_cursor_to_pane(pane)
        pane.edit_separator()
        return True

    def paste_region_from_list(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        t = 'リストから貼付け'
        m = '貼り付ける文節を選んでください．'
        c = list(reversed(self.clipboard_list))
        rd = RadiobuttonDialog(mother, self, t, m, c, 0)
        v = rd.get_value()
        if v is not None:
            pane.edit_separator()
            pane.insert('insert', v)
            p = self._get_v_position_of_insert(pane) - 1
            for i in range(v.count('\n')):
                self.paint_out_line(p + i)
            pane.edit_separator()

    def cut_rectangle(self):
        self._cut_or_copy_rectangle(True)

    def copy_rectangle(self):
        self._cut_or_copy_rectangle(False)

    def _cut_or_copy_rectangle(self, must_cut=False):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if must_cut:
            if self._is_read_only_pane(pane):
                return False
        beg, end = self._get_region(pane)
        if beg == '' or end == '':
            self._show_no_region_error()
            return False
        if must_cut:
            pane['autoseparators'] = False
            pane.edit_separator()
        beg_v = int(re.sub('\\.[0-9]+$', '', beg))
        s = pane.get(beg + ' linestart', beg)
        beg_ih = get_real_width(s)
        end_v = int(re.sub('\\.[0-9]+$', '', end))
        s = pane.get(end + ' linestart', end)
        end_ih = get_real_width(s)
        min_ih = min(beg_ih, end_ih)
        max_ih = max(beg_ih, end_ih)
        self.cancel_region(pane)
        self.rectangle_text_list = []
        for i in range(beg_v - 1, end_v):
            line = pane.get(str(i + 1) + '.0', str(i + 1) + '.end')
            line_pre, line_mid, line_pos = '', '', ''
            for c in line:
                if get_real_width(line_pre) < min_ih:
                    line_pre += c
                elif get_real_width(line_pre + line_mid) < max_ih:
                    line_mid += c
                else:
                    line_pos += c
            self.rectangle_text_list.append(line_mid)
            if must_cut:
                pane.delete(str(i + 1) + '.' + str(len(line_pre)),
                            str(i + 1) + '.' + str(len(line_pre + line_mid)))
                self.paint_out_line(i)
                self.update_toc()
        if must_cut:
            pane['autoseparators'] = True
            pane.edit_separator()
        return True

    def paste_rectangle(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return False
        if self.rectangle_text_list == []:
            return True
        pane['autoseparators'] = False
        pane.edit_separator()
        ins_v = self._get_v_position_of_insert(pane)
        max_v = self._get_max_v_position(pane)
        s = pane.get(str(ins_v) + '.0', 'insert')
        ins_ih = get_real_width(s)
        for j, line_md in enumerate(self.rectangle_text_list):
            i = ins_v + j - 1
            if i < max_v - 1:
                line = pane.get(str(i + 1) + '.0', str(i + 1) + '.end')
                line_pre, line_pos = '', ''
                for c in line:
                    if get_real_width(line_pre) < ins_ih:
                        line_pre += c
                    else:
                        break
                ins_h = str(i + 1) + '.' + str(len(line_pre))
            else:
                ins_h = 'end'
                line_md += '\n'
            pane.insert(ins_h, line_md)
            pane.mark_set('insert', ins_h)
            self.paint_out_line(i)
            self.update_toc()
        pane['autoseparators'] = True
        pane.edit_separator()
        self._put_back_cursor_to_pane(pane)
        return True

    def select_all(self):
        self.txt.tag_add('sel', '1.0', 'end-1c')

    def replace_backward(self):
        self.search_or_replace_backward(True)  # must_replace = True

    def replace_forward(self):
        self.search_or_replace_forward(True)   # must_replace = True

    def replace_all(self, focus=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return
        if focus is None:
            focus = pane
        word1 = self.stb_sor1.get()
        word2 = self.stb_sor2.get()
        t = '全置換'
        m = '検索する言葉と置換する言葉を入力してください．'
        h1, t1 = '検索', ''
        h2, t2 = '置換', ''
        sd = TwoWordsDialog(focus, self, t, m, h1, h2, t1, t2, word1, word2)
        word1, word2 = sd.get_value()
        if (word1 is None) or (word2 is None) or (word1 == ''):
            return
        self.stb_sor1.delete('0', 'end')
        self.stb_sor1.insert('0', word1)
        self.stb_sor2.delete('0', 'end')
        self.stb_sor2.insert('0', word2)
        if Makdo.search_word != word1:
            Makdo.search_word = word1
        if pane.tag_ranges('sel'):
            beg, end = pane.index('sel.first'), pane.index('sel.last')
        elif 'akauni' in pane.mark_names():
            beg, end = self._get_indices_in_order(pane, 'insert', 'akauni')
        else:
            beg, end = '1.0', 'end-1c'
        pane['autoseparators'] = False
        pane.edit_separator()
        m = 0
        res_word1 = word1
        if not self.use_regexps.get():
            res_word1 = self._escape_search_word(word1)
        res = '^((?:.|\n)*)(' + res_word1 + ')((?:.|\n)*)$'
        while True:
            tex = pane.get(beg, end)
            try:
                if not re.match(res, tex):
                    break
            except BaseException:
                pane.focus_set()
                self.set_message_on_status_bar('置換に失敗しました')
                return
            s = re.sub(res, '\\1', tex)
            w = re.sub(res, '\\2', tex)
            t = re.sub(res, '\\3', tex)
            if w == '':
                continue
            pane.delete(beg + '+' + str(len(s)) + 'c',
                        beg + '+' + str(len(s + w)) + 'c')
            pane.insert(beg + '+' + str(len(s)) + 'c', word2)
            end = beg + '+' + str(len(s)) + 'c'
            m += 1
        self.cancel_region(pane)
        pane['autoseparators'] = True
        pane.edit_separator()
        pane.focus_set()
        # MESSAGE
        self.set_message_on_status_bar(str(m) + '個を置換しました')

    def replace_lower_case_with_upper_case(self):
        self._replace_x_with_y('lower_case_with_upper_case')

    def replace_upper_case_with_lower_case(self):
        self._replace_x_with_y('upper_case_with_lower_case')

    def replace_half_width_with_full_width(self):
        self._replace_x_with_y('half_width_with_full_width')

    def replace_full_width_with_half_width(self):
        self._replace_x_with_y('full_width_with_half_width')

    def adjust_character_width(self):
        self._replace_x_with_y('adjust_character_width')

    def sort_lines(self):
        self._replace_x_with_y('sort_in_forward_order')

    def sort_lines_in_reverse_order(self):
        self._replace_x_with_y('sort_in_reverse_order')

    def _replace_x_with_y(self, mode: str) -> bool:
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return False
        if not self._is_region_specified(pane):
            self._show_no_region_error()
            return False
        # GET SCOPE
        beg_c, end_c = self._get_region(pane)
        beg_v = int(re.sub('\\.[0-9]+$', '', beg_c))
        end_v = int(re.sub('\\.[0-9]+$', '', end_c))
        if mode == 'sort_in_forward_order' or mode == 'sort_in_reverse_order':
            if not re.match('^[0-9]+\\.0$', beg_c):
                beg_v += 1
            end_v -= 1
            beg_c, end_c = str(beg_v) + '.0', str(end_v) + '.end'
        # GET OLD AND NEW STRING
        old_str = pane.get(beg_c, end_c)
        if mode == 'lower_case_with_upper_case':
            new_str = old_str.upper()
        elif mode == 'upper_case_with_lower_case':
            new_str = old_str.lower()
        elif mode == 'half_width_with_full_width':
            new_str = old_str
            for hf in HALF_FULL_TABLE:
                new_str = new_str.replace(hf[0], hf[1])
        elif mode == 'full_width_with_half_width':
            new_str = old_str
            for hf in HALF_FULL_TABLE:
                new_str = new_str.replace(hf[1], hf[0])
        elif mode == 'adjust_character_width':
            new_str = old_str
            # NUMBER
            for i in range(0, 10):
                new_str = new_str.replace(chr(i + 65296), chr(i + 48))
            new_str = re.sub('＋([0-9]+)', '+\\1', new_str)
            new_str = re.sub('－([0-9]+)', '-\\1', new_str)
            new_str = re.sub('([0-9]+)．([0-9]+)', '\\1.\\2', new_str)
            new_str = re.sub('([0-9]+)，([0-9]{3})', '\\1,\\2', new_str)
            new_str = re.sub('([0-9]+)，([0-9]{3})', '\\1,\\2', new_str)
            # ALPHABET
            for i in range(0, 26):
                new_str = new_str.replace(chr(i + 65313), chr(i + 65))
                new_str = new_str.replace(chr(i + 65345), chr(i + 97))
            # COMMA
            new_str = new_str.replace('，', '、')
        elif mode == 'sort_in_forward_order':
            old_lst = old_str.split('\n')
            new_lst = sorted(old_lst)
            new_str = '\n'.join(new_lst)
        elif mode == 'sort_in_reverse_order':
            old_lst = old_str.split('\n')
            new_lst = sorted(old_lst)
            new_lst = list(reversed(sorted(old_lst)))
            new_str = '\n'.join(new_lst)
        if old_str == new_str:
            return True
        # APPLY
        pane['autoseparators'] = False
        pane.edit_separator()
        pane.delete(beg_c, end_c)
        pane.insert(beg_c, new_str)
        self.cancel_region(pane)
        if self.current_pane == 'txt':
            for i in range(beg_v - 1, end_v):
                self.paint_out_line(i)
            self.update_toc()
        pane['autoseparators'] = True
        pane.edit_separator()
        return True

    def calculate(self, must_show_message=True) -> bool:
        line = self.txt.get('insert linestart', 'insert lineend')
        line_head = ''
        line_math = line
        line_rslt = ''
        line_tail = ''
        res = '^(.*(?:<!--|@))(.*)$'
        if re.match(res, line_math):
            line_head = re.sub(res, '\\1', line_math)
            line_math = re.sub(res, '\\2', line_math)
        res = '^(.*)((?:-->|#).*)$'
        if re.match(res, line_math):
            line_tail = re.sub(res, '\\2', line_math)
            line_math = re.sub(res, '\\1', line_math)
        res = '^(.*)(=.*)$'
        if re.match(res, line_math):
            line_rslt = re.sub(res, '\\2', line_math)
            line_math = re.sub(res, '\\1', line_math)
        if line_math == '':
            return
        math = line_math
        res = '^(.*)days\\(([MTSHR]?[0-9]+-[0-9]+-[0-9]+)\\)(.*)$'
        while re.match(res, math):
            pre = re.sub(res, '\\1', math)
            dat = re.sub(res, '\\2', math)
            pos = re.sub(res, '\\3', math)
            math = pre + str(count_days(dat)) + pos
        math = math.replace('\t', ' ').replace('\u3000', ' ')
        math = math.replace('，', ',').replace('．', '.')
        math = math.replace('０', '0').replace('１', '1').replace('２', '2')
        math = math.replace('３', '3').replace('４', '4').replace('５', '5')
        math = math.replace('６', '6').replace('７', '7').replace('８', '8')
        math = math.replace('９', '9')
        math = math.replace('〇', '0').replace('一', '1').replace('二', '2')
        math = math.replace('三', '3').replace('四', '4').replace('五', '5')
        math = math.replace('六', '6').replace('七', '7').replace('八', '8')
        math = math.replace('九', '9')
        math = math.replace('（', '(').replace('）', ')')
        math = math.replace('｛', '{').replace('｝', '}')
        math = math.replace('［', '[').replace('］', ']')
        math = math.replace('｜', '|').replace('！', '!').replace('＾', '^')
        math = math.replace('＊', '*').replace('／', '/').replace('％', '%')
        math = math.replace('＋', '+').replace('−', '-')
        math = math.replace('×', '*').replace('÷', '/').replace('ー', '-')
        math = math.replace('△', '-').replace('▲', '-')
        math = math.replace('パ-セント', '%')
        # ' ', ','
        math = math.replace(' ', '').replace(',', '')
        # {, }, [, ]
        math = math.replace('{', '(').replace('}', ')')
        math = math.replace('[', '(').replace(']', ')')
        # 千 -> 1千
        unit = ['千', '百', '十']
        new_math = ''
        res1 = '^(.*?)([京兆億万千百十0-9]+)(.*)$'
        while re.match(res1, math):
            head = re.sub(res1, '\\1', math)
            numb = re.sub(res1, '\\2', math)
            math = re.sub(res1, '\\3', math)
            tmp = ''
            for i, u in enumerate(unit):
                res2 = '^([^' + u + ']*' + u + ')(.*)$'
                if re.match(res2, numb):
                    t1 = re.sub(res2, '\\1', numb)  # [^千]*千
                    t2 = re.sub(res2, '\\2', numb)  # .*
                    if not re.match('^.*[0-9]' + u + '$', t1):
                        t1 = re.sub(u + '$', '1' + u, t1)  # 千 -> 1千
                    tmp += t1
                    numb = t2
            new_math += head + tmp + numb
        math = new_math + math
        # 1千 -> 1千0百0十0
        unit, vnit = ['千', '百', '十'], ['百', '十', '']
        new_math = ''
        res1 = '^(.*?)([京兆億万千百十0-9]+)(.*)$'
        while re.match(res1, math):
            head = re.sub(res1, '\\1', math)
            numb = re.sub(res1, '\\2', math)
            math = re.sub(res1, '\\3', math)
            tmp = ''
            for i, u in enumerate(unit):
                v = vnit[i]
                res2 = '^([^' + u + ']*' + u + ')(.*)$'
                if re.match(res2, numb):
                    t1 = re.sub(res2, '\\1', numb)  # [^千]*千
                    t2 = re.sub(res2, '\\2', numb)  # .*
                    tmp += t1
                    if not re.match('^[0-9]' + v, t2):
                        t2 = '0' + v + t2
                    numb = t2
            new_math += head + tmp + numb
        math = new_math + math
        math = math.replace('千', '').replace('百', '').replace('十', '')
        # 1兆2万 -> 1兆0000億0002万0000
        new_math = ''
        unit, vnit = ['京', '兆', '億', '万'], ['兆', '億', '万', '']
        res1 = '^(.*?)([京兆億万千百十0-9]+)(.*)$'
        while re.match(res1, math):
            head = re.sub(res1, '\\1', math)
            numb = re.sub(res1, '\\2', math)
            math = re.sub(res1, '\\3', math)
            tmp = ''
            for i, u in enumerate(unit):
                v = vnit[i]
                res2 = '^([^' + u + ']*' + u + ')(.*)$'
                if re.match(res2, numb):
                    t1 = re.sub(res2, '\\1', numb)  # [^京]*京
                    t2 = re.sub(res2, '\\2', numb)  # .*
                    tmp += t1
                    if re.match('[0-9]{,4}' + v, t2):
                        t2 = '0000' + t2
                        numb = re.sub('^[0-9]*([0-9]{4})', '\\1', t2)  # 0012兆
                    else:
                        numb = '0000' + v + t2  # 0000兆
            new_math += head + tmp + numb
        math = new_math + math
        math = math.replace('京', '').replace('兆', '')
        math = math.replace('億', '').replace('万', '')
        # %, 割, 分, 厘
        math = re.sub('([0-9\\.]+)%', '(\\1/100)', math)
        math = re.sub('([0-9\\.]+)割', '(\\1/10)', math)
        math = re.sub('([0-9\\.]+)分', '(\\1/100)', math)
        math = re.sub('([0-9\\.]+)厘', '(\\1/1000)', math)
        # FRACTION
        res = '^(.*?)' \
            + '([0-9]+|\\([^\\(\\)]+\\))分の([0-9]+|\\([^\\(\\)]+\\))' \
            + '(.*?)$'
        while re.match(res, math):
            math = re.sub(res, '\\1(\\3/\\2)\\4', math)
        # POWER
        math = re.sub('\\^', '**', math)
        # REMOVE
        math = re.sub('pi', '3.141592653589793', math)
        math = re.sub('e', '2.718281828459045', math)
        math = re.sub('[^\\(\\)\\|\\*/%\\-\\+0-9\\.]', '', math)
        # EVAL
        try:
            r = str(round(eval(math), 10))
        except BaseException:
            if must_show_message:
                n, m = 'エラー', '計算できませんでした．'
                tkinter.messagebox.showerror(n, m)
            return False
        r = re.sub('\\.0$', '', r)
        # REPLACE
        digit_separator = self.digit_separator.get()
        if '.' in r:
            i = re.sub('^(.*)(\\..*)$', '\\1', r)
            f = re.sub('^(.*)(\\..*)$', '\\2', r)
        else:
            i = r
            f = ''
        if digit_separator == '3':
            if re.match('^.*[0-9]{19}$', i):
                i = re.sub('([0-9]{18})$', ',\\1', i)
            if re.match('^.*[0-9]{16}$', i):
                i = re.sub('([0-9]{15})$', ',\\1', i)
            if re.match('^.*[0-9]{13}$', i):
                i = re.sub('([0-9]{12})$', ',\\1', i)
            if re.match('^.*[0-9]{10}$', i):
                i = re.sub('([0-9]{9})$', ',\\1', i)
            if re.match('^.*[0-9]{7}$', i):
                i = re.sub('([0-9]{6})$', ',\\1', i)
            if re.match('^.*[0-9]{4}$', i):
                i = re.sub('([0-9]{3})$', ',\\1', i)
        elif digit_separator == '4':
            if re.match('^.*[0-9]{17}$', i):
                i = re.sub('([0-9]{16})$', '京\\1', i)
            if re.match('^.*[0-9]{13}$', i):
                i = re.sub('([0-9]{12})$', '兆\\1', i)
            if re.match('^.*[0-9]{9}$', i):
                i = re.sub('([0-9]{8})$', '億\\1', i)
            if re.match('^.*[0-9]{5}$', i):
                i = re.sub('([0-9]{4})$', '万\\1', i)
        r = i + f
        v_number = self._get_v_position_of_insert(self.txt)
        beg = str(v_number) + '.' + str(len(line_head + line_math))
        end = str(v_number) + '.' + str(len(line_head + line_math + line_rslt))
        if r != line_rslt:
            self.txt['autoseparators'] = False
            self.txt.edit_separator()
            self.txt.delete(beg, end)
            self.txt.insert(beg, '=' + r)
            self.paint_out_line(v_number - 1)
            self.txt['autoseparators'] = True
            self.txt.edit_separator()
        self.win.clipboard_clear()
        self.win.clipboard_append(r)
        if self.clipboard_list[-1] != '':
            self.clipboard_list.append('')
        self.clipboard_list[-1] += r
        return True

    def change_typeface(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        c = self.txt.get('insert', 'insert+1c')
        for tf in TYPEFACES:
            if c in tf:
                t = '字体を変える'
                m = '字体を選んでください．'
                rd = RadiobuttonDialog(mother, self, t, m, tf, tf.index(c))
                v = rd.get_value()
                if v is not None:
                    pane.edit_separator()
                    pane.delete('insert', 'insert+1c')
                    pane.insert('insert', v)
                    pane.mark_set('insert', 'insert-1c')
                    p = self._get_v_position_of_insert(pane) - 1
                    self.paint_out_line(p)
                    pane.edit_separator()
                return True
        else:
            n = '警告'
            m = '"' + c + '"に異字体は登録されていません．'
            tkinter.messagebox.showwarning(n, m)
            return False

    def comment_out_region(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return
        if pane.tag_ranges('sel'):
            beg, end = pane.index('sel.first'), pane.index('sel.last')
        elif 'akauni' in pane.mark_names():
            beg, end = self._get_indices_in_order(pane, 'insert', 'akauni')
        else:
            n = 'エラー'
            m = 'コメントアウトする範囲が選択されていません．'
            tkinter.messagebox.showerror(n, m)
            return
        pane['autoseparators'] = False
        pane.edit_separator()
        tex = pane.get(beg, end)
        for i in ['8', '7', '6', '5', '4', '3', '2', '1', '-']:
            if i == '-':
                j = '1'
            else:
                j = str(int(i) + 1)
            for t in (('<!' + i + '-', '<!' + j + '-'),
                      ('-' + i + '>', '-' + j + '>')):
                res = '^((?:.|\n)*?)' + t[0] + '((?:.|\n)*)$'
                while re.match(res, tex):
                    sub = re.sub(res, '\\1', tex)
                    tex = re.sub(res, '\\1' + t[1] + '\\2', tex)
                    pane.delete(beg + '+' + str(len(sub)) + 'c',
                                beg + '+' + str(len(sub + t[0])) + 'c')
                    pane.insert(beg + '+' + str(len(sub)) + 'c', t[1])
        pane.insert(end, '-->')
        pane.insert(beg, '<!--')
        self.cancel_region(pane)
        beg_v = int(re.sub('\\.[0-9]+$', '', beg))
        end_v = int(re.sub('\\.[0-9]+$', '', end))
        for i in range(beg_v - 1, end_v):
            self.paint_out_line(i)
        pane['autoseparators'] = True
        pane.edit_separator()
        self.update_toc()

    def uncomment_in_region(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return
        #
        if pane.tag_ranges('sel'):
            beg, end = pane.index('sel.first'), pane.index('sel.last')
        elif 'akauni' in pane.mark_names():
            beg, end = self._get_indices_in_order(pane, 'insert', 'akauni')
        else:
            n = 'エラー'
            m = 'コメントアウトを解除する範囲が選択されていません．'
            tkinter.messagebox.showerror(n, m)
            return
        pane['autoseparators'] = False
        pane.edit_separator()
        tex = pane.get(beg, end)
        is_in_comment = False
        tmp = ''
        for c in tex:
            tmp += c
            if re.match('^((?:.|\n)*)<!--$', tmp) and not is_in_comment:
                tmp = re.sub('<!--$', '', tmp)
                pane.delete(beg + '+' + str(len(tmp)) + 'c',
                            beg + '+' + str(len(tmp) + 4) + 'c')
                is_in_comment = True
            if re.match('^((?:.|\n)*)-->$', tmp) and is_in_comment:
                tmp = re.sub('-->$', '', tmp)
                pane.delete(beg + '+' + str(len(tmp)) + 'c',
                            beg + '+' + str(len(tmp) + 3) + 'c')
                is_in_comment = False
        tex = tmp
        for i in ['-', '1', '2', '3', '4', '5', '6', '7', '8']:
            if i == '-':
                j = '1'
            else:
                j = str(int(i) + 1)
            for t in (('<!' + i + '-', '<!' + j + '-'),
                      ('-' + i + '>', '-' + j + '>')):
                res = '^((?:.|\n)*?)' + t[1] + '((?:.|\n)*)$'
                while re.match(res, tex):
                    sub = re.sub(res, '\\1', tex)
                    tex = re.sub(res, '\\1' + t[0] + '\\2', tex)
                    pane.delete(beg + '+' + str(len(sub)) + 'c',
                                beg + '+' + str(len(sub + t[1])) + 'c')
                    pane.insert(beg + '+' + str(len(sub)) + 'c', t[0])
        self.cancel_region(pane)
        beg_v = int(re.sub('\\.[0-9]+$', '', beg))
        end_v = int(re.sub('\\.[0-9]+$', '', end))
        for i in range(beg_v - 1, end_v):
            self.paint_out_line(i)
        pane['autoseparators'] = True
        pane.edit_separator()
        self.update_toc()

    ##########################
    # MENU INSERT

    def _make_menu_insert(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='挿入(I)', menu=menu, underline=3)
        #
        self._make_submenu_insert_sample(menu)
        menu.add_separator()
        #
        menu.add_command(label='コメントを挿入',
                         command=self.insert_comment)
        menu.add_command(label='空白を挿入',
                         command=self.insert_space)
        menu.add_command(label='改行を挿入',
                         command=self.insert_line_break)
        menu.add_command(label='画像を挿入',
                         command=self.insert_images)
        self._make_submenu_insert_font_change(menu)
        self._make_submenu_insert_font_size_change(menu)
        self._make_submenu_insert_font_width_change(menu)
        self._make_submenu_insert_underline(menu)
        self._make_submenu_insert_font_color_change(menu)
        self._make_submenu_insert_highlight_color_change(menu)
        self._make_submenu_insert_math_expression(menu)
        menu.add_command(label='代語句を挿入',
                         command=self.insert_substitute_phrases)
        menu.add_separator()
        #
        menu.add_command(label='別のファイルの内容を挿入',
                         command=self.insert_file)
        menu.add_separator()
        #
        self._make_submenu_insert_script(menu)
        menu.add_separator()
        #
        self._make_submenu_insert_file_name(menu)
        self._make_submenu_insert_time(menu)
        menu.add_separator()
        #
        menu.add_command(label='コード番号から文字を挿入',
                         command=self.insert_character_by_code)
        self._make_submenu_insert_ivs_character(menu)
        self._make_submenu_insert_horizontal_line(menu)
        menu.add_command(label='記号を挿入',
                         command=self.insert_symbol)
        # menu.add_separator()

    ################
    # SUBMENU INSERT SAMPLE

    def _make_submenu_insert_sample(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='文書のサンプルを挿入', menu=submenu)
        #
        submenu.add_command(label='基本',
                            command=self.insert_basis_sample)
        submenu.add_command(label='民法',
                            command=self.insert_law_sample)
        submenu.add_command(label='和解契約書',
                            command=self.insert_settlement_sample)
        submenu.add_command(label='訴状',
                            command=self.insert_petition_sample)
        submenu.add_command(label='証拠説明書',
                            command=self.insert_evidence_sample)
        submenu.add_command(label='判決（民事事件）',
                            command=self.insert_civil_judgement_sample)
        submenu.add_command(label='起訴状',
                            command=self.insert_indictment_sample)
        submenu.add_command(label='判決（刑事事件）',
                            command=self.insert_criminal_judgement_sample)

    ######
    # COMMAND

    def insert_basis_sample(self):               # 基本
        document = self.insert_configuration_sample('普通', '0.0') + \
            SAMPLE_BASIS
        self.insert_sample(document)

    def insert_law_sample(self):                 # 民法
        document = self.insert_configuration_sample('条文', '0.0') + \
            SAMPLE_LAW
        self.insert_sample(document)

    def insert_settlement_sample(self):          # 和解契約書
        document = self.insert_configuration_sample('契約', '1.0') + \
            SAMPLE_SETTLEMENT
        self.insert_sample(document)

    def insert_petition_sample(self):            # 訴状
        document = self.insert_configuration_sample('普通', '1.0') + \
            SAMPLE_PETITION
        self.insert_sample(document)

    def insert_evidence_sample(self):            # 証拠説明書
        document = self.insert_configuration_sample('普通', '0.0') + \
            SAMPLE_EVIDENCE
        self.insert_sample(document)

    def insert_civil_judgement_sample(self):     # 判決（民事事件）
        document = self.insert_configuration_sample('普通', '0.0') + \
            SAMPLE_CIVIL_JUDGEMENT
        self.insert_sample(document)

    def insert_indictment_sample(self):          # 起訴状
        document = self.insert_configuration_sample('普通', '0.0') + \
            SAMPLE_INDICTMENT
        self.insert_sample(document)

    def insert_criminal_judgement_sample(self):  # 判決（刑事事件）
        document = self.insert_configuration_sample('普通', '0.0') + \
            SAMPLE_CRIMINAL_JUDGEMENT
        self.insert_sample(document)

    def insert_configuration_sample(self, document_style, space_before):
        document = '''\
<!--------------------------【設定】-----------------------------

# プロパティに表示される文書のタイトルを指定できます。
書題名:

# 3つの書式（普通、契約、条文）を指定できます。
文書式: ''' + document_style + '''

# 用紙のサイズ（A3横、A3縦、A4横、A4縦）を指定できます。
用紙サ: A4縦

# 用紙の上下左右の余白をセンチメートル単位で指定できます。
上余白: 3.5 cm
下余白: 2.2 cm
左余白: 3.0 cm
右余白: 2.0 cm

# ページのヘッダーに表示する文字列（別紙 :等）を指定できます。
頭書き:

# ページ番号の書式（無、有、n :、-n-、n/N等）を指定できます。
頁番号: 有

# 行番号の記載（無、有）を指定できます。
行番号: 無

# 明朝体とゴシック体と異字体（IVS）のフォントを指定できます。
明朝体: Times New Roman / ＭＳ 明朝
ゴシ体: = / ＭＳ ゴシック
異字体: IPAmj明朝

# 基本の文字の大きさをポイント単位で指定できます。
文字サ: 12 pt

# 行間隔を基本の文字の高さの何倍にするかを指定できます。
行間隔: 2.14 倍

# セクションタイトル前後の余白を行間隔の倍数で指定できます。
前余白: 0.0 倍, ''' + space_before + ''' 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍
後余白: 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍

# 半角文字と全角文字の間の間隔調整（無、有）を指定できます。
字間整: 無

# 備考書（コメント）などを消して完成させます。
完成稿: 偽

# 原稿の作成日時と更新日時が自動で記録されます。
作成時: - USER
更新時: - USER

---------------------------------------------------------------->
'''
        return document

    def insert_sample(self, sample_document):
        txt_text = self.txt.get('1.0', 'end-1c')
        if txt_text != '':
            n, m = 'エラー', 'テキストが空ではありません．'
            tkinter.messagebox.showerror(n, m)
            return
        self.file_lines = sample_document.split('\n')
        self.txt.insert('1.0', sample_document)
        self.txt.focus_set()
        self.current_pane = 'txt'
        self.txt.mark_set('insert', '1.0')
        # PAINT
        self.paint_all_lines(self.txt)
        # CLEAR THE UNDO STACK
        self.txt.edit_reset()

    ################
    # COMMAND

    def insert_comment(self):
        self.txt.insert('insert', '<!--（ここにコメントを書く）-->')

    def insert_space(self):
        b = '空白の幅'
        p = '空白の幅を文字数（整数又は小数）で入力してください．'
        h, t = '', '文字'
        f = ''
        while not re.match('^([0-9]*\\.)?[0-9]+$', f):
            f = OneWordDialog(self.txt, self, b, p, h, t, f).get_value()
            if f is None:
                return
        self.txt.insert('insert', '< ' + f + ' >')

    def insert_line_break(self):
        self.txt.insert('insert', '<br>')

    def insert_images(self):
        ti = '画像を挿入'
        ty = [('画像', '.jpg .jpeg .png .gif .tif .tiff .bmp'),
              ('全てのファイル', '*')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        image_paths = tkinter.filedialog.askopenfilenames(
                title=ti, filetypes=ty, initialdir=_d)
        for i in image_paths:
            image_md_text = '![代替テキスト:横x縦](' + i + ' "説明")'
            self.txt.insert('insert', image_md_text)

    ################
    # SUBMENU INSERT FONT CHANGE

    def _make_submenu_insert_font_change(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='フォントの変更を挿入', menu=submenu)
        #
        # self.mincho.set(MS_INCHO_FONT)
        submenu.add_command(label='明朝体を変える',
                            command=self.insert_selected_mincho_font)
        submenu.add_command(label='欧文フォントを変える',
                            command=self.insert_selected_alphanumeric_font)
        submenu.add_separator()
        submenu.add_command(label='ゴシック体に変える',
                            command=self.insert_gothic_font)
        submenu.add_separator()
        submenu.add_command(label='手動入力',
                            command=self.insert_font_manually)

    ######
    # COMMAND

    def insert_selected_mincho_font(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        mincho_font_list = []
        for f in tkinter.font.families():
            if f not in mincho_font_list:
                if not re.match('^@', f):  # ROTATED FONT
                    if ('明朝' in f) or (f == 'Noto Serif CJK JP'):
                        mincho_font_list.append(f)
                    if ('ゴシック' in f) or (f == 'Noto Sans CJK JP'):
                        mincho_font_list.append(f)
        mincho_font_list.sort()
        t = '明朝フォントを変える'
        m = 'フォントを選んでください．'
        n = -1
        if DOCX_MINCHO_FONT in mincho_font_list:
            n = mincho_font_list.index(DOCX_MINCHO_FONT)
        rd = RadiobuttonDialog(mother, self, t, m, mincho_font_list, n)
        v = rd.get_value()
        if v is not None:
            pane.edit_separator()
            d = '@' + v + '@（ここはフォントが変わる）@' + v + '@'
            pane.insert('insert', d)
            pane.mark_set('insert', 'insert-' + str(len(v) + 2) + 'c')
            p = self._get_v_position_of_insert(pane) - 1
            self.paint_out_line(p)
            pane.edit_separator()

    def insert_selected_mincho_font(self, mother=None):
        mincho_font_list = []
        for f in tkinter.font.families():
            if f not in mincho_font_list:
                if not re.match('^@', f):  # ROTATED FONT
                    if ('明朝' in f) or (f == 'Noto Serif CJK JP'):
                        mincho_font_list.append(f)
        mincho_font_list.sort()
        self._insert_selected_x_font(mother,
                                     '明朝フォントを変える',
                                     'フォントを選んでください．',
                                     mincho_font_list,
                                     DOCX_MINCHO_FONT)

    def insert_selected_alphanumeric_font(self, mother=None):
        alphanumeric_font_list_candidates = [
            'Times New Roman',
            'Cambria',
            'Century',
            'Contantia',
            'Garamond',
            'Georgia',
            'Platino',
        ]
        alphanumeric_font_list = []
        for fc in alphanumeric_font_list_candidates:
            for f in tkinter.font.families():
                if re.match('^' + fc, f) and (f not in alphanumeric_font_list):
                    alphanumeric_font_list.append(f)
        alphanumeric_font_list.sort()
        self._insert_selected_x_font(mother,
                                     '欧文フォントを変える',
                                     'フォントを選んでください．',
                                     alphanumeric_font_list,
                                     DOCX_ALPHANUMERIC_FONT)

    def _insert_selected_x_font(self, mother, title, prompt,
                                candidates, default):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        candidates.sort()
        n = -1
        if default in candidates:
            n = candidates.index(default)
        rd = RadiobuttonDialog(mother, self, title, prompt, candidates, n)
        v = rd.get_value()
        if v is not None:
            pane.edit_separator()
            d = '@' + v + '@（ここはフォントが変わる）@' + v + '@'
            pane.insert('insert', d)
            pane.mark_set('insert', 'insert-' + str(len(v) + 2) + 'c')
            p = self._get_v_position_of_insert(pane) - 1
            self.paint_out_line(p)
            pane.edit_separator()

    def insert_gothic_font(self):
        self.txt.insert('insert', '`（ここはゴシック体）`')
        self.txt.mark_set('insert', 'insert-1c')

    def insert_font_manually(self):
        b = 'フォント'
        p = 'フォント名を入力してください．'
        h, t = '', ''
        s = OneWordDialog(self.txt, self, b, p, h, t).get_value()
        if s is None:
            return
        d = '@' + s + '@（ここはフォントが変わる）@' + s + '@'
        self.txt.insert('insert', d)
        self.txt.mark_set('insert', 'insert-' + str(len(s) + 2) + 'c')

    ################
    # SUBMENU INSERT FONT SIZE CHANGE

    def _make_submenu_insert_font_size_change(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='文字の大きさの変更を挿入', menu=submenu)
        #
        submenu.add_command(label='特小サイズ',
                            command=self.insert_ss_font_size)
        submenu.add_command(label='小サイズ',
                            command=self.insert_s_font_size)
        submenu.add_command(label='大サイズ',
                            command=self.insert_l_font_size)
        submenu.add_command(label='特大サイズ',
                            command=self.insert_ll_font_size)
        submenu.add_separator()
        submenu.add_command(label='手動入力',
                            command=self.insert_font_size_manually)

    ######
    # COMMAND

    def insert_ss_font_size(self):
        self.txt.insert('insert', '---（ここは文字が特に小さい）---')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_s_font_size(self):
        self.txt.insert('insert', '--（ここは文字が小さい）--')
        self.txt.mark_set('insert', 'insert-2c')

    def insert_l_font_size(self):
        self.txt.insert('insert', '++（ここは文字が大きい）++')
        self.txt.mark_set('insert', 'insert-2c')

    def insert_ll_font_size(self):
        self.txt.insert('insert', '+++（ここは文字が特に大きい）+++')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_font_size_manually(self):
        b = '文字の大きさ'
        p = '文字の大きさを1から100までの数字を入力してください．'
        h, t = '', 'px'
        f = ''
        while not re.match('^([0-9]*\\.)?[0-9]+$', f):
            f = OneWordDialog(self.txt, self, b, p, h, t, f).get_value()
            if f is None:
                return
        f = re.sub('\\.0+$', '', f)
        d = '@' + f + '@（ここは文字の大きさが変わる）@' + f + '@'
        self.txt.insert('insert', d)
        self.txt.mark_set('insert', 'insert-' + str(len(f) + 2) + 'c')

    ################
    # SUBMENU INSERT FONT WIDTH CHANGE

    def _make_submenu_insert_font_width_change(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='文字の幅の変更を挿入', menu=submenu)
        #
        submenu.add_command(label='特細サイズ',
                            command=self.insert_ss_font_width)
        submenu.add_command(label='細サイズ',
                            command=self.insert_s_font_width)
        submenu.add_command(label='太サイズ',
                            command=self.insert_l_font_width)
        submenu.add_command(label='特太サイズ',
                            command=self.insert_ll_font_width)

    ######
    # COMMAND

    def insert_ss_font_width(self):
        self.txt.insert('insert', '>>>（ここは文字が特に細い）<<<')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_s_font_width(self):
        self.txt.insert('insert', '>>（ここは文字が細い）<<')
        self.txt.mark_set('insert', 'insert-2c')

    def insert_l_font_width(self):
        self.txt.insert('insert', '<<（ここは文字が太い）>>')
        self.txt.mark_set('insert', 'insert-2c')

    def insert_ll_font_width(self):
        self.txt.insert('insert', '<<<（ここは文字が特に太い）>>>')
        self.txt.mark_set('insert', 'insert-3c')

    ################
    # SUBMENU INSERT UNDERLINE

    def _make_submenu_insert_underline(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='文字に下線をを引く', menu=submenu)
        #
        submenu.add_command(label='単線',
                            command=self.insert_single_underline)
        submenu.add_command(label='二重線',
                            command=self.insert_double_underline)
        submenu.add_command(label='波線',
                            command=self.insert_wave_underline)
        submenu.add_command(label='破線',
                            command=self.insert_dash_underline)
        submenu.add_command(label='点線',
                            command=self.insert_dot_underline)

    ######
    # COMMAND

    def insert_single_underline(self):
        self.txt.insert('insert', '__（ここは下線が引かれる）__')
        self.txt.mark_set('insert', 'insert-2c')

    def insert_double_underline(self):
        self.txt.insert('insert', '_=_（ここは下線が引かれる）_=_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_wave_underline(self):
        self.txt.insert('insert', '_~_（ここは下線が引かれる）_~_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_dash_underline(self):
        self.txt.insert('insert', '_-_（ここは下線が引かれる）_-_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_dot_underline(self):
        self.txt.insert('insert', '_._（ここは下線が引かれる）_._')
        self.txt.mark_set('insert', 'insert-3c')

    ################
    # SUBMENU INSERT FONT COLOR CHANGE

    def _make_submenu_insert_font_color_change(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='文字色を変える', menu=submenu)
        #
        submenu.add_command(label='赤色',
                            command=self.insert_r_font_color)
        submenu.add_command(label='黄色',
                            command=self.insert_y_font_color)
        submenu.add_command(label='緑色',
                            command=self.insert_g_font_color)
        submenu.add_command(label='シアン',
                            command=self.insert_c_font_color)
        submenu.add_command(label='青色',
                            command=self.insert_b_font_color)
        submenu.add_command(label='マゼンタ',
                            command=self.insert_m_font_color)
        submenu.add_command(label='白色',
                            command=self.insert_w_font_color)

    ######
    # COMMAND

    def insert_r_font_color(self):
        self.txt.insert('insert', '^R^（ここは文字が赤色）^R^')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_y_font_color(self):
        self.txt.insert('insert', '^Y^（ここは文字が黄色）^Y^')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_g_font_color(self):
        self.txt.insert('insert', '^G^（ここは文字が緑色）^G^')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_c_font_color(self):
        self.txt.insert('insert', '^C^（ここは文字がシアン）^C^')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_b_font_color(self):
        self.txt.insert('insert', '^B^（ここは文字が青色）^B^')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_m_font_color(self):
        self.txt.insert('insert', '^M^（ここは文字がマゼンタ）^M^')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_w_font_color(self):
        self.txt.insert('insert', '^^（ここは文字が白色）^^')
        self.txt.mark_set('insert', 'insert-2c')

    ################
    # SUBMENU INSERT HIGHLIGHT COLOR CHANGE

    def _make_submenu_insert_highlight_color_change(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='下地色を変える', menu=submenu)
        #
        submenu.add_command(label='赤色',
                            command=self.insert_r_highlight_color)
        submenu.add_command(label='黄色',
                            command=self.insert_y_highlight_color)
        submenu.add_command(label='緑色',
                            command=self.insert_g_highlight_color)
        submenu.add_command(label='シアン',
                            command=self.insert_c_highlight_color)
        submenu.add_command(label='青色',
                            command=self.insert_b_highlight_color)
        submenu.add_command(label='マゼンタ',
                            command=self.insert_m_highlight_color)

    ######
    # COMMAND

    def insert_r_highlight_color(self):
        self.txt.insert('insert', '_R_（ここは下地が赤色）_R_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_y_highlight_color(self):
        self.txt.insert('insert', '_Y_（ここは下地が黄色）_Y_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_g_highlight_color(self):
        self.txt.insert('insert', '_G_（ここは下地が緑色）_G_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_c_highlight_color(self):
        self.txt.insert('insert', '_C_（ここは下地がシアン）_C_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_b_highlight_color(self):
        self.txt.insert('insert', '_B_（ここは下地が青色）_B_')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_m_highlight_color(self):
        self.txt.insert('insert', '_M_（ここは下地がマゼンタ）_M_')
        self.txt.mark_set('insert', 'insert-3c')

    ################
    # SUBMENU INSERT MATH EXPRESSINO

    def _make_submenu_insert_math_expression(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='数式を挿入', menu=submenu)
        #
        submenu.add_command(label='基本を挿入',
                            command=self.insert_math_expression_basis)
        submenu.add_command(label='分数を挿入',
                            command=self.insert_math_fraction)
        submenu.add_command(label='上付き文字を挿入',
                            command=self.insert_math_superscript)
        submenu.add_command(label='下付き文字を挿入',
                            command=self.insert_math_subscript)
        submenu.add_command(label='平方根を挿入',
                            command=self.insert_math_square_root)
        submenu.add_command(label='累乗根を挿入',
                            command=self.insert_math_power_root)
        submenu.add_command(label='ベクトルを挿入',
                            command=self.insert_math_vector)
        submenu.add_command(label='総和を挿入',
                            command=self.insert_math_summation)
        submenu.add_command(label='総乗を挿入',
                            command=self.insert_math_product)
        submenu.add_command(label='微分（分数）を挿入',
                            command=self.insert_math_differentiation_frac)
        submenu.add_command(label='微分（ドット）を挿入',
                            command=self.insert_math_differentiation_dot)
        submenu.add_command(label='偏微分を挿入',
                            command=self.insert_math_partial_differentiation)
        submenu.add_command(label='積分を挿入',
                            command=self.insert_math_integral)

    ######
    # COMMAND

    def insert_math_expression_basis(self):
        self._insert_math_expression('（ここに"LaTeX"形式の数式を挿入）')

    def insert_math_fraction(self):
        self._insert_math_expression('\\frac{A}{B}')

    def insert_math_superscript(self):
        self._insert_math_expression('{A}^{B}')

    def insert_math_subscript(self):
        self._insert_math_expression('{A}_{B}')

    def insert_math_square_root(self):
        self._insert_math_expression('\\sqrt{A}')

    def insert_math_power_root(self):
        self._insert_math_expression('\\sqrt[A]{B}')

    def insert_math_vector(self):
        self._insert_math_expression('\\vec{A}')

    def insert_math_summation(self):
        self._insert_math_expression('\\sum_{n=A}^{B}{C_n}')

    def insert_math_product(self):
        self._insert_math_expression('\\prod_{n=A}^{B}{C_n}')

    def insert_math_differentiation_frac(self):
        self._insert_math_expression('\\frac{\\mathrm{d}A}{\\mathrm{d}B}')

    def insert_math_differentiation_dot(self):
        self._insert_math_expression('\\dot{A}')

    def insert_math_partial_differentiation(self):
        self._insert_math_expression('\\frac{{\\partial}A}{{\\partial}B}')

    def insert_math_integral(self):
        self._insert_math_expression('\\int_{A}^{B}{f(x)}dx')

    def _insert_math_expression(self, inline_text):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        #
        doc_up = pane.get('1.0', 'insert')
        while re.match(NOT_ESCAPED + '\\\\\\]', doc_up):
            doc_up = re.sub(NOT_ESCAPED + '\\\\\\]', '', doc_up)
        if not re.match(NOT_ESCAPED + '\\\\\\[', doc_up):
            inline_text = '\\[' + inline_text
        #
        doc_dn = pane.get('insert', 'end-1c')
        doc_dn = re.sub(NOT_ESCAPED + '\\\\\\[(.|\n)*$', '\\1', doc_dn)
        if not re.match(NOT_ESCAPED + '\\\\\\]', doc_dn):
            inline_text = inline_text + '\\]'
        #
        self._insert_inline_text(inline_text, -2)

    ################
    # COMMAND

    def insert_substitute_phrases(self):
        self.txt.insert('insert',
                        '\n%[（代語句名）]% = "（内容）"\n%[（代語句名）]%\n')

    ################
    # COMMAND

    def insert_file(self):
        ti = 'ファイルの内容を挿入'
        ty = [('読み込み可能なファイル', '.docx .md .txt .xlsx .csv')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        file_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d)
        if file_path == () or file_path == '':
            return
        if re.match('^(?:.|\n)+.xlsx$', file_path):
            document = self._read_xlsx_file(file_path)
        elif re.match('^(?:.|\n)+.csv$', file_path):
            document = self._read_csv_file(file_path)
        elif re.match('^(?:.|\n)+.docx$', file_path):
            document = self._read_docx_file(file_path)
            if re.match('^<!--', document):
                document = re.sub('^<!--(.|\n)*?-->\n*', '', document)
        elif re.match('^(?:.|\n)+.md$', file_path):
            document = self._read_md_file(file_path)
        else:
            document = self._read_txt_file(file_path)
        if document is None:
            return
        self.txt.insert('insert', document)

    ################
    # SUBMENU INSERT SCRIPT

    def _make_submenu_insert_script(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='スクリプトを挿入', menu=submenu)
        #
        submenu.add_command(label='1回目に実行するスクリプトを挿入',
                            command=self.insert_script_to_exec_1st_time)
        submenu.add_command(label='2回目に実行するスクリプトを挿入',
                            command=self.insert_script_to_exec_2nd_time)
        submenu.add_command(label='3回目に実行するスクリプトを挿入',
                            command=self.insert_script_to_exec_3rd_time)

    ######
    # COMMAND

    def insert_script_to_exec_1st_time(self):
        msg = '（ここにスクリプトを挿入（サンプルはTabを押す））'
        self.txt.insert('insert', '{{' + msg + '}}')
        self.txt.mark_set('insert', 'insert-2c')

    def insert_script_to_exec_2nd_time(self):
        msg = '（ここにスクリプトを挿入（サンプルはTabを押す））'
        self.txt.insert('insert', '{2{' + msg + '}2}')
        self.txt.mark_set('insert', 'insert-3c')

    def insert_script_to_exec_3rd_time(self):
        msg = '（ここにスクリプトを挿入（サンプルはTabを押す））'
        self.txt.insert('insert', '{3{' + msg + '}3}')
        self.txt.mark_set('insert', 'insert-3c')

    ################
    # SUBMENU INSERT FILE

    def _make_submenu_insert_file_name(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='ファイル名を挿入', menu=submenu)
        #
        submenu.add_command(label='ファイル名をフルパスで挿入',
                            command=self.insert_file_paths)
        submenu.add_command(label='ファイル名をファイル名のみで挿入',
                            command=self.insert_file_names)
        submenu.add_command(label='編集中のファイルと同じフォルダにある全ファイルのファイル名のみを挿入',
                            command=self.insert_file_names_in_same_folder)

    ######
    # COMMAND

    def insert_file_paths(self):
        ti = 'ファイル名をフルパスで挿入'
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        file_paths \
            = tkinter.filedialog.askopenfilenames(title=ti, initialdir=_d)
        for f in file_paths:
            self.txt.insert('insert', f + '\n')

    def insert_file_names(self):
        ti = 'ファイル名をファイル名のみで挿入'
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        file_paths \
            = tkinter.filedialog.askopenfilenames(title=ti, initialdir=_d)
        for f in file_paths:
            f = re.sub('^(.|\n)*/', '', f)
            self.txt.insert('insert', f + '\n')

    def insert_file_names_in_same_folder(self):
        file_path = self.file_path
        if file_path is None:
            return
        elif re.match('^(.*)[/\\\\](.*)$', file_path):
            dir_path = re.sub('^(.*)[/\\\\](.*)$', '\\1', file_path)
        else:
            dir_path = os.getcwd()
        files = os.listdir(dir_path)
        for f in sorted(files):
            if not re.match('^\\.', f) and os.path.isfile(f):
                if not re.match('^~\\$.*\\.zip$', f):
                    self.txt.insert('insert', f + '\n')

    ################
    # SUBMENU INSERT TIME

    def _make_submenu_insert_time(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='日時を挿入', menu=submenu)
        #
        submenu.add_command(label='YY年M月D日',
                            command=self.insert_date_YYMD)
        submenu.add_command(label='令和Y年M月D日',
                            command=self.insert_date_GYMD)
        submenu.add_command(label='yy年m月d日',
                            command=self.insert_date_yymd)
        submenu.add_command(label='令和y年m月d日',
                            command=self.insert_date_Gymd)
        submenu.add_command(label='yyyy-mm-dd',
                            command=self.insert_date_iso)
        submenu.add_command(label='gyy-mm-dd',
                            command=self.insert_date_giso)
        submenu.add_separator()
        #
        submenu.add_command(label='H時M分S秒',
                            command=self.insert_time_HHMS)
        submenu.add_command(label='午前H時M分S秒',
                            command=self.insert_time_GHMS)
        submenu.add_command(label='h時m分s秒',
                            command=self.insert_time_hhms)
        submenu.add_command(label='午前h時m分s秒',
                            command=self.insert_time_Ghms)
        submenu.add_command(label='hh:mm:ss',
                            command=self.insert_time_iso)
        submenu.add_command(label='AMhh:mm:ss',
                            command=self.insert_time_giso)
        submenu.add_separator()
        #
        submenu.add_command(label='yyyy-mm-ddThh:mm:ss+09:00',
                            command=self.insert_datetime)
        submenu.add_command(label='yy-mm-ddThh:mm:ss',
                            command=self.insert_datetime_simple)

    ######
    # COMMAND

    def insert_date_YYMD(self):
        now = self._get_now()
        date = now.strftime('%Y年%m月%d日')
        date = self._remove_zero(date)
        date = self._convert_half_to_full(date)
        self.txt.insert('insert', date)

    def insert_date_GYMD(self):
        now = self._get_now()
        year = int(now.strftime('%Y')) - 2018
        date = '令和' + str(year) + '年' + now.strftime('%m月%d日')
        date = self._remove_zero(date)
        date = self._convert_half_to_full(date)
        self.txt.insert('insert', date)

    def insert_date_yymd(self):
        now = self._get_now()
        date = now.strftime('%Y年%m月%d日')
        date = self._remove_zero(date)
        self.txt.insert('insert', date)

    def insert_date_Gymd(self):
        now = self._get_now()
        year = int(now.strftime('%Y')) - 2018
        date = '令和' + str(year) + '年' + now.strftime('%m月%d日')
        date = self._remove_zero(date)
        self.txt.insert('insert', date)

    def insert_date_iso(self):
        now = self._get_now()
        date = now.strftime('%Y-%m-%d')
        self.txt.insert('insert', date)

    def insert_date_giso(self):
        now = self._get_now()
        year = int(now.strftime('%Y')) - 2018
        if year < 10:
            date = 'R0' + str(year) + '-' + now.strftime('%m-%d')
        else:
            date = 'R' + str(year) + '-' + now.strftime('%m-%d')
        self.txt.insert('insert', date)

    def insert_time_HHMS(self):
        now = self._get_now()
        time = now.strftime('%H時%M分%S秒')
        time = self._remove_zero(time)
        time = self._convert_half_to_full(time)
        self.txt.insert('insert', time)

    def insert_time_GHMS(self):
        now = self._get_now()
        hour = int(now.strftime('%H'))
        if hour < 12:
            time = '午前' + str(hour) + '時' + now.strftime('%M分%S秒')
        else:
            time = '午後' + str(hour - 12) + '時' + now.strftime('%M分%S秒')
        time = self._remove_zero(time)
        time = self._convert_half_to_full(time)
        self.txt.insert('insert', time)

    def insert_time_hhms(self):
        now = self._get_now()
        time = now.strftime('%H時%M分%S秒')
        time = self._remove_zero(time)
        self.txt.insert('insert', time)

    def insert_time_Ghms(self):
        now = self._get_now()
        hour = int(now.strftime('%H'))
        if hour < 12:
            time = '午前' + str(hour) + '時' + now.strftime('%M分%S秒')
        else:
            time = '午後' + str(hour - 12) + '時' + now.strftime('%M分%S秒')
        time = self._remove_zero(time)
        self.txt.insert('insert', time)

    def insert_time_iso(self):
        now = self._get_now()
        time = now.strftime('%H:%M:%S')
        self.txt.insert('insert', time)

    def insert_time_giso(self):
        now = self._get_now()
        hour = int(now.strftime('%H'))
        if hour < 12:
            time = 'AM' + str(hour) + ':' + now.strftime('%M:%S')
        else:
            time = 'PM' + str(hour - 12) + ':' + now.strftime('%M:%S')
        self.txt.insert('insert', time)

    def insert_datetime(self):
        now = self._get_now()
        self.txt.insert('insert', now.isoformat(timespec='seconds'))

    def insert_datetime_simple(self):
        now = self._get_now()
        self.txt.insert('insert', now.strftime('%y-%m-%dT%H:%M:%S'))

    @staticmethod
    def _remove_zero(text):
        text = re.sub('^0', '', text)
        text = re.sub('年0', '年', text)
        text = re.sub('月0', '月', text)
        text = re.sub('時0', '時', text)
        text = re.sub('分0', '分', text)
        return text

    ################
    # COMMAND

    def insert_character_by_code(self):
        b = 'コード番号'
        p = 'コード番号を入力してください．'
        h, t = '', ''
        s = ''
        while not re.match('^[0-9a-fA-F]{4}$', s):
            s = OneWordDialog(self.txt, self, b, p, h, t, s).get_value()
            if s is None:
                return
        self.txt.insert('insert', chr(int(s, 16)))

    ################
    # SUBMENU INSERT IVS CHARACTER

    def _make_submenu_insert_ivs_character(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='人名・地名の字体を挿入', menu=submenu)
        #
        submenu.add_command(label='文字コードから人名・地名の字体を挿入',
                            command=self.insert_ivs)
        submenu.add_separator()
        submenu.add_command(label='"兼"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_517c)
        submenu.add_command(label='"化"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_5316)
        submenu.add_command(label='"啄"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_5544)
        submenu.add_command(label='"崩"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_5d29)
        submenu.add_command(label='"廣"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_5ee3)
        submenu.add_command(label='"愉"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_6109)
        submenu.add_command(label='"拳"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_62f3)
        submenu.add_command(label='"曙"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_66d9)
        submenu.add_command(label='"榊"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_698a)
        submenu.add_command(label='"浩"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_6d69)
        submenu.add_command(label='"浮"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_6d6e)
        submenu.add_command(label='"漢"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_6f22)
        submenu.add_command(label='"琢"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_7422)
        submenu.add_command(label='"社"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_793e)
        submenu.add_command(label='"祇"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_7947)
        submenu.add_command(label='"空"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_7a7a)
        submenu.add_command(label='"範"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_7bc4)
        submenu.add_command(label='"花"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_82b1)
        submenu.add_command(label='"芸"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_82b8)
        submenu.add_command(label='"菅"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_83c5)
        submenu.add_command(label='"葛"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_845b)
        submenu.add_command(label='"藏"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_85cf)
        submenu.add_command(label='"藤"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_85e4)
        submenu.add_command(label='"覇"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_8987)
        submenu.add_command(label='"角"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_89d2)
        submenu.add_command(label='"諭"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_8aed)
        submenu.add_command(label='"辻"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_8fbb)
        submenu.add_command(label='"邉"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_9089)
        submenu.add_command(label='"邊"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_908a)
        submenu.add_command(label='"餅"の人名・地名の字体の候補を全て挿入',
                            command=self.insert_ivs_of_9905)

    ######
    # COMMAND

    def insert_ivs(self):
        c = ''
        if self.txt.tag_ranges('sel'):
            c = self.txt.get('sel.first', 'sel.last')
        elif 'akauni' in self.txt.mark_names():
            c = ''
            c += self.txt.get('akauni', 'insert')
            c += self.txt.get('insert', 'akauni')
        if len(c) == 1:
            i = self.IvsDialog(self.txt, self, c)
        else:
            i = self.IvsDialog(self.txt, self)
        if len(c) == 1 and i.has_inserted:
            if self.txt.tag_ranges('sel'):
                self.txt.delete('sel.first', 'sel.first+1c')
            elif 'akauni' in self.txt.mark_names():
                if self.txt.get('akauni', 'insert') != '':
                    self.txt.delete('akauni', 'akauni+1c')
                elif self.txt.get('insert', 'akauni') != '':
                    self.txt.delete('akauni-1c', 'akauni')

    class IvsDialog(tkinter.simpledialog.Dialog):

        def __init__(self, pane, mother, char=None):
            self.pane = pane
            self.mother = mother
            self.char = None
            self.code = None
            if char is not None:
                self.char = char
                self.code = re.sub('^0x', '', hex(ord(char))).upper()
            self.has_inserted = False
            super().__init__(pane, title='文字コードから人名・地名漢字を挿入')

        def body(self, pane):
            fon = self.mother.gothic_font
            t = '下記のURLで漢字を検索してください．\n' + \
                'https://moji.or.jp/mojikibansearch/basic\n\n' + \
                '「対応するUCS」の下の段を下に入力してください．\n' + \
                '例：花の場合→<82B1,E0102>\n'
            frm = tkinter.Frame(pane)
            frm.pack(side='top')
            txt = tkinter.Label(frm, text=t, justify='left')
            txt.pack(side='left')
            frm = tkinter.Frame(pane)
            frm.pack(side='top')
            txt = tkinter.Label(frm, text='<')
            txt.pack(side='left')
            self.entry1 = tkinter.Entry(frm, width=7, font=fon)
            self.entry1.pack(side='left')
            if self.code is not None:
                self.entry1.insert(0, self.code)
            txt = tkinter.Label(frm, text=',E01', font=fon)  # E0100-E01EF
            txt.pack(side='left')
            self.entry2 = tkinter.Entry(frm, width=7, font=fon)
            self.entry2.pack(side='left')
            txt = tkinter.Label(frm, text='>')
            txt.pack(side='left')
            # self.bind('<Key-Return>', self.ok)
            # self.bind('<Key-Escape>', self.cancel)
            # super().body(pane)
            if self.code is None:
                return self.entry1
            else:
                return self.entry2

        # def buttonbox(self):
        #     btn = tkinter.Frame(self)
        #     self.btn1 = tkinter.Button(btn, text='OK', width=6,
        #                                command=self.ok)
        #     self.btn1.pack(side=tkinter.LEFT, padx=3, pady=3)
        #     self.btn2 = tkinter.Button(btn, text='Cancel', width=6,
        #                                command=self.cancel)
        #     self.btn2.pack(side=tkinter.LEFT, padx=3, pady=3)
        #     btn.pack()

        def apply(self):
            ucs = self.entry1.get()
            ivs = self.entry2.get()
            if re.match('^[0-9a-fA-F]{4}$', ucs):
                self.pane.insert('insert', chr(int(ucs, 16)))
                if re.match('^[0-9a-eA-E][0-9a-fA-F]$', ivs):
                    i = int('E01' + ivs, 16) - 917760
                    self.pane.insert('insert', str(i) + ';')
                    self.has_inserted = True

    def insert_ivs_of_517c(self):
        self.txt.insert('insert',
                        'A兼2;' +  # E0102 MJ007297
                        'B兼3;' +  # E0103 MJ007298
                        'C兼4;' +  # E0104 MJ007296
                        'D兼5;' +  # E0105 MJ056985
                        'E兼6;')   # E0106 MJ056989

    def insert_ivs_of_5316(self):
        self.txt.insert('insert',
                        'A化2;' +  # E0102 MJ007779
                        'B化3;')   # E0103 MJ007778

    def insert_ivs_of_5544(self):
        self.txt.insert('insert',
                        'A啄2;' +  # E0102 MJ008370
                        'B啄3;' +  # E0103 MJ008374
                        'C啄4;' +  # E0104 MJ008372
                        'D啄5;' +  # E0105 MJ008371
                        'E啄6;')   # E0106 MJ008373

    def insert_ivs_of_5d29(self):
        self.txt.insert('insert',
                        'A崩2;' +  # E0102 MJ010574
                        'B崩3;')   # E0103 MJ010573

    def insert_ivs_of_5ee3(self):
        self.txt.insert('insert',
                        'A廣3;' +  # E0103 MJ011077
                        'B廣4;' +  # E0104 MJ011075
                        'C廣5;' +  # E0105 MJ011076
                        'D廣12;')  # E010C MJ011078

    def insert_ivs_of_6109(self):
        self.txt.insert('insert',
                        'A愉2;' +  # E0102 MJ011726
                        'B愉3;')   # E0103 MJ011725

    def insert_ivs_of_62f3(self):
        self.txt.insert('insert',
                        'A拳2;' +  # E0102 MJ012304
                        'B拳3;')   # E0103 MJ012303

    def insert_ivs_of_66d9(self):
        self.txt.insert('insert',
                        'A曙2;' +  # E0102 MJ013447
                        'B曙3;')   # E0103 MJ013448

    def insert_ivs_of_698a(self):
        self.txt.insert('insert',
                        'A榊2;' +  # E0102 MJ014255
                        'B榊3;')   # E0103 MJ014256

    def insert_ivs_of_6d69(self):
        self.txt.insert('insert',
                        'A浩2;' +  # E0102 MJ015356
                        'B浩3;')   # E0103 MJ015355

    def insert_ivs_of_6d6e(self):
        self.txt.insert('insert',
                        'A浮2;' +  # E0102 MJ015362
                        'B浮3;')   # E0103 MJ015361

    def insert_ivs_of_6f22(self):
        self.txt.insert('insert',
                        'A漢2;' +  # E0102 MJ015841
                        'B漢3;' +  # E0102 MJ030268
                        'C漢7;')   # E0107 MJ015844

    def insert_ivs_of_7422(self):
        self.txt.insert('insert',
                        'A琢2;' +  # E0102 MJ017282
                        'B琢3;' +  # E0103 MJ030273
                        'C琢4;' +  # E0104 MJ017283
                        'D琢5;')   # E0105 MJ030271

    def insert_ivs_of_793e(self):
        self.txt.insert('insert',
                        'A社2;' +  # E0102 MJ018753
                        'B社3;' +  # E0103 MJ030274
                        'C社4;')   # E0104 MJ058201

    def insert_ivs_of_7947(self):
        self.txt.insert('insert',
                        'A祇2;' +  # E0102 MJ018770
                        'B祇3;')   # E0103 MJ018771

    def insert_ivs_of_7a7a(self):
        self.txt.insert('insert',
                        'A空2;' +  # E0102 MJ019210
                        'B空3;')   # E0103 MJ039211

    def insert_ivs_of_7bc4(self):
        self.txt.insert('insert',
                        'A範1;' +  # E0101 MJ019582
                        'B範2;' +  # E0102 MJ019583
                        'C範3;')   # E0103 MJ019584

    def insert_ivs_of_82b1(self):
        self.txt.insert('insert',
                        'A花2;' +  # E0102 MJ021591
                        'B花3;' +  # E0103 MJ021592
                        'C花4;' +  # E0104 MJ021593
                        'D花6;')   # E0106 MJ021594

    def insert_ivs_of_82b8(self):
        self.txt.insert('insert',
                        'A芸1;' +  # E0101 MJ021606
                        'B芸2;')   # E0102 MJ021607

    def insert_ivs_of_83c5(self):
        self.txt.insert('insert',
                        'A菅1;' +  # E0101 MJ022070
                        'B菅2;')   # E0102 MJ022071

    def insert_ivs_of_845b(self):
        self.txt.insert('insert',
                        'A葛2;' +  # E0102 MJ022335
                        'B葛3;' +  # E0103 MJ022336
                        'C葛4;' +  # E0104 MJ022340
                        'D葛5;' +  # E0105 MJ022341
                        'E葛6;' +  # E0106 MJ022338
                        'F葛7;' +  # E0107 MJ022337
                        'G葛8;')   # E0108 MJ022339

    def insert_ivs_of_85cf(self):
        self.txt.insert('insert',
                        'A藏2;' +  # E0102 MJ023044
                        'B藏3;' +  # E0103 MJ023046
                        'C藏4;' +  # E0104 MJ023047
                        'D藏5;')   # E0105 MJ023045

    def insert_ivs_of_85e4(self):
        self.txt.insert('insert',
                        'A藤2;' +  # E0102 MJ023079
                        'B藤3;' +  # E0103 MJ023080
                        'C藤4;' +  # E0104 MJ023081
                        'D藤5;' +  # E0105 MJ023082
                        'E藤6;')   # E0106 MJ060144

    def insert_ivs_of_8987(self):
        self.txt.insert('insert',
                        'A覇2;' +  # E0102 MJ024210
                        'B覇3;')   # E0103 MJ024209

    def insert_ivs_of_89d2(self):
        self.txt.insert('insert',
                        'A角2;' +  # E0102 MJ024281
                        'B角3;' +  # E0103 MJ024283
                        'C角4;')   # E0104 MJ024282

    def insert_ivs_of_8aed(self):
        self.txt.insert('insert',
                        'A諭2;' +  # E0102 MJ024620
                        'B諭3;' +  # E0103 MJ024621
                        'C諭4;')   # E0104 MJ024619

    def insert_ivs_of_8fbb(self):
        self.txt.insert('insert',
                        'A辻2;' +  # E0102 MJ025760
                        'B辻3;')   # E0103 MJ025761

    def insert_ivs_of_9089(self):
        self.txt.insert('insert',
                        'A邉15;' +  # E010F MJ026190
                        'B邉16;' +  # E0110 MJ060248
                        'C邉17;' +  # E0111 MJ060239
                        'D邉18;' +  # E0112 MJ060238
                        'E邉19;' +  # E0113 MJ060237
                        'F邉20;' +  # E0114 MJ060235
                        'G邉21;' +  # E0115 MJ060234
                        'H邉22;' +  # E0116 MJ058866
                        'I邉23;' +  # E0117 MJ026197
                        'J邉24;' +  # E0118 MJ060236
                        'K邉25;' +  # E0119 MJ026191
                        'L邉26;' +  # E011A MJ026194
                        'M邉27;' +  # E011B MJ026192
                        'N邉28;' +  # E011C MJ026195
                        'O邉29;' +  # E011D MJ026196
                        'P邉31;')   # E011F MJ026193

    def insert_ivs_of_908a(self):
        self.txt.insert('insert',
                        'A邊8;' +   # E0108 MJ026200
                        'B邊9;' +   # E0109 MJ060240
                        'C邊10;' +  # E010A MJ026205
                        'D邊11;' +  # E010B MJ026204
                        'E邊12;' +  # E010C MJ026203
                        'F邊13;' +  # E010D MJ026202
                        'G邊14;' +  # E010E MJ026201
                        'H邊15;' +  # E010F MJ026199
                        'I邊16;' +  # E0110 MJ026206
                        'J邊17;' +  # E0111 MJ058870
                        'K邊18;')   # E0112 MJ026207

    def insert_ivs_of_9905(self):
        self.txt.insert('insert',
                        'A餅2;' +  # E0102 MJ028397
                        'B餅3;')   # E0103 MJ028398

    ################
    # SUBMENU INSERT HORIZONTAL LINE

    def _make_submenu_insert_horizontal_line(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='横棒を挿入', menu=submenu)
        #
        submenu.add_command(label='"-"（U+002D）半角ハイフンマイナス',
                            command=self.insert_hline_002d)
        submenu.add_command(label='"­"（U+00AD）改行時だけに表示されるハイフン',
                            command=self.insert_hline_00ad)
        submenu.add_command(label='"֊"（U+058A）アメリカンハイフン',
                            command=self.insert_hline_058a)
        submenu.add_command(label='"־"（U+05BE）ヘブライ語のマカフ',
                            command=self.insert_hline_05be)
        submenu.add_command(label='"᠆"（U+1806）モンゴル語のソフトハイフン',
                            command=self.insert_hline_1806)
        submenu.add_command(label='"᠊"（U+180A）モンゴル語のニルグ',
                            command=self.insert_hline_180a)
        submenu.add_command(label='"‐"（U+2010）ハイフン',
                            command=self.insert_hline_2010)
        submenu.add_command(label='"‑"（U+2011）改行を禁止するハイフン',
                            command=self.insert_hline_2011)
        submenu.add_command(label='"‒"（U+2012）数字幅のダッシュ',
                            command=self.insert_hline_2012)
        submenu.add_command(label='"–"（U+2013）Ｎ幅のダッシュ',
                            command=self.insert_hline_2013)
        submenu.add_command(label='"—"（U+2014）Ｍ幅のダッシュ',
                            command=self.insert_hline_2014)
        submenu.add_command(label='"―"（U+2015）水平線',
                            command=self.insert_hline_2015)
        submenu.add_command(label='"⁃"（U+2043）箇条書きの記号',
                            command=self.insert_hline_2043)
        submenu.add_command(label='"⁻"（U+207B）上付きマイナス',
                            command=self.insert_hline_207b)
        submenu.add_command(label='"₋"（U+208B）下付きマイナス',
                            command=self.insert_hline_208b)
        submenu.add_command(label='"−"（U+2212）マイナスサイン',
                            command=self.insert_hline_2212)
        submenu.add_command(label='"─"（U+2500）罫線',
                            command=self.insert_hline_2500)
        submenu.add_command(label='"━"（U+2501）太字の罫線',
                            command=self.insert_hline_2501)
        submenu.add_command(label='"➖"（U+2796）太字マイナス記号',
                            command=self.insert_hline_2796)
        submenu.add_command(label='"⸺"（U+2E3A）2倍幅のＭ幅ダッシュ',
                            command=self.insert_hline_2e3a)
        submenu.add_command(label='"⸻"（U+2E3B）3倍幅のＭ幅ダッシュ',
                            command=self.insert_hline_2e3b)
        submenu.add_command(label='"ㄧ"（U+3127）注音符号の「Ｉ」の発音',
                            command=self.insert_hline_3127)
        submenu.add_command(label='"ㅡ"（U+3161）ハングルの「ウ」',
                            command=self.insert_hline_3161)
        submenu.add_command(label='"﹘"（U+FE58）小さいＭ幅ダッシュ',
                            command=self.insert_hline_fe58)
        submenu.add_command(label='"﹣"（U+FE63）小さいハイフンマイナス',
                            command=self.insert_hline_fe63)
        submenu.add_command(label='"－"（U+FF0D）全角ハイフンマイナス',
                            command=self.insert_hline_ff0d)
        submenu.add_command(label='"ｰ"（U+FF70）半角カナの長音記号',
                            command=self.insert_hline_ff70)

    ######
    # COMMAND

    # "-"（U+002D）HYPHEN-MINUS（半角ハイフンマイナス）
    # "­"（U+00AD）SOFT HYPHEN（改行時だけに表示されるハイフン',）
    # "֊"（U+058A）ARMENIAN HYPHEN（アメリカンハイフン）
    # "־"（U+05BE）HEBREW PUNCTUATION MAQAF（ヘブライ語のマカフ）
    # "᠆"（U+1806）MONGOLIAN TODO SOFT HYPHEN（モンゴル語のソフトハイフン）
    # "᠊"（U+180A）MONGOLIAN NIRUGU（モンゴル語のニルグ）
    # "‐"（U+2010）HYPHEN（ハイフン）（EUC:A1BE）
    # "‑"（U+2011）NON-BREAKING HYPHEN（改行しないハイフン）
    # "‒"（U+2012）FIGURE DASH（数字幅のダッシュ）
    # "–"（U+2013）EN DASH（Ｎ幅ダッシュ）
    # "—"（U+2014）EM DASH（Ｍ幅ダッシュ）
    # "―"（U+2015）HORIZONTAL BAR（水平線）（EUC:A1BD）
    # "⁃"（U+2043）HYPHEN BULLET（箇条書きの記号）
    # "⁻"（U+207B）SUPERSCRIPT MINUS（上付きマイナス）
    # "₋"（U+208B）SUBSCRIPT MINUS（下付きマイナス）
    # "−"（U+2212）MINUS SIGN（マイナス記号）（EUC:A1DD）
    # "─"（U+2500）BOX DRAWINGS LIGHT HORIZONTAL（罫線）（EUC:A8A1）
    # "━"（U+2501）BOX DRAWINGS HEAVY HORIZONTAL（太字罫線）（EUC:A8AC）
    # "➖"（U+2796）HEAVY MINUS SIGN（太字マイナス記号）
    # "⸺"（U+2E3A）TWO-EM DASH（2倍幅のＭ幅ダッシュ）
    # "⸻"（U+2E3B）THREE-EM DASH（3倍幅のＭ幅ダッシュ）
    # "ー"（U+30FC）KATAKANA-HIRAGANA PROLONGED SOUND MARK（日本語の長音記号）
    # "ㄧ"（U+3127）BOPOMOFO LETTER I（注音符号の「Ｉ」の発音）
    # "ㅡ"（U+3161）HANGUL LETTER EU（ハングルの「ウ」）
    # "一"（U+4E00）CJK IDEOGRAPH FIRST（漢数字の「１」）
    # "﹘"（U+FE58）SMALL EM DASH（小さいＭ幅ダッシュ）
    # "﹣"（U+FE63）SMALL HYPHEN-MINUS（小さいハイフンマイナス）
    # "－"（U+FF0D）FULLWIDTH HYPHEN-MINUS（全角ハイフンマイナス）
    # "ｰ"（U+FF70）HALFWIDTH KATAKANA-HIRAGANA PROLONGED SOUND MARK
    #                                        （半角カナの長音記号）（EUC:8EB0）

    def insert_hline_002d(self):
        self._insert_hline('\u002D')  # 半角ハイフンマイナス

    def insert_hline_00ad(self):
        self._insert_hline('\u00AD')  # 改行時だけに表示されるハイフン',

    def insert_hline_058a(self):
        self._insert_hline('\u058A')  # アメリカンハイフン

    def insert_hline_05be(self):
        self._insert_hline('\u05BE')  # ヘブライ語のマカフ

    def insert_hline_1806(self):
        self._insert_hline('\u1806')  # モンゴル語のソフトハイフン

    def insert_hline_180a(self):
        self._insert_hline('\u180A')  # モンゴル語のニルグ

    def insert_hline_2010(self):
        self._insert_hline('\u2010')  # ハイフン

    def insert_hline_2011(self):
        self._insert_hline('\u2011')  # 改行しないハイフン

    def insert_hline_2012(self):
        self._insert_hline('\u2012')  # 数字幅のダッシュ

    def insert_hline_2013(self):
        self._insert_hline('\u2013')  # Ｎ幅ダッシュ

    def insert_hline_2014(self):
        self._insert_hline('\u2014')  # Ｍ幅ダッシュ

    def insert_hline_2015(self):
        self._insert_hline('\u2015')  # 水平線

    def insert_hline_2043(self):
        self._insert_hline('\u2043')  # 箇条書きの記号

    def insert_hline_207b(self):
        self._insert_hline('\u207B')  # 上付きマイナス

    def insert_hline_208b(self):
        self._insert_hline('\u208B')  # 下付きマイナス

    def insert_hline_2212(self):
        self._insert_hline('\u2212')  # マイナス記号

    def insert_hline_2500(self):
        self._insert_hline('\u2500')  # 罫線

    def insert_hline_2501(self):
        self._insert_hline('\u2501')  # 太字罫線

    def insert_hline_2796(self):
        self._insert_hline('\u2796')  # 太字マイナス記号

    def insert_hline_2e3a(self):
        self._insert_hline('\u2E3A')  # 2倍幅のＭ幅ダッシュ

    def insert_hline_2e3b(self):
        self._insert_hline('\u2E3B')  # 3倍幅のＭ幅ダッシュ

    def insert_hline_3127(self):
        self._insert_hline('\u3127')  # 注音符号のIの発音

    def insert_hline_3161(self):
        self._insert_hline('\u3161')  # ハングルの「ウ」

    def insert_hline_fe58(self):
        self._insert_hline('\uFE58')  # 小さいＭ幅ダッシュ

    def insert_hline_fe63(self):
        self._insert_hline('\uFE63')  # 小さいハイフンマイナス

    def insert_hline_ff0d(self):
        self._insert_hline('\uFF0D')  # 全角ハイフンマイナス

    def insert_hline_ff70(self):
        self._insert_hline('\uFF70')  # 半角カナの長音記号

    def _insert_hline(self, char):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if self._is_read_only_pane(pane):
            return
        pane.insert('insert', char)
        self.paint_out_line(self._get_v_position_of_insert(pane) - 1)

    ################
    # COMMAND

    def insert_symbol(self):
        candidates = ['⑴', '⑵', '⑶', '⑷', '⑸', '⑹', '⑺', '⑻', '⑼', '⑽',
                      '⑾', '⑿', '⒀', '⒁', '⒂', '⒃', '⒄', '⒅', '⒆', '⒇',
                      '⓪',
                      '①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩',
                      '⑪', '⑫', '⑬', '⑭', '⑮', '⑯', '⑰', '⑱', '⑲', '⑳',
                      '²', '³',
                      'Α', 'Β', 'Γ', 'Δ', 'Ε', 'Ζ', 'Η', 'Θ', 'Ι', 'Κ',
                      'Λ', 'Μ', 'Ν', 'Ξ', 'Ο', 'Π', 'Ρ', 'Σ', 'Τ', 'Υ',
                      'Φ', 'Χ', 'Ψ', 'Ω',
                      'α', 'β', 'γ', 'δ', 'ε', 'ζ', 'η', 'θ', 'ι', 'κ',
                      'λ', 'μ', 'ν', 'ξ', 'ο', 'π', 'ρ', 'ς', 'σ', 'τ',
                      'υ', 'φ', 'χ', 'ψ', 'ω',
                      '∞', '√', '∛', '∜', '∀', '∃', '∴', '∵',
                      '±', '∓', '≠', '≡', '≒', '≈', '≦', '≧', '≤', '≥',
                      '⊂', '⊃', '⊆', '⊇', '∈', '∋', '∩', '∪',
                      '⋮', '⋯', '⋰', '⋱', '∂', '∫', '∮', '∑', '∏',
                      '㊞',
                      '♠', '♡', '♢', '♣', '♤', '♥', '♦', '♧',
                      '☹', '☺', '☻',
                      '✊', '✋', '✌',
                      '✿', '❀',
                      '☯']
        self.SymbolDialog(self.txt, self, candidates)

    class SymbolDialog(tkinter.simpledialog.Dialog):

        def __init__(self, pane, mother, candidates):
            self.pane = pane
            self.mother = mother
            self.candidates = candidates
            super().__init__(pane, title='記号を挿入')

        def body(self, pane):
            fon = self.mother.gothic_font
            self.symbol = tkinter.StringVar()
            for i, cnd in enumerate(self.candidates):
                rd = tkinter.Radiobutton(pane, text=cnd, font=fon,
                                         variable=self.symbol, value=cnd)
                y, x = int(i / 10), (i % 10)
                rd.grid(row=y, column=x, columnspan=1, padx=3, pady=3,
                        sticky='w')
            # self.bind('<Key-Return>', self.ok)
            # self.bind('<Key-Escape>', self.cancel)
            # super().body(pane)

        # def buttonbox(self):
        #     btn = tkinter.Frame(self)
        #     self.btn1 = tkinter.Button(btn, text='OK', width=6,
        #                                command=self.ok)
        #     self.btn1.pack(side=tkinter.LEFT, padx=3, pady=3)
        #     self.btn2 = tkinter.Button(btn, text='Cancel', width=6,
        #                                command=self.cancel)
        #     self.btn2.pack(side=tkinter.LEFT, padx=3, pady=3)
        #     btn.pack()

        def apply(self):
            symbol = self.symbol.get()
            self.pane.insert('insert', symbol)
            # self.pane.mark_set('insert', 'insert-1c')
            self.pane.focus_set()

    ##########################
    # MENU PARAGRAPH

    def _make_menu_paragraph(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='段落(P)', menu=menu, underline=3)
        #
        menu.add_command(label='段落の余白の長さを設定',
                         command=self.set_paragraph_length)
        menu.add_separator()
        #
        menu.add_command(label='設定を挿入',
                         command=self.insert_config)
        self._make_submenu_insert_multicolumns(menu)
        self._make_submenu_insert_chapter(menu)
        self._make_submenu_insert_section(menu)
        self._make_submenu_insert_list(menu)
        self._make_submenu_insert_table(menu)
        self._make_submenu_insert_image(menu)
        menu.add_command(label='改ページを挿入',
                         command=self.insert_page_break)
        self._make_submenu_insert_math_paragraph_expression(menu)
        menu.add_separator()
        #
        menu.add_command(label='チャプターの番号を変更',
                         command=self.set_chapter_number)
        menu.add_command(label='セクションの番号を変更',
                         command=self.set_section_number)
        menu.add_command(label='箇条書きの番号を変更',
                         command=self.set_list_number)
        menu.add_separator()
        #
        menu.add_command(label='段落を整形',
                         command=self.tidy_up_paragraph)
        # menu.add_separator()

    ################
    # COMMAND

    def set_paragraph_length(self):
        self.LengthRevisersDialog(self.txt, self)

    class LengthRevisersDialog(tkinter.simpledialog.Dialog):

        def __init__(self, pane, mother, length=None):
            self.pane = pane
            self.mother = mother
            bef_text = self.pane.get('1.0', 'insert')
            aft_text = self.pane.get('insert', 'end-1c')
            self.head_text \
                = re.sub('^((?:.|\n)*\n\n)((?:.|\n)*)?', '\\1', bef_text)
            bef_para = re.sub('^(.|\n)*\n\n', '', bef_text)
            aft_para = re.sub('\n\n(.|\n)*$', '', aft_text)
            paragraph = bef_para + aft_para
            res_length_reviser = '(?:v|V|X|<<|<|>)=[-\\+]?(?:[0-9]*\\.)?[0-9]+'
            res = '^((?:' + res_length_reviser + '(?:\\s|\n)*)*)((?:.|\n)*)$'
            self.length_revisers = re.sub(res, '\\1', paragraph)
            if length is not None:
                self.length = length
            else:
                self.length = {'space before': '0.0', 'space after': '0.0',
                               'line spacing': '0.0', 'first indent': '0.0',
                               'left indent': '0.0', 'right indent': '0.0'}
                res_bef = '(?:.|\n)*'
                res_aft = '=([-\\+]?(?:[0-9]*\\.)?[0-9]+)' + '(?:.|\n)*'
                res = res_bef + 'v' + res_aft
                if re.match(res, self.length_revisers):
                    self.length['space before'] \
                        = str(float(re.sub(res, '\\1', self.length_revisers)))
                res = res_bef + 'V' + res_aft
                if re.match(res, self.length_revisers):
                    self.length['space after'] \
                        = str(float(re.sub(res, '\\1', self.length_revisers)))
                res = res_bef + 'X' + res_aft
                if re.match(res, self.length_revisers):
                    self.length['line spacing'] \
                        = str(float(re.sub(res, '\\1', self.length_revisers)))
                res = res_bef + '<<' + res_aft
                if re.match(res, self.length_revisers):
                    self.length['first indent'] \
                        = str(float(re.sub(res, '\\1', self.length_revisers))
                              * -1)
                res = res_bef + '<' + res_aft
                if re.match(res, self.length_revisers):
                    self.length['left indent'] \
                        = str(float(re.sub(res, '\\1', self.length_revisers))
                              * -1)
                res = res_bef + '>' + res_aft
                if re.match(res, self.length_revisers):
                    self.length['right indent'] \
                        = str(float(re.sub(res, '\\1', self.length_revisers))
                              * -1)
            super().__init__(pane, title='段落の長さを設定')

        def body(self, pane):
            f = self.mother.gothic_font
            self.title1 = tkinter.Label(pane, text='前の段落との間の幅')
            self.title1.grid(row=0, column=0)
            self.entry1 = tkinter.Entry(pane, width=7, font=f, justify='right')
            self.entry1.insert(0, self.length['space before'])
            self.entry1.grid(row=0, column=1)
            self.unit1 = tkinter.Label(pane, text='行間')
            self.unit1.grid(row=0, column=2)
            self.title2 = tkinter.Label(pane, text='次の段落との間の幅')
            self.title2.grid(row=1, column=0)
            self.entry2 = tkinter.Entry(pane, width=7, font=f, justify='right')
            self.entry2.insert(0, self.length['space after'])
            self.entry2.grid(row=1, column=1)
            self.unit2 = tkinter.Label(pane, text='行間')
            self.unit2.grid(row=1, column=2)
            self.title3 = tkinter.Label(pane, text='段落内の改行の幅　')
            self.title3.grid(row=2, column=0)
            self.entry3 = tkinter.Entry(pane, width=7, font=f, justify='right')
            self.entry3.insert(0, self.length['line spacing'])
            self.entry3.grid(row=2, column=1)
            self.unit3 = tkinter.Label(pane, text='行間')
            self.unit3.grid(row=2, column=2)
            self.title4 = tkinter.Label(pane, text='一行目の字下げの幅')
            self.title4.grid(row=3, column=0)
            self.entry4 = tkinter.Entry(pane, width=7, font=f, justify='right')
            self.entry4.insert(0, self.length['first indent'])
            self.entry4.grid(row=3, column=1)
            self.unit4 = tkinter.Label(pane, text='文字')
            self.unit4.grid(row=3, column=2)
            self.title5 = tkinter.Label(pane, text='左の字下げの幅　　')
            self.title5.grid(row=4, column=0)
            self.entry5 = tkinter.Entry(pane, width=7, font=f, justify='right')
            self.entry5.insert(0, self.length['left indent'])
            self.entry5.grid(row=4, column=1)
            self.unit5 = tkinter.Label(pane, text='文字')
            self.unit5.grid(row=4, column=2)
            self.title6 = tkinter.Label(pane, text='右の字下げの幅　　')
            self.title6.grid(row=5, column=0)
            self.entry6 = tkinter.Entry(pane, width=7, font=f, justify='right')
            self.entry6.insert(0, self.length['right indent'])
            self.entry6.grid(row=5, column=1)
            self.unit6 = tkinter.Label(pane, text='文字')
            self.unit6.grid(row=5, column=2)
            return self.entry1

        def apply(self):
            has_error = False
            res = '^[-\\+]?(?:[0-9]*\\.)?[0-9]+$'
            space_before = re.sub('\\s', '', self.entry1.get())
            if re.match(res, space_before):
                self.length['space before'] = space_before
            else:
                has_error = True
            space_after = re.sub('\\s', '', self.entry2.get())
            if re.match(res, space_after):
                self.length['space after'] = space_after
            else:
                has_error = True
            line_spacing = re.sub('\\s', '', self.entry3.get())
            if re.match(res, line_spacing):
                self.length['line spacing'] = line_spacing
            else:
                has_error = True
            first_indent = re.sub('\\s', '', self.entry4.get())
            if re.match(res, first_indent):
                self.length['first indent'] = first_indent
            else:
                has_error = True
            left_indent = re.sub('\\s', '', self.entry5.get())
            if re.match(res, left_indent):
                self.length['left indent'] = left_indent
            else:
                has_error = True
            right_indent = re.sub('\\s', '', self.entry6.get())
            if re.match(res, right_indent):
                self.length['right indent'] = right_indent
            else:
                has_error = True
            if has_error:
                n = 'エラー'
                m = '値に正負の小数以外が含まれています．'
                tkinter.messagebox.showerror(n, m)
                Makdo.LengthRevisersDialog(self.pane, self.length)
            else:
                len_beg = len(self.head_text)
                len_end = len(self.head_text + self.length_revisers)
                beg = '1.0+' + str(len_beg) + 'c'
                end = '1.0+' + str(len_end) + 'c'
                self.pane.delete(beg, end)
                leng_revs = ''
                leng = float(self.length['space before'])
                if leng > 0:
                    leng_revs += 'v=+' + re.sub('\\.0+$', '', str(leng)) + ' '
                else:
                    leng_revs += 'v=' + re.sub('\\.0+$', '', str(leng)) + ' '
                leng = float(self.length['space after'])
                if leng > 0:
                    leng_revs += 'V=+' + re.sub('\\.0+$', '', str(leng)) + ' '
                else:
                    leng_revs += 'V=' + re.sub('\\.0+$', '', str(leng)) + ' '
                leng = float(self.length['line spacing'])
                if leng > 0:
                    leng_revs += 'X=+' + re.sub('\\.0+$', '', str(leng)) + ' '
                elif leng < 0:
                    leng_revs += 'X=' + re.sub('\\.0+$', '', str(leng)) + ' '
                leng = float(self.length['first indent']) * -1
                if leng > 0:
                    leng_revs += '<<=+' + re.sub('\\.0+$', '', str(leng)) + ' '
                elif leng < 0:
                    leng_revs += '<<=' + re.sub('\\.0+$', '', str(leng)) + ' '
                leng = float(self.length['left indent']) * -1
                if leng > 0:
                    leng_revs += '<=+' + re.sub('\\.0+$', '', str(leng)) + ' '
                elif leng < 0:
                    leng_revs += '<=' + re.sub('\\.0+$', '', str(leng)) + ' '
                leng = float(self.length['right indent']) * -1
                if leng > 0:
                    leng_revs += '>=+' + re.sub('\\.0+$', '', str(leng)) + ' '
                elif leng < 0:
                    leng_revs += '>=' + re.sub('\\.0+$', '', str(leng)) + ' '
                leng_revs = re.sub(' $', '', leng_revs)
                self.pane.insert(beg, leng_revs + '\n')

    def insert_config(self):
        config = '''
<!--------------------------【設定】-----------------------------

# プロパティに表示される文書のタイトルを指定できます。
書題名:

# 3つの書式（普通、契約、条文）を指定できます。
文書式: 普通

# 用紙のサイズ（A3横、A3縦、A4横、A4縦）を指定できます。
用紙サ: A4縦

# 用紙の上下左右の余白をセンチメートル単位で指定できます。
上余白: 3.5 cm
下余白: 2.2 cm
左余白: 3.0 cm
右余白: 2.3 cm

# ページのヘッダーに表示する文字列（別紙 :等）を指定できます。
頭書き:

# ページ番号の書式（無、有、n :、-n-、n/N等）を指定できます。
頁番号: 有

# 行番号の記載（無、有）を指定できます。
行番号: 無

# 明朝体とゴシック体と異字体（IVS）のフォントを指定できます。
明朝体: Times New Roman / ＭＳ 明朝
ゴシ体: = / ＭＳ ゴシック
異字体: IPAmj明朝

# 基本の文字の大きさをポイント単位で指定できます。
文字サ: 12 pt

# 行間隔を基本の文字の高さの何倍にするかを指定できます。
行間隔: 2.14 倍

# セクションタイトル前後の余白を行間隔の倍数で指定できます。
前余白: 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍
後余白: 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍, 0.0 倍

# 半角文字と全角文字の間の間隔調整（無、有）を指定できます。
字間整: 無

# 備考書（コメント）などを消して完成させます。
完成稿: 偽

# 原稿の作成日時と更新日時が自動で記録されます。
作成時: - USER
更新時: - USER

---------------------------------------------------------------->

'''
        config = re.sub('^\n+', '', config)
        self.txt.insert('1.0', config)

    ################
    # SUBMENU INSERT MULTICOLUMNS

    def _make_submenu_insert_multicolumns(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='段組を挿入', menu=submenu)
        #
        submenu.add_command(label='一段組（元に戻す）',
                            command=self.insert_1_columns)
        submenu.add_command(label='二段組（同じ大きさ）',
                            command=self.insert_2_columns_mm)
        submenu.add_command(label='二段組（左が大きい）',
                            command=self.insert_2_columns_ls)
        submenu.add_command(label='二段組（右が大きい）',
                            command=self.insert_2_columns_sl)
        submenu.add_command(label='三段組',
                            command=self.insert_3_columns)

    ######
    # COMMAND

    def insert_1_columns(self):
        self._insert_paragraph_text('|-|')

    def insert_2_columns_mm(self):
        self._insert_paragraph_text('|-|-|')

    def insert_2_columns_ls(self):
        self._insert_paragraph_text('|--|-|')

    def insert_2_columns_sl(self):
        self._insert_paragraph_text('|-|--|')

    def insert_3_columns(self):
        self._insert_paragraph_text('|-|-|-|')

    ################
    # SUBMENU INSERT CHAPTER

    def _make_submenu_insert_chapter(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='チャプターを挿入', menu=submenu)
        #
        submenu.add_command(label='第１編　…',
                            command=self.insert_chap_1)
        submenu.add_command(label='　第１章　…',
                            command=self.insert_chap_2)
        submenu.add_command(label='　　第１節　…',
                            command=self.insert_chap_3)
        submenu.add_command(label='　　　第１款　…',
                            command=self.insert_chap_4)
        submenu.add_command(label='　　　　第１目　…',
                            command=self.insert_chap_5)

    ######
    # COMMAND

    def insert_chap_1(self):
        self._insert_paragraph_text('$ ')  # 第1編

    def insert_chap_2(self):
        self._insert_paragraph_text('$$ ')  # 第1章

    def insert_chap_3(self):
        self._insert_paragraph_text('$$$ ')  # 第1節

    def insert_chap_4(self):
        self._insert_paragraph_text('$$$$ ')  # 第1款

    def insert_chap_5(self):
        self._insert_paragraph_text('$$$$$ ')  # 第1目

    ################
    # SUBMENU INSERT SECTION

    def _make_submenu_insert_section(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='セクションを挿入', menu=submenu)
        #
        submenu.add_command(label='（書面のタイトル）',
                            command=self.insert_sect_1)
        submenu.add_command(label='第１　…',
                            command=self.insert_sect_2)
        submenu.add_command(label='　１　…',
                            command=self.insert_sect_3)
        submenu.add_command(label='　　(1) …',
                            command=self.insert_sect_4)
        submenu.add_command(label='　　　ア　…',
                            command=self.insert_sect_5)
        submenu.add_command(label='　　　　(ｱ) …',
                            command=self.insert_sect_6)
        submenu.add_command(label='　　　　　ａ　…',
                            command=self.insert_sect_7)
        submenu.add_command(label='　　　　　　(a) …',
                            command=self.insert_sect_8)

    ######
    # COMMAND

    def insert_sect_1(self):
        self._insert_paragraph_text('# ')  # タイトル

    def insert_sect_2(self):
        self._insert_paragraph_text('## ')  # 第1

    def insert_sect_3(self):
        self._insert_paragraph_text('### ')  # 1

    def insert_sect_4(self):
        self._insert_paragraph_text('#### ')  # (1)

    def insert_sect_5(self):
        self._insert_paragraph_text('##### ')  # ア

    def insert_sect_6(self):
        self._insert_paragraph_text('###### ')  # (ｱ)

    def insert_sect_7(self):
        self._insert_paragraph_text('####### ')  # ａ

    def insert_sect_8(self):
        self._insert_paragraph_text('######## ')  # (a)

    ################
    # SUBMENU INSERT LIST

    def _make_submenu_insert_list(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='箇条書きを挿入', menu=submenu)
        #
        submenu.add_command(label='①　…',
                            command=self.insert_nlist_1)
        submenu.add_command(label='　㋐　…',
                            command=self.insert_nlist_2)
        submenu.add_command(label='　　ⓐ　…',
                            command=self.insert_nlist_3)
        submenu.add_command(label='　　　㊀　…',
                            command=self.insert_nlist_4)
        submenu.add_separator()
        #
        submenu.add_command(label='・　…',
                            command=self.insert_blist_1)
        submenu.add_command(label='　○　…',
                            command=self.insert_blist_2)
        submenu.add_command(label='　　△　…',
                            command=self.insert_blist_3)
        submenu.add_command(label='　　　◇　…',
                            command=self.insert_blist_4)

    ######
    # COMMAND

    def insert_nlist_1(self):
        self._insert_paragraph_text('1. ')  # ①

    def insert_nlist_2(self):
        self._insert_paragraph_text('  1. ')  # ㋐

    def insert_nlist_3(self):
        self._insert_paragraph_text('    1. ')  # ⓐ

    def insert_nlist_4(self):
        self._insert_paragraph_text('      1. ')  # ㊀

    def insert_blist_1(self):
        self._insert_paragraph_text('- ')  # ・

    def insert_blist_2(self):
        self._insert_paragraph_text('  - ')  # ○

    def insert_blist_3(self):
        self._insert_paragraph_text('    - ')  # △

    def insert_blist_4(self):
        self._insert_paragraph_text('      - ')  # ◇

    ################
    # SUBMENU INSERT TABLE

    def _make_submenu_insert_table(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='表を挿入', menu=submenu)
        submenu.add_command(label='表をエクセルのファイルから挿入',
                            command=self.insert_table_from_excel)
        submenu.add_command(label='表を書式で挿入',
                            command=self.insert_table_format)

    ######
    # COMMAND

    def insert_table_from_excel(self, file_path=None):
        if file_path is None:
            ti = '表をエクセルのファイルから挿入'
            ty = [('エクセル', '.xlsx .csv')]
            _d = None
            if self.file_path is not None:
                _d = os.path.dirname(self.file_path)
            file_path = tkinter.filedialog.askopenfilename(
                    title=ti, filetypes=ty, initialdir=_d)
        if file_path == () or file_path == '':
            return
        if re.match('^(?:.|\n)+.xlsx$', file_path):
            table = self._read_xlsx_file(file_path)
        else:
            table = self._read_csv_file(file_path)
        if table is not None:
            self._insert_paragraph_text(table)

    def insert_table_format(self):
        md_text = ''
        md_text += '|タイトル  |タイトル  |タイトル  |\n'
        md_text += '==================================\n'
        md_text += '|:---------|:--------:|---------:|\n'
        md_text += '|左寄せセル|中寄せセル|右寄せセル|\n'
        md_text += '|左寄せセル|中寄せセル|右寄せセル|'
        self._insert_paragraph_text(md_text)

    ################
    # SUBMENU INSERT IMAGE

    def _make_submenu_insert_image(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='画像を挿入', menu=submenu)
        submenu.add_command(label='画像を単独で挿入',
                            command=self.insert_image_paragraph)
        submenu.add_command(label='画像と文章の段組を挿入',
                            command=self.insert_image_and_text_paragraph)
        submenu.add_command(label='文章と画像の段組を挿入',
                            command=self.insert_text_and_image_paragraph)

    ######
    # COMMAND

    def insert_image_paragraph(self):
        ti = '画像を挿入'
        ty = [('画像', '.jpg .jpeg .png .gif .tif .tiff .bmp'),
              ('全てのファイル', '*')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        image_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d)
        if image_path != () and image_path != '':
            md_text = '![代替テキスト:横x縦](' + image_path + ' "説明")'
            self._insert_paragraph_text(md_text)

    def insert_image_and_text_paragraph(self):
        ti = '画像を挿入'
        ty = [('画像', '.jpg .jpeg .png .gif .tif .tiff .bmp'),
              ('全てのファイル', '*')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        image_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d)
        if image_path != () and image_path != '':
            md_text = ''
            md_text += '<!--'
            md_text += '二段組にします．"-"の数で幅を設定してください．'
            md_text += '-->\n'
            md_text += '|-|--|\n\n'
            md_text += '<!--画像の位置の調整のために入れています．-->\n'
            md_text += '^^-----^^\n\n'
            md_text += '<!--画像の大きさをセンチメートルで設定してください．-->\n'
            md_text += '![代替テキスト:横x縦](' + image_path + ' "説明")\n\n'
            md_text += '<!--ここに文章を書きます．-->\n'
            md_text += '（ここに文章を書く）\n\n'
            md_text += '<!--一段組に戻します．-->\n'
            md_text += '|-|'
            self._insert_paragraph_text(md_text)

    def insert_text_and_image_paragraph(self):
        ti = '画像を挿入'
        ty = [('画像', '.jpg .jpeg .png .gif .tif .tiff .bmp'),
              ('全てのファイル', '*')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        image_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d)
        if image_path != () and image_path != '':
            md_text = ''
            md_text += '<!--'
            md_text += '二段組にします．"-"の数で幅を設定してください．'
            md_text += '-->\n'
            md_text += '|--|-|\n\n'
            md_text += '<!--ここに文章を書きます．-->\n'
            md_text += '（ここに文章を書く）\n\n'
            md_text += '<!--画像の位置の調整のために入れています．-->\n'
            md_text += '^^-----^^\n\n'
            md_text += '<!--画像の大きさをセンチメートルで設定してください．-->\n'
            md_text += '![代替テキスト:横x縦](' + image_path + ' "説明")\n\n'
            md_text += '<!--一段組に戻します．-->\n'
            md_text += '|-|'
            self._insert_paragraph_text(md_text)

    ################
    # COMMAND

    def insert_page_break(self):
        self._insert_paragraph_text('<pgbr>')

    ################
    # SUBMENU INSERT MATH SAMPLE

    def _make_submenu_insert_math_paragraph_expression(self, menu):
        submenu = tkinter.Menu(menu, tearoff=False)
        menu.add_cascade(label='数式を挿入', menu=submenu)
        #
        submenu.add_command(label='基本',
                            command=self.insert_math_paragraph_expression)
        submenu.add_command(label='テイラー展開',
                            command=self.insert_taylor_expansion)
        submenu.add_command(label='ピタゴラスの定理',
                            command=self.insert_pythagorean_theorem)
        submenu.add_command(label='フーリエ級数',
                            command=self.insert_fourier_series)
        submenu.add_command(label='円の面積',
                            command=self.insert_area_of_circle)
        submenu.add_command(label='三角比の公式1',
                            command=self.insert_trigonometric_formulas_1)
        submenu.add_command(label='三角比の公式2',
                            command=self.insert_trigonometric_formulas_2)
        submenu.add_command(label='二項定理',
                            command=self.insert_binomial_theorem)
        submenu.add_command(label='二次関数の解の公式',
                            command=self.insert_quadratic_formula)
        submenu.add_command(label='和の展開',
                            command=self.insert_expansion_of_addition)

    ######
    # COMMAND

    def insert_math_paragraph_expression(self):
        self._insert_paragraph_text('\\[\n（ここに"LaTeX"形式の数式を挿入）\n\\]')

    def insert_taylor_expansion(self):
        tex = 'e^x ' + '= 1 ' \
            + '+ \\frac{x}{1!} + \\frac{x^2}{2!} + \\frac{x^3}{3!} + …, ' \
            + '-∞<x<∞'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_pythagorean_theorem(self):
        tex = 'a^2 + b^2 = c^2'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_fourier_series(self):
        tex = 'f(x) ' + '= a_0 ' + '+ \\sum_{n=1}^{∞}{' \
            + '(a_n\\cos{\\frac{nπx}{L}} + b_n\\sin{\\frac{nπx}{L}})' \
            + '}'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_area_of_circle(self):
        tex = 'A = π r^2'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_trigonometric_formulas_1(self):
        tex = '\\sin{α} ± \\sin{β} ' \
            + '= 2 \\sin{\\frac{1}{2}({α±β})} \\cos{\\frac{1}{2}({α∓β})}'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_trigonometric_formulas_2(self):
        tex = '\\cos{α} + \\cos{β} ' \
            + '= 2 \\cos{\\frac{1}{2}({α+β})} \\cos{\\frac{1}{2}({α-β})}'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_binomial_theorem(self):
        tex = '{(x + a)}^{n} ' \
            + '= \\sum_{k=0}^{n}{_{n}\\mathrm{C}_{k} a^{k} x^{n-k}}'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_quadratic_formula(self):
        tex = 'x = \\frac{-b ± \\sqrt{b^2 - 4ac}}{2a}'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    def insert_expansion_of_addition(self):
        tex = '(1 + x)^n = 1 + \\frac{n x}{1!} + \\frac{n(n - 1) x^2}{2!} + …'
        self._insert_paragraph_text('X=+0.5\n\\[\n' + tex + '\n\\]')

    ################
    # COMMAND

    def set_chapter_number(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        t = 'チャプター番号を変更'
        m = '変更後のチャプター番号を入力してください．'
        confs = [['第', '１', '編', '第', '', '編'],
                 ['第', '１', '章', '第', '', '章'],
                 ['第', '１', '節', '第', '', '節'],
                 ['第', '１', '款', '第', '', '款'],
                 ['第', '１', '目', '第', '', '目']]
        while True:
            must_break = True
            rd = NumberRevisersDialog(mother, self, t, m, confs)
            val = rd.get_values()
            if val == []:
                return False
            for i in range(5):
                if val[i] == '' or c2n_n_arab(val[i]) >= 0:
                    confs[i][3] = val[i]
                else:
                    must_break = False
            if must_break:
                break
        revs = ''
        for i in range(5):
            if val[i] != '':
                n = c2n_n_arab(val[i])
                if n > 0:
                    revs += ' ' + ('$' * (i + 1)) + '=' + str(n)
        revs = re.sub('^\\s+', '', revs)
        if revs != '':
            self._insert_number_revisers(pane, revs)
        return True

    def set_section_number(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        t = 'セクション番号を変更'
        m = '変更後のセクション番号を入力してください．'
        confs = [['第', '１', '',   '第', '', ''],
                 ['',   '１', '',   '',   '', ''],
                 ['（', '１', '）', '（', '', '）'],
                 ['',   'ア', '',   '',   '', ''],
                 ['（', 'ア', '）', '（', '', '）'],
                 ['',   'ａ', '',   '',   '', ''],
                 ['（', 'ａ', '）', '（', '', '）']]
        while True:
            must_break = True
            rd = NumberRevisersDialog(mother, self, t, m, confs)
            val = rd.get_values()
            if val == []:
                return False
            for i in range(3):
                if val[i] == '' or c2n_n_arab(val[i]) >= 0:
                    confs[i][3] = val[i]
                else:
                    must_break = False
            for i in range(2):
                if val[i + 3] == '' or c2n_n_kata(val[i + 3]) >= 0:
                    confs[i + 3][3] = val[i + 3]
                else:
                    must_break = False
            for i in range(2):
                if val[i + 5] == '' or c2n_n_alph(val[i + 5]) >= 0:
                    confs[i + 5][3] = val[i + 5]
                else:
                    must_break = False
            if must_break:
                break
        revs = ''
        for i in range(7):
            if val[i] != '':
                if i <= 2:
                    n = c2n_n_arab(val[i])
                elif i <= 4:
                    n = c2n_n_kata(val[i])
                else:
                    n = c2n_n_alph(val[i])
                if n > 0:
                    revs += ' ' + ('#' * (i + 1)) + '=' + str(n)
        revs = re.sub('^\\s+', '', revs)
        if revs != '':
            self._insert_number_revisers(pane, revs)

    def set_list_number(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        t = 'リスト番号を変更'
        m = '変更後のリスト番号を入力してください．\n' + \
            '丸囲みは不要ですので、"ア"や"ａ"をそのまま入力してください．'
        confs = [['', '①', '', '（', '', '）'],
                 ['', '㋐', '', '（', '', '）'],
                 ['', 'ⓐ', '', '（', '', '）'],
                 ['', '㊀', '', '（', '', '）']]
        while True:
            must_break = True
            rd = NumberRevisersDialog(mother, self, t, m, confs)
            val = rd.get_values()
            if val == []:
                return False
            if val[0] == '' or c2n_n_arab(val[0]) >= 0:
                confs[0][3] = val[0]
            else:
                must_break = False
            if val[1] == '' or c2n_n_kata(val[1]) >= 0:
                confs[1][3] = val[1]
            else:
                must_break = False
            if val[2] == '' or c2n_n_alph(val[2]) >= 0:
                confs[2][3] = val[2]
            else:
                must_break = False
            if val[3] == '' or c2n_n_kanj(val[3]) >= 0:
                confs[3][3] = val[3]
            else:
                must_break = False
            if must_break:
                break
        revs = ''
        for i in range(4):
            if val[i] != '':
                if i == 0:
                    n = c2n_n_arab(val[0])
                elif i == 1:
                    n = c2n_n_kata(val[0])
                elif i == 2:
                    n = c2n_n_alph(val[0])
                elif i == 3:
                    n = c2n_n_kanj(val[0])
                if n > 0:
                    revs += (' ' * 2 * i) + '1.=' + str(n) + '\n'
        revs = re.sub('^\\s+', '', revs)
        if revs != '':
            self._insert_number_revisers(pane, revs)
        return True

    def _insert_number_revisers(self, pane, revisers):
        doc = pane.get('1.0', 'insert')
        res = '^(' \
            + '((.|\n)*\n\n)?' \
            + '(((v|V|X|<<|<|>)=[-\\+]?[0-9]+\\s*)*\n)?' \
            + ')(.|\n)*$'
        doc = re.sub(res, '\\1', doc)
        pane.insert('1.0+' + str(len(doc)) + 'c', revisers + '\n')

    def tidy_up_paragraph(self) -> bool:
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        pane['autoseparators'] = False
        pane.edit_separator()
        pre_text, bare_par, pos_text = self.get_bare_paragraph(pane)
        par_class = self.get_par_class(bare_par)
        if par_class == 'chapter':
            self.tidy_up_chapter(pane, pre_text, bare_par, pos_text)
        elif par_class == 'section':
            self.tidy_up_section(pane, pre_text, bare_par, pos_text)
        elif par_class == 'list':
            self.tidy_up_list(pane, pre_text, bare_par, pos_text)
        elif par_class == 'alignment':
            self.tidy_up_alignment(pane, pre_text, bare_par, pos_text)
        elif par_class == 'table':
            good_table = self.tidy_up_table(pane, pre_text, bare_par, pos_text)
            if not good_table:
                return True
        _, tidied_par, _ = self.get_bare_paragraph(pane)
        if bare_par == tidied_par:
            return False
        pre, par, pos = self.get_paragraph(pane)
        pre_n, par_n = pre.count('\n'), par.count('\n')
        for i in range(pre_n, (pre_n + par_n)):
            if self.current_pane == 'txt':
                self.paint_out_line(i)
        pane['autoseparators'] = True
        pane.edit_separator()
        return True

    def get_bare_paragraph(self, pane) -> (str):
        pre_pars, cur_par, pos_pars = self.get_paragraph(pane)
        pre, par, pos = pre_pars, cur_par, pos_pars
        res_sp = '(\\s+)'
        res_cm = '(<!--(?:.|\n)*?-->)'
        res_ln = '((?:v|V|x|X|<<|<|>)=[\\-\\+]?(?:[0-9]+\\.)?[0-9]+)'
        res_f1 = '(\\*{1,3}|~~|`|//|\\-{2,3}|\\+{2,3}|>{2,3}|<{2,3})'
        res_f2 = '(\\^[0-9A-Za-z]{0,11}\\^|_[0-9A-Za-z]{0,11}_)'
        res_f3 = '(_[\\$=\\.#\\-~\\+]{,4}_)'
        res_f4 = '(@(?:[0-9]*\\.)?[0-9]+@|@[^@]{1,66}@)'
        # "res_ln(<<=n)" should be before "res_f1(<<)".
        res_head = [res_sp, res_cm, res_ln, res_f1, res_f2, res_f3, res_f4]
        res_tail = [res_sp, res_cm, res_f1, res_f2, res_f3, res_f4]
        tmp = ''
        while par != tmp:
            tmp = par
            for res_xx in res_head:
                res = '^' + res_xx + '((?:.|\n)*)$'
                if re.match(res, par):
                    pre = pre + re.sub(res, '\\1', par)
                    par = re.sub(res, '\\2', par)
            for res_xx in res_tail:
                res = '^((?:.|\n)*?)' + res_xx + '$'
                if re.match(res, par):
                    pos = re.sub(res, '\\2', par) + pos
                    par = re.sub(res, '\\1', par)
        res = '^((?:.|\n)*\n)?([ \t\u3000]+)$'
        if re.match(res, pre):
            t = re.sub(res, '\\1', pre)
            s = re.sub(res, '\\2', pre)
            pre, par = t, s + par
        res = '^([ \t\u3000]*\n)((?:.|\n)*)?$'
        if re.match(res, pos):
            s = re.sub(res, '\\1', pos)
            t = re.sub(res, '\\2', pos)
            par, pos = par + s, t
        pre_text, bare_par, pos_text = pre, par, pos
        return pre_text, bare_par, pos_text

    @staticmethod
    def get_paragraph(pane) -> (str):
        pre = pane.get('1.0', 'insert')
        res = '^((?:.|\n)*\n\n)((?:.|\n)*)$'
        if re.match(res, pre):
            pre_pars = re.sub(res, '\\1', pre)
            par_head = re.sub(res, '\\2', pre)
        else:
            pre_pars = ''
            par_head = pre
        pos = pane.get('insert', 'end-1c')
        res = '^((?:.|\n)*?\n)(\n(?:.|\n)*)$'
        if re.match(res, pos):
            par_tail = re.sub(res, '\\1', pos)
            pos_pars = re.sub(res, '\\2', pos)
        else:
            par_tail = pos
            pos_pars = ''
        cur_par = par_head + par_tail
        return pre_pars, cur_par, pos_pars

    @staticmethod
    def get_par_class(par) -> str:
        par = re.sub('<!--(.|\n)*?-->', '', par)
        par = re.sub('^\\s+', '', par)
        par = re.sub('\\s+$', '', par)
        par = re.sub('\n\\s+', '\n', par)
        par = re.sub('\\s+\n', '\n', par)
        full_text = par.replace('\n', ' ')
        # CHAPTER
        if re.match('^\\$+(-\\$+)*\\s.*$', full_text):
            return 'chapter'
        # SECTION
        if re.match('^#+(-#+)*\\s.*$', full_text):
            return 'section'
        # LIST
        if re.match('^([0-9]+\\.|-)\\s.*$', full_text):
            return 'list'
        # TABLE (Should be before ALIGNMENT)
        if re.match('^(: )?\\s*\\|.*\\|(:?-*:?(\n?(\\^+|=+))?)?( :)?$',
                    full_text):
            return 'table'
        # ALIGNMENT
        if re.match('^:\\s.*$', full_text) or re.match('^.*\\s:$', full_text):
            return 'alignment'
        return None

    def tidy_up_chapter(self, pane, pre_text, bare_par, pos_text):
        self._tidy_up_general_par(pane, pre_text, bare_par, pos_text,
                                  '\\s*\\$+[\\-|\\s]')

    def tidy_up_section(self, pane, pre_text, bare_par, pos_text):
        self._tidy_up_general_par(pane, pre_text, bare_par, pos_text,
                                  '\\s*#+[\\-|\\s]')

    def tidy_up_list(self, pane, pre_text, bare_par, pos_text):
        self._tidy_up_general_par(pane, pre_text, bare_par, pos_text,
                                  '\\s*(?:[0-9]+\\.|-)\\s')

    def tidy_up_alignment(self, pane, pre_text, bare_par, pos_text):
        self._tidy_up_general_par(pane, pre_text, bare_par, pos_text,
                                  '\\s*:\\s')

    def _tidy_up_general_par(self, pane, pre_text, bare_par, pos_text,
                             res_indent):
        text, par = pre_text, bare_par
        # PRE SPACE
        res = '^(\\s+)((?:.|\n)*)$'
        if re.match(res, par):
            spc = re.sub(res, '\\1', par)
            bdy = re.sub(res, '\\2', par)
            new = spc.replace('\t', (' ' * 4)).replace('\u3000', (' ' * 2))
            spc = self._replace_spaces(pane, text, spc, new)
            par = spc + bdy
        # GLUE SPACE
        res = '^(\\s*\\S+)(\\s+)((?:.|\n)*)$'
        if re.match(res, par):
            sym = re.sub(res, '\\1', par)
            spc = re.sub(res, '\\2', par)
            tit = re.sub(res, '\\3', par)
            spc = self._replace_spaces(pane, text + sym, spc, ' ')
            par = sym + spc + tit
        # INDENT WIDTH
        head_string = re.sub('^(' + res_indent + ')(.|\n)*$', '\\1', par)
        indent = ' ' * len(head_string)
        # SENTENCE SECTION
        if '#' in res_indent:
            if re.match('^(.|\n)*[.．。]\\s*$', par):
                res = '^(.*\\S+? \\\\?)((.|\n)*)$'
                sym = re.sub(res, '\\1', par)
                tmp = re.sub(res, '\\2', par)
                if not re.match('^.*\\\\$', sym):
                    beg = '1.0+' + str(len(text + sym)) + 'c'
                    pane.insert(beg, '\\')
                    sym = sym + '\\'
                if not re.match('^\n', tmp):
                    beg = '1.0+' + str(len(text + sym)) + 'c'
                    pane.insert(beg, '\n')
                    tmp = '\n' + tmp
                par = sym + tmp
                head_string = re.sub('^(\\s*)(.|\n)*$', '\\1', par)
                indent = ' ' * len(head_string)
        # TIDY UP
        lines = par.split('\n')
        for i, line in enumerate(lines):
            lin = line
            res = '^(\\s*)((?:.|\n)*)$'
            if i > 0:
                spc = re.sub(res, '\\1', lin)
                bdy = re.sub(res, '\\2', lin)
                spc = self._replace_spaces(pane, text, spc, indent)
                lin = spc + bdy
            res = '^(.*?)(\\s+)$'
            if re.match(res, lin):
                spc = re.sub(res, '\\2', lin)
                bdy = re.sub(res, '\\1', lin)
                spc = self._replace_spaces(pane, text + bdy, spc, '')
                lin = bdy + spc
            text += lin + '\n'

    def tidy_up_table(self, pane, pre_text, bare_par, pos_text):
        table = self._get_table(bare_par)
        should_tidy_up = self.get_should_tidy_up(table)
        if not should_tidy_up:
            return False
        table = self._prepare_table(pane, pre_text, table, pos_text)
        cr_number = self._get_conf_row_number(table)
        alignment = self._get_alignment_of_cell(cr_number, table)
        widths, borders = self._get_cell_widths_and_borders(cr_number, table)
        self._tidy_up_table(pane, pre_text, bare_par, pos_text,
                            alignment, widths, borders, table)
        return True

    @staticmethod
    def _get_table(bare_par: str) -> [[str]]:
        indent = ''
        res = '^(\\s*:)(\\s.*)$'
        for line in bare_par.split('\n'):
            if re.match(res, line):
                s = re.sub(res, '\\1', line)
                s = s.replace('\t', (' ' * 4))
                s = s.replace('\u3000', (' ' * 2))
                indent = '\\s{,' + str(len(s)) + '}'
        table = []
        row = []
        cell = ''
        for i, c in enumerate(bare_par + '\0'):
            if c == '|':
                if re.match('^.*\\\\$', cell) and \
                   not re.match(NOT_ESCAPED + '\\|$', cell + c):
                    # "..\|..."
                    cell += c
                    continue
                elif len(row) == 0:
                    # "^: |..."
                    row.append(cell + c)
                else:
                    # "...|..."
                    row.append(cell)
                    row.append(c)
                cell = ''
            elif c == '\n':
                if re.match('^(\\^|=)+$', cell) and \
                   len(row) == 0 and \
                   len(table) > 0 and len(table[-1]) > 0:
                    # "|...|\n^^^^^\n" or "|...|\n=====\n"
                    table[-1][-1] += cell + c
                    cell = ''
                elif (i < len(bare_par) - 1 and
                      re.match('^\\s*(\\|[:-]-*:?(\\^|=)?)+\n?(\\|(\\s*:)?)',
                               bare_par[i + 1:]) and
                      len(table) > 0 and len(table[-1]) > 0):
                    # "|:--\n|:--"
                    cell += c
                elif (i < len(bare_par) - 1 and
                      not re.match('^' + indent + '\\|', bare_par[i + 1:]) and
                      not re.match('^\\s*:\\s+\\|', bare_par[i + 1:])):
                    # "|...\n...|"
                    cell += c
                else:
                    cell += c
                    if len(row) > 0:
                        # "|...|\n|...|"
                        row[-1] += cell
                    else:
                        # ERROR PREVENTION
                        row.append(cell)
                    table.append(row)
                    row = []
                    cell = ''
            elif c == '\0':
                if row != []:
                    table.append(row)
                    row = []
            else:
                cell += c
        return table

    @staticmethod
    def get_should_tidy_up(table):
        m = len(table[0])
        for row in table:
            if len(row) != m:
                n = '警告'
                m = '各行のセルの数が不均一です．\n' \
                    + '本当に表を整形しますか？'
                if tkinter.messagebox.askyesno(n, m, default='no'):
                    break
                else:
                    return False
        return True

    def _prepare_table(self, pane, pre_text, table, pos_text):
        m = 0
        for row in table:
            if m < len(row):
                m = len(row)
        text = pre_text
        for row in table:
            if len(row) < m:
                row[-1] = re.sub('\n$', '', row[-1])
            for j in range(m):
                if j < len(row):
                    cell = row[j]
                else:
                    cell = ''
                    if (j % 2) == 0:
                        cell += '|'
                        beg = '1.0+' + str(len(text)) + 'c'
                        pane.insert(beg, cell)
                    if j == m - 1:
                        cell += '\n'
                    row.append(cell)
                if j == 0:
                    res = '^(\\s*:)(\\s+)(\\|)$'
                    if re.match(res, cell):
                        sym = re.sub(res, '\\1', cell)
                        spc = re.sub(res, '\\2', cell)
                        vln = re.sub(res, '\\3', cell)
                        spc = self._replace_spaces(pane, text + sym, spc, ' ')
                        cell = sym + spc + vln
                        row[j] = cell
                if j == m - 1:
                    res = '^(\\|)(\\s+)(:\\s*\n)$'
                    if re.match(res, cell):
                        vln = re.sub(res, '\\1', cell)
                        spc = re.sub(res, '\\2', cell)
                        sym = re.sub(res, '\\3', cell)
                        spc = self._replace_spaces(pane, text + vln, spc, ' ')
                        cell = vln + spc + sym
                        row[j] = cell
                text += cell
        return table

    @staticmethod
    def _get_conf_row_number(table: [[str]]) -> int:
        for i, row in enumerate(table):
            m = len(row) - 1
            for j, cell in enumerate(row):
                if (j % 2) == 0:
                    pass
                elif re.match('^:?-*:?(\\^|=)?(\\s*\\\\?\n\\s*)?$', cell):
                    pass
                else:
                    break
            else:
                return i
        return -1

    @staticmethod
    def _get_alignment_of_cell(conf_row_number, table):
        b_row = []
        for j, cell in enumerate(table[conf_row_number]):
            if (j % 2) == 0:
                continue
            if re.match('^:-*:(\\^|=)?$', cell):
                b_row.append('c')
            elif re.match('^-*:(\\^|=)?$', cell):
                b_row.append('r')
            else:
                b_row.append('l')
        a_tab = []
        for i, row in enumerate(table):
            a_row = []
            for j, cell in enumerate(row):
                if (j % 2) == 0:
                    continue
                k = int(j / 2)
                if i < conf_row_number:
                    a_row.append('c')
                else:
                    a_row.append(b_row[k])
                if re.match('^:\\s.*\\s:(\\s*@([0-9]*x)?[0-9]+)?$', cell):
                    a_row[k] = 'c'
                elif re.match('^.*\\s:(\\s*@([0-9]*x)?[0-9]+)?$', cell):
                    a_row[k] = 'r'
                elif re.match('^:\\s.*$', cell):
                    a_row[k] = 'l'
            a_tab.append(a_row)
        alignment = a_tab
        return alignment

    @staticmethod
    def _get_cell_widths_and_borders(conf_row_number, table):
        cell_widths = []
        for row in table:
            for j, cell in enumerate(row):
                if (j + 1) > len(cell_widths):
                    cell_widths.append(0)
                c = re.sub('\\s*\\\\?\n\\s*', '', cell)
                w = get_real_width(c)
                if cell_widths[j] < w:
                    cell_widths[j] = w
        if conf_row_number >= 0:
            for j, cell in enumerate(table[conf_row_number]):
                if cell != '':
                    c = re.sub('\\s*\\\\?\n\\s*', '', cell)
                    w = get_real_width(c)
                    cell_widths[j] = w
        right_borders = []
        for i in range(len(cell_widths)):
            if i == 0:
                b = 0
            else:
                b = right_borders[-1]
            if i < len(table[conf_row_number]) and \
               '\n' in table[conf_row_number][i]:
                s = re.sub('^((.|\n)*\n)', '', table[conf_row_number][i])
                b = len(s)
            else:
                b += cell_widths[i]
            right_borders.append(b)
        return cell_widths, right_borders

    def _tidy_up_table(self, pane, pre_text, bare_par, pos_text,
                       alignment, widths, borders, table):
        text = pre_text
        res = '^(\\s*)(.*?)(\\s*)$'
        for i, row in enumerate(table):
            m = len(row) - 1
            r_width = 0
            for j, cell in enumerate(row):
                c = cell
                r_width += widths[j]
                k = int(j / 2)
                if j == 0:
                    # BEGINNING OF A ROW
                    res = '^(\\s*)((?::\\s+)?\\|)$'
                    if re.match(res, c):
                        spc = re.sub(res, '\\1', c)
                        sym = re.sub(res, '\\2', c)
                        w = r_width - get_real_width(sym)
                        spc = self._replace_spaces(pane, text, spc, ' ' * w)
                        c = spc + sym
                elif j < m:
                    if (j % 2) == 0:
                        pass  # cell = '|'
                    else:
                        b0 = borders[j - 1] if j > 0 else 0
                        b1 = borders[j]
                        bdrs = (b0, b1)
                        if '\n' in c:
                            c = self._tupt_linebreaks(pane, text, c, bdrs)
                        elif alignment[i][k] == 'c':
                            c = self._tupt_center(pane, text, c, r_width)
                        elif alignment[i][k] == 'r':
                            c = self._tupt_right(pane, text, c, r_width)
                        else:
                            c = self._tupt_left(pane, text, c, r_width)
                else:
                    # END OF A ROW
                    res = '^(\\|(?:\\s+:)?)(\\s+)(\n)$'
                    if re.match(res, c):
                        sym = re.sub(res, '\\1', c)
                        spc = re.sub(res, '\\2', c)
                        nln = re.sub(res, '\\3', c)
                        spc = self._replace_spaces(pane, text + sym, spc, '')
                        c = sym + spc + nln
                r_width -= get_real_width(c)
                if '\n' in c:
                    r_width = 0
                text += c

    def _tupt_linebreaks(self, pane, text, cell, borders):
        # RIGHT SPACES
        res_nl = '^((?:.|\n)*?)(\\s*\n\\s*)$'
        res_nn = '^((?:.|\n)*\n\\s*)(.*?)(\\s*)(:?)$'
        rgt_spc = ''
        if re.match(res_nl, cell):
            bdy = re.sub(res_nl, '\\1', cell)
            spc = re.sub(res_nl, '\\2', cell)
            new_spc = '\n' + (' ' * borders[1])
            spc = self._replace_spaces(pane, text + bdy, spc, new_spc)
            cell, rgt_spc = bdy, spc
        elif re.match(res_nn, cell):
            bdy = re.sub(res_nn, '\\1', cell)
            lin = re.sub(res_nn, '\\2', cell)
            spc = re.sub(res_nn, '\\3', cell)
            sym = re.sub(res_nn, '\\4', cell)
            if sym != ':':
                new_spc \
                    = ' ' * (borders[1] - borders[0] - get_real_width(lin))
            else:
                new_spc \
                    = ' ' * (borders[1] - borders[0] - get_real_width(lin) - 2)
                new_spc += ' '
            spc = self._replace_spaces(pane, text + bdy + lin, spc, new_spc)
            cell = bdy + lin + spc + sym
        # CENTER SPACES
        if '\n' in cell:
            bdy = cell
            new_bdy = ''
            res = '^((?:.|\n)*)(\\s*\n\\s*)(.*\\S.*)$'
            while '\n' in bdy:
                pos = re.sub(res, '\\3', bdy)
                spc = re.sub(res, '\\2', bdy)
                bdy = re.sub(res, '\\1', bdy)
                new_spc = '\n' + (' ' * borders[0])
                spc = self._replace_spaces(pane, text + bdy, spc, new_spc)
                new_bdy = new_spc + pos + new_bdy
            cell = bdy + new_bdy
        # LEFT SPACE
        res = '^(:?)(\\s+)((?:.|\n)*)$'
        if re.match(res, cell):
            sym = re.sub(res, '\\1', cell)
            spc = re.sub(res, '\\2', cell)
            bdy = re.sub(res, '\\3', cell)
            new_spc = ' ' if sym == ':' else ''
            spc = self._replace_spaces(pane, text + sym, spc, new_spc)
            cell = sym + spc + bdy
        cell += rgt_spc
        return cell

    def _tupt_left(self, pane, text, cell, width):
        if not re.match('^:\\s.*$', cell):
            res = '^()(\\s*)(.*?)(\\s*)((?:\\s@[0-9]*(?:x[0-9]+)?)?)$'
        else:
            res = '^(:\\s)(\\s*)(.*?)(\\s*)((?:\\s@[0-9]*(?:x[0-9]+)?)?)$'
        cell = self._tupt_basis(pane, text, cell, res, 'l', width)
        return cell

    def _tupt_center(self, pane, text, cell, width):
        if not re.match('^:\\s.*\\s:(@[0-9]*(x[0-9]+)?)?$', cell):
            res = '^()(\\s*)(.*?)(\\s*)((?:\\s@[0-9]*(?:x[0-9]+)?)?)$'
        else:
            res = '^(:\\s)(\\s*)(.*?)(\\s*)(\\s:(?:@[0-9]*(?:x[0-9]+)?)?)$'
        cell = self._tupt_basis(pane, text, cell, res, 'c', width)
        return cell

    def _tupt_right(self, pane, text, cell, width):
        if not re.match('^.*\\s:(@[0-9]*(x[0-9]+)?)?$', cell):
            res = '^()(\\s*)(.*?)(\\s*)((?:\\s@[0-9]*(?:x[0-9]+)?)?)$'
        else:
            res = '^()(\\s*)(.*?)(\\s*)(\\s:(?:@[0-9]*(?:x[0-9]+)?)?)$'
        cell = self._tupt_basis(pane, text, cell, res, 'r', width)
        return cell

    def _tupt_basis(self, pane, text, cell, res, alignment, width):
        symbol_l, spaces_l_from, body, spaces_r_from, symbol_r \
            = self._split_cell(res, cell)
        w = width - len(symbol_l) - get_real_width(body) - len(symbol_r)
        if alignment == 'r':
            spaces_l_to, spaces_r_to = (' ' * w), ('')
        elif alignment == 'c':
            w_r = int(w / 2)
            w_l = w - w_r
            spaces_l_to, spaces_r_to = (' ' * w_l), (' ' * w_r)
        else:
            spaces_l_to, spaces_r_to = (''), (' ' * w)
        if cell != (symbol_l + spaces_l_to + body + spaces_r_to + symbol_r):
            t = text + symbol_l + spaces_l_from + body
            sp_r = self._replace_spaces(pane, t, spaces_r_from, spaces_r_to)
            t = text + symbol_l
            sp_l = self._replace_spaces(pane, t, spaces_l_from, spaces_l_to)
            cell = symbol_l + sp_l + body + sp_r + symbol_r
        return cell

    @staticmethod
    def _split_cell(res, cell):
        sy_l = re.sub(res, '\\1', cell)
        sp_l = re.sub(res, '\\2', cell)
        body = re.sub(res, '\\3', cell)
        sp_r = re.sub(res, '\\4', cell)
        sy_r = re.sub(res, '\\5', cell)
        return sy_l, sp_l, body, sp_r, sy_r

    @staticmethod
    def _replace_spaces(pane, text, spaces_from, spaces_to):
        if spaces_from != spaces_to:
            beg = '1.0+' + str(len(text)) + 'c'
            end = '1.0+' + str(len(text + spaces_from)) + 'c'
            spaces_tmp = pane.get(beg, end)
            if re.match('^\\s*$', spaces_tmp):
                pane.delete(beg, end)
                pane.insert(beg, spaces_to)
                return spaces_to
        return spaces_from

    ##########################
    # MENU MOVE

    def _make_menu_move(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='移動(M)', menu=menu, underline=3)
        #
        menu.add_command(label='文頭に移動',
                         command=self.goto_beg_of_doc)
        menu.add_command(label='文末に移動',
                         command=self.goto_end_of_doc)
        menu.add_command(label='行頭に移動',
                         command=self.goto_beg_of_line)
        menu.add_command(label='行末に移動',
                         command=self.goto_end_of_line)
        menu.add_separator()
        #
        menu.add_command(label='前を検索',
                         command=self.search_backward)
        menu.add_command(label='後を検索',
                         command=self.search_forward, accelerator='Ctrl+F')
        menu.add_separator()
        #
        self._make_submenu_place_flag(menu)
        self._make_submenu_goto_flag(menu)
        menu.add_separator()
        #
        menu.add_command(label='サブカーソルに移動',
                         command=self.goto_sub_cursor)
        menu.add_separator()
        #
        menu.add_command(label='行数と文字数を指定して移動',
                         command=self.goto_by_position)
        # menu.add_separator()

    ################
    # COMMAND

    def goto_beg_of_doc(self):
        self.txt.mark_set('insert', '1.0')
        self._put_back_cursor_to_pane(self.txt)

    def goto_end_of_doc(self):
        self.txt.mark_set('insert', 'end-1c')
        self._put_back_cursor_to_pane(self.txt)

    def goto_beg_of_line(self):
        self.txt.mark_set('insert', 'insert linestart')
        self._put_back_cursor_to_pane(self.txt)

    def goto_end_of_line(self):
        self.txt.mark_set('insert', 'insert lineend')
        self._put_back_cursor_to_pane(self.txt)

    def search_backward(self):
        word1 = self.stb_sor1.get()
        if word1 == '':
            pane = self.txt
            if self.current_pane == 'sub':
                pane = self.sub
            self.search_backward_from_dialog(pane)
        else:
            self.search_or_replace_backward(False)  # must_replace = False

    def search_forward(self):
        word1 = self.stb_sor1.get()
        if word1 == '':
            pane = self.txt
            if self.current_pane == 'sub':
                pane = self.sub
            self.search_forward_from_dialog(pane)
        else:
            self.search_or_replace_forward(False)   # must_replace = False

    ################
    # SUBMENU PLACE FLAG

    def _make_submenu_place_flag(self, menu):
        submenu = tkinter.Menu(self.mnb, tearoff=False)
        menu.add_cascade(label='フラグを設置', menu=submenu)
        #
        submenu.add_command(label='フラグ1を設置',
                            command=self.place_flag1)
        submenu.add_command(label='フラグ2を設置',
                            command=self.place_flag2)
        submenu.add_command(label='フラグ3を設置',
                            command=self.place_flag3)
        submenu.add_command(label='フラグ4を設置',
                            command=self.place_flag4)
        submenu.add_command(label='フラグ5を設置',
                            command=self.place_flag5)

    #######
    # COMMAND

    def place_flag1(self):
        self._place_flag(1)

    def place_flag2(self):
        self._place_flag(2)

    def place_flag3(self):
        self._place_flag(3)

    def place_flag4(self):
        self._place_flag(4)

    def place_flag5(self):
        self._place_flag(5)

    def _place_flag(self, n):
        self.txt.mark_set('flag' + str(n), 'insert')
        self.txt.mark_gravity('flag' + str(n), 'left')

    ################
    # SUBMENU GOTO FLAG

    def _make_submenu_goto_flag(self, menu):
        submenu = tkinter.Menu(self.mnb, tearoff=False)
        menu.add_cascade(label='フラグに移動', menu=submenu)
        #
        submenu.add_command(label='フラグ1に移動',
                            command=self.goto_flag1)
        submenu.add_command(label='フラグ2に移動',
                            command=self.goto_flag2)
        submenu.add_command(label='フラグ3に移動',
                            command=self.goto_flag3)
        submenu.add_command(label='フラグ4に移動',
                            command=self.goto_flag4)
        submenu.add_command(label='フラグ5に移動',
                            command=self.goto_flag5)

    #######
    # COMMAND

    def goto_flag1(self):
        self._goto_flag(1)

    def goto_flag2(self):
        self._goto_flag(2)

    def goto_flag3(self):
        self._goto_flag(3)

    def goto_flag4(self):
        self._goto_flag(4)

    def goto_flag5(self):
        self._goto_flag(5)

    def _goto_flag(self, n) -> bool:
        if 'flag' + str(n) not in self.txt.mark_names():
            t, m = 'エラー', 'フラグ' + str(n) + 'は設定されていません．'
            tkinter.messagebox.showerror(t, m)
            return False
        self.txt.mark_set('insert', 'flag' + str(n))
        self._put_back_cursor_to_pane(self.txt)
        return True

    ################
    # COMMAND

    def goto_sub_cursor(self):
        m = self.txt.index('insert')
        s = '1.0'
        if 'sub_insert' in self.txt.mark_names():
            s = self.txt.index('sub_insert')
        self.txt.mark_set('insert', s)
        self._put_back_cursor_to_pane(self.txt)
        self.txt.mark_set('sub_insert', m)
        self.txt.mark_gravity('sub_insert', 'left')

    def goto_by_position(self, father=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if father is None:
            father = pane
        self.PositionDialog(pane, father, self)

    class PositionDialog(tkinter.simpledialog.Dialog):

        def __init__(self, pane, father, mother):
            self.pane = pane      # "self.txt" or "self.sub"
            self.father = father  # parent object
            self.mother = mother  # Makdo()
            super().__init__(father, title='行数と文字数を指定して移動')

        def body(self, father):
            i = self.pane.index('insert')
            v = int(re.sub('\\.[0-9]+$', '', i))
            h = int(re.sub('^[0-9]+\\.', '', i)) + 1
            fon = self.mother.gothic_font
            t = '移動先の行数と文字数を入力してください．\n'
            self.text1 = tkinter.Label(father, text=t)
            self.text1.pack(side='top', anchor='w')
            self.frame = tkinter.Frame(father)
            self.frame.pack(side='top')
            self.entry1 = tkinter.Entry(self.frame, width=5, justify='center',
                                        font=fon)
            self.entry1.pack(side='left')
            self.entry1.insert(0, str(v))
            tkinter.Label(self.frame, text='行目の').pack(side='left')
            self.entry2 = tkinter.Entry(self.frame, width=5, justify='center',
                                        font=fon)
            self.entry2.pack(side='left')
            self.entry2.insert(0, str(h))
            tkinter.Label(self.frame, text='文字目').pack(side='left')
            # self.bind('<Key-Return>', self.ok)
            # self.bind('<Key-Escape>', self.cancel)
            # super().body(father)
            return self.entry1

        def apply(self):
            line = self.entry1.get()
            char = self.entry2.get()
            if re.match('^[0-9]+$', line) and re.match('^[0-9]+$', char):
                il, ic = int(line), int(char)
                if il == 0:
                    il = 1
                if ic == 0:
                    ic = 1
                ic -= 1
                self.pane.mark_set('insert', str(il) + '.' + str(ic))
                self.mother._put_back_cursor_to_pane(self.pane)
                self.pane.focus_force()

    ##########################
    # MENU TOOL

    def _make_menu_tool(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='ツール(T)', menu=menu, underline=4)
        #
        menu.add_command(label='文字数と行数を確認',
                         command=self.count_chars)
        menu.add_separator()
        #
        menu.add_command(label='定型句を挿入',
                         command=self.insert_formula)
        menu.add_command(label='定型句を編集',
                         command=self.edit_formula)
        menu.add_separator()
        #
        menu.add_command(label='メモ帳を開く',
                         command=self.open_memo_pad)
        menu.add_separator()
        #
        menu.add_command(label='画面を二つに分割',
                         command=self.split_window)
        menu.add_command(label='二つの画面を入替え',
                         command=self.swap_windows)
        menu.add_separator()
        #
        menu.add_command(label='別ファイルの内容を見る',
                         command=self.show_file)
        menu.add_separator()
        #
        menu.add_command(label='編集前の原稿と比較して元に戻す',
                         command=self.compare_with_previous_draft)
        menu.add_command(label='別ファイルの内容と比較して反映',
                         command=self.compare_files)
        menu.add_separator()
        #
        menu.add_command(label='編集前の原稿との違いを変更履歴にする',
                         command=self.insert_track_change_tag)
        menu.add_command(label='前の変更履歴に移動',
                         command=self.search_previous_track_change_tag)
        menu.add_command(label='次の変更履歴に移動',
                         command=self.search_next_track_change_tag)
        menu.add_command(label='カーソル位置の変更履歴を反映',
                         command=self.accept_current_track_change)
        menu.add_command(label='全ての変更履歴を反映',
                         command=self.accept_all_track_changes)
        menu.add_separator()
        #
        menu.add_command(label='セクションを折り畳む・展開',
                         command=self.fold_or_unfold_section)
        menu.add_command(label='セクションを全て展開',
                         command=self.unfold_section_fully)
        menu.add_separator()
        #
        menu.add_command(label='キーボードマクロを実行',
                         command=self.execute_keyboard_macro,
                         accelerator='Ctrl+E')
        menu.add_separator()
        #
        menu.add_command(label='サブウィンドウを閉じる',
                         command=self._close_sub_pane)
        menu.add_separator()
        #
        menu.add_command(label='コマンドを入力して実行',
                         command=self.start_minibuffer,
                         accelerator='Esc+X')
        # menu.add_separator()

    ################
    # COMMAND

    def count_chars(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if pane.tag_ranges('sel'):
            beg, end = pane.index('sel.first'), pane.index('sel.last')
        elif 'akauni' in pane.mark_names():
            beg, end = self._get_indices_in_order(pane, 'insert', 'akauni')
        else:
            beg, end = '1.0', 'end-1c'
        tex = pane.get(beg, end)
        n_nl = tex.count('\n') + 1
        tex = tex.replace('\n', '')
        n_ch = len(tex)
        t = '文字数と行数'
        m = "{:,}".format(n_ch) + '文字\n' + \
            "{:,}".format(n_nl) + '行'
        tkinter.messagebox.showinfo(t, m)

    # INSERT AND EDIT FORMULA

    def insert_formula(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        t = '定型句を挿入'
        m = '挿入する定型句を選んでください．'
        formulas = self._get_formulas()
        fs = []
        for i, f in enumerate(formulas):
            f = re.sub('タイトル:\\s*', '', f)
            f = re.sub('\n(.|\n)*$', '', f)
            if f == '':
                f = '（空）'
            fs.append(str(i + 1) + '. ' + f)
        rd = RadiobuttonDialog(mother, self, t, m, fs, fs[0])
        v = rd.get_value()
        if v is not None:
            self.formula_number = fs.index(v) + 1
            self._insert_formula()

    @staticmethod
    def _get_formulas():
        formulas = ['' for i in range(9)]
        for i in range(9):
            try:
                path = CONFIG_DIR + '/formula' + str(i + 1) + '.md'
                with open(path, 'r') as f:
                    formulas[i] = f.read()
            except BaseException:
                pass
        return formulas

    def _insert_formula(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        formulas = self._get_formulas()
        n = self.formula_number - 1
        v = formulas[n]
        v = re.sub('^タイトル:\\s*.*\n', '', v)
        pane.edit_separator()
        pane.insert('insert', v)
        p = self._get_v_position_of_insert(pane) - 1
        for i in range(v.count('\n')):
            self.paint_out_line(p + i)
        pane.edit_separator()
        self.formula_number = -1

    def insert_formula1(self):
        self.formula_number = 1
        self._insert_formula()

    def insert_formula2(self):
        self.formula_number = 2
        self._insert_formula()

    def insert_formula3(self):
        self.formula_number = 3
        self._insert_formula()

    def insert_formula4(self):
        self.formula_number = 4
        self._insert_formula()

    def insert_formula5(self):
        self.formula_number = 5
        self._insert_formula()

    def insert_formula6(self):
        self.formula_number = 6
        self._insert_formula()

    def insert_formula7(self):
        self.formula_number = 7
        self._insert_formula()

    def insert_formula8(self):
        self.formula_number = 8
        self._insert_formula()

    def insert_formula9(self):
        self.formula_number = 9
        self._insert_formula()

    def edit_formula(self, mother=None):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        if mother is None:
            mother = pane
        self.quit_editing_formula()
        t = '定型句を編集'
        m = '編集する定型句を選んでください．'
        formulas = self._get_formulas()
        fs = []
        for i, f in enumerate(formulas):
            f = re.sub('タイトル:\\s+', '', f)
            f = re.sub('\n(.|\n)*$', '', f)
            if f == '':
                f = '（空）'
            fs.append(str(i + 1) + '. ' + f)
        rd = RadiobuttonDialog(mother, self, t, m, fs, fs[0])
        v = rd.get_value()
        if v is not None:
            self.formula_number = fs.index(v) + 1
            self._edit_formula()

    def _edit_formula(self):
        n = self.formula_number
        formula_path = CONFIG_DIR + '/formula' + str(n) + '.md'
        if not os.path.exists(formula_path):
            open(formula_path, 'w').close()
        try:
            with open(formula_path, 'r') as f:
                formula = f.read()
        except BaseException:
            return
        #
        if not re.match('^タイトル:\\s*', formula):
            formula = 'タイトル: （タイトル）\n' + formula
        self._open_sub_pane(formula, False)

    def edit_formula1(self):
        self.quit_editing_formula()
        self.formula_number = 1
        self._edit_formula()

    def edit_formula2(self):
        self.quit_editing_formula()
        self.formula_number = 2
        self._edit_formula()

    def edit_formula3(self):
        self.quit_editing_formula()
        self.formula_number = 3
        self._edit_formula()

    def edit_formula4(self):
        self.quit_editing_formula()
        self.formula_number = 4
        self._edit_formula()

    def edit_formula5(self):
        self.quit_editing_formula()
        self.formula_number = 5
        self._edit_formula()

    def edit_formula6(self):
        self.quit_editing_formula()
        self.formula_number = 6
        self._edit_formula()

    def edit_formula7(self):
        self.quit_editing_formula()
        self.formula_number = 7
        self._edit_formula()

    def edit_formula8(self):
        self.quit_editing_formula()
        self.formula_number = 8
        self._edit_formula()

    def edit_formula9(self):
        self.quit_editing_formula()
        self.formula_number = 9
        self._edit_formula()

    def quit_editing_formula(self) -> bool:
        n = self.formula_number
        self.formula_number = -1
        if n > 0:
            formula_path = CONFIG_DIR + '/formula' + str(n) + '.md'
            contents = self.sub.get('1.0', 'end-1c')
            self._save_config_file(formula_path, contents)
            return True
        return False

    # OPEN MEMO PAD

    def open_memo_pad(self):
        if CONFIG_DIR is None:
            return False
        memo_pad_path = CONFIG_DIR + '/memo.md'
        if not os.path.exists(memo_pad_path):
            try:
                open(memo_pad_path, 'w').close()
            except BaseException:
                return False
        if not os.path.exists(memo_pad_path):
            return False
        try:
            with open(memo_pad_path, 'r') as f:
                self.memo_pad_memory = f.read()
        except BaseException:
            return False
        #
        self._open_sub_pane(self.memo_pad_memory, False)

    def update_memo_pad(self):
        memo_pad_memory = self.memo_pad_memory
        if self.memo_pad_memory is None:
            return False
        memo_pad_path = CONFIG_DIR + '/memo.md'
        # DISPLAY
        memo_pad_display = self.sub.get('1.0', 'end-1c')
        if memo_pad_display != memo_pad_memory:
            self.memo_pad_memory = memo_pad_display
            self._save_config_file(memo_pad_path, memo_pad_display)
            return True
        # FILE
        if not os.path.exists(memo_pad_path):
            return False
        try:
            with open(memo_pad_path, 'r') as f:
                memo_pad_file = f.read()
        except BaseException:
            return False
        if memo_pad_file != memo_pad_memory:
            # MEMORY
            self.memo_pad_memory = memo_pad_file
            # DISPLAY
            self.sub.delete('1.0', 'end')
            self.sub.insert('1.0', memo_pad_file)

    def close_memo_pad(self):
        if self.memo_pad_memory is not None:
            self.update_memo_pad()
            self.memo_pad_memory = None

    # SPLIT WINDOW

    def split_window(self) -> bool:
        if len(self.pnd_r.panes()) > 1:
            return False
        self._close_sub_pane()
        document = self.txt.get('1.0', 'end-1c')
        self._open_sub_pane(document, True, '入替')
        # SUB CURSOR
        for i in range(17):
            cs = round((len(document) * i) / 16)
            ind = 'x' + str(i) + 'sixteenth'
            pos = '1.0+' + str(cs) + 'c'
            self.txt.mark_set(ind, pos)
            self.sub.mark_set(ind, pos)
        if 'sub_insert' not in self.txt.mark_names():
            self.txt.mark_set('sub_insert', '1.0')
        s = self.txt.index('sub_insert')
        self.sub.mark_set('insert', s)
        self._put_back_cursor_to_pane(self.sub, must_place_in_center=True)
        # PAINT
        self.file_lines = document.split('\n')
        self.paint_all_lines(self.sub)
        # RETURN
        self.sub.focus_force()
        return True

    def swap_windows(self) -> bool:
        if len(self.pnd_r.panes()) == 1:
            return False
        if 'x0sixteenth' not in self.sub.mark_names():
            return False
        if self.current_pane == 'sub':
            current_pane = 'sub'
        else:
            current_pane = 'txt'
        self._close_sub_pane()
        self.goto_sub_cursor()
        self.split_window()
        if current_pane == 'sub':
            self.sub.focus_set()
            self.current_pane = 'sub'
        else:
            self.txt.focus_set()
            self.current_pane = 'txt'
        return True

    def show_file(self):
        ti = '別のファイルの内容を見る'
        ty = [('読み込み可能なファイル', '.docx .md .txt .xlsx .csv')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        file_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d)
        if file_path == () or file_path == '':
            return
        if re.match('^(?:.|\n)+.xlsx$', file_path):
            document = self._read_xlsx_file(file_path)
        elif re.match('^(?:.|\n)+.csv$', file_path):
            document = self._read_csv_file(file_path)
        elif re.match('^(?:.|\n)+.docx$', file_path):
            document = self._read_docx_file(file_path)
        elif re.match('^(?:.|\n)+.md$', file_path):
            document = self._read_md_file(file_path)
        else:
            document = self._read_txt_file(file_path)
        if document is None:
            return
        #
        self._open_sub_pane(document, True)

    # COMPARE

    # MDDIFF>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

    def compare_with_previous_draft(self) -> bool:
        if len(self.pnd_r.panes()) > 1:
            return False
        importlib.reload(makdo.makdo_mddiff)
        text2 = self.init_text
        file2 = makdo.makdo_mddiff.File()
        file2.set_up_from_text(text2)
        file2.cmp_paragraphs \
            = makdo.makdo_mddiff.File.reset_configs(file2.cmp_paragraphs)
        para2 = file2.cmp_paragraphs
        self._compare_files_loop(para2)
        return True

    def compare_files(self) -> bool:
        if len(self.pnd_r.panes()) > 1:
            return False
        importlib.reload(makdo.makdo_mddiff)
        text2 = self._get_text_to_compare()
        if text2 is None:
            return False
        file2 = makdo.makdo_mddiff.File()
        file2.set_up_from_text(text2)
        file2.cmp_paragraphs \
            = makdo.makdo_mddiff.File.reset_configs(file2.cmp_paragraphs)
        para2 = file2.cmp_paragraphs
        self._compare_files_loop(para2)
        return True

    def _get_text_to_compare(self) -> str:
        ti = '別のファイルの内容と比較して反映'
        ty = [('可能な形式', '.md .docx'),
              ('Markdown', '.md'), ('MS Word', '.docx'),
              ('全てのファイル', '*')]
        _d = None
        if self.file_path is not None:
            _d = os.path.dirname(self.file_path)
        file_path = tkinter.filedialog.askopenfilename(
            title=ti, filetypes=ty, initialdir=_d)
        if file_path == () or file_path == '':
            return None
        # DOCX OR MD
        if re.match('^(?:.|\n)+.docx$', file_path):
            md_path = self.temp_dir.name + '/doc.md'
        else:
            md_path = file_path
        # OPEN DOCX FILE
        if re.match('^(?:.|\n)+.docx$', file_path):
            stderr = sys.stderr
            sys.stderr = tempfile.TemporaryFile(mode='w+')
            importlib.reload(makdo.makdo_docx2md)
            try:
                d2m = makdo.makdo_docx2md.Docx2Md(file_path)
                d2m.save(md_path)
            except BaseException:
                pass
            sys.stderr.seek(0)
            msg = sys.stderr.read()
            sys.stderr = stderr
            if msg != '':
                n = 'エラー'
                tkinter.messagebox.showerror(n, msg)
                return
        # OPEN MD FILE
        document = self._read_md_file(md_path)
        if document is None:
            return None
        return document

    def _compare_files_loop(self, para2):
        cols = self.colors
        text1 = self.txt.get('1.0', 'end-1c')
        file1 = makdo.makdo_mddiff.File()
        file1.set_up_from_text(text1)
        #
        configs = makdo.makdo_mddiff.File.get_configs(file1.raw_paragraphs)
        file1.cmp_paragraphs \
            = makdo.makdo_mddiff.File.reset_configs(file1.cmp_paragraphs)
        #
        para1 = file1.cmp_paragraphs
        comp = makdo.makdo_mddiff.Comparison(para1, para2)
        #
        p = [comp.paragraphs[0].main_paragraph]
        comp.paragraphs[0].main_paragraph \
            = makdo.makdo_mddiff.File.set_configs(p, configs)[0]
        p = [comp.paragraphs[0].sub_paragraph]
        comp.paragraphs[0].sub_paragraph \
            = makdo.makdo_mddiff.File.set_configs(p, configs)[0]
        #
        # self.quit_editing_formula()
        # self.close_memo_pad()
        self.pnd_r.update()
        half_height = int(self.pnd_r.winfo_height() / 2) - 5
        self.pnd_r.remove(self.pnd1)
        self.pnd_r.remove(self.pnd2)
        self.pnd_r.remove(self.pnd3)
        self.pnd_r.remove(self.pnd4)
        self.pnd_r.remove(self.pnd5)
        self.pnd_r.remove(self.pnd6)
        self.pnd_r.forget(self.pnd3)
        self.pnd3 = tkinter.PanedWindow(self.pnd_r, bd=0, bg='#758F00')  # 070
        self.pnd_r.add(self.pnd1, height=half_height, minsize=100)
        self.pnd_r.add(self.pnd3, height=half_height, minsize=100)
        # self.pnd_r.update()
        #
        cvs = tkinter.Canvas(self.pnd3, bg=cols['bg'])
        cvs_frm = tkinter.Frame(cvs, bg=cols['bg'])
        cvs.pack(expand=True, fill='both', anchor='w')
        scb = tkinter.Scrollbar(cvs, orient='vertical', command=cvs.yview)
        scb.pack(side='right', fill='y')
        cvs['yscrollcommand'] = scb.set
        cvs.create_window((0, 0), window=cvs_frm, anchor='nw')
        cvs_frm.bind(
            '<Configure>',
            lambda e: cvs.configure(scrollregion=cvs.bbox('all')))
        cvs.bind('<Button-1>', self.cmp_process_button1)
        cvs.bind('<Button-2>', self.close_mouse_menu)
        cvs.bind('<Button-3>', self.close_mouse_menu)
        cvs_frm.bind('<Escape>', lambda e: self._quit_diff())
        cvs_frm.bind('<Up>', lambda e: cvs.yview_scroll(-1, 'units'))
        cvs_frm.bind('<Down>', lambda e: cvs.yview_scroll(1, 'units'))
        cvs_frm.bind('<Prior>', lambda e: cvs.yview_scroll(-10, 'units'))
        cvs_frm.bind('<Next>', lambda e: cvs.yview_scroll(10, 'units'))
        if sys.platform == 'win32':
            cvs_frm.bind_all('<MouseWheel>', self.scroll_by_mouse_wheel)
        elif sys.platform == 'darwin':
            cvs_frm.bind_all('<MouseWheel>', self.scroll_by_mouse_wheel)
        elif sys.platform == 'linux':
            cvs_frm.bind_all('<Button-4>', self.scroll_by_mouse_wheel)
            cvs_frm.bind_all('<Button-5>', self.scroll_by_mouse_wheel)
        # if sys.platform == 'win32':
        #     cvs_frm.bind_all(
        #         '<MouseWheel>',
        #         lambda e: cvs.yview_scroll(- int(e.delta / 100), 'units'))
        # elif sys.platform == 'darwin':
        #     cvs_frm.bind_all(
        #         '<MouseWheel>',
        #         lambda e: cvs.yview_scroll(- int(e.delta / 120), 'units'))
        # elif sys.platform == 'linux':
        #     cvs_frm.bind_all('<4>', lambda e: cvs.yview_scroll(-1, 'units'))
        #     cvs_frm.bind_all('<5>', lambda e: cvs.yview_scroll(+1, 'units'))
        btn = tkinter.Button(self.pnd3, text='終了', command=self._quit_diff)
        btn.pack(side='bottom')
        self.btns = []
        for p in comp.paragraphs:
            if p.ses_symbol == '.':
                continue
            frm0 = tkinter.Frame(cvs_frm)
            frm1 = tkinter.Frame(frm0)
            frm2 = tkinter.Frame(frm0)
            frm3 = tkinter.Frame(frm0)
            btn1 = tkinter.Button(frm1, text='適用',
                                  command=self._apply_diff(frm0,
                                                           p.diff_id, comp))
            self.btns.append(btn1)
            btn2 = tkinter.Button(frm1, text='除外',
                                  command=self._exclude_diff(frm0))
            self.btns.append(btn2)
            btn3 = tkinter.Button(frm1, text='移動',
                                  command=self._goto_diff(p.diff_id, comp))
            self.btns.append(btn3)
            btn9 = tkinter.Label(frm1, font=self.gothic_font, text='\n')
            res = '^(.*)\n((?:.|\n)*?)\n*$'
            head_text = re.sub(res, '\\1', p.diff_text)
            diff_text = re.sub(res, '\\2', p.diff_text)
            head_font = self.gothic_font.copy()
            head_font['weight'] = 'bold'
            diff_font = self.gothic_font.copy()
            head_lbl = tkinter.Label(frm2, text=('## ' + head_text),
                                     font=head_font, justify='left')
            diff_lbl = tkinter.Label(frm3, text=diff_text,
                                     font=diff_font, justify='left')
            frm0.configure(bg=cols['bg'])
            frm1.configure(bg=cols['bg'])
            frm2.configure(bg=cols['bg'])
            frm3.configure(bg=cols['bg'])
            btn9.configure(bg=cols['bg'])
            head_lbl.configure(bg=cols['bg'], fg=cols['sc'][1])
            diff_lbl.configure(bg=cols['bg'], fg=cols['fg'])
            frm0.pack(expand=True, side='top', anchor='w', fill='x')
            # frm1.pack(expand=True, side='top', anchor='w', fill='x')
            # btn1.pack(side='left')
            # btn2.pack(side='left')
            # btn3.pack(side='left')
            frm2.pack(expand=True, side='top', anchor='w', fill='x')
            head_lbl.pack(expand=True, side='left', anchor='w')
            frm3.pack(expand=True, side='top', anchor='w', fill='x')
            diff_lbl.pack(expand=True, side='left', anchor='w')
            frm1.pack(expand=True, side='top', anchor='w', fill='x')
            btn1.pack(side='left', anchor='n')
            btn2.pack(side='left', anchor='n')
            btn3.pack(side='left', anchor='n')
            btn9.pack(side='left', anchor='n')
            frm0.bind('<Button-1>', self.cmp_process_button1)
            frm0.bind('<Button-2>', self.close_mouse_menu)
            frm0.bind('<Button-3>', self.close_mouse_menu)
            frm1.bind('<Button-1>', self.cmp_process_button1)
            frm1.bind('<Button-2>', self.close_mouse_menu)
            frm1.bind('<Button-3>', self.close_mouse_menu)
            frm2.bind('<Button-1>', self.cmp_process_button1)
            frm2.bind('<Button-2>', self.close_mouse_menu)
            frm2.bind('<Button-3>', self.close_mouse_menu)
            frm3.bind('<Button-1>', self.cmp_process_button1)
            frm3.bind('<Button-2>', self.close_mouse_menu)
            frm3.bind('<Button-3>', self.close_mouse_menu)
            head_lbl.bind('<Button-1>', self.cmp_process_button1)
            head_lbl.bind('<Button-2>', self.close_mouse_menu)
            head_lbl.bind('<Button-3>', self.close_mouse_menu)
            diff_lbl.bind('<Button-1>', self.cmp_process_button1)
            diff_lbl.bind('<Button-2>', self.close_mouse_menu)
            diff_lbl.bind('<Button-3>', self.close_mouse_menu)
        self._put_back_cursor_to_pane(self.txt)
        cvs_frm.focus_force()
        self.cmp_cvs = cvs
        self.cmp_cvs_frm = cvs_frm

    def cmp_process_button1(self, event):
        self.close_mouse_menu()
        self.cmp_cvs.focus_force()

    def _apply_diff(self, frame, diff_id, comp):
        def x():
            txt = self.txt.get('1.0', 'end-1c')
            beg, end = self._get_diff_position(diff_id, comp, txt)
            if beg < 0 or end < 0:
                return False
            for cp in comp.paragraphs:
                if cp.diff_id != diff_id:
                    continue
                self.txt['autoseparators'] = False
                self.txt.edit_separator()
                if cp.ses_symbol == '&':
                    self.txt.delete('1.0+' + str(beg) + 'c',
                                    '1.0+' + str(end) + 'c')
                    if cp.sub_paragraph != '':  # for empty configuration
                        insert_text = cp.sub_paragraph + '\n\n'
                        self.txt.insert('1.0+' + str(beg) + 'c', insert_text)
                        t = self.txt.get('1.0', '1.0+' + str(beg) + 'c')
                        beg_line = t.count('\n')
                        end_line = beg_line + insert_text.count('\n')
                        for i in range(beg_line, end_line):
                            self.paint_out_line(i)
                        self.update_toc()
                elif cp.ses_symbol == '-':
                    self.txt.delete('1.0+' + str(beg) + 'c',
                                    '1.0+' + str(end) + 'c')
                elif cp.ses_symbol == '+':
                    if cp.sub_paragraph != '':  # for empty configuration
                        if beg == 0:
                            insert_text = cp.sub_paragraph + '\n\n'
                        elif beg == 1:
                            beg = 0
                            insert_text = cp.sub_paragraph + '\n'
                        elif (beg >= len(txt) and
                              not re.match('^(.|\n)*\n$', txt)):
                            insert_text = '\n\n' + cp.sub_paragraph + '\n'
                        else:
                            insert_text = '\n' + cp.sub_paragraph + '\n'
                        self.txt.insert('1.0+' + str(beg) + 'c', insert_text)
                        t = self.txt.get('1.0', '1.0+' + str(beg) + 'c')
                        beg_line = t.count('\n')
                        end_line = beg_line + insert_text.count('\n')
                        for i in range(beg_line, end_line + 1):
                            self.paint_out_line(i)
                        self.update_toc()
                cp.has_applied = True
                frame.destroy()
                self.txt['autoseparators'] = True
                self.txt.edit_separator()
                return True
        return x

    def _exclude_diff(self, frame):
        def x():
            frame.destroy()
            return True
        return x

    def _goto_diff(self, diff_id, comp):
        def x():
            txt = self.txt.get('1.0', 'end-1c')
            beg, end = self._get_diff_position(diff_id, comp, txt)
            if beg < 0 or end < 0:
                return False
            self.txt.mark_set('insert', '1.0+' + str(beg) + 'c')
            self._put_back_cursor_to_pane(self.txt)
            return True
        return x

    @staticmethod
    def _get_diff_position(diff_id, comp, txt):
        pars = makdo.makdo_mddiff.File.get_raw_paragraphs(txt)
        if pars[0] == '':
            pars.pop(0)  # for empty configuration
        p = ''
        n = 0
        s = ''
        for cp in comp.paragraphs:
            if cp.get_current_paragraph() != '':
                p = cp.get_current_paragraph()
                n += 1
            if cp.diff_id == diff_id:
                s = cp.ses_symbol
                break
        if len(pars) > 0 and re.match('^\n+$', pars[0]):
            n += 1
        if len(pars) == 0:
            pars = ['']
        elif n > 0:
            par = pars[n - 1]
            if p != re.sub('\n+$', '', par):
                n = 'エラー'
                m = '編集場所が見当たりません．'
                tkinter.messagebox.showerror(n, m)
                return -1, -1
        if s != '+':
            pre = ''.join(pars[:n - 1])
        else:
            pre = ''.join(pars[:n])
        beg = len(pre)
        if re.match('^(.|\n)*\n\n$', pre):
            if s == '+':
                beg -= 1
        if s == '+':
            end = beg
        else:
            end = beg + len(par)
        return beg, end

    def _quit_diff(self):
        self.pnd_r.remove(self.pnd3)
        self.txt.focus_set()
        self.current_pane = 'txt'

    # MDDIFF<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

    def insert_track_change_tag(self):
        text1 = self.txt.get('1.0', 'end-1c')
        text2 = self.init_text
        file1 = makdo.makdo_mddiff.File()
        file2 = makdo.makdo_mddiff.File()
        file1.set_up_from_text(text1)
        file2.set_up_from_text(text2)
        para1 = file1.cmp_paragraphs
        para2 = file2.cmp_paragraphs
        comp = makdo.makdo_mddiff.Comparison(para2, para1)
        # REMAKE DOC
        doc = ''
        for cp in comp.paragraphs:
            if cp.track_change_text != '':
                doc += cp.track_change_text + '\n\n'
        doc = makdo.makdo_mddiff.TrackChange.repair_tc_text(doc)
        doc = re.sub('\n+$', '\n', doc)
        self.txt.delete('1.0', 'end-1c')
        self.txt.insert('1.0', doc)
        # PAINT
        self.file_lines = doc.split('\n')
        self.paint_all_lines(self.txt)

    def search_previous_track_change_tag(self):
        doc = self.txt.get('1.0', 'end-1c')
        p = len(self.txt.get('1.0', 'insert'))
        beg = makdo.makdo_mddiff.TrackChange.get_previous_track_change(doc, p)
        if beg > 0:
            self.txt.mark_set('insert', '1.0+' + str(beg) + 'c')
            return True
        return False

    def search_next_track_change_tag(self):
        doc = self.txt.get('1.0', 'end-1c')
        p = len(self.txt.get('1.0', 'insert'))
        beg = makdo.makdo_mddiff.TrackChange.get_next_track_change(doc, p)
        if beg > 0:
            self.txt.mark_set('insert', '1.0+' + str(beg) + 'c')
            return True
        return False

    def accept_current_track_change(self):
        doc = self.txt.get('1.0', 'end-1c')
        p = len(self.txt.get('1.0', 'insert'))
        del_or_ins, beg_num, end_num \
            = makdo.makdo_mddiff.TrackChange.get_current_track_change(doc, p)
        if del_or_ins == 'del':
            self.txt.delete('1.0+' + str(beg_num) + 'c',
                            '1.0+' + str(end_num) + 'c')
        elif del_or_ins == 'ins':
            self.txt.delete('1.0+' + str(beg_num + 0) + 'c',
                            '1.0+' + str(beg_num + 2) + 'c')
            self.txt.delete('1.0+' + str(end_num - 4) + 'c',
                            '1.0+' + str(end_num - 2) + 'c')
            # PAINT
            beg_line = doc[:beg_num].count('\n')
            end_line = doc[:end_num].count('\n')
            for i in range(beg_line, end_line + 1):
                self.paint_out_line(i)
        else:
            n = 'エラー'
            m = 'カーソルが変更履歴の範囲にありません．'
            tkinter.messagebox.showerror(n, m)

    def accept_all_track_changes(self):
        self.txt.mark_set('_tmp', 'insert')
        self.txt.mark_set('insert', '1.0')
        while self.search_next_track_change_tag():
            self.accept_current_track_change()
        self.txt.mark_set('insert', '_tmp')
        self.txt.mark_unset('_tmp')

    # FOLD

    def fold_section(self):
        sub_document = self.txt.get('insert linestart', 'end-1c')
        tmp_document = self._remove_head_and_tail_fds(sub_document)
        # CHECK THAT THE LINE IS SECITION
        res = '^#+(?:-#+)*(?:\\s.*)?\n'
        if not re.match(res, tmp_document):
            n = 'エラー'
            m = '行がセクションの見出し（"#"から始まる行）ではありません．'
            tkinter.messagebox.showerror(n, m)
            return
        # CHECK THAT HEADING IS NOT EMPTY
        res = '^#+(?:-#+)*\\s*\n\n'
        if re.match(res, tmp_document):
            n = 'エラー'
            m = 'セクションの見出しがありません（字下げの調整です）．'
            tkinter.messagebox.showerror(n, m)
            return
        # CHECK THAT THE END OF LINE IS NOT ESCAPED
        fln = sub_document.split('\n')[0]
        if not re.match(NOT_ESCAPED + 'X$', fln + 'X'):
            n = 'エラー'
            m = 'セクションの見出しがエスケープされています' + \
                '（バックスラッシュで終わっています）．'
            tkinter.messagebox.showerror(n, m)
            return
        # CHECK THAT SECITION IS NOT FOLDED
        res = '^#+(?:-#+)*(?:\\s.*)?\\.\\.\\.\\[([0-9]+)\\]\n(?:.|\n)*$'
        if re.match(res, sub_document):
            n, m = 'エラー', 'セクションは折り畳まれています．'
            tkinter.messagebox.showerror(n, m)
            return
        # SHOW MESSAGE
        self.show_folding_help_message()
        # GET FOLDING NUMBER
        folding_number = 1
        all_document = self.txt.get('1.0', 'end-1c')
        res = '^\\.\\.\\.\\[([0-9]+)\\].*$'
        for line in all_document.split('\n'):
            if re.match(res, line):
                n = int(re.sub(res, '\\1', line))
                if folding_number <= n:
                    folding_number = n + 1
        # GET SECTION LINE
        sub_lines = sub_document.split('\n')
        section_line = sub_lines[0]
        # GET SECTION LEVEL
        res = '^(#+).*$'
        section_level = len(re.sub(res, '\\1', section_line))
        # GET TEXT TO FOLD
        text_to_fold = ''
        is_end_of_document = False
        m = len(sub_lines) - 1
        for i in range(1, m + 1):
            line = sub_lines[i]
            if re.match('^(#+)(?:-#+)*(?:\\s.*)?$', line):
                # SECTION
                level = len(re.sub(res, '\\1', line))
                if level <= section_level:
                    if not re.match('^#+(?:-#+)*\\s*$', line) or \
                       not (i < m and sub_lines[i + 1] == ''):
                        tmp = re.sub('<!--(.|\n)*?-->', '', text_to_fold)
                        if re.match('^(.|\n)*\n<!--(.|\n)*$', tmp):
                            # "\n<!--\n## xxx"
                            text_to_fold \
                                = re.sub('<!--(.|\n)*$', '', text_to_fold)
                            break
                        if not re.match('^(.|\n)*<!--(.|\n)*$', tmp):
                            # not "yyy<!--\n## xxx"
                            break
            if re.match('^\\.\\.\\.\\[[0-9]+\\]#+(?:-#+)*(?:\\s.*)?$', line):
                # FOLDED SECTION
                text_to_fold \
                    = re.sub(DONT_EDIT_MESSAGE + '\n\n$', '', text_to_fold)
                break
            text_to_fold += line + '\n'
        else:
            is_end_of_document = True
        # INSERT FOLDING MARK
        self.txt.insert('insert lineend', '...[' + str(folding_number) + ']')
        # INSERT TEXT TO FOLD
        if is_end_of_document:
            self.txt.insert('end', '\n')
        else:
            if not re.match('^(.|\n)*\n$', sub_document):
                self.txt.insert('end', '\n\n')
            elif not re.match('^(.|\n)*\n\n$', sub_document):
                self.txt.insert('end', '\n')
        self.txt.insert('end', DONT_EDIT_MESSAGE + '\n\n')
        self.txt.insert('end', '...[' + str(folding_number) + ']')
        self.txt.insert('end', section_line + '\n')
        self.txt.insert('end', text_to_fold)
        if re.match('^(.|\n)*\n\n\n', text_to_fold):
            self.txt.delete('end-1c', 'end')
        # DELETE FOLDING TEXT
        beg = 'insert lineend + 1c'
        end = 'insert lineend +' + str(len(text_to_fold)) + 'c'
        self.txt.delete(beg, end)
        # MOVE
        # self.txt.mark_set('insert', 'insert linestart')

    def unfold_section(self):
        sub_document = self.txt.get('insert linestart', 'end-1c')
        tmp_document = self._remove_head_and_tail_fds(sub_document)
        # CHECK THAT THE LINE IS SECITION
        res = '^#+(?:-#+)*(?:\\s.*)?\n'
        if not re.match(res, tmp_document):
            n = 'エラー'
            m = '行がセクションの見出し（"#"から始まる行）ではありません．'
            tkinter.messagebox.showerror(n, m)
            return
        # CHECK THAT SECITION IS FOLDED
        res = '^#+(?:-#+)*(?:\\s.*)?\\.\\.\\.\\[([0-9]+)\\]\n(?:.|\n)*$'
        if not re.match(res, tmp_document):
            n, m = 'エラー', 'セクションが折り畳まれていません．'
            tkinter.messagebox.showerror(n, m)
            return
        # CHECK THAT TEXT TO UNFOLD EXISTS
        folding_number = re.sub(res, '\\1', tmp_document)
        res_mark = '\\.\\.\\.\\[' + folding_number + '\\]'
        res = '^' + '((?:.|\n)*?\n)' \
            + '((?:' + DONT_EDIT_MESSAGE + '\n+)?)' \
            + '(' + res_mark + '.*#+(?:-#+)*(?:\\s.*)?\n)' \
            + '((?:.|\n)*)$'
        if not re.match(res, sub_document):
            n, m = 'エラー', '折り畳み先が見付かりません．'
            tkinter.messagebox.showerror(n, m)
            return
        # DISPLAY MESSAGE
        self.show_folding_help_message()
        # GET TEXT
        text_a = re.sub(res, '\\1', sub_document)  # unconcerned
        text_b = re.sub(res, '\\2', sub_document)  # dont edit message
        text_c = re.sub(res, '\\3', sub_document)  # folding mark line
        text_d = re.sub(res, '\\4', sub_document)  # text to unfold
        res = '^' + '((?:.|\n)*?\n)' \
            + '((?:' + DONT_EDIT_MESSAGE + '\n+)?)' \
            + '(\\.\\.\\.\\[[0-9]+\\].*#+(?:-#+)*(?:\\s.*)?\n)' \
            + '((?:.|\n)*)$'
        if re.match(res, text_d):
            text_d = re.sub(res, '\\1', text_d)
        # ADJUST LINE BREAK
        number_of_line_break_to_insert = 0
        if self.txt.get('insert lineend +1c', 'insert lineend +2c') == '\n':
            number_of_line_break_to_insert -= 1
        if not re.match('^(.|\n)*\n$', text_d):
            number_of_line_break_to_insert += 2
        elif not re.match('^(.|\n)*\n\n$', text_d):
            number_of_line_break_to_insert += 1
        # INSERT TEXT TO UNFOLD
        self.txt.insert('insert lineend +1c', text_d)
        # PAINT LINES
        beg = self._get_v_position_of_insert(self.txt)
        end = beg + text_d.count('\n')
        for i in range(beg - 1, end):
            self.paint_out_line(i)
        # UPDATE TOC
        self.update_toc()
        # REMOVE TEXT TO UNFOLD
        text_e = text_a + text_b + text_c + text_d
        beg = 'insert linestart +' + str(len(text_d + text_a)) + 'c'
        end = 'insert linestart +' + str(len(text_d + text_e)) + 'c'
        self.txt.delete(beg, end)
        # ADJUST LINE BREAK
        if number_of_line_break_to_insert == -1:
            beg = 'insert lineend +' + str(len(text_d)) + 'c'
            end = 'insert lineend +' + str(len(text_d) + 1) + 'c'
            self.txt.delete(beg, end)
        elif number_of_line_break_to_insert > 0:
            ins = 'insert lineend +' + str(len(text_d) + 1) + 'c'
            self.txt.insert(ins, '\n' * number_of_line_break_to_insert)
        # REMOVE FOLDING MARK
        text_f = '...[' + folding_number + ']'
        beg = 'insert lineend -' + str(len(text_f)) + 'c'
        end = 'insert lineend'
        self.txt.delete(beg, end)
        # MOVE
        # self.txt.mark_set('insert', 'insert linestart')

    def unfold_section_fully(self):
        old_document = self.txt.get('1.0', 'end-1c')
        if old_document == '':
            return
        new_document = self.get_fully_unfolded_document(old_document)
        self.file_lines = new_document.split('\n')
        self.txt.insert('1.0', new_document)
        self.txt.delete('1.0+' + str(len(new_document)) + 'c', 'end')
        self.txt.focus_set()
        self.current_pane = 'txt'
        self.txt.mark_set('insert', '1.0')
        # PAINT
        self.paint_all_lines(self.txt)

    def get_fully_unfolded_document(self, old_document: str, must_warn=True):
        # |                ->  |
        # |## www...[3]    ->  |## www
        # |                ->  |
        # |...[1]#### yyy  ->  |### xxx
        # |                ->  |
        # |zzz             ->  |#### yyy
        # |                ->  |
        # |...[2]### xxx   ->  |zzz
        # |                ->  |
        # |#### yyy...[1]  ->  |
        # |                ->  |
        # |...[3]## www    ->  |
        # |                ->  |
        # |### xxx...[2]   ->  |
        # |                ->  |
        if old_document == '':
            return ''
        old_lines = old_document.split('\n')
        new_lines = []
        remain_lines = [True for i in old_lines]
        m = len(old_lines) - 1
        line_numbers = [0]
        res_mark = '\\.\\.\\.\\[([0-9]+)\\]'
        res_from = '^(#+(?:-#+)*(?:\\s.*)?)' + res_mark + '$'
        res_to = '^' + res_mark + '#+(-#+)*(\\s|$)'
        while line_numbers != []:
            i = line_numbers[-1]
            if i > m:
                line_numbers.pop(-1)
                continue
            if not remain_lines[i]:
                line_numbers.pop(-1)
                continue
            if re.match(res_to, old_lines[i]):
                line_numbers.pop(-1)
                if new_lines[-2] == DONT_EDIT_MESSAGE and \
                   new_lines[-1] == '':
                    new_lines.pop(-1)
                    new_lines.pop(-1)
                continue
            if re.match(res_from, old_lines[i]) and \
               re.match(NOT_ESCAPED + res_mark + '$', old_lines[i]):
                folding_number \
                    = re.sub(res_from, '\\2', old_lines[i])
                old_lines[i] \
                    = re.sub(res_from, '\\1', old_lines[i])
                # APPEND "FROM LINE"
                new_lines.append(old_lines[i])
                remain_lines[i] = False
                line_numbers[-1] += 1
                if i < m and old_lines[i + 1] == '':
                    # SKIP "NEXT EMPTY LINE"
                    # new_lines.append(old_lines[i])
                    remain_lines[i + 1] = False
                    line_numbers[-1] += 1
                for j, line in enumerate(old_lines):
                    if not remain_lines[j]:
                        continue
                    res = '^\\.\\.\\.\\[' + folding_number + '\\]'
                    if re.match(res, line):
                        if j >= 2:
                            if old_lines[j - 2] == DONT_EDIT_MESSAGE and \
                               old_lines[j - 1] == '':
                                # SKIP "DONT EDIT MESSAGE"
                                remain_lines[j - 2] = False
                                remain_lines[j - 1] = False
                        line_numbers.append(j)
                        # JUMP TO "TO LINE"
                        # new_lines.append(old_lines[j])
                        remain_lines[j] = False
                        line_numbers[-1] += 1
            else:
                # APPEND "USUAL LINE"
                new_lines.append(old_lines[i])
                remain_lines[i] = False
                line_numbers[-1] += 1
        for i, ml in enumerate(old_lines):
            if remain_lines[i]:
                if must_warn:
                    n, m = 'エラー', '折り畳まれたセクションが残っています．'
                    tkinter.messagebox.showerror(n, m)
                    must_warn = False
                new_lines.append(old_lines[i])
        new_document = '\n'.join(new_lines) + '\n\n'
        new_document = re.sub('\n\n+', '\n\n', new_document)
        new_document = re.sub('\n+$', '\n', new_document)
        return new_document

    def fold_or_unfold_section(self):
        sub_document = self.txt.get('insert linestart', 'end-1c')
        sub_document = self._remove_head_and_tail_fds(sub_document)
        # CHECK THAT THE LINE IS SECITION
        res = '^#+(?:-#+)*(?:\\s.*)?\n'
        if not re.match(res, sub_document):
            n = 'エラー'
            m = '行がセクションの見出し（"#"から始まる行）ではありません．'
            tkinter.messagebox.showerror(n, m)
            return
        # CHECK THAT SECITION IS FOLDED
        res = '^#+(?:-#+)*(?:\\s.*)?\\.\\.\\.\\[([0-9]+)\\]\n(?:.|\n)*$'
        if not re.match(res, sub_document):
            self.fold_section()
        else:
            self.unfold_section()
        # TABLE OF CONTENTS
        self.update_toc()

    def execute_keyboard_macro(self):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        self.show_keyboard_macro_help_message()
        reversed_history = list(reversed(self.key_history))
        k1, k2 = reversed_history[0], reversed_history[1]
        if not self._is_key(k2, 'F17', 'C-}', 'C-]', 'C-e'):
            if self._is_key(k1, 'F17', 'C-}', 'C-]', 'C-e'):
                reversed_history.pop(0)
            layout = self.shortcut_key_layout.get()
            for i in range(10, -1, -1):
                kh1 = []
                for j in range(i):
                    kh1.append(reversed_history[j])
                kh2 = []
                for j in range(i, i * 2):
                    kh2.append(reversed_history[j])
                if 'F17' in kh1 or \
                   (layout == 'akauni' and 'C-}' in kh1) or \
                   (layout == 'qwerty' and 'C-]' in kh1) or \
                   (layout == 'normal' and 'C-e' in kh1) or \
                   'F17' in kh2 or \
                   (layout == 'akauni' and 'C-}' in kh2) or \
                   (layout == 'qwerty' and 'C-]' in kh2) or \
                   (layout == 'normal' and 'C-e' in kh2):
                    continue
                if kh1 == kh2:
                    self.keyboard_macro = list(reversed(kh1))
                    self.ideal_h_position \
                        = self._get_ideal_h_position_of_insert(pane)
                    break
            else:
                self.keyboard_macro = []
                return False
        pane['autoseparators'] = False
        pane.edit_separator()
        for i, key in enumerate(self.keyboard_macro):
            if key in ASCII_SYMBOLS:
                key = ASCII_SYMBOLS[key]
            if self._is_key(key, 'BackSpace', 'C-h', 'C-j'):
                pane.delete('insert-1c', 'insert')
                self.paint_out_line(self._get_v_position_of_insert(pane) - 1)
                self.update_toc()
            elif self._is_key(key, 'Delete', 'C-d', 'C-h', 'C-x'):
                if i > 0:
                    k = self.keyboard_macro[i - 1]
                    if not self._is_key(k, 'Delete', 'C-d', 'C-h', 'C-x'):
                        self.win.clipboard_clear()
                        if self.clipboard_list[-1] != '':
                            self.clipboard_list.append('')
                self._execute_when_delete_is_pressed(pane)
                self.paint_out_line(self._get_v_position_of_insert(pane) - 1)
                self.update_toc()
            elif self._is_key(key, 'Return', 'C-m', 'C-m'):
                pane.insert('insert', '\n')
                if pane == self.txt:
                    vp = self._get_v_position_of_insert(pane)
                    self.paint_out_line(vp - 2)
                    self.paint_out_line(vp - 1)
                    self.update_toc()
            elif self._is_key(key, 'F15', 'C-g', 'C-u', 'C-v'):
                self.paste_region()
            elif self._is_key(key, 'Home', 'C-l', 'C-p'):
                width = self._get_width_of_pane(pane)
                self._move_horizontally(pane, -width)
            elif self._is_key(key, 'End', 'C-{', 'C-['):
                width = self._get_width_of_pane(pane)
                self._move_horizontally(pane, +width)
            elif self._is_key(key, 'Up', 'C-r', 'C-o'):
                self._move_vertically(pane, self.ideal_h_position, -1)
            elif self._is_key(key, 'Down', 'C-n', 'C-l'):
                self._move_vertically(pane, self.ideal_h_position, +1)
            elif self._is_key(key, 'Left', 'C-t', 'C-k'):
                pane.mark_set('insert', 'insert-1c')
            elif self._is_key(key, 'Right', 'C-s', 'C-;'):
                pane.mark_set('insert', 'insert+1c')
            elif self._is_key(key, 'F22', 'C-f', 'C-.'):
                pane.mark_set('akauni', 'insert')
            else:
                pane.insert('insert', key)
                self.paint_out_line(self._get_v_position_of_insert(pane) - 1)
                self.update_toc()
            if not self._is_key(key, 'Up', 'C-r', 'C-o') and \
               not self._is_key(key, 'Down', 'C-n', 'C-l'):
                self.ideal_h_position \
                    = self._get_ideal_h_position_of_insert(pane)
        pane['autoseparators'] = True
        pane.edit_separator()
        self._put_back_cursor_to_pane(pane)
        return True

    # MINIBUFFER

    def start_minibuffer(self):
        mb = self.Minibuffer(self.txt, self)
        rt = mb.get_return_to()
        if rt == self.sub:
            self.sub.focus_force()
            self.current_pane = 'sub'
        elif 'toc_cvs_frm' in vars(self) and rt == self.toc_cvs_frm:
            self.toc_cvs_frm.focus_force()
            self.current_pane = 'toc'
        else:
            self.txt.focus_force()
            self.current_pane = 'txt'

    class Minibuffer(tkinter.simpledialog.Dialog):

        class MinibufferCommand:

            def __init__(self, command_text, help_texts, method_texts):
                self.command_text = command_text
                self.help_texts = help_texts
                self.method_texts = method_texts

        minibuffer_commands = []

        mc = MinibufferCommand(
            'help',
            [None, 'このメッセージを表示'],
            ['self.Help(self, self.mother, self.help_message)',
             'Makdo.Minibuffer(self.pane, self.mother)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'toggle-read-only',
            [None, '読取専用を指定又は解除'],
            ['self.toggle_read_only()'])
        minibuffer_commands.append(mc)

        def toggle_read_only(self):
            is_read_only = self.mother.is_read_only.get()
            if is_read_only:
                self.mother.is_read_only.set(False)
            else:
                self.mother.is_read_only.set(True)

        mc = MinibufferCommand(
            'goto-by-position',
            [None, '行数と文字数を指定して移動'],
            ['self.mother.goto_by_position(self)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'goto-sub-cursor',
            [None, 'サブカーソルに移動'],
            ['self.mother.goto_sub_cursor()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'place-flag',
            ['place-flagX(X=1..5)', 'フラグXを設置'],
            ['self.mother.place_flag1()'])
        minibuffer_commands.append(mc)
        for i in range(1, 10):
            mc = MinibufferCommand(
                'place-flag' + str(i),
                None,
                ['self.mother.place_flag' + str(i) + '()'])
            minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'goto-flag',
            ['goto-flagX(X=1..5)', 'フラグXに移動'],
            ['self.mother.goto_flag1()'])
        minibuffer_commands.append(mc)
        for i in range(1, 10):
            mc = MinibufferCommand(
                'goto-flag' + str(i),
                None,
                ['self.mother.goto_flag' + str(i) + '()'])
            minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'search-backward',
            [None, '上方を検索'],
            ['self.mother.search_backward_from_dialog(self)'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'search-forward',
            [None, '下方を検索'],
            ['self.mother.search_forward_from_dialog(self)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'replace-string',
            [None, '文章全体又は選択範囲を全置換'],
            ['self.mother.replace_all(self)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'replace-backward',
            [None, '上方を置換'],
            ['self.mother.replace_forward_from_dialog(self)'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'replace-forward',
            [None, '下方を置換'],
            ['self.mother.replace_backward_from_dialog(self)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'clear-search-and-replace-words',
            [None, '検索語と置換語をリセットする'],
            ['self.mother.clear_search_and_replace()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'upcase-region',
            [None, '選択範囲を大文字に変換'],
            ['self.mother.replace_lower_case_with_upper_case()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'downcase-region',
            [None, '選択範囲を小文字に変換'],
            ['self.mother.replace_upper_case_with_lower_case()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'japanese-zenkaku-region',
            [None, '選択範囲を全角文字に変換'],
            ['self.mother.replace_half_width_with_full_width()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'japanese-hankaku-region',
            [None, '選択範囲を半角文字に変換'],
            ['self.mother.replace_full_width_with_half_width()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'adjust-character-width',
            [None, '選択範囲の半角全角を推奨の形に変換'],
            ['self.mother.adjust_character_width()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'sort-lines',
            [None, '選択範囲の行を正順にソート'],
            ['self.mother.sort_lines()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'sort-lines-in-reverse-order',
            [None, '選択範囲の行を逆順にソート'],
            ['self.mother.sort_lines_in_reverse_order()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'comment-out-region',
            [None, '選択範囲をコメントアウト'],
            ['self.mother.comment_out_region()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'uncomment-in-region',
            [None, '選択範囲のコメントアウトを解除'],
            ['self.mother.uncomment_in_region()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'fold-or-unfold-section',
            [None, 'セクションの折畳又は展開'],
            ['self.mother.fold_or_unfold_section()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'unfold-section-fully',
            [None, 'セクションの折畳を全て展開'],
            ['self.mother.unfold_section_fully()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'what-cursor-position',
            [None, '入力位置の文字情報を表示'],
            ['self.mother.show_char_info()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'change-typeface',
            [None, '字形を変える'],
            ['self.mother.change_typeface(self)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'paste-region-from-list',
            [None, 'リストから選んで貼り付け'],
            ['self.mother.paste_region_from_list(self)'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-file',
            [None, 'ファイルの内容を挿入'],
            ['self.mother.insert_file()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-file-names',
            [None, 'ファイル名のみを一括挿入'],
            ['self.mother.insert_file_names_in_same_folder()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-current-time',
            [None, '現在の日時を挿入'],
            ['self.mother.insert_datetime_simple()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-current-date',
            [None, '今日の日付を挿入'],
            ['self.mother.insert_date_Gymd()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-symbol',
            [None, '記号を挿入'],
            ['self.mother.insert_symbol()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-hyphen',
            [None, 'ハイフンを挿入'],
            ['self.mother.insert_hline_2010()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-minus-sign',
            [None, 'マイナスサインを挿入'],
            ['self.mother.insert_hline_2212()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-small-hyphen-minus',
            [None, '小さいハイフンマイナスを挿入'],
            ['self.mother.insert_hline_fe63()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-fullwidth-hyphen-minus',
            [None, '全角ハイフンマイナスを挿入'],
            ['self.mother.insert_hline_ff0d()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'insert-comment',
            [None, 'コメントを挿入'],
            ['self.mother.txt.insert("insert", "<!---->")',
             'self.mother.txt.mark_set("insert", "insert-3c")'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'edit-formula',
            ['edit-formulaX(X=1..9)', '定型句Xを編集'],
            ['self.mother.edit_formula1()'])
        minibuffer_commands.append(mc)
        for i in range(1, 10):
            mc = MinibufferCommand(
                'edit-formula' + str(i),
                None,
                ['self.mother.edit_formula' + str(i) + '()'])
            minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'insert-formula',
            ['insert-formulaX(X=1..9)', '定型句Xを挿入'],
            ['self.mother.insert_formula1()'])
        minibuffer_commands.append(mc)
        for i in range(1, 10):
            mc = MinibufferCommand(
                'insert-formula' + str(i),
                None,
                ['self.mother.insert_formula' + str(i) + '()'])
            minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'split-window',
            [None, '画面を分割'],
            ['self.mother.split_window()',
             'self.set_return_to()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'open-memo-pad',
            [None, 'メモ帳を開く'],
            ['self.mother.open_memo_pad()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'compare-with-previous-draft',
            [None, '編集前の原稿と比較'],
            ['self.mother.compare_with_previous_draft()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'previous_window',
            [None, '前のウィンドウに移動'],
            ['self.mother._jump_to_prev_pane()',
             'self.set_return_to()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'next_window',
            [None, '次のウィンドウに移動'],
            ['self.mother._jump_to_next_pane()',
             'self.set_return_to()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'close-sub-window',
            [None, 'サブウィンドウを閉じる'],
            ['self.mother._close_sub_pane()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'undo',
            [None, '元に戻す'],
            ['self.mother.edit_modified_undo()'])
        minibuffer_commands.append(mc)
        mc = MinibufferCommand(
            'redo',
            [None, 'やり直す'],
            ['self.mother.edit_modified_redo()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'save-buffer',
            [None, 'ファイルを保存'],
            ['self.mother.save_file()'])
        minibuffer_commands.append(mc)

        mc = MinibufferCommand(
            'kill-makdo',
            [None, 'Makdoを終了'],
            ['self.mother.quit_makdo()'])  # 2 ERRORS OCCUR
        minibuffer_commands.append(mc)

        history = []

        def __init__(self, pane, mother, init=''):
            self.pane = pane
            self.mother = mother
            self.init = init
            self.commands = self.get_commands()
            self.help_message = self.get_help_message()
            self.return_to = pane
            self.history_number = 0
            if len(self.history) == 0:
                Makdo.Minibuffer.history.append('')
            elif self.history[-1] in self.commands:
                Makdo.Minibuffer.history.append('')
            else:
                Makdo.Minibuffer.history[-1] = ''
            super().__init__(pane, title='ミニバッファ')

        def get_commands(self):
            commands = []
            for mc in Makdo.Minibuffer.minibuffer_commands:
                commands.append(mc.command_text)
                commands.sort()
            return commands

        def get_help_message(self):
            help_message = ''
            for mc in Makdo.Minibuffer.minibuffer_commands:
                if mc.help_texts is None:
                    continue
                help_command = mc.help_texts[0]
                help_explanation = mc.help_texts[1]
                if help_command is None:
                    help_command = mc.command_text
                help_message += \
                    help_command \
                    + (' ' * (32 - len(help_command)))
                help_message += \
                    help_explanation \
                    + (' ' * (32 - get_real_width(help_explanation))) + '\n'
            return help_message

        def body(self, pane):
            fon = self.mother.gothic_font
            t = 'コマンドを入力してください．\n' \
                + '分からなければ"help"と入力してください．'
            lbl = tkinter.Label(pane, text=t, justify='left')
            lbl.pack(side='top', anchor='w')
            self.etr = tkinter.Entry(pane, font=fon, width=50)
            self.etr.pack(side='top')
            self.etr.insert(0, self.init)
            self.bind('<Key>', self.process_key)
            # self.bind('<Key-Tab>', self.process_key_tab)
            # self.bind('<Key-Up>', self.process_key_up)
            # self.bind('<Key-Down>', self.process_key_down)
            super().body(pane)
            # self.grab_set()
            return self.etr

        def apply(self):
            com = self.etr.get()
            Makdo.Minibuffer.history[-1] = com
            if len(self.history) > 1:
                if Makdo.Minibuffer.history[-2] == com:
                    Makdo.Minibuffer.history.pop(-1)
            if com == '':
                return False
            for mc in self.minibuffer_commands:
                if mc.command_text == com:
                    for mt in mc.method_texts:
                        eval(mt)
                    return True
            Makdo.Minibuffer(self, self.mother, com)

        def set_return_to(self):
            self.return_to = self.mother.txt
            if self.mother.current_pane == 'sub':
                self.return_to = self.mother.sub
            elif self.mother.current_pane == 'toc':
                self.return_to = self.mother.toc_cvs_frm

        def get_return_to(self):
            return self.return_to

        class Help(tkinter.simpledialog.Dialog):

            def __init__(self, minibuffer, mother, message):
                self.minibuffer = minibuffer
                self.mother = mother
                self.message = message
                super().__init__(minibuffer, title='ヘルプ')

            def body(self, minibuffer):
                res = '^((?:.*\n){40})((?:.|\n)*)$'
                mes1 = re.sub(res, '\\1', self.message)
                mes2 = re.sub(res, '\\2', self.message)
                fon = self.mother.gothic_font.copy()
                fon['size'] -= 6
                fon['weight'] = 'bold'
                frm = tkinter.Frame(minibuffer)
                frm.pack()
                self.frm_u = tkinter.Frame(frm)
                self.frm_d = tkinter.Frame(frm)
                self.frm_u.pack(side='top', anchor='n')
                self.frm_d.pack(side='bottom', anchor='s')
                frm1 = tkinter.Frame(self.frm_u)
                # frm2 = tkinter.Frame(self.frm_u)
                frm3 = tkinter.Frame(self.frm_u)
                frm1.pack(side='left', anchor='n')
                # frm2.pack(side='left', anchor='n')
                frm3.pack(side='left', anchor='n')
                lbl1 = tkinter.Label(frm1, font=fon, text=mes1, justify='left')
                # lbl2 = tkinter.Label(frm2, font=fon, text='　')
                lbl3 = tkinter.Label(frm3, font=fon, text=mes2, justify='left')
                lbl1.pack(side='left', anchor='n')
                # lbl2.pack(side='left', anchor='n')
                lbl3.pack(side='left', anchor='n')
                self.bind('<Key-Return>', self.ok)
                self.bind('<Key-Escape>', self.ok)
                btn = tkinter.Button(self.frm_d, text='OK',
                                     width=10, command=self.ok)
                btn.pack(side='bottom', anchor='s')
                # self.grab_set()
                return btn

            def buttonbox(self):
                # btn = tkinter.Button(self.frm_d, text='OK',
                #                           width=10, command=self.ok)
                # btn.pack(side='bottom', anchor='s')
                return

        def process_key(self, key):
            if 'prev_key' not in vars(self):
                self.prev_key = None
            k1 = self.mother._get_key(key)
            k2 = self.prev_key
            if k1 is None:
                return 'break'  # shift, control, alt, mode_switch
            # TAB
            if self.mother._is_key(k1, 'Tab', 'C-i', 'C-g'):
                com = self.etr.get()
                if self.mother._is_key(k2, 'Tab', 'C-i', 'C-g'):
                    if len(self.command_candidates) == 0:
                        return 'break'
                    if len(self.command_candidates) == 1:
                        return  # Entry -> OK -> Cancel -> Entry
                    self.etr.delete(0, 'end')
                    for i in range(len(self.command_candidates)):
                        if com != self.command_candidates[i]:
                            continue
                        if i < len(self.command_candidates) - 1:
                            self.etr.insert(0, self.command_candidates[i + 1])
                        else:
                            self.etr.insert(0, self.command_candidates[0])
                        break
                    else:
                        self.etr.insert(0, self.command_candidates[0])
                    return 'break'
                else:
                    self.process_key_tab(key)
                    self.prev_key = 'Tab'
                    if com == self.etr.get() and \
                       com in self.command_candidates:
                        return  # Entry -> OK -> Cancel -> Entry
                    return 'break'
            self.prev_key = None
            if self.mother._is_key(k1, 'Up', 'C-r', 'C-o'):
                self.process_key_up(key)
                return 'break'
            if self.mother._is_key(k1, 'Down', 'C-n', 'C-l'):
                self.process_key_down(key)
                return 'break'

        def process_key_tab(self, key):
            com = self.etr.get()
            if com == '':
                return  # empty
            self.command_candidates = []
            for c in self.commands:
                if com == c:
                    self.command_candidates.append(c)
                    return  # completed
                if re.match('^' + com, c):
                    self.command_candidates.append(c)
            x = ''
            for y in self.command_candidates:
                if x == '':
                    x = y
                else:
                    nx, ny = len(x), len(y)
                    nz = min(nx, ny)
                    for n in range(nz):
                        if x[:n+1] != y[:n+1]:
                            if n == 0:
                                x = ''
                            else:
                                x = x[:n]
                            break
                    else:
                        x = x[:nz]
            if x != '':
                self.etr.delete(0, 'end')
                self.etr.insert(0, x)

        def process_key_up(self, event):
            if self.history_number == 0:
                Makdo.Minibuffer.history[-1] = self.etr.get()
            if self.history_number < len(self.history) - 1:
                self.history_number += 1
                self.etr.delete(0, 'end')
                self.etr.insert(0, self.history[-self.history_number - 1])

        def process_key_down(self, event):
            if self.history_number > 0:
                self.history_number -= 1
                self.etr.delete(0, 'end')
                self.etr.insert(0, self.history[-self.history_number - 1])

    ##########################
    # MENU CONFIGURATION

    def _make_menu_configuration(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='設定(S)', menu=menu, underline=3)
        #
        self.dont_show_help = tkinter.BooleanVar(value=False)
        if self.args_dont_show_help:
            self.dont_show_help.set(True)
        elif self.file_dont_show_help:
            self.dont_show_help.set(True)
        menu.add_checkbutton(label='ヘルプを表示しない',
                             variable=self.dont_show_help,
                             command=self.show_config_help_message)
        menu.add_separator()
        #
        self.is_toc_display_mode = tkinter.BooleanVar(value=False)
        if Makdo.file_is_toc_display_mode:
            self.is_toc_display_mode = tkinter.BooleanVar(value=True)
        menu.add_checkbutton(label='目次を表示',
                             variable=self.is_toc_display_mode,
                             command=self.settle_or_remove_toc)
        menu.add_separator()
        #
        self.is_read_only = tkinter.BooleanVar(value=False)
        if self.args_read_only:
            self.is_read_only.set(True)
        menu.add_checkbutton(label='読取専用',
                             variable=self.is_read_only)
        menu.add_separator()
        #
        self._make_submenu_background_color(menu)
        self._make_submenu_character_size(menu)
        menu.add_separator()
        #
        self.paint_keywords = tkinter.BooleanVar(value=False)
        if self.args_paint_keywords:
            self.paint_keywords.set(True)
        elif self.file_paint_keywords:
            self.paint_keywords.set(True)
        menu.add_checkbutton(label='キーワードに色付け',
                             variable=self.paint_keywords,
                             command=self.show_config_help_message)
        Makdo.keywords_to_paint = ''
        if self.args_keywords_to_paint is not None:
            Makdo.keywords_to_paint = self.args_keywords_to_paint
        elif self.file_keywords_to_paint is not None:
            Makdo.keywords_to_paint = self.file_keywords_to_paint
        menu.add_command(label='色付けするキーワードを設定',
                         command=self.set_keywords_to_paint)
        menu.add_separator()
        #
        self.use_regexps = tkinter.BooleanVar(value=False)
        if self.args_use_regexps:
            self.use_regexps.set(True)
        elif self.file_use_regexps:
            self.use_regexps.set(True)
        menu.add_checkbutton(label='検索・置換に正規表現を使う',
                             variable=self.use_regexps,
                             command=self.args_use_regexps)
        menu.add_separator()
        #
        self._make_submenu_digit_separator(menu)
        menu.add_separator()
        #
        self.make_backup_file = tkinter.BooleanVar(value=False)
        if self.args_make_backup_file:
            self.make_backup_file.set(True)
        elif self.file_make_backup_file:
            self.make_backup_file.set(True)
        menu.add_checkbutton(label='バックアップファイルを残す',
                             variable=self.make_backup_file,
                             command=self.show_config_help_message)
        menu.add_separator()
        #
        menu.add_command(label='OneDriveフォルダを設定',
                         command=self.set_onedrive_directory)
        menu.add_separator()
        #
        menu.add_command(label='設定を保存',
                         command=self.save_configurations)
        # menu.add_separator()

    ################
    # COMMAND

    def settle_or_remove_toc(self):
        is_toc_display_mode = self.is_toc_display_mode.get()
        if is_toc_display_mode:
            self.settle_toc()
        else:
            self.remove_toc()

    def settle_toc(self):
        if 'toc_cvs' in vars(self):
            return
        if 'colors' not in vars(self):
            return
        cols = self.colors
        self.pnd.remove(self.pnd_l)
        self.pnd.remove(self.pnd_r)
        self.pnd.update()
        width = int(self.pnd.winfo_width() / 5)
        if width > 256:
            width = 256
        self.pnd.add(self.pnd_l, minsize=100, width=width)
        background_color = self.background_color.get()
        cvs = tkinter.Canvas(self.pnd_l, bg=cols['bg'])
        cvs_frm = tkinter.Frame(cvs, bg=cols['bg'])
        cvs.pack(expand=True, fill='both', anchor='w')
        scb = tkinter.Scrollbar(cvs, orient='vertical', command=cvs.yview)
        scb.pack(side='right', fill='y')
        cvs['yscrollcommand'] = scb.set
        cvs.create_window((0, 0), window=cvs_frm, anchor='nw')
        cvs_frm.bind(
            '<Configure>',
            lambda e: cvs.configure(scrollregion=cvs.bbox('all')))
        cvs.bind('<Button-1>', self.toc_process_button1)
        cvs.bind('<Button-2>', self.close_mouse_menu)
        cvs.bind('<Button-3>', self.close_mouse_menu)
        # cvs_frm.bind('<Up>', lambda e: cvs.yview_scroll(-1, 'units'))
        # cvs_frm.bind('<Down>', lambda e: cvs.yview_scroll(1, 'units'))
        cvs_frm.bind('<Prior>', lambda e: cvs.yview_scroll(-10, 'units'))
        cvs_frm.bind('<Next>', lambda e: cvs.yview_scroll(10, 'units'))
        if sys.platform == 'win32':
            cvs_frm.bind_all('<MouseWheel>', self.scroll_by_mouse_wheel)
        elif sys.platform == 'darwin':
            cvs_frm.bind_all('<MouseWheel>', self.scroll_by_mouse_wheel)
        elif sys.platform == 'linux':
            cvs_frm.bind_all('<Button-4>', self.scroll_by_mouse_wheel)
            cvs_frm.bind_all('<Button-5>', self.scroll_by_mouse_wheel)
        # if sys.platform == 'win32':
        #     cvs_frm.bind_all(
        #         '<MouseWheel>',
        #         lambda e: cvs.yview_scroll(- int(e.delta / 100), 'units'))
        # elif sys.platform == 'darwin':
        #     cvs_frm.bind_all(
        #         '<MouseWheel>',
        #         lambda e: cvs.yview_scroll(- int(e.delta / 120), 'units'))
        # elif sys.platform == 'linux':
        #     cvs_frm.bind_all('<4>', lambda e: cvs.yview_scroll(-1, 'units'))
        #     cvs_frm.bind_all('<5>', lambda e: cvs.yview_scroll(+1, 'units'))
        cvs_frm.bind('<Key>', self.toc_process_key)
        self.toc_cvs = cvs
        self.toc_cvs_frm = cvs_frm
        self.toc_lines = []
        self.toc_screen_data = []
        self.update_toc()
        self.pnd.add(self.pnd_r, minsize=100)

    def toc_process_key(self, event):
        k1 = self._get_key(event)
        k2 = self.key_history[-1]
        if k1 is None:
            return 'break'  # shift, control, alt, mode_switch
        self.key_history.append(k1)
        self.key_history.pop(0)
        if self._is_key(k1, 'F19', 'C-x', 'C-b'):
            self._any_process_escp1()
            return 'break'
        elif self._is_key(k1, 'Up', 'C-r', 'C-o'):
            if self._is_key(k2, 'F19', 'C-x', 'C-b'):
                self._jump_to_prev_pane()
            else:
                self.toc_cvs.yview_scroll(-1, 'units')
            return 'break'
        elif self._is_key(k1, 'Down', 'C-n', 'C-l'):
            if self._is_key(k2, 'F19', 'C-x', 'C-b'):
                self._jump_to_next_pane()
            else:
                self.toc_cvs.yview_scroll(1, 'units')
            return 'break'
        elif self._is_key(k1, 'Tab', 'C-i', 'C-g'):
            self._jump_to_next_pane()
            return 'break'
        self.set_message_on_status_bar('目次ウィンドウです')

    def toc_process_button1(self, event):
        self.close_mouse_menu()
        self.toc_cvs_frm.focus_force()

    def remove_toc(self):
        if 'toc_cvs' not in vars(self):
            return
        self.pnd.remove(self.pnd_l)
        self.pnd.remove(self.pnd_r)
        self.toc_cvs.destroy()
        del self.toc_cvs
        self.toc_lines = []
        self.toc_screen_data = []
        self.pnd.add(self.pnd_r, minsize=100)
        self.txt.focus_set()
        self.current_pane = 'txt'

    def update_toc(self):
        res_chapter = '^(\\${1,5})(-\\$+)*(\\s.*)?$'
        res_section = '^(#{1,8})(-#+)*(\\s.*)?$'
        # CONFIRM
        is_toc_display_mode = self.is_toc_display_mode.get()
        if not is_toc_display_mode:
            return
        # COLOR
        if 'colors' not in vars(self):
            return
        cols = self.colors
        # TOC LINES
        file_text = self.txt.get('1.0', 'end-1c')
        toc_lines = []
        c4, c3, c2, n, line, is_in_comment = '', '', '', 1, '', False
        for c1 in file_text + '\n':
            if is_in_comment:
                if c3 == '-' and c2 == '-' and c1 == '>':
                    is_in_comment = False
                    continue
            else:
                if c4 == '<' and c3 == '!' and c2 == '-' and c1 == '-':
                    is_in_comment = True
                    line = line[:-3]
                    continue
            if c1 == '\n':
                if line != '':
                    if line[0] == '.':
                        res = '^\\.{3}\\[[0-9]+\\]#+(-#+)*\\s+.*$'
                        if re.match(res, line):
                            break
                    elif line[0] == '$':
                        if re.match(res_chapter, line):
                            toc_lines.append([n, line])
                    elif line[0] == '#':
                        if re.match(res_section, line):
                            if re.match('^.*\\.{3}\\[[0-9]\\]+', line):
                                res = '^(\\S+)\\s+(.*)\\.{3}\\[[0-9]\\]+'
                                line = re.sub(res, '\\1 > \\2', line)
                            toc_lines.append([n, line])
                    elif (line[0] == '`' or line[0] == '*' or
                          line[0] == '-' or line[0] == '+' or
                          line[0] == '>' or line[0] == '<' or
                          line[0] == '^' or line[0] == '_' or
                          line[0] == '@' or line[0] == '|'):
                        line = self._remove_head_and_tail_fds(line)
                        if re.match(res_chapter, line):
                            toc_lines.append([n, line])
                        elif re.match(res_section, line):
                            toc_lines.append([n, line])
                    if len(toc_lines) > 0:
                        t = toc_lines[-1]
                        if n == t[0] + 1:
                            if re.match('^\\S+(\\s+\\\\?)?$', t[1]):
                                t[1] = re.sub('\\s+\\\\?$', '', t[1]) \
                                    + ' / ' + re.sub('^\\s+', '', line)
                line = ''
                n += 1
            else:
                if not is_in_comment:
                    line += c1
            c4, c3, c2 = c3, c2, c1
        # REWRITE
        if self.toc_lines != toc_lines:
            fon = self.gothic_font.copy()
            fon['size'] -= 3
            fon['weight'] = 'bold'
            bc, lfc, bfc = cols['bg'], cols['fg'], cols['fg']
            for i, line in enumerate(toc_lines):
                n, t = line[0], line[1]
                if i < len(self.toc_lines):
                    if self.toc_lines[i][0] == n and self.toc_lines[i][1] == t:
                        continue
                if re.match(res_chapter, t):
                    h = re.sub(res_chapter, '\\1', t)
                    if len(h) <= 5:
                        bfc = cols['cp'][len(h) - 1]
                elif re.match(res_section, t):
                    h = re.sub(res_section, '\\1', t)
                    if len(h) <= 8:
                        bfc = cols['sc'][len(h) - 1]
                if i > len(self.toc_screen_data) - 1:
                    tsd = self.TocScreenDatum(self, self.toc_cvs_frm,
                                              fon, bc, lfc)
                    self.toc_screen_data.append(tsd)
                self.toc_screen_data[i].settle(n, t, bfc)
            while len(toc_lines) < len(self.toc_screen_data):
                self.toc_screen_data[-1].remove()
                self.toc_screen_data.pop(-1)
            self.toc_lines = toc_lines
        # CURRENT LINE
        m = len(self.toc_screen_data) - 1
        if m >= 0:
            v_pos = self._get_v_position_of_insert(self.txt)
            if v_pos < self.toc_screen_data[0].number:
                v_pos = -1
            for i, curr_tsd in enumerate(self.toc_screen_data):
                if i < m:
                    next_tsd = self.toc_screen_data[i + 1]
                is_here = False
                if v_pos > 0:
                    if i < m:
                        if v_pos < next_tsd.number:
                            is_here, v_pos = True, -1
                    else:
                        if True:
                            is_here, v_pos = True, -1
                if is_here:
                    curr_tsd.set_mark()
                else:
                    curr_tsd.unset_mark()

    class TocScreenDatum:

        def __init__(self, makdo, frame, font, bg, lfg):
            self.makdo = makdo
            self.mother = frame
            self.frame = tkinter.Frame(frame, bg=bg)
            self.frame.pack(side='top', anchor='w')
            self.checkbox = tkinter.Label(self.frame, font=font, bg=bg, fg=lfg)
            self.checkbox.pack(side='left')
            self.button = tkinter.Label(self.frame, font=font, bg=bg)
            self.button.pack(side='left')
            self.button.bind('<Button-1>', self.goto_toc_line)

        def settle(self, number, text, bfg):
            self.number = number
            # self.text = text
            self.checkbox['text'] = ' '
            self.button['text'] = text + ' (' + str(number) + ')'
            self.button['fg'] = bfg

        def remove(self):
            self.checkbox.destroy()
            self.button.destroy()
            self.frame.destroy()

        def set_mark(self):
            if self.checkbox['text'] != '*':
                self.checkbox['text'] = '*'

        def unset_mark(self):
            if self.checkbox['text'] != ' ':
                self.checkbox['text'] = ' '

        def goto_toc_line(self, event):
            self.makdo.close_mouse_menu()
            self.makdo.txt.mark_set('insert', str(self.number) + '.0')
            self.makdo.update_toc()
            self.makdo.txt.focus_force()
            self.makdo._put_back_cursor_to_pane(self.makdo.txt)

    ################
    # SUBMENU BACKGROUND COLOR

    def _make_submenu_background_color(self, menu):
        submenu = tkinter.Menu(self.mnb, tearoff=False)
        menu.add_cascade(label='背景色', menu=submenu)
        self.background_color \
            = tkinter.StringVar(value='W')
        if self.args_background_color is not None:
            self.background_color.set(self.args_background_color)
        elif self.file_background_color is not None:
            self.background_color.set(self.file_background_color)
        color_themes = {'W': '白色', 'B': '黒色', 'G': '緑色'}
        for c in color_themes:
            submenu.add_radiobutton(label=color_themes[c],
                                    variable=self.background_color, value=c,
                                    command=self.set_background_color)

    ################
    # COMMAND

    def set_keywords_to_paint(self):
        t = '色付けするキーワードを設定'
        m = '色付けするキーワードを設定してください．'
        i = Makdo.keywords_to_paint
        ktp = self.KeywordsToPaintDialog(self.txt, self, t, m, i)
        v = ktp.get_value()
        if v is not None:
            Makdo.keywords_to_paint = ktp.get_value()

    class KeywordsToPaintDialog(tkinter.simpledialog.Dialog):

        def __init__(self, pane, mother, title, prompt, init):
            self.pane = pane
            self.mother = mother
            self.prompt = prompt
            self.init = init
            self.inits = ['' for i in range(20)]
            self.value = None
            self.values = ['' for i in range(20)]
            super().__init__(pane, title=title)

        def body(self, pane):
            fon = self.mother.gothic_font
            prompt \
                = tkinter.Label(pane, text=self.prompt + '\n', justify='left')
            prompt.pack(side='top', anchor='w')
            kws = []
            kw = ''
            for c in Makdo.keywords_to_paint + '|':
                if re.match(NOT_ESCAPED + '\\|$', kw + c):
                    kw = kw.replace('\\|', '|')
                    kw = kw.replace('\\\\', '\\')
                    kws.append(kw)
                    kw = ''
                else:
                    kw += c
            for i in range(min(20, len(kws))):
                if kws[i] != '':
                    self.inits[i] = kws[i]
            self.entries = []
            for i in range(4):
                frm = tkinter.Frame(pane)
                frm.pack(side='left')
                for j in range(5):
                    self.entry = tkinter.Entry(frm, width=15, font=fon)
                    self.entry.pack(side='top')
                    self.entry.insert(0, self.inits[(i * 5) + j])
                    self.entries.append(self.entry)
            super().body(pane)
            return self.entries[0]

        def apply(self):
            for i in range(20):
                self.values[i] = self.entries[i].get()
                self.values[i] = self.values[i].replace('\\', '\\\\')
                self.values[i] = self.values[i].replace('|', '\\|')
            self.value = '|'.join(self.values)
            self.value = re.sub('\\|+', '|', self.value)
            self.value = re.sub('\\|$', '', self.value)

        def get_value(self):
            return self.value

    ################
    # SUBMENU CHARACTER SIZE

    def _make_submenu_character_size(self, menu):
        submenu = tkinter.Menu(self.mnb, tearoff=False)
        menu.add_cascade(label='文字サイズ', menu=submenu)
        sizes = [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36,
                 42, 48, 54, 60, 66, 72, 78, 84, 90, 96, 102, 108]
        self.font_size = tkinter.IntVar(value=18)
        if self.args_font_size is not None:
            self.font_size.set(self.args_font_size)
        elif self.file_font_size is not None:
            self.font_size.set(self.file_font_size)
        for s in sizes:
            submenu.add_radiobutton(label=str(s) + 'px',
                                    variable=self.font_size, value=s,
                                    command=self.set_character_size)

    ######
    # COMMAND

    def set_background_color(self):
        self.show_config_help_message()
        self.set_font()

    def set_character_size(self):
        self.show_config_help_message()
        self.set_font()

    def set_font(self):
        cs = COLOR_SPACE
        background_color = self.background_color.get()
        size = self.font_size.get()
        self.gothic_font['size'] = size
        self.mincho_font['size'] = size
        # BASIC FONT
        self.txt['font'] = self.gothic_font
        self.stb_sor1['font'] = self.gothic_font
        self.stb_sor2['font'] = self.gothic_font
        self.txt.tag_config('error_tag', foreground='#FF0000')
        self.sub.tag_config('error_tag', foreground='#FF0000')
        self.txt.tag_config('search_tag', background='#777777')
        self.sub.tag_config('search_tag', background='#777777')
        # TAB
        font = tkinter.font.Font(font=self.txt['font'])
        tab_width = font.measure(' ' * TAB_WIDTH)
        self.txt.config(tabs=tab_width)
        self.sub.config(tabs=tab_width)
        # COLOR FONT
        if background_color == 'W':
            self.txt.config(bg='white', fg='black')
            self.txt.tag_config('eol_tag', background='#EEEEEE')
            self.txt.tag_config('eof_tag', background='#EEEEEE')
            self.txt.tag_config('eol_and_eof_tag', background='#CCCCCC')
            # self.sub.tag_config('eol_tag', background='#EEEEEE')
            self.sub.tag_config('eof_tag', background='#EEEEEE')
            # self.sub.tag_config('eol_and_eof_tag', background='#CCCCCC')
            self.txt.tag_config('akauni_tag', background='#CCCCCC')
            self.sub.tag_config('akauni_tag', background='#CCCCCC')
            self.txt.tag_config('hsp_tag', foreground='#C8C8FF',
                                underline=True)                   # (0.80, 240)
            self.txt.tag_config('fsp_tag', foreground='#90D9FF',
                                underline=True)                   # (0.80, 200)
            self.txt.tag_config('tab_tag', background='#C9FFEC')  # (0.95, 160)
            self.colors = {
                'bg': 'white', 'cg': '#CCCCCC', 'fg': 'black',
                'cp': [cs[21][1], cs[22][1], cs[23][1], cs[24][1], cs[25][1]],
                'sc': [cs[3][1], cs[4][1], cs[5][1], cs[6][1],
                       cs[7][1], cs[8][1], cs[9][1], cs[10][1]]
            }
        elif background_color == 'B':
            self.txt.config(bg='black', fg='white')
            self.txt.tag_config('eol_tag', background='#333333')
            self.txt.tag_config('eof_tag', background='#333333')
            self.txt.tag_config('eol_and_eof_tag', background='#666666')
            # self.sub.tag_config('eol_tag', background='#333333')
            self.sub.tag_config('eof_tag', background='#333333')
            # self.sub.tag_config('eol_and_eof_tag', background='#666666')
            self.txt.tag_config('akauni_tag', background='#666666')
            self.sub.tag_config('akauni_tag', background='#666666')
            self.txt.tag_config('hsp_tag', foreground='#7676FF',
                                underline=True)                   # (0.50, 240)
            self.txt.tag_config('fsp_tag', foreground='#009AED',
                                underline=True)                   # (0.50, 200)
            self.txt.tag_config('tab_tag', background='#005437')  # (0.25, 160)
            self.colors = {
                'bg': 'black', 'cg': '#666666', 'fg': 'white',
                'cp': [cs[21][2], cs[22][2], cs[23][2], cs[24][2], cs[25][2]],
                'sc': [cs[3][2], cs[4][2], cs[5][2], cs[6][2],
                       cs[7][2], cs[8][2], cs[9][2], cs[10][2]]
            }
        elif background_color == 'G':
            self.txt.config(bg='darkgreen', fg='lightyellow')  # 006400/FFFFE0
            self.txt.tag_config('eol_tag', background='#117511')
            self.txt.tag_config('eof_tag', background='#117511')
            self.txt.tag_config('eol_and_eof_tag', background='#339733')
            # self.sub.tag_config('eol_tag', background='#117511')
            self.sub.tag_config('eof_tag', background='#117511')
            # self.sub.tag_config('eol_and_eof_tag', background='#339733')
            self.txt.tag_config('akauni_tag', background='#888888')
            self.sub.tag_config('akauni_tag', background='#888888')
            self.txt.tag_config('hsp_tag', foreground='#7676FF',
                                underline=True)                   # (0.50, 240)
            self.txt.tag_config('fsp_tag', foreground='#009AED',
                                underline=True)                   # (0.50, 200)
            self.txt.tag_config('tab_tag', background='#00754C')  # (0.35, 160)
            self.colors = {
                'bg': 'darkgreen', 'cg': '#339733', 'fg': 'lightyellow',
                'cp': [cs[21][2], cs[22][2], cs[23][2], cs[24][2], cs[25][2]],
                'sc': [cs[3][2], cs[4][2], cs[5][2], cs[6][2],
                       cs[7][2], cs[8][2], cs[9][2], cs[10][2]]
            }
        for u in ['-x', '-u']:
            und = False if u == '-x' else True
            for f in ['-g', '-m']:
                # WHITE
                if f == '-g':
                    fon = self.gothic_font.copy()
                else:
                    fon = self.mincho_font.copy()
                for i in range(3):
                    a = '-XXX'
                    y = '-' + str(i)
                    tag = 'c' + a + y + f + u
                    if background_color == 'W':
                        col = BLACK_SPACE[i]
                    elif background_color == 'B':
                        col = WHITE_SPACE[i]
                    elif background_color == 'G':
                        col = LIGHTYELLOW_SPACE[i]
                    self.txt.tag_config(tag, font=fon,
                                        foreground=col, underline=und)
                # COLOR
                if f == '-g':
                    fon = self.gothic_font.copy()
                else:
                    fon = self.mincho_font.copy()
                fon['weight'] = 'bold'
                for i in range(3):  # lightness
                    y = '-' + str(i)
                    for j, c in enumerate(COLOR_SPACE):  # angle
                        a = '-' + str(j * 10)
                        tag = 'c' + a + y + f + u  # example: c-120-1-g-x
                        if background_color == 'W':
                            col = c[i]
                        elif background_color == 'B':
                            col = c[i + 1]
                        elif background_color == 'G':
                            col = c[i + 1]
                        self.txt.tag_config(tag, font=fon,
                                            foreground=col, underline=und)
                        self.sub.tag_config(tag, font=fon,
                                            foreground=col, underline=und)
        is_toc_display_mode = self.is_toc_display_mode.get()
        if is_toc_display_mode:
            if 'toc_cvs' in vars(self):
                self.remove_toc()
            if 'toc_cvs' not in vars(self):
                self.settle_toc()

    ################
    # SUBMENU DIGIT SEPARATOR

    def _make_submenu_digit_separator(self, menu):
        submenu = tkinter.Menu(self.mnb, tearoff=False)
        menu.add_cascade(label='計算結果', menu=submenu)
        #
        self.digit_separator = tkinter.StringVar(value='4')
        submenu.add_radiobutton(label='桁区切りなし（12345678）',
                                variable=self.digit_separator, value='0',
                                command=self.show_config_help_message)
        submenu.add_radiobutton(label='3桁区切り（12,345,678）',
                                variable=self.digit_separator, value='3',
                                command=self.show_config_help_message)
        submenu.add_radiobutton(label='4桁区切り（1234万5678）',
                                variable=self.digit_separator, value='4',
                                command=self.show_config_help_message)
        # menu.add_separator()

    ################
    # COMMAND

    def set_onedrive_directory(self) -> bool:
        od = self.onedrive_directory
        if od is None:
            d = os.path.expanduser('~/OneDrive')
            if os.path.exists(d):
                if os.path.exists(d) and os.path.isdir(d):
                    od = d
        ti = 'OneDriveフォルダを設定'
        if od is None:
            od = tkinter.filedialog.askdirectory(title=ti)
        else:
            od = tkinter.filedialog.askdirectory(title=ti, initialdir=od)
        if od == () or od == '':
            return False
        self.onedrive_directory = od
        return True

    ################
    # CONFIGURATION FILE

    def get_and_set_configurations(self) -> bool:
        if not os.path.exists(CONFIG_DIR):
            os.mkdir(CONFIG_DIR)
        if not os.path.exists(CONFIG_FILE):
            open(CONFIG_FILE, 'w').close()
        try:
            with open(CONFIG_FILE, 'r') as f:
                lines = f.read().replace('\r', '')
        except BaseException:
            return False
        for line in lines.split('\n'):
            line = line.rstrip()
            item = re.sub('^\\s*(\\S*)\\s*:\\s*(.*)\\s*$', '\\1', line)
            valu = re.sub('^\\s*(\\S*)\\s*:\\s*(.*)\\s*$', '\\2', line)
            if item == 'dont_show_help':
                if valu == 'True':
                    Makdo.file_dont_show_help = True
                else:
                    Makdo.file_dont_show_help = False
            elif item == 'make_backup_file':
                if valu == 'True':
                    Makdo.file_make_backup_file = True
                elif valu == 'False':
                    Makdo.file_make_backup_file = False
            elif item == 'is_toc_display_mode':
                if valu == 'True':
                    Makdo.file_is_toc_display_mode = True
                else:
                    Makdo.file_is_toc_display_mode = False
            elif item == 'background_color':
                if valu == 'W' or valu == 'B' or valu == 'G':
                    Makdo.file_background_color = valu
            elif item == 'font_size':
                if re.match('^[0-9]+$', valu) and (int(valu) % 3) == 0:
                    Makdo.file_font_size = int(valu)
            elif item == 'paint_keywords':
                if valu == 'True':
                    Makdo.file_paint_keywords = True
                elif valu == 'False':
                    Makdo.file_paint_keywords = False
            elif item == 'keywords_to_paint':
                Makdo.file_keywords_to_paint = valu
            elif item == 'shortcut_key_layout':
                Makdo.file_shortcut_key_layout = valu
            elif item == 'use_regexps':
                if valu == 'True':
                    Makdo.file_use_regexps = True
                elif valu == 'False':
                    Makdo.file_use_regexps = False
            elif item == 'digit_separator':
                if valu == '3' or valu == '4':
                    Makdo.file_digit_separator = valu
            elif item == 'onedrive_directory':
                if os.path.exists(valu) and os.path.isdir(valu):
                    self.onedrive_directory = valu
            elif item == 'epwing_directory':
                if os.path.exists(valu) and os.path.isdir(valu):
                    self.epwing_directory = valu
            elif item == 'openai_model':
                self.openai_model = valu
            elif item == 'openai_key':
                self.openai_key = valu
            elif item == 'llama_model_file':
                if os.path.exists(valu) and os.path.isfile(valu):
                    self.llama_model_file = valu
            elif item == 'llama_gpu_layers':
                if re.match('^(-1|[0-9]+)$', valu):
                    self.llama_gpu_layers = int(valu)
            elif item == 'llama_context_size':
                if re.match('^[0-9]+$', valu):
                    self.llama_context_size = int(valu)
            elif item == 'ollama_model':
                self.ollama_model = valu
        return True

    def save_configurations(self):
        if os.path.exists(CONFIG_FILE + '~'):
            os.chmod(CONFIG_FILE + '~', 0o600)
            os.remove(CONFIG_FILE + '~')
        if os.path.exists(CONFIG_FILE):
            os.rename(CONFIG_FILE, CONFIG_FILE + '~')
        with open(CONFIG_FILE, 'w') as f:
            f.write('dont_show_help:         '
                    + str(self.dont_show_help.get()) + '\n')
            f.write('make_backup_file:       '
                    + str(self.make_backup_file.get()) + '\n')
            f.write('is_toc_display_mode:    '
                    + str(self.is_toc_display_mode.get()) + '\n')
            f.write('background_color:       '
                    + self.background_color.get() + '\n')
            f.write('font_size:              '
                    + str(self.font_size.get()) + '\n')
            f.write('paint_keywords:         '
                    + str(self.paint_keywords.get()) + '\n')
            if self.keywords_to_paint != '':
                f.write('keywords_to_paint:      '
                        + self.keywords_to_paint + '\n')
            f.write('shortcut_key_layout:    '
                    + self.shortcut_key_layout.get() + '\n')
            f.write('use_regexps:            '
                    + str(self.use_regexps.get()) + '\n')
            f.write('digit_separator:        '
                    + str(self.digit_separator.get()) + '\n')
            if self.onedrive_directory is not None:
                f.write('onedrive_directory:     '
                        + self.onedrive_directory + '\n')
            if 'epwing_directory' in vars(self):
                f.write('epwing_directory:       '
                        + self.epwing_directory + '\n')
            if 'openai_model' in vars(self):
                f.write('openai_model:           '
                        + self.openai_model + '\n')
            if 'openai_key' in vars(self):
                f.write('openai_key:             '
                        + self.openai_key + '\n')
            if 'llama_model_file' in vars(self):
                f.write('llama_model_file:       '
                        + self.llama_model_file + '\n')
            if 'llama_gpu_layers' in vars(self):
                f.write('llama_gpu_layers:       '
                        + str(self.llama_gpu_layers) + '\n')
            if 'llama_context_size' in vars(self):
                f.write('llama_context_size:     '
                        + str(self.llama_context_size) + '\n')
            if 'ollama_model' in vars(self):
                f.write('ollama_model:           '
                        + self.ollama_model + '\n')
            self.set_message_on_status_bar('設定を保存しました')
        os.chmod(CONFIG_FILE, 0o400)

    ##########################
    # MENU INTERNET

    def _make_menu_internet(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='ネット(N)', menu=menu, underline=4)
        #
        menu.add_command(label='最新のMakdoを確認',
                         command=self.browse_makdo_home)
        menu.add_separator()
        #
        menu.add_command(label='人名・地名漢字を探す',
                         command=self.browse_ivs)
        menu.add_separator()
        #
        menu.add_command(label='goo辞書で調べる',
                         command=self.browse_goo_dictionary)
        menu.add_command(label='weblio辞書で調べる',
                         command=self.browse_weblio_dictionary)
        menu.add_command(label='Wikipediaで調べる',
                         command=self.browse_wikipedia)
        menu.add_separator()
        #
        menu.add_command(label='法律を調べる',
                         command=self.browse_law)
        menu.add_command(label='・日本国憲法',
                         command=self.browse_law_constitution_law)
        menu.add_command(label='・民法',
                         command=self.browse_law_civil_law)
        menu.add_command(label='・商法',
                         command=self.browse_law_commercial_law)
        menu.add_command(label='・会社法',
                         command=self.browse_law_corporation_law)
        menu.add_command(label='・民事訴訟法',
                         command=self.browse_law_civil_procedure)
        menu.add_command(label='・刑法',
                         command=self.browse_law_crime_law)
        menu.add_command(label='・刑事訴訟法',
                         command=self.browse_law_crime_procedure)
        menu.add_command(label='裁判所規則を調べる',
                         command=self.browse_rule_of_court)
        menu.add_separator()
        #
        menu.add_command(label='ChatGPTに接続',
                         command=self.browse_chatgpt)
        menu.add_command(label='OpenAIに接続',
                         command=self.browse_openai)
        menu.add_separator()
        #
        menu.add_command(label='Google Driveに接続',
                         command=self.browse_google_drive)
        menu.add_command(label='Microsoft OneDriveに接続',
                         command=self.browse_onedrive)
        menu.add_separator()
        #
        menu.add_command(label='mints（民事裁判書類電子提出システム）に接続',
                         command=self.browse_mints)
        # menu.add_separator()

    ################
    # COMMAND

    def browse_makdo_home(self):
        webbrowser.open('https://github.com/hata48915b/makdo/')

    def browse_ivs(self):
        c = ''
        if self.txt.tag_ranges('sel'):
            c = self.txt.get('sel.first', 'sel.last')
        elif 'akauni' in self.txt.mark_names():
            c = ''
            c += self.txt.get('akauni', 'insert')
            c += self.txt.get('insert', 'akauni')
        if len(c) == 1:
            d = re.sub('^0x', '', hex(ord(c))).upper()
            u = 'https://moji.or.jp/mojikibansearch/result' \
                + '?UCS%E6%BC%A2%E5%AD%97=' + d
            webbrowser.open(u)
            i = self.IvsDialog(self.txt, self, c)
        else:
            d = None
            u = 'https://moji.or.jp/mojikibansearch/basic'
            webbrowser.open(u)
            i = self.IvsDialog(self.txt, self)
        if len(c) == 1 and i.has_inserted:
            if self.txt.tag_ranges('sel'):
                self.txt.delete('sel.first', 'sel.first+1c')
            elif 'akauni' in self.txt.mark_names():
                if self.txt.get('akauni', 'insert') != '':
                    self.txt.delete('akauni', 'akauni+1c')
                elif self.txt.get('insert', 'akauni') != '':
                    self.txt.delete('akauni-1c', 'akauni')

    def browse_goo_dictionary(self):
        if self.txt.tag_ranges('sel'):
            w = self.txt.get('sel.first', 'sel.last')
            u = 'https://dictionary.goo.ne.jp/srch/all/' + w + '/m6u/'
            webbrowser.open(u)
        elif 'akauni' in self.txt.mark_names():
            w = ''
            w += self.txt.get('akauni', 'insert')
            w += self.txt.get('insert', 'akauni')
            u = 'https://dictionary.goo.ne.jp/srch/all/' + w + '/m6u/'
            webbrowser.open(u)
        else:
            webbrowser.open('https://dictionary.goo.ne.jp/')

    def browse_weblio_dictionary(self):
        if self.txt.tag_ranges('sel'):
            w = self.txt.get('sel.first', 'sel.last')
            u = 'https://www.weblio.jp/content/' + w
            webbrowser.open(u)
        elif 'akauni' in self.txt.mark_names():
            w = ''
            w += self.txt.get('akauni', 'insert')
            w += self.txt.get('insert', 'akauni')
            u = 'https://www.weblio.jp/content/' + w
            webbrowser.open(u)
        else:
            webbrowser.open('https://www.weblio.jp/')

    def browse_wikipedia(self):
        if self.txt.tag_ranges('sel'):
            w = self.txt.get('sel.first', 'sel.last')
            webbrowser.open('https://ja.wikipedia.org/wiki/' + w)
        if 'akauni' in self.txt.mark_names():
            w = ''
            w += self.txt.get('akauni', 'insert')
            w += self.txt.get('insert', 'akauni')
            webbrowser.open('https://ja.wikipedia.org/wiki/' + w)

    def browse_law(self):
        webbrowser.open('https://laws.e-gov.go.jp/')

    def browse_law_constitution_law(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/321CONSTITUTION')

    def browse_law_civil_law(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/129AC0000000089')

    def browse_law_commercial_law(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/132AC0000000048')

    def browse_law_corporation_law(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/417AC0000000086')

    def browse_law_civil_procedure(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/408AC0000000109')

    def browse_law_crime_law(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/140AC0000000045')

    def browse_law_crime_procedure(self):
        webbrowser.open('https://laws.e-gov.go.jp/law/323AC0000000131')

    def browse_rule_of_court(self):
        u = 'https://www.courts.go.jp/toukei_siryou/kisokusyu/index.html'
        webbrowser.open(u)

    def browse_chatgpt(self):
        webbrowser.open('https://chatgpt.com/')

    def browse_openai(self):
        webbrowser.open('https://openai.com/')

    def browse_google_drive(self):
        webbrowser.open('https://drive.google.com/drive/my-drive')

    def browse_onedrive(self):
        webbrowser.open('https://onedrive.live.com/')

    def browse_mints(self):
        webbrowser.open('https://www.mints.courts.go.jp/user/')

    ##########################

    def _make_menu_special(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='裏の技(Z)', menu=menu, underline=3)
        #
        submenu = tkinter.Menu(menu, tearoff=0)
        menu.add_cascade(label='ショートカットキーのレイアウト', menu=submenu)
        self.shortcut_key_layout = tkinter.StringVar(value='normal')
        submenu.add_radiobutton(variable=self.shortcut_key_layout,
                                label='normal', value='normal')
        submenu.add_radiobutton(variable=self.shortcut_key_layout,
                                label='qwerty', value='qwerty')
        submenu.add_radiobutton(variable=self.shortcut_key_layout,
                                label='akauni', value='akauni')
        menu.add_separator()
        #
        menu.add_command(label='取引履歴の見本を挿入',
                         command=self.insert_sample_trading_history)
        menu.add_command(label='利息・遅延損害金を計算',
                         command=self.calculate_interest_and_charge)
        menu.add_separator()
        #
        menu.add_command(label='Epwing形式の辞書で調べる',
                         command=self.look_in_epwing)
        menu.add_command(label='Epwing形式の辞書フォルダを設定',
                         command=self.set_epwing_directory)
        menu.add_separator()
        #
        menu.add_command(label='OpenAIに質問（外部処理、有料）',
                         command=self.open_openai)
        menu.add_command(label='OpenAIのモデルを設定',
                         command=self.set_openai_model)
        menu.add_command(label='OpenAIのキーを設定',
                         command=self.set_openai_key)
        menu.add_separator()
        #
        menu.add_command(label='LlamaにRAGなしで質問（内部処理、無料）',
                         command=self.open_llama_without_rag)
        menu.add_command(label='LlamaにRAGありで質問（内部処理、無料）',
                         command=self.open_llama_with_rag)
        menu.add_command(label='Llamaのモデルファイルを設定',
                         command=self.set_llama_model_file)
        menu.add_command(label='LlamaのGPUで処理するレイヤーの数を設定',
                         command=self.set_llama_gpu_layers)
        menu.add_command(label='Llamaのコンテクストサイズを設定',
                         command=self.set_llama_context_size)
        menu.add_command(label='LlamaのRAGデータを編集',
                         command=self.edit_llama_rag_data)
        menu.add_separator()
        #
        menu.add_command(label='Ollamaに質問（内部処理、無料）',
                         command=self.open_ollama)
        menu.add_command(label='Ollamaのモデルを設定',
                         command=self.set_ollama_model)
        # menu.add_separator()

    @staticmethod
    def _show_message_reducing_functions():
        n = '警告'
        m = 'この機能は使用できません．\n\n' \
            + '実行ファイル形式は、\n' \
            + '起動を早くするため、\n' \
            + '機能の一部を落としています．'
        tkinter.messagebox.showwarning(n, m)

    def insert_sample_trading_history(self) -> None:
        self._show_message_reducing_functions()

    def calculate_interest_and_charge(self) -> None:
        self._show_message_reducing_functions()

    def look_in_epwing(self) -> None:
        self._show_message_reducing_functions()

    def set_epwing_directory(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def open_openai(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def set_openai_model(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def set_openai_key(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def open_llama_without_rag(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def open_llama_with_rag(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def set_llama_model_file(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def set_llama_gpu_layers(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def set_llama_context_size(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def edit_llama_rag_data(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def open_ollama(self) -> bool:
        self._show_message_reducing_functions()
        return False

    def set_ollama_model(self) -> bool:
        self._show_message_reducing_functions()
        return False

    ##########################
    # MENU HELP

    def _make_menu_help(self):
        menu = tkinter.Menu(self.mnb, tearoff=False,
                            postcommand=self.close_mouse_menu)
        self.mnb.add_cascade(label='ヘルプ(H)', menu=menu, underline=4)
        #
        menu.add_command(label='文字情報',
                         command=self.show_char_info)
        menu.add_separator()
        #
        menu.add_command(label='ヘルプ(H)', underline=4,
                         command=self.show_help)
        menu.add_separator()
        #
        menu.add_command(label='ライセンス情報(F)', underline=8,
                         command=self.show_license_info)
        menu.add_separator()
        #
        menu.add_command(label='Makdoについて(A)', underline=10,
                         command=self.show_about_makdo)
        # menu.add_separator()

    ################
    # COMMAND

    def show_char_info(self):
        n = '文字情報'
        c = self.txt.get('insert', 'insert+1c')
        if c != '' and c != '\n':
            m = ''
            if c == ' ':
                m += '文字：\t（半角スペース）\n'
            elif c == '\t':
                m += '文字：\t（水平タブ）\n'
            elif c == '\u3000':
                m += '文字：\t（全角スペース）\n'
            else:
                m += '文字：\t' + c + '\n'
            m += 'UTF-8：\t' + re.sub('^0x', '', hex(ord(c))).upper() + '\n\n'
            for jk in JOYOKANJI:
                if c in jk[1]:
                    m += '常用漢字です．\n'
                    m += '字体：\u3000' + jk[1] + '\n'
                    m += '読み：\u3000' + jk[2] + '\n'
                    if jk[3] != '':
                        m += '用例：' + jk[3] + '\n'
                    break
            else:
                m += '常用漢字ではありません．\n'
                if re.match('^[ -~]$', c):
                    m += '半角英数記号\n'
                if re.match('^[ｦ-ﾟ]$', c):
                    m += '半角カタカナ\n'
                if re.match('^[ぁ-ゖ]$', c):
                    m += 'ひらがな\n'
                if re.match('^[ァ-ヺ]$', c) or re.match('^[ㇰ-ㇿ]$', c):
                    m += 'カタカナ\n'
                if re.match('^[０-９]$', c):
                    m += '数字\n'
            m = re.sub('\n+$', '', m)
            tkinter.messagebox.showinfo(n, m)

    def show_help(self):
        n = 'ヘルプ'
        m = 'このダイアログを閉じた後、' + \
            'ウィンドウにMS_Wordのファイル（拡張子docx）を' + \
            'ドラッグアンドドロップしてみてください．'
        tkinter.messagebox.showinfo(n, m)

    def show_license_info(self):
        n = 'ライセンス情報'
        m = 'Copyright (C) 2022-2024  Seiichiro HATA\n\n' + \
            'このソフトウェアは、\n' + \
            '"GNU GENERAL PUBLIC LICENSE\n' + \
            'Version 3 (GPLv3)"という\n' + \
            'ライセンスで開発されています．\n\n' + \
            'このソフトウェアは、\n' + \
            '次のモジュールを利用しており、\n' + \
            'それぞれ付記したライセンスで\n' + \
            '配布されています．\n'
        m += \
            '- argparse: PSF License\n'
        # PYTHON DOCX
        m += \
            '- python-docx: MIT License\n' + \
            '- lxml: BSD License (3-Clause)\n' + \
            '- typing_extensions: PSF License\n'
        # CHARDET
        m += \
            '- chardet: LGPLv2+\n'
        # TKINTERDND2
        m += \
            '- tkinterdnd2: MIT License\n'
        # PYWIN32
        m += \
            '- pywin32: PSF License\n'
        # OPENPYXL (EXCEL)
        m += \
            '- openpyxl: MIT License\n' + \
            '- et_xmlfile: MIT License\n'
        # OPENAI
        m += \
            '- openai: Apache Software License\n' + \
            '- annotated-types: MIT License\n' + \
            '- anyio: MIT License\n' + \
            '- certifi: Mozilla Public License 2.0\n' + \
            '- distro: Apache Software License\n' + \
            '- exceptiongroup: MIT License\n' + \
            '- h11: MIT License\n' + \
            '- httpcore: BSD License\n' + \
            '- httpx: BSD License\n' + \
            '- idna: BSD License\n' + \
            '- jiter: MIT License\n' + \
            '- pydantic: MIT License\n' + \
            '- pydantic_core: MIT License\n' + \
            '- sniffio: Apache Software License;\n' + \
            '　　MIT License\n' + \
            '- tqdm: MIT License;\n' + \
            '　　Mozilla Public License 2.0\n'
        # LLAMA
        m += \
            '- llama_cpp_python: MIT License\n' + \
            '- Jinja2: BSD License\n' + \
            '- MarkupSafe: BSD License\n' + \
            '- diskcache:\n' + \
            '　　Apache Software License\n' + \
            '- numpy: BSD License\n'
        # '- typing_extensions: PSF License\n'
        # LEVENSHTEIN (MDDIFF)
        m += \
            '- Levenshtein: GPLv2+\n'
        #
        m += \
            '\n利用、改変、再配布等をする場合には、\n' + \
            'ライセンスに十分ご注意ください．\n' + \
            'スクリプトファイルは、\n' + \
            '外部のモジュールを読み込みますが、\n' + \
            'バイナリファイルは、\n' + \
            '内部にモジュールを含んでいますので、\n' + \
            '特にご注意ください．'

        tkinter.messagebox.showinfo(n, m)

    def show_about_makdo(self):
        n = 'バージョン情報'
        m = 'makdo ' + __version__ + '\n\n' + \
            '秦誠一郎により開発されています．'
        tkinter.messagebox.showinfo(n, m)

    ####################################
    # KEY CONFIGURATION

    # event.char
    # Ctrl+A '\x01' select all          # Ctrl+N '\x0e' new document
    # Ctrl+B '\x02' bold                # Ctrl+O '\x0f' open document
    # Ctrl+C '\x03' copy                # Ctrl+P '\x10' print
    # Ctrl+D '\x04' font                # Ctrl+Q '\x11' quit
    # Ctrl+E '\x05' centered            # Ctrl+R '\x12' right
    # Ctrl+F '\x06' search              # Ctrl+S '\x13' save
    # Ctrl+G '\x07' move                # Ctrl+T '\x14' hanging indent
    # Ctrl+H '\x08' replace             # Ctrl+U '\x15' underline
    # Ctrl+I '\x09' italic              # Ctrl+V '\x16' paste
    # Ctrl+J '\x0a' justified           # Ctrl+W '\x17' close document
    # Ctrl+K '\x0b' hyper link          # Ctrl+X '\x18' cut
    # Ctrl+L '\x0c' left                # Ctrl+Y '\x19' redo
    # Ctrl+M '\x0d' indent              # Ctrl+Z '\x1a' undo

    def _make_txt_key_configuration(self):
        self.txt.bind('<Key>', self.txt_process_key)
        self.txt.bind('<KeyRelease>', self.txt_process_key_release)
        self.txt.bind('<Button-1>', self.txt_process_button1)
        self.txt.bind('<Button-2>', self.txt_process_button2)
        self.txt.bind('<Button-3>', self.txt_process_button3)
        self.txt.bind('<ButtonRelease-1>', self.txt_process_button1_release)
        self.txt.bind('<ButtonRelease-2>', self.txt_process_button2_release)
        self.txt.bind('<ButtonRelease-3>', self.txt_process_button3_release)

    def _make_sub_key_configuration(self):
        self.sub.bind('<Key>', self.sub_process_key)
        self.sub.bind('<KeyRelease>', self.sub_process_key_release)
        self.sub.bind('<Button-1>', self.sub_process_button1)
        self.sub.bind('<Button-2>', self.sub_process_button2)
        self.sub.bind('<Button-3>', self.sub_process_button3)
        self.sub.bind('<ButtonRelease-1>', self.sub_process_button1_release)
        self.sub.bind('<ButtonRelease-2>', self.sub_process_button2_release)
        self.sub.bind('<ButtonRelease-3>', self.sub_process_button3_release)

    ##########################
    # COMMAND

    def txt_process_key(self, key):
        self.destroy_splash_screen()
        self.current_pane = 'txt'
        is_read_only = self.is_read_only.get()
        if is_read_only:
            return self.read_only_process_key(self.txt, key)
        else:
            if key.keysym != 'Return' and \
               not (key.state & 4 and key.keysym == 'm') and \
               not (key.state & 8192 and key.keysym == 'm'):
                self.win.after(5, self._paint_one_line_after_key_pressed)
            else:
                self.win.after(5, self._paint_two_lines_after_key_pressed)
            return self.read_and_write_process_key(self.txt, key)

    def _paint_one_line_after_key_pressed(self):
        vp = self._get_v_position_of_insert(self.txt)
        self.paint_out_line(vp - 1)
        self.update_toc()

    def _paint_two_lines_after_key_pressed(self):
        vp = self._get_v_position_of_insert(self.txt)
        self.paint_out_line(vp - 2)
        self.paint_out_line(vp - 1)
        self.update_toc()

    def sub_process_key(self, key):
        self.current_pane = 'sub'
        if key.keysym == 'Escape':
            self._close_sub_pane()
            return 'break'
        if self.sub_pane_is_read_only:
            return self.read_only_process_key(self.sub, key)
        else:
            return self.read_and_write_process_key(self.sub, key)

    def txt_process_key_release(self, key):
        self.set_position_info_on_status_bar()
        # FOR AKAUNI
        self._paint_akauni_region(self.txt, '')

    def sub_process_key_release(self, key):
        # FOR AKAUNI
        self._paint_akauni_region(self.sub, '')
        return 'break'

    def read_and_write_process_key(self, pane, key):
        self.set_message_on_status_bar('')
        # HISTORY
        k1 = self._get_key(key)
        k2 = self.key_history[-1]
        if k1 is None:
            return 'break'  # shift, control, alt, mode_switch
        self.key_history.append(k1)
        self.key_history.pop(0)
        unix_time = datetime.datetime.now().timestamp()
        self.key_pressed_time.append(unix_time)
        self.key_pressed_time.pop(0)
        # KEY
        if k1 == 'Escape':
            if k2 == 'Escape':
                self.key_history[-1] = ''
                return 'break'
            self.set_message_on_status_bar('"Esc"が押されました．')
            return 'break'
        elif self._is_key(k1, 'F19', 'C-x', 'C-b'):            # C-x
            self._any_process_escp1()
            return 'break'
        elif self._is_key(k1, 'F20', 'C-b', 'C-n'):            # C-b
            #
            return 'break'
        elif self._is_key(k1, 'F13', 'C--', 'C-q'):            # C--
            #
            return 'break'
        elif self._is_key(k1, 'Left', 'C-t', 'C-k'):           # C-t
            self._any_process_left(pane)
            return 'break'
        elif self._is_key(k1, 'Right', 'C-s', 'C-;'):          # C-s
            self._any_process_right(pane)
            return 'break'
        elif self._is_key(k1, 'Up', 'C-r', 'C-o'):             # C-r
            self._any_process_up(pane)
            return 'break'
        elif self._is_key(k1, 'Down', 'C-n', 'C-l'):           # C-n
            self._any_process_down(pane)
            return 'break'
        elif self._is_key(k1, 'Home', 'C-l', 'C-p'):           # C-l
            self._any_process_home(pane)
            return 'break'
        elif self._is_key(k1, 'End', 'C-{', 'C-['):            # C-{
            self._any_process_end(pane)
            return 'break'
        elif self._is_key(k1, 'Prior', 'C-[', 'C-@'):          # C-[
            self._any_process_prior(pane)
            return 'break'
        elif self._is_key(k1, 'Next', 'C-]', 'C-:'):           # C-]
            if self._is_key(k2, 'F13', 'C--', 'C-q'):
                if self.current_pane == 'sub':
                    self._execute_sub_pane()
                    return 'break'
            self._any_process_next(pane)
            return 'break'
        elif self._is_key(k1, 'Delete', 'C-d', 'C-h', 'C-x'):  # C-d
            if self._any_process_delete():
                return 'break'
            # FOR PAINTING
            if pane == self.txt and \
               not pane.tag_ranges('sel') and \
               'akauni' not in pane.mark_names():
                if pane.index('insert') != pane.index('end-1c'):
                    c = self.txt.get('insert', 'insert+1c')
                    if c == '\n':
                        vp = self._get_v_position_of_insert(pane)
                        self.line_data.pop(vp)
                        for i, ld in enumerate(self.line_data):
                            ld.line_number = i
            self._execute_when_delete_is_pressed(pane)
            return 'break'
        elif self._is_key(k1, 'BackSpace', 'C-h', 'C-j'):      # C-h
            if self._is_key(k2, 'F19', 'C-x', 'C-b'):
                self.split_window()
                return 'break'
            # FOR PAINTING
            if pane == self.txt:
                c = pane.get('insert-1c', 'insert')
                if c == '\n':
                    vp = self._get_v_position_of_insert(pane)
                    self.line_data.pop(vp - 1)
                    for i, ld in enumerate(self.line_data):
                        ld.line_number = i
            pane.delete('insert-1c', 'insert')
            return 'break'
        elif self._is_key(k1, 'Return', 'C-m', 'C-m'):         # C-m
            # FOR PAINTING
            if pane == self.txt:
                vp = self._get_v_position_of_insert(pane)
                self.line_data.insert(vp, LineDatum())
                for i, ld in enumerate(self.line_data):
                    ld.line_number = i
            return
        elif self._is_key(k1, 'Tab', 'C-i', 'C-g'):            # C-i
            if self._any_process_tab(pane):
                return 'break'
            return
        elif self._is_key(k1, 'F14', 'C-v', 'C-y'):            # C-v
            self._any_process_quit(pane)
            return 'break'
        elif self._is_key(k1, 'F15', 'C-g', 'C-u', 'C-v'):     # C-g
            self._any_process_paste()
            return 'break'
        elif self._is_key(k1, 'F16', 'C-c', 'C-i', 'C-f'):     # C-c
            self._any_process_search(pane)
            return 'break'
        elif self._is_key(k1, 'F17', 'C-}', 'C-]', 'C-e'):     # C-}
            if self._is_key(k2, 'F13', 'C--', 'C-q'):
                self.calculate()
                return 'break'
            self.execute_keyboard_macro()
            return 'break'
        elif self._is_key(k1, 'F21', 'C-w', 'C-,', 'C-z'):     # C-w
            if self._is_key(k2, 'F13', 'C--', 'C-q'):
                self.edit_modified_redo()
            else:
                self.edit_modified_undo()
            return 'break'
        elif self._is_key(k1, 'F22', 'C-f', 'C-.'):            # C-f
            self._any_process_mark(pane)
            return 'break'
        elif k1 == 'g':
            if k2 == 'Escape':
                k3, k4 = self.key_history[-3], self.key_history[-4]
                was_pasted = \
                    self._is_key(k3, 'F15', 'C-g', 'C-u', 'C-v') and \
                    not self._is_key(k4, 'F13', 'C--', 'C-q')
                was_repeated = (k3 == 'g' and k4 == 'Escape')
                if was_pasted or was_repeated:
                    if was_pasted:
                        self.clipboard_list_number \
                            = len(self.clipboard_list) - 2
                    else:
                        self.clipboard_list_number -= 1
                    if self.clipboard_list_number < 0:
                        self.set_message_on_status_bar('履歴がなくなりました')
                        return 'break'
                    prev = self.clipboard_list[self.clipboard_list_number + 1]
                    curr = self.clipboard_list[self.clipboard_list_number]
                    pane.delete('insert-' + str(len(prev)) + 'c', 'insert')
                    pane.insert('insert', curr)
                    return 'break'
            return
        elif k1 == 'x':
            if k2 == 'Escape':
                self.start_minibuffer()
                return 'break'
            return
        elif self._is_key(k1, None, None, None, 'C-a'):
            self.select_all()
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-c'):
            self.copy_region()
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-l'):
            self.replace_forward()
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-o'):
            self.open_file()
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-p'):
            self.start_writer()
            return 'break'
        # elif self._is_key(k1, None, None, None, 'C-q'):
        #     self.quit_makdo()
        #     return 'break'
        elif self._is_key(k1, None, None, None, 'C-s'):
            self.save_file()
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-y'):
            self.edit_modified_redo()
            return 'break'

    def read_only_process_key(self, pane, key):
        self.set_message_on_status_bar('')
        # HISTORY
        k1 = self._get_key(key)
        k2 = self.key_history[-1]
        if k1 is None:
            return 'break'  # shift, control, alt, mode_switch
        self.key_history.append(k1)
        self.key_history.pop(0)
        unix_time = datetime.datetime.now().timestamp()
        self.key_pressed_time.append(unix_time)
        self.key_pressed_time.pop(0)
        # KEY
        if k1 == 'Escape':
            self.set_message_on_status_bar('"Esc"が押されました．')
            return 'break'
        elif self._is_key(k1, 'F19', 'C-x', 'C-b'):
            self._any_process_escp1()
            return 'break'
        elif self._is_key(k1, 'F20', 'C-b', 'C-n'):
            #
            return 'break'
        elif self._is_key(k1, 'F13', 'C--', 'C-q'):
            #
            return 'break'
        elif self._is_key(k1, 'Left', 'C-t', 'C-k'):
            self._any_process_left(pane)
            return 'break'
        elif self._is_key(k1, 'Right', 'C-s', 'C-;'):
            self._any_process_right(pane)
            return 'break'
        elif self._is_key(k1, 'Up', 'C-r', 'C-o'):
            self._any_process_up(pane)
            return 'break'
        elif self._is_key(k1, 'Down', 'C-n', 'C-l'):
            self._any_process_down(pane)
            return 'break'
        elif self._is_key(k1, 'Home', 'C-l', 'C-p'):
            self._any_process_home(pane)
            return 'break'
        elif self._is_key(k1, 'End', 'C-{', 'C-['):
            self._any_process_end(pane)
            return 'break'
        elif self._is_key(k1, 'Prior', 'C-[', 'C-@'):
            self._any_process_prior(pane)
            return 'break'
        elif self._is_key(k1, 'Next', 'C-]', 'C-:'):
            self._any_process_next(pane)
            return 'break'
        elif self._is_key(k1, 'Delete', 'C-d', 'C-h', 'C-x'):
            if self._any_process_delete():
                return 'break'
            self._execute_when_delete_is_pressed(pane)
            return 'break'
        elif self._is_key(k1, 'F14', 'C-v', 'C-y'):
            self._any_process_quit(pane)
            return 'break'
        elif self._is_key(k1, 'F16', 'C-c', 'C-i', 'C-f'):
            self._any_process_search(pane)
            return 'break'
        elif self._is_key(k1, 'F22', 'C-f', 'C-.'):
            self._any_process_mark(pane)
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-a'):
            self.select_all()
            return 'break'
        elif self._is_key(k1, None, None, None, 'C-c'):
            self.copy_region()
            return 'break'
        self.set_message_on_status_bar('読取専用です')
        return 'break'

    @staticmethod
    def _is_modifier_key(keysym):
        if keysym == 'Shift_L' or keysym == 'Shift_R':
            return True
        if keysym == 'Control_L' or keysym == 'Control_R':
            return True
        if keysym == 'Alt_L' or keysym == 'Alt_R':
            return True
        if keysym == 'Mode_switch':
            return True
        return False

    def _any_process_escp1(self):  # F19 / Ctrl+X
        k2 = self.key_history[-2]
        if self._is_key(k2, 'F19', 'C-x', 'C-b'):
            self.key_history[-1] = ''
            self._jump_to_next_pane()

    def _any_process_left(self, pane):  # Left / Ctrl+T
        self._any_process_left_or_right(pane, 'left')

    def _any_process_right(self, pane):  # Right / Ctrl+S
        self._any_process_left_or_right(pane, 'right')

    def _any_process_left_or_right(self, pane, key):
        k2 = self.key_history[-2]
        if self._is_key(k2, 'F19', 'C-x', 'C-b'):
            if len(self.pnd_r.panes()) > 1 and self.sub_btn1['text'] == '入替':
                self.swap_windows()
                return
        if self._is_boosted_movement():
            if key == 'left':
                pane.mark_set('insert', 'insert-5c')
            else:
                pane.mark_set('insert', 'insert+5c')
        else:
            if key == 'left':
                pane.mark_set('insert', 'insert-1c')
            else:
                pane.mark_set('insert', 'insert+1c')
        self._put_back_cursor_to_pane(pane)
        self._paint_akauni_region(pane, '')
        return

    def _any_process_up(self, pane):  # Up / Ctrl+R
        self._any_process_up_or_down(pane, 'up')

    def _any_process_down(self, pane):  # Down / Ctrl+N
        self._any_process_up_or_down(pane, 'down')

    def _any_process_up_or_down(self, pane, key):
        k3 = self.key_history[-3]
        k2 = self.key_history[-2]
        if self._is_key(k2, 'F19', 'C-x', 'C-b'):
            if key == 'up':
                self._jump_to_prev_pane()
            else:
                self._jump_to_next_pane()
            self.key_history[-1] = ''
            return
        if not (self._is_key(k2, 'Up', 'C-r', 'C-o') or
                self._is_key(k2, 'Down', 'C-n', 'C-l') or
                self._is_key(k2, 'Prior', 'C-[', 'C-@') or
                self._is_key(k2, 'Next', 'C-]', 'C-:')):
            self.ideal_h_position = self._get_ideal_h_position_of_insert(pane)
        elif (self._is_key(k3, 'F13', 'C--', 'C-q') or
              self._is_key(k3, 'F19', 'C-x', 'C-b')):
            self.ideal_h_position = self._get_ideal_h_position_of_insert(pane)
        if self._is_boosted_movement():
            if key == 'up':
                self._move_vertically(pane, self.ideal_h_position, -5)
            else:
                self._move_vertically(pane, self.ideal_h_position, +5)
        else:
            if key == 'up':
                self._move_vertically(pane, self.ideal_h_position, -1)
            else:
                self._move_vertically(pane, self.ideal_h_position, +1)
        self._paint_akauni_region(pane, '')
        return

    def _any_process_prior(self, pane):  # Prior / Ctrl+[
        self._any_process_prior_or_next(pane, 'prior')

    def _any_process_next(self, pane):  # Next / Ctrl+]
        self._any_process_prior_or_next(pane, 'next')

    def _any_process_prior_or_next(self, pane, key):
        k3 = self.key_history[-3]
        k2 = self.key_history[-2]
        if not (self._is_key(k2, 'Up', 'C-r', 'C-o') or
                self._is_key(k2, 'Down', 'C-n', 'C-l') or
                self._is_key(k2, 'Prior', 'C-[', 'C-@') or
                self._is_key(k2, 'Next', 'C-]', 'C-:')):
            self.ideal_h_position = self._get_ideal_h_position_of_insert(pane)
        elif (self._is_key(k3, 'F13', 'C--', 'C-q') or
              self._is_key(k3, 'F19', 'C-x', 'C-b')):
            self.ideal_h_position = self._get_ideal_h_position_of_insert(pane)
        if self._is_boosted_movement():
            if key == 'prior':
                pane.mark_set('insert', '1.0')
            else:
                pane.mark_set('insert', 'end-1c')
            self._put_back_cursor_to_pane(pane)
        else:
            lines = self._get_lines_of_pane(pane)
            if key == 'prior':
                self._move_vertically(pane, self.ideal_h_position, -lines)
            else:
                self._move_vertically(pane, self.ideal_h_position, +lines)
        self._paint_akauni_region(pane, '')

    def _any_process_home(self, pane):  # Home / Ctrl+L
        self._any_process_home_or_end(pane, 'home')

    def _any_process_end(self, pane):  # End / Ctrl+{
        self._any_process_home_or_end(pane, 'end')

    def _any_process_home_or_end(self, pane, key):
        if self._is_boosted_movement():
            if key == 'home':
                pane.mark_set('insert', 'insert linestart')
            else:
                pane.mark_set('insert', 'insert lineend')
            self._put_back_cursor_to_pane(pane)
        else:
            width = self._get_width_of_pane(pane)
            if key == 'home':
                self._move_horizontally(pane, -width)
            else:
                self._move_horizontally(pane, +width)
        self._paint_akauni_region(pane, '')

    def _any_process_tab(self, pane):  # Tab / Ctrl+I
        text = pane.get('1.0', 'insert')
        line = pane.get('insert linestart', 'insert lineend')
        posi = pane.index('insert')
        # CONFIGURATION
        res_open = '^<!--(?:.|\n)*'
        res_close = '^(?:.|\n)*-->(?:.|\n)*'
        if re.match(res_open, text) and not re.match(res_close, text):
            for i, sample in enumerate(CONFIGURATION_SAMPLE):
                if line == sample:
                    pane.delete('insert linestart', 'insert lineend')
                    pane.insert('insert', CONFIGURATION_SAMPLE[i + 1])
                    pane.mark_set('insert', 'insert lineend')
                    return True
        # CALCULATE
        res_open = '^((?:.|\n)*)(<!--(?:.|\n)*)'
        res_close = '^((?:.|\n)*)(-->(?:.|\n)*)'
        if re.match(res_open, text):
            text = re.sub(res_open, '\\2', text)
            if not re.match(res_close, text):
                if self.calculate(False):
                    return True
        # SCRIPT
        res_open = '^((?:.|\n)*){([0-9]*){((?:.|\n)*)'
        res_close = '^((?:.|\n)*)}([0-9]*)}((?:.|\n)*)'
        if re.match(res_open, text):
            befo = re.sub(res_open, '\\1', text)
            numb = re.sub(res_open, '\\2', text)
            scri = re.sub(res_open, '\\3', text)
            if not re.match(res_close, text):
                cur_to_end = pane.get('insert', 'end-1c')
                if re.match('^}' + numb + '}', cur_to_end):
                    msg = '（ここにスクリプトを挿入（サンプルはTabを押す））'
                    if scri == msg:
                        beg_n = len(befo + '{' + numb + '{')
                        end_n = beg_n + len(scri)
                        beg = '1.0+' + str(beg_n) + 'c'
                        end = '1.0+' + str(end_n) + 'c'
                        pane.delete(beg, end)
                        pane.insert(beg, SCRIPT_SAMPLE[1])
                        return True
                    for i, sample in enumerate(SCRIPT_SAMPLE):
                        if scri == sample:
                            beg_n = len(befo + '{' + numb + '{')
                            end_n = beg_n + len(scri)
                            beg = '1.0+' + str(beg_n) + 'c'
                            end = '1.0+' + str(end_n) + 'c'
                            pane.delete(beg, end)
                            pane.insert(beg, SCRIPT_SAMPLE[i + 1])
                            return True
        # PARAGRAPH SAMPLE
        if posi == pane.index('insert lineend'):
            for i, sample in enumerate(PARAGRAPH_SAMPLE):
                if line == sample:
                    pane.delete('insert linestart', 'insert lineend')
                    pane.insert('insert', PARAGRAPH_SAMPLE[i + 1])
                    pane.mark_set('insert', 'insert lineend')
                    return True
        # TIDY UP PARAGRAPH
        has_tidied = self.tidy_up_paragraph()
        if has_tidied:
            return True
        # AUTO CORRECT
        left = pane.get('insert linestart', 'insert')
        if len(left) > 0:
            # FULL TO HALF
            res = '^(.*?)([' \
                + '\t\u3000！”＃＄％＆’（）＊＋，―−－ー．／０-９：；＜＝＞？＠Ａ-Ｚ［￥＼］＾＿｀ａ-ｚ｛｜｝〜' \
                + ']+)$'  # "―"(2015), "−"(2212), "－"(FF0D)
            if re.match(res, left):
                rest = re.sub(res, '\\1', left)
                full = re.sub(res, '\\2', left)
                if rest == '' or full != 'ー':
                    half = full
                    fhs = [['\t', ' '], ['\u3000', ' '],
                           ['！', '!'], ['”', '"'], ['＃', '#'], ['＄', '$'],
                           ['％', '%'], ['＆', '&'], ['’', "'"], ['（', '('],
                           ['）', ')'], ['＊', '*'], ['＋', '+'], ['，', ','],
                           ['−', '-'], ['―', '-'], ['ー', '-'],
                           ['．', '.'], ['／', '/'],  # 0-9
                           ['：', ':'], ['；', ';'], ['＜', '<'], ['＝', '='],
                           ['＞', '>'], ['？', '?'], ['＠', '@'],  # A-Z
                           ['［', '['], ['￥', '\\'], ['＼', '\\'], ['］', ']'],
                           ['＾', '^'], ['＿', '_'], ['｀', '`'],  # a-z
                           ['｛', '{'], ['｜', '|'], ['｝', '}'], ['〜', '~']]
                    for fh in fhs:
                        half = half.replace(fh[0], fh[1])
                    for i in range(0, 10):
                        half = half.replace(chr(i + 65296), chr(i + 48))  # 0-9
                    for i in range(0, 26):
                        half = half.replace(chr(i + 65313), chr(i + 65))  # A-Z
                        half = half.replace(chr(i + 65345), chr(i + 97))  # a-z
                    if half != full:
                        pane.delete('insert-' + str(len(full)) + 'c', 'insert')
                        pane.insert('insert', half)
                        return True
            # HALF TO FULL
            res = '^(.*?)( +)$'
            if re.match(res, left):
                rest = re.sub(res, '\\1', left)
                half = re.sub(res, '\\2', left)
                full = half.replace(' ', '\u3000')
                pane.delete('insert-' + str(len(half)) + 'c', 'insert')
                pane.insert('insert', full)
                return True
        # FONT DECORATER
        for i, sample in enumerate(FONT_DECORATOR_SAMPLE):
            if i == 0:
                continue
            if i == len(FONT_DECORATOR_SAMPLE) - 1:
                break
            sample_esc = sample
            sample_esc = sample_esc.replace('*', '\\*')
            sample_esc = sample_esc.replace('+', '\\+')
            sample_esc = sample_esc.replace('^', '\\^')
            beg_to_ins = pane.get('insert linestart', 'insert')
            if re.match('^.*' + sample_esc + '$', beg_to_ins):
                pane.delete(posi + '-' + str(len(sample)) + 'c', posi)
                pane.insert('insert', FONT_DECORATOR_SAMPLE[i + 1])
                return True
        else:
            pane.insert('insert', FONT_DECORATOR_SAMPLE[1])
            return True
        return False

    def _any_process_delete(self) -> bool:  # Delete / Ctrl+D
        k2 = self.key_history[-2]
        if self._is_key(k2, 'F19', 'C-x', 'C-b'):
            if self.current_pane == 'txt':
                self.quit_makdo()
                return True
        if self._is_key(k2, 'F13', 'C--', 'C-q'):
            self.cut_rectangle()
            return True
        if not self._is_key(k2, 'Delete', 'C-d', 'C-h', 'C-x'):
            self.win.clipboard_clear()
            if self.clipboard_list[-1] != '':
                self.clipboard_list.append('')
        return False

    def _any_process_quit(self, pane):  # F14 / Ctrl+V
        k2 = self.key_history[-2]
        # CLOSE
        if self._is_key(k2, 'F19', 'C-x', 'C-b'):
            if self.current_pane == 'txt':
                self.close_file()
                return
        # UNMARK
        if 'akauni' in pane.mark_names():
            pane.tag_remove('akauni_tag', '1.0', 'end')
            pane.mark_unset('akauni')
        return

    def _any_process_paste(self):  # F15 / Ctrl+G
        k2 = self.key_history[-2]
        # PASTE
        if self._is_key(k2, 'F13', 'C--', 'C-q'):
            self.paste_rectangle()
        else:
            self.paste_region()

    def _any_process_search(self, pane):  # F16 / Ctrl+C
        k2 = self.key_history[-2]
        k3 = self.key_history[-3]
        k4 = self.key_history[-4]
        # SEARCH
        if self._is_key(k2, 'F13', 'C--', 'C-q'):
            if self._is_key(k3, 'F16', 'C-c', 'C-i', 'C-f') and \
               self._is_key(k4, 'F13', 'C--', 'C-q') and \
               Makdo.search_word != '':
                self.search_backward()
            else:
                self.search_backward_from_dialog(pane)
        else:
            if self._is_key(k2, 'F16', 'C-c', 'C-i', 'C-f') and \
               not self._is_key(k3, 'F13', 'C--', 'C-q') and \
               Makdo.search_word != '':
                self.search_forward()
            else:
                self.search_forward_from_dialog(pane)

    def _any_process_mark(self, pane):  # F22 / Ctrl+F
        k2 = self.key_history[-2]
        if self._is_key(k2, 'F19', 'C-x', 'C-b'):
            self.save_file()
            return
        # UNMARK
        if 'akauni' in pane.mark_names():
            has_akauni = True
            pane.mark_set('prev_akauni', 'akauni')
            pane.tag_remove('akauni_tag', '1.0', 'end')
            pane.mark_unset('akauni')
        else:
            has_akauni = False
        delta = self.key_pressed_time[-1] - self.key_pressed_time[-2]
        if self._is_key(k2, 'F22', 'C-f', 'C-.') and delta < 0.32:
            # SUB CURSOR
            if len(self.pnd_r.panes()) > 1 and \
               'x0sixteenth' in self.sub.mark_names():
                self.swap_windows()
            else:
                self.goto_sub_cursor()
        else:
            # MARK
            if not has_akauni:
                pane.mark_set('akauni', 'insert')
                pane.mark_unset('prev_akauni')

    def _is_boosted_movement(self) -> bool:
        key4 = self.key_history[-4]
        key3 = self.key_history[-3]
        key2 = self.key_history[-2]
        key1 = self.key_history[-1]
        delta34 = self.key_pressed_time[-3] - self.key_pressed_time[-4]
        delta23 = self.key_pressed_time[-2] - self.key_pressed_time[-3]
        delta12 = self.key_pressed_time[-1] - self.key_pressed_time[-2]
        # --> O (t<0.2) O (0.4<t<1.0) O (t<0.1) O -->
        if '_has_boosted' not in vars(self):
            # --> X X X X -->
            self._has_boosted = False
        elif key2 != key1 or key3 != key1 or key4 != key1:
            # --> ? ? X O --> | --> ? X ? O --> | --> X ? ? O -->
            self._has_boosted = False
        else:
            if delta12 < 0.1:
                if delta23 > 0.4 and delta23 < 1.0:
                    if delta34 < 0.2:
                        self._has_boosted = True
            else:
                self._has_boosted = False
        return self._has_boosted

    @staticmethod
    def _paint_akauni_region(pane, shift=''):
        if 'akauni' in pane.mark_names():
            pane.tag_remove('akauni_tag', '1.0', 'end')
            if pane.compare('akauni', '<', 'insert' + shift):
                pane.tag_add('akauni_tag', 'akauni', 'insert' + shift)
            else:
                pane.tag_add('akauni_tag', 'insert' + shift, 'akauni')

    # MOUSE BUTTON LEFT

    def txt_process_button1(self, click):
        self.txt.focus_set()
        self.cancel_region(self.txt)
        self.current_pane = 'txt'
        self.win.after(5, self.update_toc)
        return

    def txt_process_button1_release(self, click):
        self.close_mouse_menu()
        self.set_position_info_on_status_bar()
        return 'break'

    def sub_process_button1(self, click):
        self.sub.focus_set()
        self.cancel_region(self.sub)
        self.current_pane = 'sub'
        return

    def sub_process_button1_release(self, click):
        self.close_mouse_menu()
        return 'break'

    # MOUSE BUTTON CENTER

    def txt_process_button2(self, click):
        return 'break'

    def txt_process_button2_release(self, click):
        self.close_mouse_menu()
        # self.paste_region()
        return 'break'

    def sub_process_button2(self, click):
        return 'break'

    def sub_process_button2_release(self, click):
        self.close_mouse_menu()
        # self.paste_region()
        return 'break'

    # MOUSE BUTTON RIGHT

    def txt_process_button3(self, click):
        self.any_process_button3(self.txt, click)
        return 'break'

    def txt_process_button3_release(self, click):
        return 'break'

    def sub_process_button3(self, click):
        self.any_process_button3(self.sub, click)
        return 'break'

    def sub_process_button3_release(self, click):
        return 'break'

    def any_process_button3(self, pane, click):
        # CLOSE MOUSE MENU
        self.close_mouse_menu()
        # CLIPBOARD
        try:
            cb = self.win.clipboard_get()
        except BaseException:
            cb = ''
        # SEPARATER
        needs_separater = False
        if pane.tag_ranges('sel') or 'akauni' in pane.mark_names():
            needs_separater = True
        if not self._is_read_only_pane(pane):
            if cb != '' and self.rectangle_text_list != []:
                needs_separater = True
        # BUTTON MENU
        self.bt3 = tkinter.Menu(self.win, tearoff=False)
        if pane.tag_ranges('sel') or 'akauni' in pane.mark_names():
            if not self._is_read_only_pane(pane):
                self.bt3.add_command(label='切り取り',
                                     command=self.cut_region)
            self.bt3.add_command(label='コピー',
                                 command=self.copy_region)
        if not self._is_read_only_pane(pane):
            if cb != '':
                self.bt3.add_command(label='貼り付け',
                                     command=self.paste_region)
        if needs_separater:
            self.bt3.add_separator()
        if pane.tag_ranges('sel') or 'akauni' in pane.mark_names():
            if not self._is_read_only_pane(pane):
                self.bt3.add_command(label='矩形の切り取り',
                                     command=self.cut_rectangle)
            self.bt3.add_command(label='矩形のコピー',
                                 command=self.copy_rectangle)
        if not self._is_read_only_pane(pane):
            if self.rectangle_text_list != []:
                self.bt3.add_command(label='矩形の貼り付け',
                                     command=self.paste_rectangle)
        self.bt3.post(click.x_root, click.y_root)

    def close_mouse_menu(self, event=None):
        try:
            self.bt3.destroy()
        except BaseException:
            pass

    ####################################
    # STATUS BAR

    def _make_status_bar(self):
        self._make_status_search_or_replace()
        self._make_status_file_name()
        self._make_status_position_information()
        self._make_status_message()

    ##########################
    # STATUS FILE NAME

    def _make_status_file_name(self):
        self.stb_fnm1 = tkinter.Label(self.stb_l, anchor='w', text='')
        self.stb_fnm1.pack(side='left')
        tkinter.Label(self.stb_l, text=' ').pack(side='left')

    ################
    # COMMAND

    def set_file_name_on_status_bar(self, file_name, must_update=False):
        fn = file_name
        fn = re.sub('\n', '/', fn)
        res = '^(.*)(\\..{1,4})$'
        if re.match(res, fn):
            nam = re.sub(res, '\\1', fn)
            ext = re.sub(res, '\\2', fn)
        else:
            nam = fn
            ext = ''
        if len(fn) > 15:
            nam = re.sub('^(.{' + str(14 - len(ext)) + '})(.*)', '\\1…', nam)
        self.stb_fnm1['text'] = nam + ext
        if must_update:
            self.stb_l.update()

    ##########################
    # STATUS POSITION INFORMATION

    def _make_status_position_information(self):
        self.stb_pos1 = tkinter.Label(self.stb_l, anchor='w', text='1x0/1x0')
        self.stb_pos1.pack(side='left')
        tkinter.Label(self.stb_l, text=' ').pack(side='left')

    ################
    # COMMAND

    def set_position_info_on_status_bar(self, must_update=False):
        p = self.txt.index('insert')
        cur_v = re.sub('\\.[0-9]+$', '', p)
        s = self.txt.get('insert linestart', 'insert')
        cur_h = str(get_real_width(s))
        cur_p = cur_v + 'x' + cur_h
        p = self.txt.index('end-1c')
        max_v = re.sub('\\.[0-9]+$', '', p)
        s = self.txt.get('insert linestart', 'insert lineend')
        max_h = str(get_real_width(s))
        max_p = max_v + 'x' + max_h
        self.stb_pos1['text'] = cur_p + '/' + max_p
        if must_update:
            self.stb_l.update()

    ##########################
    # STATUS MESSAGE

    def _make_status_message(self):
        self.stb_msg1 = tkinter.Label(self.stb_l, anchor='w', text='')
        self.stb_msg1.pack(side='left')

    ################
    # COMMAND

    def set_message_on_status_bar(self, msg, must_update=False):
        if msg == '':
            self.stb_msg1['text'] = ''
        else:
            self.stb_msg1['text'] = '→ ' + msg
        if must_update:
            self.stb_l.update()

    ##########################
    # STATUS SEARCH OR REPLACE

    def _make_status_search_or_replace(self):
        tkinter.Label(self.stb_r, text=' ').pack(side='left')
        self.stb_sor1 = tkinter.Entry(self.stb_r, width=20)
        self.stb_sor1.pack(side='left')
        self.stb_sor1.bind('<Key>', self.sor1_key)
        self.stb_sor1.bind('<Button-1>', self.sor1_button0)
        self.stb_sor1.bind('<Button-2>', self.sor1_button0)
        self.stb_sor1.bind('<Button-3>', self.sor1_button3)
        self.stb_sor2 = tkinter.Entry(self.stb_r, width=20)
        self.stb_sor2.pack(side='left')
        self.stb_sor2.bind('<Key>', self.sor2_key)
        self.stb_sor2.bind('<Button-1>', self.sor2_button0)
        self.stb_sor2.bind('<Button-2>', self.sor2_button0)
        self.stb_sor2.bind('<Button-3>', self.sor2_button3)
        self.stb_sor3 \
            = tkinter.Button(self.stb_r, text='前',
                             command=self.search_or_replace_backward_on_stb)
        self.stb_sor3.pack(side='left')
        self.stb_sor4 \
            = tkinter.Button(self.stb_r, text='後',
                             command=self.search_or_replace_forward_on_stb)
        self.stb_sor4.pack(side='left')
        self.stb_sor5 \
            = tkinter.Button(self.stb_r, text='消',
                             command=self.clear_search_and_replace)
        self.stb_sor5.pack(side='left')
        #
        self.search_word_history, self.search_word_history_number = [''], 0
        self.replace_word_history, self.replace_word_history_number = [''], 0

    ################
    # COMMAND

    def sor1_key(self, key):
        k = Makdo._get_key(key)
        if k is None:
            return 'break'  # shift, control, alt, mode_switch
        if self._is_key(k, 'Up', 'C-r', 'C-o'):
            h = self.search_word_history
            n = self.search_word_history_number
            if n == len(h) - 1:
                h[-1] = self.stb_sor1.get()
            if n > 0:
                n -= 1
                self.stb_sor1.delete(0, 'end')
                self.stb_sor1.insert(0, h[n])
                self.search_word_history_number = n
            return 'break'
        elif self._is_key(k, 'Down', 'C-n', 'C-l'):
            h = self.search_word_history
            n = self.search_word_history_number
            # if n == len(h) - 1:
            #     h[-1] = self.stb_sor1.get()
            if n < len(h) - 1:
                n += 1
                self.stb_sor1.delete(0, 'end')
                self.stb_sor1.insert(0, h[n])
                self.search_word_history_number = n
            return 'break'
        elif self._is_key(k, 'F15', 'C-g', 'C-g', 'C-v'):
            self.sor1_paste_word()
            return 'break'

    def sor2_key(self, key):
        k = Makdo._get_key(key)
        if k is None:
            return 'break'  # shift, control, alt, mode_switch
        if self._is_key(k, 'Up', 'C-r', 'C-o'):
            h = self.replace_word_history
            n = self.replace_word_history_number
            if n == len(h) - 1:
                h[-1] = self.stb_sor2.get()
            if n > 0:
                n -= 1
                self.stb_sor2.delete(0, 'end')
                self.stb_sor2.insert(0, h[n])
                self.replace_word_history_number = n
        elif self._is_key(k, 'Down', 'C-n', 'C-l'):
            h = self.replace_word_history
            n = self.replace_word_history_number
            # if n == len(h) - 1:
            #     h[-1] = self.stb_sor1.get()
            if n < len(h) - 1:
                n += 1
                self.stb_sor2.delete(0, 'end')
                self.stb_sor2.insert(0, h[n])
                self.replace_word_history_number = n
        elif self._is_key(k, 'F15', 'C-g', 'C-g', 'C-v'):
            self.sor2_paste_word()
            return 'break'

    def sor1_button0(self, click):
        self.close_mouse_menu()  # close mouse menu
        self.stb_sor1.focus_force()

    def sor1_button3(self, click):
        self.close_mouse_menu()  # close mouse menu
        self.stb_sor1.focus_force()
        self.bt3 = tkinter.Menu(self.win, tearoff=False)
        e = self.stb_sor1.get()
        if e != '':
            self.bt3.add_command(label='コピー',
                                 command=self.sor1_copy_word)
            self.bt3.add_separator()
        self.bt3.add_command(label='貼り付け',
                             command=self.sor1_paste_word)
        self.bt3.post(click.x_root, click.y_root)

    def sor2_button0(self, click):
        self.close_mouse_menu()  # close mouse menu
        self.stb_sor2.focus_force()

    def sor2_button3(self, click):
        self.close_mouse_menu()  # close mouse menu
        self.stb_sor2.focus_force()
        self.bt3 = tkinter.Menu(self.win, tearoff=False)
        e = self.stb_sor2.get()
        if e != '':
            self.bt3.add_command(label='コピー',
                                 command=self.sor2_copy_word)
            self.bt3.add_separator()
        self.bt3.add_command(label='貼り付け',
                             command=self.sor2_paste_word)
        self.bt3.post(click.x_root, click.y_root)

    def sor1_copy_word(self):
        e = self.stb_sor1.get()
        if e != '':
            self._clipboard_append(e)

    def sor1_paste_word(self):
        try:
            cb = self.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb != '':
            self.stb_sor1.insert('insert', cb)

    def sor2_copy_word(self):
        e = self.stb_sor2.get()
        if e != '':
            self._clipboard_append(e)

    def sor2_paste_word(self):
        try:
            cb = self.win.clipboard_get()
        except BaseException:
            cb = ''
        if cb != '':
            self.stb_sor2.insert('insert', cb)

    def search_or_replace_backward_on_stb(self):
        word2 = self.stb_sor2.get()
        if word2 == '':
            self.search_or_replace_backward(False)
        else:
            self.search_or_replace_backward(True)

    def search_or_replace_backward(self, must_replace=False):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        word1 = self.stb_sor1.get()
        word2 = self.stb_sor2.get()
        if word1 != '':
            if len(self.search_word_history) <= 1 or \
               self.search_word_history[-2] != word1:
                self.search_word_history[-1] = word1
                self.search_word_history.append(word1)
                self.search_word_history_number \
                    = len(self.search_word_history) - 1
        if word2 != '':
            if len(self.replace_word_history) <= 1 or \
               self.replace_word_history[-2] != word2:
                self.replace_word_history[-1] = word2
                self.replace_word_history.append(word2)
                self.replace_word_history_number \
                    = len(self.replace_word_history) - 1
        if word1 == '':
            return
        if Makdo.search_word != word1:
            Makdo.search_word = word1
            if word1 != '':
                self._highlight_search_word()
        tex = pane.get('1.0', 'insert')
        tex = re.sub(word1 + '$', '', tex)
        res_word1 = word1
        if not self.use_regexps.get():
            res_word1 = self._escape_search_word(word1)
        res = '^((?:.|\n)*?)(' + res_word1 + ')((?:.|\n)*)$'
        # TEST
        try:
            re.match(res, tex)
        except BaseException:
            pane.focus_set()
            self.set_message_on_status_bar('検索・置換に失敗しました')
            return
        # SEARCH AND REPLACE
        if re.match(res, tex):
            sub = ''
            while re.match(res, tex):
                s = re.sub(res, '\\1', tex)
                w = re.sub(res, '\\2', tex)
                tex = re.sub(res, '\\3', tex)
                sub += s + w
                wrd = w
                if wrd == '':
                    return
            if wrd == '':
                return
            # SEARCH
            pane.mark_set('insert', '1.0 +' + str(len(sub)) + 'c')
            self._put_back_cursor_to_pane(pane, True)
            if must_replace:
                if not self._is_read_only_pane(pane):
                    # REPLACE
                    pane.delete('insert-' + str(len(wrd)) + 'c', 'insert')
                    pane.insert('insert', word2)
        pane.focus_set()
        # MESSAGE
        n, m = self._count_word(pane, word1)
        self.set_message_on_status_bar(str(m) + '個が見付かりました' +
                                       '（' + str(n) + '/' + str(m) + '）')

    def search_or_replace_forward_on_stb(self):
        word2 = self.stb_sor2.get()
        if word2 == '':
            self.search_or_replace_forward(False)
        else:
            self.search_or_replace_forward(True)

    def search_or_replace_forward(self, must_replace=False):
        pane = self.txt
        if self.current_pane == 'sub':
            pane = self.sub
        word1 = self.stb_sor1.get()
        word2 = self.stb_sor2.get()
        if word1 != '':
            if len(self.search_word_history) <= 1 or \
               self.search_word_history[-2] != word1:
                self.search_word_history[-1] = word1
                self.search_word_history.append(word1)
                self.search_word_history_number \
                    = len(self.search_word_history) - 1
        if word2 != '':
            if len(self.replace_word_history) <= 1 or \
               self.replace_word_history[-2] != word2:
                self.replace_word_history[-1] = word2
                self.replace_word_history.append(word2)
                self.replace_word_history_number \
                    = len(self.replace_word_history) - 1
        if word1 == '':
            return
        if Makdo.search_word != word1:
            Makdo.search_word = word1
            if word1 != '':
                self._highlight_search_word()
        tex = pane.get('insert', 'end-1c')
        res_word1 = word1
        if not self.use_regexps.get():
            res_word1 = self._escape_search_word(word1)
        res = '^((?:.|\n)*?)(' + res_word1 + ')(?:.|\n)*$'
        # TEST
        try:
            re.match(res, tex)
        except BaseException:
            pane.focus_set()
            self.set_message_on_status_bar('検索・置換に失敗しました')
            return
        # SEARCH AND REPLACE
        if re.match(res, tex):
            sub = re.sub(res, '\\1\\2', tex)
            wrd = re.sub(res, '\\2', tex)
            if wrd == '':
                return
            # SEARCH
            pane.mark_set('insert', 'insert +' + str(len(sub)) + 'c')
            self._put_back_cursor_to_pane(pane, True)
            if must_replace:
                if not self._is_read_only_pane(pane):
                    # REPLACE
                    pane.delete('insert-' + str(len(wrd)) + 'c', 'insert')
                    pane.insert('insert', word2)
        pane.focus_set()
        # MESSAGE
        n, m = self._count_word(pane, word1)
        self.set_message_on_status_bar(str(m) + '個が見付かりました' +
                                       '（' + str(n) + '/' + str(m) + '）')

    def _count_word(self, pane, word):
        res = '^((?:.|\n)*?' + word + ')((?:.|\n)*)$'
        #
        x = 0
        tex = pane.get('1.0', 'insert')
        while re.match(res, tex):
            x += 1
            pre = re.sub(res, '\\1', tex)
            tex = re.sub(res, '\\2', tex)
            if pre == '':
                break
        #
        y = 0
        tex = pane.get('insert', 'end-1c')
        while re.match(res, tex):
            y += 1
            pre = re.sub(res, '\\1', tex)
            tex = re.sub(res, '\\2', tex)
            if pre == '':
                break
        return x, x + y

    def clear_search_and_replace(self):
        self.stb_sor1.delete('0', 'end')
        self.stb_sor2.delete('0', 'end')
        self.txt.tag_remove('search_tag', '1.0', 'end')
        self.sub.tag_remove('search_tag', '1.0', 'end')
        Makdo.search_word = ''

    def _highlight_search_word(self):
        word = Makdo.search_word
        for pane in (self.txt, self.sub):
            pane.tag_remove('search_tag', '1.0', 'end')
            tex = pane.get('1.0', 'end-1c')
            beg = 0
            res_word = word
            if not self.use_regexps.get():
                res_word = self._escape_search_word(word)
            res = '^((?:.|\n)*?)(' + res_word + ')((?:.|\n)*)$'
            try:
                re.match(res, tex)
            except BaseException:
                return
            while re.match(res, tex):
                pre = re.sub(res, '\\1', tex)
                wrd = re.sub(res, '\\2', tex)
                tex = re.sub(res, '\\3', tex)
                if wrd == '':
                    break
                beg += len(pre)
                end = beg + len(wrd)
                pane.tag_add('search_tag',
                             '1.0+' + str(beg) + 'c',
                             '1.0+' + str(end) + 'c',)
                beg = end

    def replace_backward_from_dialog(self, pane):
        t = '前検索又は置換'
        m = '前を検索する言葉と置換する言葉を入力してください．'
        word1, word2 = self.stb_sor1.get(), self.stb_sor2.get()
        hist1, hist2 = self.search_word_history, self.replace_word_history
        h1, t1 = '検索', ''
        h2, t2 = '置換', ''
        hist1.pop(-1)
        hist2.pop(-1)
        sd = TwoWordsDialog(pane, self, t, m, h1, h2, t1, t2,
                            word1, word2, hist1, hist2)
        word1, word2 = sd.get_value()
        if word1 is not None:
            if word1 == '':
                self.clear_search_and_replace()
            else:
                Makdo.search_word = word1
                self._highlight_search_word()
                self.stb_sor1.delete(0, 'end')
                self.stb_sor1.insert(0, word1)
                self.stb_sor2.delete(0, 'end')
                self.stb_sor2.insert(0, word2)
                self.search_or_replace_backward(True)  # must_replace = True

    def replace_forward_from_dialog(self, pane):
        t = '後検索又は置換'
        m = '後を検索する言葉と置換する言葉を入力してください．'
        word1, word2 = self.stb_sor1.get(), self.stb_sor2.get()
        hist1, hist2 = self.search_word_history, self.replace_word_history
        hist1.pop(-1)
        hist2.pop(-1)
        h1, t1 = '検索', ''
        h2, t2 = '置換', ''
        sd = TwoWordsDialog(pane, self, t, m, h1, h2, t1, t2,
                            word1, word2, hist1, hist2)
        word1, word2 = sd.get_value()
        if word1 is not None:
            if word1 == '':
                self.clear_search_and_replace()
            else:
                Makdo.search_word = word1
                self._highlight_search_word()
                self.stb_sor1.delete(0, 'end')
                self.stb_sor1.insert(0, word1)
                self.stb_sor2.delete(0, 'end')
                self.stb_sor2.insert(0, word2)
                self.search_or_replace_forward(True)   # must_replace = True

    def search_backward_from_dialog(self, pane):
        b = '前検索'
        m = '前を検索する言葉を入力してください．'
        h, t = '', ''
        word1 = self.stb_sor1.get()
        hist1 = self.search_word_history
        hist1.pop(-1)
        sd = OneWordDialog(pane, self, b, m, h, t, word1, hist1)
        word1 = sd.get_value()
        self.stb_sor2.delete(0, 'end')
        self.stb_sor2.insert(0, '')
        if word1 is not None:
            if word1 == '':
                self.clear_search_and_replace()
            else:
                Makdo.search_word = word1
                self._highlight_search_word()
                self.stb_sor1.delete(0, 'end')
                self.stb_sor1.insert(0, word1)
                self.search_or_replace_backward(False)  # must_replace = False

    def search_forward_from_dialog(self, pane):
        b = '後検索'
        m = '後を検索する言葉を入力してください．'
        h, t = '', ''
        word1 = self.stb_sor1.get()
        hist1 = self.search_word_history
        hist1.pop(-1)
        sd = OneWordDialog(pane, self, b, m, h, t, word1, hist1)
        word1 = sd.get_value()
        self.stb_sor2.delete(0, 'end')
        self.stb_sor2.insert(0, '')
        if word1 is not None:
            if word1 == '':
                self.clear_search_and_replace()
            else:
                Makdo.search_word = word1
                self._highlight_search_word()
                self.stb_sor1.delete(0, 'end')
                self.stb_sor1.insert(0, word1)
                self.search_or_replace_forward(False)   # must_replace = False

    ####################################
    # SHOW MESSAGE

    def show_font_help_message(self):
        if self.dont_show_help.get():
            return
        n = 'ご説明'
        m = 'Markdownで書かれた原稿は、\n' + \
            '文字幅が均一なフォントで表示すると、\n' + \
            '位置が揃って、読みやすくなります．\n' + \
            'また、\n' + \
            'ウェイトが複数用意されていると、\n' + \
            'ボールド（太字）にしたときも、\n' + \
            '文字幅が広がったりせず、\n' + \
            '位置が揃います．\n\n' + \
            'すなわち、\n' + \
            'ウェイトが複数ある等幅フォントが、\n' + \
            '最適です．\n\n' + \
            'この条件を満たすフォントの中で、\n' + \
            '①字体の読みやすさと、\n' + \
            '②無料で使えること（感謝！）から、\n' + \
            '"BIZ UDゴシック"をおすすめしています．\n\n' + \
            'https://fonts.google.com/specimen/BIZ+UDGothic'
        tkinter.messagebox.showinfo(n, m)

    def show_first_help_message(self):
        if self.dont_show_help.get():
            return
        n = 'ご説明'
        m = 'MS Word形式（拡張子docx）の\n' + \
            'ファイルを、この画面に\n' + \
            'ドラッグ＆ドロップしてみてください．\n' + \
            'Markdown形式に変換されて、\n' + \
            '画面に表示されます．\n\n' + \
            'その内容を編集して保存することで、\n' + \
            'MS Word形式（拡張子docx）の\n' + \
            'ファイルを編集できます．\n\n' + \
            '編集方法が分からない場合は、\n' + \
            'MS Wordで必要な編集したものを\n' + \
            'このアプリで開いてみて、\n' + \
            '編集前と見比べてください．'
        # mac doesn't support "tkinterdnd2" (drag and drop)
        m += \
            '\n\nただし、\n' + \
            'Macをお使いの方は\n' + \
            'ドラッグ＆ドロップが使えませんので、\n' + \
            '「ファイル」から「ファイルを開く」で、\n' + \
            'ファイルを開いてください．'
        tkinter.messagebox.showinfo(n, m)

    def show_folding_help_message(self):
        if self.dont_show_help.get():
            return
        if not self.must_show_folding_help_message:
            return
        n = 'ご説明'
        m = 'セクションを折り畳みます．' + \
            '（セクションの中身を一時的に文面の最後に移動させます）．\n\n' + \
            'そうすることで、' + \
            '文面の構造を視覚的に把握しやすくできます．\n\n' + \
            '他方で、' + \
            '一時的に文の順序が入れ替わってしまいますので、' + \
            'コメントや下線などの範囲を正しく把握できず、' + \
            '画面上の見た目が崩れる可能性があります．\n\n' + \
            'ファイルを保存する際には、' + \
            '全て展開した状態で保存されます．\n\n' + \
            '注）"...[n]"という記号は、' + \
            '折り畳んだことを記録したもので展開する際に必要ですので、' + \
            '絶対に書き替えたり消したりしないでください．'
        tkinter.messagebox.showinfo(n, m)
        self.must_show_folding_help_message = False

    def show_keyboard_macro_help_message(self):
        if self.dont_show_help.get():
            return
        if not self.must_show_keyboard_macro_help_message:
            return
        n = 'ご説明'
        m = 'キー入力の中から、繰り返しを探して、\n' + \
            'その繰り返しを実行します．\n\n' + \
            '同じ作業を何度も繰り返すときに、\n' + \
            '便利です．\n\n' + \
            '"Ctrl+E"でも実行できます．'
        tkinter.messagebox.showinfo(n, m)
        self.must_show_keyboard_macro_help_message = False

    def show_config_help_message(self):
        if self.dont_show_help.get():
            return
        if not self.must_show_config_help_message:
            return
        n = 'ご説明'
        m = '設定を次回以降に引き継ぐ場合は、\n' + \
            '「設定」の項目の「設定を保存」を\n' + \
            'クリックして、保存してください．'
        tkinter.messagebox.showinfo(n, m)
        self.must_show_config_help_message = False

    ####################################
    # RUN PERIODICALLY

    def run_periodically(self):
        self.footmarks = []
        self.goal_line_to_paint = 0
        self.local_line_to_paint = 0
        self.global_line_to_paint = 0
        self.save_auto_file(self.file_path)  # must execute immediately
        self.run_periodically = 0
        self.__run_periodically()

    ##########################
    # COMMAND

    def __run_periodically(self):
        # FOCUS
        try:
            focus = self.win.focus_get()
        except BaseException:
            focus = None
        # NEXT
        if focus is None:
            interval = 1_000  # 10ms
        else:
            interval = 20     # 20ms
        self.win.after(interval, self.__run_periodically)
        # NUMBER
        self.run_periodically += interval
        # if self.run_periodically >= 60_000:  # 1min
        #     self.run_periodically = 0
        # EXECUTE
        if focus == self.txt:  # if focus is not None:
            n = self.run_periodically
            # AUTO FILE
            if (n % 60_000) == 0:  # 1 / 60,000ms
                self.save_auto_file(self.file_path)
            # TABLE OF CONTENTS
            if (n % 5_000) == 0:   # 1 /  5,000ms
                self.update_toc()
            # MEMO PAD
            if (n % 1_000) == 0:   # 1 /  1,000ms
                self.update_memo_pad()
            # POSITION INFO
            if (n % 100) == 0:     # 1 /    100ms
                self.set_position_info_on_status_bar()
            # PAINT LINE LOCALLLY
            if True:               # 1 /     20ms
                self.run_periodically_to_paint_line_locally()
            # PAINT LINE GLOBALLY
            #    1,    2,    3,    4,    5,    6,    8,   10,   12,   15,
            #   16,   20,   24,   25,   30,   40,   48,   50,   60,   75,
            #   80,  100,  120,  125,  150,  200,  240,  250,  300,  375,
            #  400,  500,  600,  750, 1000, 1200, 1500, 2000, 3000, 6000
            m = len(self.file_lines)
            # if m <= 100:     # 60*2*1000/100  = 1200
            #     if (n % 1200) == 0:
            #         self.run_periodically_to_paint_line_globally()
            # elif m <= 200:   # 60*2*1000/200  =  600
            #     if (n % 600) == 0:
            #         self.run_periodically_to_paint_line_globally()
            # elif m <= 300:   # 60*2*1000/300  =  400
            if m <= 300:   # 60*2*1000/300  =  400
                if (n % 400) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 400:   # 60*2*1000/400  =  300
                if (n % 300) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 500:   # 60*2*1000/500  =  240
                if (n % 240) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 600:   # 60*2*1000/600  =  200
                if (n % 200) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 750:   # 60*2*1000/750  =  160
                if (n % 160) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 1000:  # 60*2*1000/1000 =  120
                if (n % 120) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 1200:  # 60*2*1000/1200 =  100
                if (n % 100) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 1500:  # 60*2*1000/1500 =   80
                if (n % 80) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 2000:  # 60*2*1000/2000 =   60
                if (n % 60) == 0:
                    self.run_periodically_to_paint_line_globally()
            elif m <= 3000:  # 60*2*1000/3000 =   40
                if (n % 40) == 0:
                    self.run_periodically_to_paint_line_globally()
            else:            # 60*2*1000/6000 =   20
                if (n % 30) == 0:
                    self.run_periodically_to_paint_line_globally()
        if focus == self.sub:  # if focus is not None:
            n = self.run_periodically
            if self.formula_number > 0 or \
               self.memo_pad_memory is not None:
                # CURSOR LINE
                vp = self._get_v_position_of_insert(self.sub)
                self.paint_out_line(vp - 1, self.sub)
                # PREVIOUS LINE
                c = self.sub.get('insert-1c', 'insert')
                if c == '\n':
                    self.paint_out_line(vp - 2, self.sub)
                # GLOBALLY
                self.run_periodically_to_paint_line_globally_on_sub_pane()

    # LOCAL PAINTING
    def run_periodically_to_paint_line_locally(self):
        # FOOTMARKS
        p_ind = self.txt.index('insert')
        p_lin = int(re.sub('\\.[0-9]+$', '', p_ind)) - 1
        self.footmarks.append(p_lin)
        # PAINT
        self.paint_out_line(self.local_line_to_paint)
        # NEXT
        self.local_line_to_paint += 1
        if self.local_line_to_paint > self.goal_line_to_paint:
            d_min = self.txt.index('@0,0')  # "x,y" not "y,x"
            d_max = self.txt.index('@1000000,1000000')  # "x,y" not "y,x"
            v_min = int(re.sub('\\.[0-9]+$', '', d_min)) - 1
            v_max = int(re.sub('\\.[0-9]+$', '', d_max)) - 1
            self.footmarks.append(v_min)
            self.footmarks.append(v_max)
            p_min = min(self.footmarks)
            p_max = max(self.footmarks)
            self.footmarks = []
            m = len(self.file_lines) - 1
            # MIN
            self.local_line_to_paint = p_min - 1
            if self.local_line_to_paint < 0:
                self.local_line_to_paint = 0
            if self.local_line_to_paint > m:
                self.local_line_to_paint = m
            # MAX
            self.goal_line_to_paint = p_max + 1
            if self.goal_line_to_paint < 0:
                self.goal_line_to_paint = 0
            if self.goal_line_to_paint > m:
                self.goal_line_to_paint = m
        # LINE AND EOF PAINTING
        ii = self.txt.index('insert lineend +1c')
        ei = self.txt.index('end lineend')
        self.txt.tag_remove('eol_tag', '1.0', 'end')
        self.txt.tag_remove('eof_tag', '1.0', 'end')
        self.txt.tag_remove('eol_and_eof_tag', '1.0', 'end')
        if ii != ei:
            # LINE PAINTING
            self.txt.tag_add('eol_tag',
                             'insert lineend', 'insert lineend +1c')
            # EOF PAINTING
            self.txt.tag_add('eof_tag',
                             'end-1c', 'end')
        else:
            # EOL PAINTING
            self.txt.tag_add('eol_and_eof_tag',
                             'insert lineend', 'insert lineend +1c')
        self.sub.tag_remove('eof_tag', '1.0', 'end')
        self.sub.tag_add('eof_tag', 'end-1c', 'end')

    # GLOBAL PAINTING
    def run_periodically_to_paint_line_globally(self):
        self.paint_out_line(self.global_line_to_paint)
        self.global_line_to_paint += 1
        if self.global_line_to_paint > len(self.file_lines) - 1:
            self.global_line_to_paint = 0

    # GLOBAL PAINTING ON SUB PANE
    def run_periodically_to_paint_line_globally_on_sub_pane(self):
        if 'global_line_to_paint_on_sub_pane' not in vars(self):
            self.global_line_to_paint_on_sub_pane = 0
            self.file_lines_on_sub_pane \
                = self.sub.get('1.0', 'end-1c').split('\n')
        self.paint_out_line(self.global_line_to_paint_on_sub_pane, self.sub)
        self.global_line_to_paint_on_sub_pane += 1
        if self.global_line_to_paint_on_sub_pane \
           > len(self.file_lines_on_sub_pane) - 1:
            self.global_line_to_paint_on_sub_pane = 0

    ####################################
    # NOT PYINSTALLER
    if not getattr(sys, 'frozen', False):

        # CALCULATE INTEREST OR CHARGE

        def _load_keiji(self):
            if 'keiji' not in vars(self):
                import makdo.keiji  # keiji
                self.keiji = makdo.keiji

        def insert_sample_trading_history(self) -> None:
            self._load_keiji()
            self._insert_line_break_as_necessary()
            ins = self.keiji.MANUAL_FOR_MAKDO + '\n\n' \
                + self.keiji.SAMPLE_DATA
            self.txt.insert('insert', ins)
            vp = self._get_v_position_of_insert(self.txt) - 1
            n = ins.count('\n') + 1
            for i in range(n):
                self.paint_out_line(vp - n + i + 1)
            self.txt.mark_set('insert', 'insert+1c')

        def calculate_interest_and_charge(self):
            self._load_keiji()
            # GET DATA
            upper_text = self.txt.get('1.0', 'insert')
            lower_text = self.txt.get('insert', 'end-1c')
            res = '^((?:.|\n)*\n\n)((?:.|\n)*)$'
            if re.match(res, upper_text):
                prev_par = re.sub(res, '\\1', upper_text)
            else:
                prev_par = ''
            res = '^((?:.|\n)*?\n)(\n(?:.|\n)*)$'
            lower_par = re.sub(res, '\\1', lower_text)
            beg = '1.0+' + str(len(prev_par)) + 'c'
            end = '1.0+' + str(len(upper_text + lower_par)) + 'c'
            par = self.txt.get(beg, end)
            # CALCULATE
            output = self.keiji.main(par)
            # WRITE
            self.txt.edit_separator()
            self.txt.insert(end, '\n' + output)
            self.txt.edit_separator()

        # EPWING

        mc = Minibuffer.MinibufferCommand(
            'lookup',
            [None, 'Epwing形式の辞書で調べる'],
            ['self.mother.look_in_epwing(self)',
             'self.set_return_to()'])
        Minibuffer.minibuffer_commands.append(mc)

        def _load_eblook(self):
            if 'eblook' not in vars(self):
                import makdo.eblook  # epwing
                self.eblook = makdo.eblook.Eblook()

        def look_in_epwing(self, pane=None) -> bool:
            if pane is None:
                pane = self.txt
            self._load_eblook()
            if 'epwing_directory' not in vars(self):
                self.set_epwing_directory()
            if 'epwing_directory' not in vars(self):
                return False
            if 'epwing_history' not in vars(self):
                self.epwing_history = []
            w = ''
            if self.txt.tag_ranges('sel'):
                w = self.txt.get('sel.first', 'sel.last')
            if 'akauni' in self.txt.mark_names():
                w = ''
                w += self.txt.get('akauni', 'insert')
                w += self.txt.get('insert', 'akauni')
            # LOOK IN DICTIONARY
            b = '辞書で調べる'
            p = '調べる言葉を入力してください．'
            h, t = '', ''
            e = self.epwing_history
            s = OneWordDialog(pane, self, b, p, h, t, w, e).get_value()
            if s is None:
                return
            msg = '辞書で検索しています'
            self.set_message_on_status_bar(msg, True)
            if self.epwing_directory is None:
                return
            self.eblook.set_dictionaries(self.epwing_directory)
            self.eblook.set_search_word(s)
            dic = ''
            if len(self.eblook.items) == 0:
                msg = '辞書に登録がありません'
                self.set_message_on_status_bar(msg)
                return
            msg = ''
            self.set_message_on_status_bar(msg, True)
            for ei in self.eblook.items:
                dic += '## 【' + ei.dictionary.k_name \
                    + '\u3000' + ei.title + '】\n\n'
                dic += ei.content + '\n\n'
            self._open_sub_pane(dic, True)
            # TITLE
            n = 0
            pos = dic
            res = '^((?:.|\n)*?)(## .*)(\n(?:.|\n)*)$'
            while re.match(res, pos):
                pre = re.sub(res, '\\1', pos)
                key = re.sub(res, '\\2', pos)
                pos = re.sub(res, '\\3', pos)
                beg = '1.0+' + str(n + len(pre)) + 'c'
                end = '1.0+' + str(n + len(pre) + len(key)) + 'c'
                n += len(pre) + len(key)
                self.sub.tag_add('c-40-1-g-x', beg, end)
            # ERROR
            n = 0
            pos = dic
            res = '^((?:.|\n)*?)(<gaiji=[^<>]+>)((?:.|\n)*)$'
            while re.match(res, pos):
                pre = re.sub(res, '\\1', pos)
                key = re.sub(res, '\\2', pos)
                pos = re.sub(res, '\\3', pos)
                beg = '1.0+' + str(n + len(pre)) + 'c'
                end = '1.0+' + str(n + len(pre) + len(key)) + 'c'
                n += len(pre) + len(key)
                self.sub.tag_add('error_tag', beg, end)
            #
            # self.sub.focus_force()
            # self.current_pane = 'sub'
            return True

        def set_epwing_directory(self):
            ed = ''
            if 'epwing_directory' in vars(self):
                ed = self.epwing_directory
            ti = 'Epwing形式の辞書フォルダの設定'
            ed = tkinter.filedialog.askdirectory(title=ti, initialdir=ed)
            if ed == () or ed == '':
                return False
            self.epwing_directory = ed
            self.show_config_help_message()
            return True

        # OPENAI

        mc = Minibuffer.MinibufferCommand(
            'ask-openai',
            [None, 'OpenAIに質問する'],
            ['self.mother.open_openai()',
             'self.set_return_to()'])
        Minibuffer.minibuffer_commands.append(mc)

        def open_openai(self) -> bool:
            # CONFIGURATION
            if 'openai_model' not in vars(self):
                self.set_openai_model()
            if 'openai_model' not in vars(self):
                n = 'エラー'
                m = 'OpenAIのモデルが設定されていません．'
                tkinter.messagebox.showerror(n, m)
                return False
            if 'openai_key' not in vars(self):
                self.openai_key()
            if 'openai_key' not in vars(self):
                n = 'エラー'
                m = 'OpenAIのキーが設定されていません．'
                tkinter.messagebox.showerror(n, m)
                return False
            # LOAD MODULE
            if 'openai' not in vars(self):
                self.set_message_on_status_bar('openaiを起動しています', True)
                try:
                    import openai  # Apache Software License
                except ImportError:
                    n = 'エラー'
                    m = '"openai"を\n' \
                        + 'インポートできませんでした．\n\n' \
                        + '次のコマンドを実行して、\n' \
                        + 'インストールしてください．\n\n' \
                        + 'pip install openai'
                    tkinter.messagebox.showerror(n, m)
                    return False
                self.openai = openai
            om = self.openai_model
            m = 'モデルは"' + om + '"が設定されています'
            self.set_message_on_status_bar(m)
            # PROMPT
            if 'openai_qanda' not in vars(self):
                n = MD_TEXT_WIDTH - get_real_width('## 【OpenAIにＸＸ】')
                self.openai_qanda \
                    = '- 外部処理ですので、個人情報の流出に注意してください。\n' \
                    + '- 有料ですので、料金に注意してください。\n\n' \
                    + '## 【OpenAIの設定】' + ('-' * n) + '\n\n' \
                    + 'あなたは誠実で優秀な日本人のアシスタントです。\n' \
                    + '特に指示が無い場合は、常に日本語で回答してください。\n\n' \
                    + '## 【OpenAIに質問】' + ('-' * n) + '\n\n'
            self.txt.focus_force()
            self._execute_sub_pane = self.ask_openai
            self._close_sub_pane = self.close_openai
            self._open_sub_pane(self.openai_qanda, False, '質問')
            self.sub.mark_set('insert', 'end-1c')
            self._paint_geneai_lines('OpenAI')
            self.sub.edit_separator()
            return True

        def ask_openai(self) -> None:
            messages = self._get_message('OpenAI')
            self.set_message_on_status_bar('OpenAIに質問しています', True)
            ok = Witch.dechant(self.openai_key)
            output = self.openai.OpenAI(api_key=ok).chat.completions.create(
                model=self.openai_model,
                n=1, max_tokens=1000,
                messages=messages,
            )
            self.set_message_on_status_bar('', True)
            answer = output.choices[0].message.content
            answer = adjust_line(answer)
            self._write_answer('OpenAI', answer)
            self.openai_qanda = self.sub.get('1.0', 'end-1c')

        def close_openai(self) -> None:
            del self._execute_sub_pane
            del self._close_sub_pane
            # file_path = CONFIG_DIR + '/' + 'openai.md'
            # contents = self.sub.get('1.0', 'end-1c')
            # self._save_config_file(file_path, contents)
            self.set_message_on_status_bar('')
            self._close_sub_pane()

        def set_openai_model(self) -> bool:
            b = 'OpenAIのモデル'
            m = 'OpenAIのモデルを入力してください．'
            h, t = '', ''
            if 'openai_model' not in vars(self):
                self.openai_model = DEFAULT_OPENAI_MODEL
            om = self.openai_model
            ca = []
            for c in OPENAI_MODELS:
                if c != om:
                    ca.append(c)
            om = OneWordDialog(self.txt, self, b, m, h, t, om, ca).get_value()
            if om is None:
                return False
            self.openai_model = om
            self.show_config_help_message()
            return True

        def set_openai_key(self) -> bool:
            t = 'OpenAIのキー'
            m = 'OpenAIのキーを入力してください．'
            ok = PasswordDialog(self.txt, self, t, m).get_value()
            if ok is None:
                return False
            self.openai_key = Witch.enchant(ok)
            self.show_config_help_message()
            return True

        # LLAMA

        # RAG (Retrieval-Augmented Generation)
        embeded_model = 'intfloat/multilingual-e5-large'
        # embeded_model = 'all-MiniLM-L6-v2'
        llama_rag_file = CONFIG_DIR + '/rag.md'
        is_editing_llama_rag_data = False

        mc = Minibuffer.MinibufferCommand(
            'ask-llama-without-rag',
            [None, 'LlamaにRAGなし質問する'],
            ['self.mother.open_llama_without_rag()',
             'self.set_return_to()'])
        Minibuffer.minibuffer_commands.append(mc)

        mc = Minibuffer.MinibufferCommand(
            'ask-llama-with-rag',
            [None, 'LlamaにRAGありで質問する'],
            ['self.mother.open_llama_with_rag()',
             'self.set_return_to()'])
        Minibuffer.minibuffer_commands.append(mc)

        mc = Minibuffer.MinibufferCommand(
            'edit-llama-rag-data',
            [None, 'RAG用のデータを編集する'],
            ['self.mother.edit_llama_rag_data()',
             'self.set_return_to()'])
        Minibuffer.minibuffer_commands.append(mc)

        def open_llama_without_rag(self) -> bool:
            # CONFIGURATION
            if 'llama_model_file' not in vars(self):
                self.set_llama_model_file()
            if 'llama_model_file' not in vars(self):
                n = 'エラー'
                m = 'Llamaのモデルファイルが設定されていません．'
                tkinter.messagebox.showerror(n, m)
                return False
            if 'llama_gpu_layers' not in vars(self):
                self.llama_gpu_layers = 0
            if 'llama_context_size' not in vars(self):
                self.llama_context_size = 512
            # LOAD MODULE
            if 'llama_with_rag' in vars(self):
                del self.llama_with_rag
            if 'llama_without_rag' not in vars(self):
                self.set_message_on_status_bar('LlamaをRAGなしで起動しています', True)
                try:
                    from llama_cpp import Llama  # pip install llama-cpp-python
                except ImportError:
                    n = 'エラー'
                    m = '"Llama"を\n' \
                        + 'インポートできませんでした．\n\n' \
                        + '次のコマンドを実行して、\n' \
                        + 'インストールしてください．\n\n' \
                        + 'pip install llama_cpp_python'
                    tkinter.messagebox.showerror(n, m)
                    return False
                self.llama_without_rag = Llama(
                    model_path=self.llama_model_file,
                    n_gpu_layers=self.llama_gpu_layers,
                    n_ctx=self.llama_context_size,
                )
            mf = os.path.basename(self.llama_model_file)
            m = 'モデルファイルは"' + mf + '"が設定されています'
            self.set_message_on_status_bar(m)
            # PROMPT
            if 'llama_qanda' not in vars(self):
                n = MD_TEXT_WIDTH - get_real_width('## 【LlamaにＸＸ】')
                self.llama_qanda \
                    = '- 内部処理ですので、情報を外部に送信しません。\n' \
                    + '- 無料ですので、料金は発生しません。\n\n' \
                    + '## 【Llamaの設定】' + ('-' * n) + '\n\n' \
                    + 'あなたは誠実で優秀な日本人のアシスタントです。\n' \
                    + '特に指示が無い場合は、常に日本語で回答してください。\n\n' \
                    + '## 【Llamaに質問】' + ('-' * n) + '\n\n'
            self.txt.focus_force()
            self._execute_sub_pane = self.ask_llama_without_rag
            self._close_sub_pane = self.close_llama_without_rag
            self._open_sub_pane(self.llama_qanda, False, '質問')
            self.sub.mark_set('insert', 'end-1c')
            self._paint_geneai_lines('Llama')
            self.sub.edit_separator()
            return True

        def open_llama_with_rag(self) -> bool:
            # CONFIGURATION
            if 'llama_model_file' not in vars(self):
                self.set_llama_model_file()
            if 'llama_model_file' not in vars(self):
                n = 'エラー'
                m = 'Llamaのモデルファイルが設定されていません．'
                tkinter.messagebox.showerror(n, m)
                return False
            # LOAD MODULE
            if 'llama_without_rag' in vars(self):
                del self.llama_without_rag
            if 'llama_with_rag' not in vars(self):
                self.set_message_on_status_bar('LlamaをRAGありで起動しています', True)
                try:
                    # pip install numpy
                    # pip install llama-index-llms-llama-cpp
                    from llama_index.llms.llama_cpp import LlamaCPP
                    from llama_index.core \
                        import SimpleDirectoryReader, VectorStoreIndex
                    # pip install llama-index-embeddings-huggingface
                    from llama_index.embeddings.huggingface \
                        import HuggingFaceEmbedding
                except ImportError:
                    n = 'エラー'
                    m = '"Llama"を\n' \
                        + 'インポートできませんでした．\n\n' \
                        + '次のコマンドを実行して、\n' \
                        + 'インストールしてください．\n\n' \
                        + 'pip install llama_cpp_python'
                    tkinter.messagebox.showerror(n, m)
                    return False
                if 'llama_cpp' not in vars(self):
                    self.llama_cpp = LlamaCPP(
                        model_path=self.llama_model_file,
                        model_kwargs={
                            'n_gpu_layers': self.llama_gpu_layers,
                            'n_ctx': self.llama_context_size,
                        }
                    )
                if 'llama_embed_model' not in vars(self):
                    self.llama_embed_model = HuggingFaceEmbedding(
                        model_name=self.embeded_model)
                if 'llama_reader' not in vars(self):
                    self.llama_reader = SimpleDirectoryReader(
                        input_files=[self.llama_rag_file])
                llama_rag_data = self.llama_reader.load_data()
                index = VectorStoreIndex.from_documents(
                    llama_rag_data, embed_model=self.llama_embed_model)
                self.llama_with_rag \
                    = index.as_query_engine(llm=self.llama_cpp,
                                            streaming=False,
                                            similarity_top_k=3)
            mf = os.path.basename(self.llama_model_file)
            m = 'モデルファイルは"' + mf + '"が設定されています'
            self.set_message_on_status_bar(m)
            # PROMPT
            if 'llama_qanda' not in vars(self):
                n = MD_TEXT_WIDTH - get_real_width('## 【LlamaにＸＸ】')
                self.llama_qanda \
                    = '- 内部処理ですので、情報を外部に送信しません。\n' \
                    + '- 無料ですので、料金は発生しません。\n\n' \
                    + '## 【Llamaの設定】' + ('-' * n) + '\n\n' \
                    + 'あなたは誠実で優秀な日本人のアシスタントです。\n' \
                    + '特に指示が無い場合は、常に日本語で回答してください。\n\n' \
                    + '## 【Llamaに質問】' + ('-' * n) + '\n\n'
            self.txt.focus_force()
            self._execute_sub_pane = self.ask_llama_with_rag
            self._close_sub_pane = self.close_llama_with_rag
            self._open_sub_pane(self.llama_qanda, False, '質問')
            self.sub.mark_set('insert', 'end-1c')
            self._paint_geneai_lines('Llama')
            self.sub.edit_separator()
            return True

        def ask_llama_without_rag(self) -> None:
            messages = self._get_message('Llama')
            self.set_message_on_status_bar('LlamaにRAGなしで質問しています', True)
            output = self.llama_without_rag.create_chat_completion(
                # temperature=0.8,
                # frequency_penalty=1.000000001,
                # presence_penalty=1.1,
                # repeat_penalty=1.1,
                # top_k=40,
                # top_p=0.5,
                # num_beams=1,
                messages=messages)
            self.set_message_on_status_bar('', True)
            answer = output['choices'][0]['message']['content']
            answer = adjust_line(answer)
            self._write_answer('Llama', answer)
            self.llama_qanda = self.sub.get('1.0', 'end-1c')

        def ask_llama_with_rag(self) -> None:
            messages = self._get_message('Llama')
            self.set_message_on_status_bar('LlamaにRAGありで質問しています', True)
            q = ''
            for m in messages:
                if m['role'] == 'user':
                    q = m['content']
            output = self.llama_with_rag.query(q)
            self.set_message_on_status_bar('', True)
            answer = output.response
            answer = re.sub('^\\s+', '', answer)
            answer = re.sub('\\s+$', '', answer)
            answer = re.sub('\\\\n', '\n', answer)
            answer = adjust_line(answer)
            self._write_answer('Llama', answer)
            self.llama_qanda = self.sub.get('1.0', 'end-1c')

        def close_llama_without_rag(self) -> None:
            del self._execute_sub_pane
            del self._close_sub_pane
            self.llama_qanda = self.sub.get('1.0', 'end-1c')
            self.set_message_on_status_bar('')
            self._close_sub_pane()

        def close_llama_with_rag(self) -> None:
            del self._execute_sub_pane
            del self._close_sub_pane
            del self.llama_with_rag
            del self.llama_qanda
            self.set_message_on_status_bar('')
            self._close_sub_pane()

        def set_llama_model_file(self) -> bool:
            mf, md = '', ''
            if 'llama_model_file' in vars(self):
                mf = self.llama_model_file
                md = os.path.dirname(mf)
            ti = 'Llamaのモデルファイルを設定'
            lmf = tkinter.filedialog.askopenfilename(
                title=ti, initialdir=md, initialfile=mf)
            if lmf == () or lmf == '':
                return False
            self.llama_model_file = lmf
            # CLEAR
            if 'llama' in vars(self):
                del self.llama
            self.show_config_help_message()
            return True

        def set_llama_gpu_layers(self) -> bool:
            default_size: int = 0
            b = 'GPUで処理するレイヤーの数'
            p = 'GPUで処理するレイヤーの数を整数で入力してください．\n' \
                + '（GPUなし:0、全て:-1）'
            h, t = '', ''
            if 'llama_gpu_layers' not in vars(self):
                self.llama_gpu_layers = default_size
            gl = str(self.llama_gpu_layers)
            while True:
                gl = OneWordDialog(self.txt, self, b, p, h, t, gl).get_value()
                if gl is None:
                    return False
                if re.match('^(-1|[0-9]+)$', gl):
                    break
            self.llama_gpu_layers = int(gl)
            # CLEAR
            if 'llama' in vars(self):
                del self.llama
            self.show_config_help_message()
            return True

        def set_llama_context_size(self) -> bool:
            default_size: int = 512
            b = 'コンテクストサイズ'
            p = 'コンテクストサイズを整数で入力してください．\n' \
                + '（初期値:512、推奨値:2048）'
            h, t = '', ''
            if 'llama_context_size' not in vars(self):
                self.llama_context_size = default_size
            cs = str(self.llama_context_size)
            while True:
                cs = OneWordDialog(self.txt, self, b, p, h, t, cs).get_value()
                if cs is None:
                    return False
                if re.match('^[0-9]+$', cs):
                    break
            self.llama_context_size = int(cs)
            # CLEAR
            if 'llama' in vars(self):
                del self.llama
            self.show_config_help_message()
            return True

        def edit_llama_rag_data(self) -> bool:
            try:
                with open(self.llama_rag_file, 'r') as f:
                    self.llama_rag_data = f.read()
            except BaseException:
                return False
            self._open_sub_pane(self.llama_rag_data, False)
            self.is_editing_llama_rag_data = True
            return True

        def quit_editing_llama_rag_data(self) -> bool:
            self.is_editing_llama_rag_data = False
            llama_rag_data = self.sub.get('1.0', 'end-1c')
            self._save_config_file(self.llama_rag_file, llama_rag_data)
            return True

        # OLLAMA

        mc = Minibuffer.MinibufferCommand(
            'ask-ollama',
            [None, 'Ollamaに質問する'],
            ['self.mother.open_ollama()',
             'self.set_return_to()'])
        Minibuffer.minibuffer_commands.append(mc)

        def open_ollama(self) -> bool:
            if 'ollama_model' not in vars(self):
                self.set_ollama_model()
            if 'ollama_model' not in vars(self):
                n, m = 'エラー', 'oLlamaのモデルが設定されていません．'
                tkinter.messagebox.showerror(n, m)
                return False
            self.set_message_on_status_bar('Ollamaを起動しています', True)
            m = 'モデルは"' + self.ollama_model + '"が設定されています'
            self.set_message_on_status_bar(m, True)
            # PROMPT
            if 'ollama_qanda' not in vars(self):
                n = MD_TEXT_WIDTH - get_real_width('## 【OllamaにＸＸ】')
                self.ollama_qanda \
                    = '- 内部処理ですので、情報を外部に送信しません。\n' \
                    + '- 無料ですので、料金は発生しません。\n\n' \
                    + '## 【Ollamaの設定】' + ('-' * n) + '\n\n' \
                    + 'あなたは誠実で優秀な日本人のアシスタントです。\n' \
                    + '特に指示が無い場合は、常に日本語で回答してください。\n\n' \
                    + '## 【Ollamaに質問】' + ('-' * n) + '\n\n'
            self.txt.focus_force()
            self._execute_sub_pane = self.ask_ollama
            self._close_sub_pane = self.close_ollama
            self._open_sub_pane(self.ollama_qanda, False, '質問')
            self.sub.mark_set('insert', 'end-1c')
            self._paint_geneai_lines('Ollama')
            self.sub.edit_separator()
            return True

        def ask_ollama(self):
            thread_1 = threading.Thread(target=self._ask_ollama)
            thread_1.start()
            return True

        def _ask_ollama(self):
            model = self.ollama_model
            messages = self._get_message('Ollama')
            self.set_message_on_status_bar('Ollamaに質問しています', True)
            if 'ollama' not in sys.modules:
                try:
                    import ollama
                except ImportError:
                    n = 'エラー'
                    m = '"ollama"を\n' \
                        + 'インポートできませんでした．\n\n' \
                        + '次のコマンドを実行して、\n' \
                        + 'インストールしてください．\n\n' \
                        + 'pip install ollama'
                    tkinter.messagebox.showerror(n, m)
                    return False
                self.ollama = ollama
            try:
                response = self.ollama.chat(
                    model=model, messages=messages,
                    think=False,  # for reasoning model
                    # options={ "temperature": 0, "num_ctx": 512 }
                )
            except BaseException:
                n = 'エラー'
                m = '"ollama"に\n' \
                    + '接続できませんでした．\n\n' \
                    + '次のコマンドを実行して、\n' \
                    + '起動しておいてください．\n' \
                    + 'ollama serve'
                tkinter.messagebox.showerror(n, m)
                return False
            answer = response.message.content
            answer = adjust_line(answer)
            self.set_message_on_status_bar('', True)
            self._write_answer('Ollama', answer)
            self.ollama_qanda = self.sub.get('1.0', 'end-1c')

        def close_ollama(self) -> None:
            del self._execute_sub_pane
            del self._close_sub_pane
            self.ollama_qanda = self.sub.get('1.0', 'end-1c')
            self.set_message_on_status_bar('')
            self._close_sub_pane()

        def set_ollama_model(self) -> bool:
            if 'ollama_model' not in vars(self):
                self.ollama_model = DEFAULT_OLLAMA_MODEL
            b, p = 'Ollamaのモデル', 'Ollamaのモデルを入力してください．'
            h, t, i = '', '', self.ollama_model
            om = OneWordDialog(self.txt, self, b, p, h, t, i).get_value()
            if om is None:
                return False
            self.ollama_model = om
            self.show_config_help_message()
            return True

        # TOOLS

        def _get_geneai_head(self, geneai):
            n = MD_TEXT_WIDTH - get_real_width('## 【' + geneai + 'にＸＸ】')
            cnf_head = '## 【' + geneai + 'の設定】' + ('-' * n)
            que_head = '## 【' + geneai + 'に質問】' + ('-' * n)
            ans_head = '## 【' + geneai + 'の回答】' + ('-' * n)
            return cnf_head, que_head, ans_head

        def _get_message(self, geneai):
            cnf_head, que_head, ans_head = self._get_geneai_head(geneai)
            messages = []
            role, mc = '', ''
            doc = self.sub.get('1.0', 'end-1c') + '\n\n' + ans_head
            for line in doc.split('\n'):
                if line == cnf_head or \
                   line == que_head or \
                   line == ans_head:
                    if role != '' and mc != '':
                        mc = re.sub('^\n+', '', mc)
                        mc = re.sub('\n+$', '', mc)
                        messages.append({'role': role, 'content': mc})
                    mc = ''
                if line == cnf_head:
                    role = 'system'
                elif line == que_head:
                    role = 'user'
                elif line == ans_head:
                    role = 'assistant'
                else:
                    mc += line + '\n'
            return messages

        def _write_answer(self, geneai, answer):
            if answer == '':
                return -1
            cnf_head, que_head, ans_head = self._get_geneai_head(geneai)
            self.sub['autoseparators'] = False
            self.sub.edit_separator()
            doc = self.sub.get('1.0', 'end-1c')
            if not re.match('^(.|\n)*\n$', doc):
                self.sub.insert('end', '\n')
            if not re.match('^(.|\n)*\n\n$', doc):
                self.sub.insert('end', '\n')
            self.sub.insert('end', ans_head + '\n\n')
            self.sub.insert('end', answer + '\n\n')
            self.sub.edit_separator()
            self.sub.insert('end', que_head + '\n\n')
            self.sub.mark_set('insert', 'end-1c')
            self._put_back_cursor_to_pane(self.sub)
            self._paint_geneai_lines(geneai)
            self.sub['autoseparators'] = True
            self.sub.edit_separator()

        def _paint_geneai_lines(self, geneai):
            for tag in self.sub.tag_names():
                self.sub.tag_remove(tag, '1.0', 'end-1c')
            n = 0
            pos = self.sub.get('1.0', 'end-1c')
            res = '^((?:.|\n)*?)' + \
                '(## 【' + geneai + '...】-+)' + \
                '(\n(?:.|\n)*)$'
            while re.match(res, pos):
                pre = re.sub(res, '\\1', pos)
                key = re.sub(res, '\\2', pos)
                pos = re.sub(res, '\\3', pos)
                beg = '1.0+' + str(n + len(pre)) + 'c'
                end = '1.0+' + str(n + len(pre) + len(key)) + 'c'
                n += len(pre) + len(key)
                self.sub.tag_add('c-40-1-g-x', beg, end)


######################################################################
# MAIN


if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description='Markdownファイルを編集します',
        add_help=False)
    parser.add_argument(
        '-h', '--help',
        action='help',
        help='ヘルプメッセージを表示します')
    parser.add_argument(
        '-v', '--version',
        action='version',
        version=('%(prog)s ' + __version__),
        help='バージョン番号を表示します')
    parser.add_argument(
        '-H', '--dont-show-help',
        action='store_true',
        help='ヘルプを表示します')
    parser.add_argument(
        '-c', '--background-color',
        type=str,
        choices=['W', 'B', 'G'],
        help='背景の色（白、黒、緑）を設定します')
    parser.add_argument(
        '-s', '--font-size',
        type=int,
        choices=[12, 15, 18, 21, 24, 27, 30, 33, 36, 42, 48],
        help='文字の大きさをピクセル単位で設定します')
    parser.add_argument(
        '-R', '--use-regexps',
        action='store_true',
        help='検索・置換に正規表現を使う')
    parser.add_argument(
        '-p', '--paint-keywords',
        action='store_true',
        help='キーワードに色を付けます')
    parser.add_argument(
        '-k', '--keywords-to-paint',
        type=str,
        help='色付けするキーワードを設定します')
    parser.add_argument(
        '-d', '--digit-separator',
        type=str,
        choices=['3', '4'],
        help='計算結果の区切りを設定します')
    parser.add_argument(
        '-r', '--read-only',
        action='store_true',
        help='読み取り専用で開きます')
    parser.add_argument(
        '-b', '--make-backup-file',
        action='store_true',
        help='バックアップファイルを残します')
    parser.add_argument(
        'input_file',
        nargs='?',
        help='Markdownファイル or MS Wordファイル')
    args = parser.parse_args()

    Makdo.args_dont_show_help = args.dont_show_help
    Makdo.args_background_color = args.background_color
    Makdo.args_font_size = args.font_size
    Makdo.args_paint_keywords = args.paint_keywords
    Makdo.args_keywords_to_paint = args.keywords_to_paint
    Makdo.args_digit_separator = args.digit_separator
    Makdo.args_read_only = args.read_only
    Makdo.args_make_backup_file = args.make_backup_file
    Makdo.args_input_file = args.input_file

    Makdo()
